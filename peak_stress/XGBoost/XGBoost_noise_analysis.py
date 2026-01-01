#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XGBoost model noise analysis and uncertainty analysis

Function:
1. Load already trained XGBoost model
2. Noise robustness analysis: Test model robustness to Gaussian noise and outliers
3. Prediction interval analysis: Generate confidence intervals using Bootstrap and Quantile Regression methods

Note: This script needs to run XGBoost.py first to train the model, then load the saved model for analysis
"""

import os
import sys
import numpy as np
import pandas as pd
# Set matplotlib backend to Agg (non-interactive, avoid GUI related errors）
import matplotlib
matplotlib.use('Agg')  # Must be set before importing pyplot
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score, KFold, train_test_split
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
import math
from scipy import stats
import xgboost as xgb
import optuna
from optuna.visualization import plot_optimization_history, plot_param_importances
import shap
from sklearn.inspection import partial_dependence
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import interp1d, UnivariateSpline, RegularGridInterpolator, griddata
import warnings
import joblib
warnings.filterwarnings('ignore')

# Set Chinese font and LaTeX rendering
def configure_plot_fonts(fonts=None):
    """Uniformly set Matplotlib font, ensure negative sign is displayed correctly"""
    if fonts is None:
        fonts = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
    plt.rcParams['font.sans-serif'] = fonts
    plt.rcParams['axes.unicode_minus'] = False

configure_plot_fonts()
plt.rcParams['mathtext.default'] = 'regular'  # Use regular font to render mathematical symbols

# Work directory setting - automatically find project root directory
def find_project_root():
    """Find project root directory"""
    # 先尝试从脚本位置定位
    try:
        script_path = os.path.abspath(__file__)
        script_dir = os.path.dirname(script_path)
        # If script is in subdirectory, find directory containing dataset or SAVE
        current_dir = script_dir
        search_depth = 0
        while current_dir != os.path.dirname(current_dir) and search_depth < 10:
            if os.path.exists(os.path.join(current_dir, 'dataset')) or \
               os.path.exists(os.path.join(current_dir, 'SAVE')):
                return current_dir
            current_dir = os.path.dirname(current_dir)
            search_depth += 1
    except NameError:
        # Run in Jupyter, search from current directory upwards
        pass
    
    # Search from current working directory upwards
    current_dir = os.path.abspath(os.getcwd())
    search_limit = 0
    while current_dir != os.path.dirname(current_dir) and search_limit < 10:
        # Check if contains dataset folder or specific project file
        if os.path.exists(os.path.join(current_dir, 'dataset')) or \
           os.path.exists(os.path.join(current_dir, 'SAVE')):
            return current_dir
        current_dir = os.path.dirname(current_dir)
        search_limit += 1
    
    # If still not found, return current directory itself
    return os.path.abspath(os.getcwd())

# Set work directory
PROJECT_ROOT = find_project_root()
if not os.path.exists(os.path.join(PROJECT_ROOT, 'dataset')):
    # If still not found, use current directory
    PROJECT_ROOT = os.getcwd()

os.chdir(PROJECT_ROOT)
print(f"Work directory set to: {PROJECT_ROOT}")

# Save directory root path
SAVE_ROOT = os.path.join(PROJECT_ROOT, 'peak_stress', 'XGBoost', 'save')
os.makedirs(SAVE_ROOT, exist_ok=True)

def mean_absolute_percentage_error(y_true, y_pred):
    """Calculate mean absolute percentage error (MAPE)
    
    Parameters:
        y_true: True value array
        y_pred: Predicted value array
    
    Returns:
        MAPE value (percentage)

    Note:
        When true value is 0, skip this sample to avoid division by zero error
    """
    y_true = np.array(y_true).flatten()
    y_pred = np.array(y_pred).flatten()
    
    # Filter out samples with true value 0 (avoid division by zero）
    mask = np.abs(y_true) > 1e-10
    if np.sum(mask) == 0:
        return np.nan  # If all true values are 0, return NaN
    
    y_true_filtered = y_true[mask]
    y_pred_filtered = y_pred[mask]
    
    # Calculate MAPE: (1/n) * Σ|y_true - y_pred| / |y_true| * 100
    mape = np.mean(np.abs((y_true_filtered - y_pred_filtered) / y_true_filtered)) * 100
    
    return mape

def load_data():
    """Load data - use 10 material parameters and 5 specimen parameters to regress peak stress"""
    print("=== Load data ===")
    
    # Use specified data file
    data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
    
    if not os.path.exists(data_file):
        print(f"Error: Data file not found: {data_file}")
        return None
    
    print(f"Found data file: {data_file}")
    
    # Read first worksheet (material parameter table) of Excel file
    df = pd.read_excel(data_file, sheet_name=0)
    print(f"Data shape: {df.shape}")
    print(f"Column names: {list(df.columns)}")
    
    # Define features: 10 material parameters + 5 specimen parameters
    # 10 material features
    material_features = ['water', 'cement', 'w/c', 'CS', 'sand', 'CA', 'r', 'WA', 'S', 'CI']
    # 5 specimen parameters
    specimen_features = ['age', 'μe', 'DJB', 'side', 'GJB']
    
    feature_names = material_features + specimen_features
    
    # Target variable: peak stress
    target_column = 'fc'
    
    # Check column names
    missing_cols = [col for col in feature_names if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns {missing_cols}")
        feature_names = [col for col in feature_names if col in df.columns]
    
    missing_target = target_column not in df.columns
    if missing_target:
        print(f"Error: Missing target variable column '{target_column}'")
        print(f"Available columns: {list(df.columns)}")
        return None
    
    # Extract data
    X = df[feature_names].values
    y = df[target_column].values
    
    # Special check w/c feature (if exists）
    if 'w/c' in feature_names:
        wc_idx = feature_names.index('w/c')
        wc_values = X[:, wc_idx]
        print(f"\n🔍 Check w/c feature (original data):")
        print(f"  Unique value number: {len(np.unique(wc_values))}")
        print(f"  Minimum value: {np.min(wc_values):.10f}")
        print(f"  Maximum value: {np.max(wc_values):.10f}")
        print(f"  Range: {np.max(wc_values) - np.min(wc_values):.10f}")
        print(f"  Mean: {np.mean(wc_values):.10f}")
        print(f"  Standard deviation: {np.std(wc_values):.10f}")
        if len(np.unique(wc_values)) <= 5:
            print(f"  All unique values: {np.unique(wc_values)}")
        else:
            print(f"  First 5 unique values: {np.unique(wc_values)[:5]}")
        if np.max(wc_values) - np.min(wc_values) < 1e-6:
            print(f"  ⚠ Warning: w/c feature has a range close to 0 in original data!")
    
    # Check if there are missing values
    if np.isnan(X).any():
        print(f"Warning: Feature data contains NaN values, will be filled")
        from sklearn.impute import SimpleImputer
        imputer = SimpleImputer(strategy='mean')
        X = imputer.fit_transform(X)
    
    if np.isnan(y).any():
        print(f"Warning: Target variable contains NaN values, will be removed")
        valid_mask = ~np.isnan(y)
        X = X[valid_mask]
        y = y[valid_mask]
        df = df.iloc[valid_mask].reset_index(drop=True)
    
    # Extract sample division information (using DataSlice column）
    sample_divisions = []
    if 'DataSlice' in df.columns:
        sample_divisions = df['DataSlice'].values
        print(f"Found data set division information: {np.unique(sample_divisions, return_counts=True)}")
    else:
        print("DataSlice column not found, will use random division")
        sample_divisions = None
    
    # Extract sample IDs
    sample_ids = []
    if 'No_Customized' in df.columns:
        sample_ids = df['No_Customized'].values
    else:
        sample_ids = [f"sample_{i}" for i in range(len(X))]
    
    print(f"Feature number: {X.shape[1]} (10 material parameters + 5 specimen parameters)")
    print(f"Sample number: {X.shape[0]}")
    print(f"Target variable '{target_column}' range: {np.min(y):.2f} - {np.max(y):.2f}")
    
    # Save original DataFrame indices (for subsequent mapping prediction results）
    original_df_indices = df.index.values
    
    return X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df

def split_data_by_divisions(X, y, sample_divisions, sample_ids, test_ratio=0.2, val_ratio=0.2, random_state=42):
    """Split data based on sample division information"""
    print("=== Split data based on sample division information ===")
    
    if sample_divisions is None:
        from sklearn.model_selection import train_test_split
        X_train, X_temp, y_train, y_temp = train_test_split(
            X, y, test_size=test_ratio + val_ratio, random_state=random_state
        )
        val_ratio_adjusted = val_ratio / (test_ratio + val_ratio)
        X_val, X_test, y_val, y_test = train_test_split(
            X_temp, y_temp, test_size=1-val_ratio_adjusted, random_state=random_state
        )
        
        train_ids = [f"train_{i}" for i in range(len(X_train))]
        val_ids = [f"val_{i}" for i in range(len(X_val))]
        test_ids = [f"test_{i}" for i in range(len(X_test))]
        
        return X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids
    
    # Classify samples based on sample division information (support DataSlice format: subset1, subset2, subset3, test）
    train_indices = []
    val_indices = []
    test_indices = []
    
    for i, division in enumerate(sample_divisions):
        division_str = str(division).strip()
        if division_str.lower().startswith('test'):
            test_indices.append(i)
        elif division_str.startswith('subset'):
            # subset1, subset2, subset3 are all used as training set (or can be allocated according to needs）
            # Here all are used as training set by default, if validation set is needed, it can be divided from subset
            train_indices.append(i)
        elif division == 'train' or division == 'train' or division_str.startswith('train'):
            train_indices.append(i)
        elif division == 'val' or division == 'val' or division == 'validation' or division_str.startswith('val'):
            val_indices.append(i)
        else:
            # Default to training set
            train_indices.append(i)
    
    # If there is no validation set but there is a training set, divide validation set from training set
    # If test set exists, keep it unchanged
    if train_indices and not val_indices:
        print("Only training set samples, divide validation set from training set...")
        from sklearn.model_selection import train_test_split
        
        train_indices = np.array(train_indices)
        
        if len(train_indices) > 1:
            # If test set exists, only divide validation set from training set
            # If test set does not exist, divide validation set and test set from training set
            if len(test_indices) > 0:
                # Test set already exists, only divide validation set
                val_ratio_adjusted = val_ratio
                train_idx, val_idx = train_test_split(
                    train_indices, test_size=val_ratio_adjusted, random_state=random_state
                )
                train_indices = train_idx.tolist()
                val_indices = val_idx.tolist()
            else:
                # Test set does not exist, divide validation set and test set from training set
                train_val_idx, test_idx = train_test_split(
                    train_indices, test_size=test_ratio, random_state=random_state
                )
                val_ratio_adjusted = val_ratio / (1 - test_ratio)
                train_idx, val_idx = train_test_split(
                    train_val_idx, test_size=val_ratio_adjusted, random_state=random_state
                )
                train_indices = train_idx.tolist()
                val_indices = val_idx.tolist()
                test_indices = test_idx.tolist()
        else:
            val_indices = []
    
    # Convert to numpy array
    train_indices = np.array(train_indices)
    val_indices = np.array(val_indices)
    test_indices = np.array(test_indices)
    
    # Split data
    X_train = X[train_indices] if len(train_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    X_val = X[val_indices] if len(val_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    X_test = X[test_indices] if len(test_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    
    y_train = y[train_indices] if len(train_indices) > 0 else np.array([])
    y_val = y[val_indices] if len(val_indices) > 0 else np.array([])
    y_test = y[test_indices] if len(test_indices) > 0 else np.array([])
    
    # Get corresponding sample IDs
    train_ids = [sample_ids[i] for i in train_indices] if len(train_indices) > 0 else []
    val_ids = [sample_ids[i] for i in val_indices] if len(val_indices) > 0 else []
    test_ids = [sample_ids[i] for i in test_indices] if len(test_indices) > 0 else []
    
    print(f"Data split results:")
    print(f"  Training set: {len(X_train)} samples")
    print(f"  Validation set: {len(X_val)} samples")
    print(f"  Test set: {len(X_test)} samples")
    
    return X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids

def plot_results(y_true, y_pred, save_dir, model_name='xgboost'):
    """Plot results"""
    print("\n=== Plot results ===")
    
    os.makedirs(save_dir, exist_ok=True)
    
    y_true = np.array(y_true).flatten()
    y_pred = np.array(y_pred).flatten()
    
    r2 = r2_score(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    mape = mean_absolute_percentage_error(y_true, y_pred)
    
    # Create subplots
    fig, axes = plt.subplots(2, 2, figsize=(15, 12))
    fig.suptitle(f'XGBoost Results ($R^2$={r2:.4f})', fontsize=16, fontweight='bold')
    
    # 1. Predicted vs true values
    axes[0,0].scatter(y_true, y_pred, alpha=0.7, s=60, color='green')
    axes[0,0].plot([y_true.min(), y_true.max()], [y_true.min(), y_true.max()], 'r--', lw=2)
    axes[0,0].set_xlabel('True Peak Stress fc (MPa)')
    axes[0,0].set_ylabel('Predicted Peak Stress fc (MPa)')
    axes[0,0].set_title(f'Prediction vs True Values\n$R^2$={r2:.4f}, MAE={mae:.3f}, RMSE={rmse:.3f}')
    axes[0,0].grid(True, alpha=0.3)
    
    # 2. Residual plot
    residuals = y_pred - y_true
    axes[0,1].scatter(y_pred, residuals, alpha=0.7, s=60, color='green')
    axes[0,1].axhline(y=0, color='r', linestyle='--', lw=2)
    axes[0,1].set_xlabel('Predicted Peak Stress fc (MPa)')
    axes[0,1].set_ylabel('Residuals (MPa)')
    axes[0,1].set_title('Residual Plot')
    axes[0,1].grid(True, alpha=0.3)
    
    # 3. Residual distribution
    axes[1,0].hist(residuals, bins=15, alpha=0.7, color='green', density=True)
    axes[1,0].axvline(x=0, color='r', linestyle='--', lw=2)
    axes[1,0].set_xlabel('Residuals (MPa)')
    axes[1,0].set_ylabel('Density')
    axes[1,0].set_title('Residual Distribution')
    axes[1,0].grid(True, alpha=0.3)
    
    # 4. Feature importance
    # Note: Feature importance needs to be obtained from the model object after model training
    stats_text = f"""Model Performance Statistics:
    
Sample Count: {len(y_true)}
R2: {r2:.4f}
MAE: {mae:.3f} MPa
RMSE: {rmse:.3f} MPa

XGBoost Features:
- Gradient Boosting
- Optuna Hyperparameter Optimization
- Cross-validation evaluation
- Feature Importance Analysis"""
    
    axes[1,1].text(0.1, 0.9, stats_text, transform=axes[1,1].transAxes, 
                   fontsize=11, verticalalignment='top', fontfamily='monospace')
    axes[1,1].set_xlim(0, 1)
    axes[1,1].set_ylim(0, 1)
    axes[1,1].axis('off')
    axes[1,1].set_title('Performance Statistics')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, f'xgboost_results.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    return r2, mae, rmse

def plot_train_test_comparison(model, X_train, y_train, X_test, y_test, save_dir):
    """Plot train and test set comparison plot (with marginal distribution) - optimized version
    
    Returns:
        DataFrame: DataFrame containing training and test set prediction results
    """
    print("\n=== Plot train and test set comparison plot (with marginal distribution) - optimized version ===")
    
    # Predict
    y_train_pred = model.predict(X_train)
    y_test_pred = model.predict(X_test)
    
    # Calculate metrics
    r2_train = r2_score(y_train, y_train_pred)
    mae_train = mean_absolute_error(y_train, y_train_pred)
    rmse_train = np.sqrt(mean_squared_error(y_train, y_train_pred))
    r2_test = r2_score(y_test, y_test_pred)
    mae_test = mean_absolute_error(y_test, y_test_pred)
    rmse_test = np.sqrt(mean_squared_error(y_test, y_test_pred))
    
    # Create figure (scatter plot with marginal distribution）
    fig = plt.figure(figsize=(13, 11))
    
    # Define grid layout (adjust ratio）
    gs = fig.add_gridspec(3, 3, hspace=0.08, wspace=0.08,
                         height_ratios=[0.8, 4, 0.1], width_ratios=[4, 0.8, 0.1])
    
    # Main scatter plot
    ax_main = fig.add_subplot(gs[1, 0])
    
    # Top histogram
    ax_top = fig.add_subplot(gs[0, 0], sharex=ax_main)
    
    # Right histogram
    ax_right = fig.add_subplot(gs[1, 1], sharey=ax_main)
    
    # Plot main scatter plot
    # Training set - use more professional blue
    ax_main.scatter(y_train, y_train_pred, alpha=0.5, s=50, color='#2E86AB', 
                   label='Training', edgecolors='white', linewidths=0.5, zorder=3)
    
    # Test set - use more突出的橙红色
    ax_main.scatter(y_test, y_test_pred, alpha=0.7, s=60, color='#E63946',
                   label='Testing', edgecolors='white', linewidths=0.5, zorder=4)
    
    # 1:1 diagonal line
    all_values = np.concatenate([y_train, y_test, y_train_pred, y_test_pred])
    min_val = all_values.min()
    max_val = all_values.max()
    # Add padding
    padding = (max_val - min_val) * 0.05
    min_val -= padding
    max_val += padding
    
    ax_main.plot([min_val, max_val], [min_val, max_val], 'k--', lw=2, 
                alpha=0.4, label='Perfect Prediction', zorder=1)
    
    # Calculate and plot best fit line and confidence interval (training set）
    slope_train, intercept_train, r_value_train, p_value_train, std_err_train = stats.linregress(y_train, y_train_pred)
    line_train = slope_train * y_train + intercept_train
    predict_train = slope_train * np.sort(y_train) + intercept_train
    
    # Calculate confidence interval
    residuals_train = y_train_pred - line_train
    std_residuals_train = np.std(residuals_train)
    ci_train = 1.96 * std_residuals_train  # 95% CI
    
    ax_main.plot(np.sort(y_train), predict_train, color='#2E86AB', lw=3, alpha=0.9,
                linestyle='-', zorder=5)
    ax_main.fill_between(np.sort(y_train), predict_train - ci_train, predict_train + ci_train,
                         alpha=0.2, color='#2E86AB', zorder=2)
    
    # Best fit line and confidence interval (test set）
    slope_test, intercept_test, r_value_test, p_value_test, std_err_test = stats.linregress(y_test, y_test_pred)
    line_test = slope_test * y_test + intercept_test
    predict_test = slope_test * np.sort(y_test) + intercept_test
    
    residuals_test = y_test_pred - line_test
    std_residuals_test = np.std(residuals_test)
    ci_test = 1.96 * std_residuals_test
    
    ax_main.plot(np.sort(y_test), predict_test, color='#E63946', lw=3, alpha=0.9,
                linestyle='-', zorder=6)
    ax_main.fill_between(np.sort(y_test), predict_test - ci_test, predict_test + ci_test,
                         alpha=0.2, color='#E63946', zorder=2)
    
    # Set axis labels
    ax_main.set_xlabel('Observed Peak Stress fc (MPa)', fontsize=13, fontweight='bold')
    ax_main.set_ylabel('Predicted Peak Stress fc (MPa)', fontsize=13, fontweight='bold')
    
    # Optimized legend - placed in lower right corner, using LaTeX format to display R²
    legend_elements = [
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#2E86AB', 
                  markersize=8, alpha=0.7, label=f'Training ($R^2$={r2_train:.3f}, MAE={mae_train:.2f})'),
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#E63946', 
                  markersize=9, alpha=0.8, label=f'Testing ($R^2$={r2_test:.3f}, MAE={mae_test:.2f})'),
        plt.Line2D([0], [0], color='k', linestyle='--', lw=2, alpha=0.4, label='Perfect Prediction'),
        plt.Line2D([0], [0], color='gray', lw=3, label='Best Fit Line'),
    ]
    ax_main.legend(handles=legend_elements, loc='lower right', fontsize=9.5, 
                  framealpha=0.95, edgecolor='gray', fancybox=True, shadow=True)
    
    ax_main.grid(True, alpha=0.25, linestyle='--', linewidth=0.5)
    ax_main.set_axisbelow(True)
    
    # Top histogram (observed value distribution）- stacked display
    bins = np.linspace(min_val, max_val, 20)
    # Use numpy.histogram to calculate frequency
    hist_train_obs, _ = np.histogram(y_train, bins=bins)
    hist_test_obs, _ = np.histogram(y_test, bins=bins)
    
    # Plot stacked histogram
    bin_centers = (bins[:-1] + bins[1:]) / 2
    width = bins[1] - bins[0]
    ax_top.bar(bin_centers, hist_train_obs, width=width, alpha=0.7, 
              color='#2E86AB', label='Training', edgecolor='white', linewidth=0.5)
    ax_top.bar(bin_centers, hist_test_obs, width=width, alpha=0.8, 
              color='#E63946', label='Testing', edgecolor='white', linewidth=0.5, bottom=hist_train_obs)
    ax_top.set_ylabel('Count', fontsize=11, fontweight='bold')
    ax_top.legend(loc='upper right', fontsize=9, framealpha=0.9)
    ax_top.tick_params(labelbottom=False, labelsize=10)
    ax_top.grid(True, alpha=0.25, axis='y', linestyle='--', linewidth=0.5)
    ax_top.set_axisbelow(True)
    ax_top.spines['top'].set_visible(False)
    ax_top.spines['right'].set_visible(False)
    
    # Right histogram (predicted value distribution）- stacked display
    hist_train_pred, _ = np.histogram(y_train_pred, bins=bins)
    hist_test_pred, _ = np.histogram(y_test_pred, bins=bins)
    
    # Plot horizontal stacked histogram
    ax_right.barh(bin_centers, hist_train_pred, height=width, alpha=0.7, 
                 color='#2E86AB', edgecolor='white', linewidth=0.5)
    ax_right.barh(bin_centers, hist_test_pred, height=width, alpha=0.8, 
                 color='#E63946', edgecolor='white', linewidth=0.5, left=hist_train_pred)
    ax_right.set_xlabel('Count', fontsize=11, fontweight='bold')
    ax_right.tick_params(labelleft=False, labelsize=10)
    ax_right.grid(True, alpha=0.25, axis='x', linestyle='--', linewidth=0.5)
    ax_right.set_axisbelow(True)
    ax_right.spines['top'].set_visible(False)
    ax_right.spines['right'].set_visible(False)
    
    # Add total title - include sample number information
    title = f'XGBoost Model Performance\nTraining: n={len(y_train)} | Testing: n={len(y_test)}'
    fig.suptitle(title, fontsize=15, fontweight='bold', y=0.985)
    
    # Set overall background
    fig.patch.set_facecolor('white')
    
    plt.savefig(os.path.join(save_dir, 'train_test_comparison_with_margins.png'), 
               dpi=300, bbox_inches='tight', facecolor='white')
    # plt.show()
    plt.close()
    
    print(f"\nTraining set performance - R2: {r2_train:.4f}, MAE: {mae_train:.3f} MPa, RMSE: {rmse_train:.3f} MPa")
    print(f"Test set performance - R2: {r2_test:.4f}, MAE: {mae_test:.3f} MPa, RMSE: {rmse_test:.3f} MPa")
    print(f"Performance difference - ΔR2: {abs(r2_train - r2_test):.4f}, ΔMAE: {abs(mae_train - mae_test):.3f} MPa")
    
    # Create DataFrame containing training and test set prediction results
    train_df = pd.DataFrame({
        'dataset': ['Training set'] * len(y_train),
        'observed': y_train,
        'predicted': y_train_pred,
        'residual': y_train - y_train_pred,
        'absolute_error': np.abs(y_train - y_train_pred)
    })
    
    test_df = pd.DataFrame({
        'dataset': ['Test set'] * len(y_test),
        'observed': y_test,
        'predicted': y_test_pred,
        'residual': y_test - y_test_pred,
        'absolute_error': np.abs(y_test - y_test_pred)
    })
    
    # Merge training and test set data
    prediction_df = pd.concat([train_df, test_df], ignore_index=True)
    
    return prediction_df

def plot_feature_importance(model, feature_names, save_dir, X=None):
    """Plot feature importance
    
    Parameters:
        model: Trained XGBoost model
        feature_names: Feature name list
        save_dir: Save directory
        X: Feature data (optional, for calculating coefficient of variation）
    """
    print("\n=== Plot feature importance ===")
    
    # Get feature importance
    importance = model.feature_importances_
    
    # Create DataFrame
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': importance
    })
    
    # If feature data is provided, calculate coefficient of variation (CV）
    if X is not None:
        cv_values = []
        for i in range(X.shape[1]):
            feature_mean = np.mean(X[:, i])
            feature_std = np.std(X[:, i])
            # Coefficient of variation = standard deviation / mean (to avoid division by zero）
            cv = feature_std / (feature_mean + 1e-10) if abs(feature_mean) > 1e-10 else 0.0
            cv_values.append(cv)
        importance_df['CV'] = cv_values
        print("✓ Coefficient of variation (CV）calculated")
    else:
        importance_df['CV'] = np.nan
        print("⚠ No feature data provided, cannot calculate coefficient of variation")
    
    # Sort by importance
    importance_df = importance_df.sort_values('importance', ascending=False)
    
    # Plot bar chart
    plt.figure(figsize=(12, 15))
    plt.barh(importance_df['feature'], importance_df['importance'], color='skyblue')
    plt.xlabel('Feature Importance')
    plt.title('XGBoost Feature Importance')
    plt.gca().invert_yaxis()
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'feature_importance.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    # Save to Excel（将在主函数中统一保存到综合Excel文件）
    # Here we do not save directly, return DataFrame for main function to save
    
    # Print statistics
    print(f"\nFeature importance statistics:")
    print(f"  Most important feature: {importance_df.iloc[0]['feature']} (importance={importance_df.iloc[0]['importance']:.6f})")
    if X is not None:
        print(f"  Highest CV feature: {importance_df.loc[importance_df['CV'].idxmax(), 'feature']} (CV={importance_df['CV'].max():.4f})")
        print(f"  Lowest CV feature: {importance_df.loc[importance_df['CV'].idxmin(), 'feature']} (CV={importance_df['CV'].min():.4f})")
    
    return importance_df

def fix_xgboost_model_for_shap(model):
    """Fix XGBoost model to be compatible with SHAP analysis
    
    Some XGBoost models save base_score as a string format (e.g. '[3.1165977E1]'）
    Need to convert it to a float number
    """
    try:
        # Check if the model has get_booster method (XGBoost model）
        if hasattr(model, 'get_booster'):
            booster = model.get_booster()
            config = booster.save_config()
            import json
            config_dict = json.loads(config)
            
            # Fix base_score
            if 'learner' in config_dict and 'learner_model_param' in config_dict['learner']:
                learner_model_param = config_dict['learner']['learner_model_param']
                if 'base_score' in learner_model_param:
                    base_score_str = learner_model_param['base_score']
                    # If it is a string format, try to parse
                    if isinstance(base_score_str, str):
                        # Remove square brackets and spaces
                        base_score_clean = base_score_str.strip('[]').strip()
                        try:
                            base_score_float = float(base_score_clean)
                            # Update configuration
                            learner_model_param['base_score'] = str(base_score_float)
                            # Reload configuration
                            new_config = json.dumps(config_dict)
                            booster.load_config(new_config)
                            print(f"✓ Fix base_score: '{base_score_str}' -> {base_score_float}")
                        except ValueError as ve:
                            print(f"⚠ Cannot parse base_score '{base_score_str}': {ve}")
                            # Try to set base_score property directly
                            try:
                                if hasattr(booster, 'set_param'):
                                    booster.set_param('base_score', base_score_clean)
                                    print(f"✓ Fix base_score through set_param")
                            except:
                                pass
        
        return model
    except Exception as e:
        print(f"⚠ Error fixing model (may not affect usage): {e}")
        import traceback
        traceback.print_exc()
        return model

def plot_shap_analysis(model, X_background, X_explain, feature_names, save_dir):
    """SHAP value analysis
    
    Parameters:
        model: Trained XGBoost model
        X_background: Background dataset (used to initialize explainer, usually using training set）
        X_explain: Dataset to explain (data for calculating SHAP values, usually using test set）
        feature_names: Feature name list
        save_dir: Save directory
    
    Traditional做法说明:
        - Background dataset (X_background): Using training set (or training set + validation set）
          Used to initialize SHAP explainer, representing the "baseline" distribution of the model
        - Explanation dataset (X_explain): Using test set
          Explain the data that has not been seen, evaluate the interpretability of model generalization ability
        
        Advantages:
        1. Conforms to machine learning best practices (training/test separation）
        2. Evaluates the interpretability of the model in real application scenarios
        3. Avoids data leakage (do not use test set to train explainer）
    """
    print("\n=== Perform SHAP analysis ===")
    print(f"Background dataset sample number: {len(X_background)} (used to initialize explainer)")
    print(f"Explanation dataset sample number: {len(X_explain)} (calculate SHAP values)")
    
    # Fix model to be compatible with SHAP (handle base_score format problem）
    model = fix_xgboost_model_for_shap(model)
    
    # Use TreeExplainer (XGBoost专用）
    # Note: Some XGBoost model save format may cause explainer initialization error
    # If the background data set is passed in error, it can be passed in without background data set (using default value）
    try:
        # Try to use background data set to initialize explainer
        explainer = shap.TreeExplainer(model, X_background)
        print("✓ Successfully initialized explainer using background data set")
    except (ValueError, TypeError, AttributeError) as e:
        print(f"⚠ Failed to initialize explainer using background data set: {e}")
        print("  Try to initialize explainer without background data set...")
        try:
            # Do not pass in background data set, use default value
            explainer = shap.TreeExplainer(model)
            print("✓ Successfully initialized explainer without background data set")
        except Exception as e2:
            print(f"⚠ Failed to initialize explainer: {e2}")
            print("  Try to use KernelExplainer as a backup solution...")
            # If TreeExplainer fails, try to use KernelExplainer (slower but more general）
            try:
                # Sample background data to improve speed
                if len(X_background) > 100:
                    background_sample = shap.sample(X_background, 100)
                else:
                    background_sample = X_background
                explainer = shap.KernelExplainer(model.predict, background_sample)
                print("✓ Successfully initialized explainer using KernelExplainer (note: slower calculation speed）")
            except Exception as e3:
                print(f"⚠ All explainer initialization methods failed: {e3}")
                raise ValueError(f"Failed to initialize SHAP explainer: {e3}")
    
    # Calculate SHAP values (for explanation dataset）
    print("  Calculating SHAP values...")
    shap_values = explainer.shap_values(X_explain)
    
    # Calculate feature importance (average absolute SHAP value）
    shap_importance_values = np.abs(shap_values).mean(axis=0)
    
    # Set importance threshold (filter out features with importance 0 or close to 0）
    importance_threshold = 1e-6  # Very small threshold, filter out truly 0 features
    
    # Find the indices of features with importance greater than the threshold
    important_feature_indices = np.where(shap_importance_values > importance_threshold)[0]
    
    print(f"\nFeature importance statistics:")
    print(f"  Total number of features: {len(feature_names)}")
    print(f"  Number of features with importance > 0: {len(important_feature_indices)}")
    print(f"  Number of features with importance = 0: {len(feature_names) - len(important_feature_indices)}")
    
    if len(important_feature_indices) < len(feature_names):
        zero_importance_features = [feature_names[i] for i in range(len(feature_names)) if i not in important_feature_indices]
        print(f"  Features with importance = 0: {zero_importance_features}")
    
    # Set font, avoid Chinese characters, and ensure negative sign is displayed normally
    original_font = list(plt.rcParams['font.sans-serif'])
    original_unicode_minus = plt.rcParams['axes.unicode_minus']
    
    # For SHAP plot, use DejaVu Sans to ensure negative sign is displayed normally
    configure_plot_fonts(['DejaVu Sans', 'SimHei', 'Microsoft YaHei'])
    plt.rcParams['axes.unicode_minus'] = False  # Ensure negative sign is displayed normally
    
    # 1. Summary plot (complete version - including all features） - save separately
    # Increase width to 30, significantly lengthen the horizontal axis ratio
    fig = plt.figure(figsize=(30, 12))  # Increase width to 30, significantly lengthen the horizontal axis ratio
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, show=False)
    # Get current axes
    ax = plt.gca()
    # Do not change data range, only increase graphic width to lengthen the physical ratio of the horizontal axis
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_summary.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # 2. Bar plot (complete version - including all features） - save separately
    plt.figure(figsize=(10, 8))
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, plot_type="bar", show=False)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_bar.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # 3. Combined plot: left bar plot + right summary plot
    # First draw summary plot to get feature order
    fig_temp = plt.figure(figsize=(30, 12))
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, show=False)
    ax_temp = plt.gca()
    # Get the feature order displayed in the summary plot (from y-axis labels, from top to bottom）
    y_tick_labels = [label.get_text() for label in ax_temp.get_yticklabels()]
    plt.close(fig_temp)
    
    # Reorder data according to the feature order of the summary plot
    if len(y_tick_labels) > 0 and len(y_tick_labels) == len(feature_names):
        # Create a mapping from feature name to index
        feature_to_idx = {name: idx for idx, name in enumerate(feature_names)}
        sorted_indices = [feature_to_idx[name] for name in y_tick_labels if name in feature_to_idx]
        if len(sorted_indices) == len(feature_names):
            mean_abs_shap = np.abs(shap_values).mean(axis=0)
            mean_abs_shap_sorted = mean_abs_shap[sorted_indices]
            feature_names_sorted = y_tick_labels
        else:
            # If the order does not match, use default sorting
            mean_abs_shap = np.abs(shap_values).mean(axis=0)
        sorted_indices = np.argsort(mean_abs_shap)[::-1]
        mean_abs_shap_sorted = mean_abs_shap[sorted_indices]
        feature_names_sorted = [feature_names[i] for i in sorted_indices]
    else:
        # Use default sorting
        mean_abs_shap = np.abs(shap_values).mean(axis=0)
        sorted_indices = np.argsort(mean_abs_shap)[::-1]
        mean_abs_shap_sorted = mean_abs_shap[sorted_indices]
        feature_names_sorted = [feature_names[i] for i in sorted_indices]
    
    # Create combined plot, using GridSpec layout: left bar plot, right summary plot
    from matplotlib.gridspec import GridSpec
    fig = plt.figure(figsize=(30, 12))
    gs = GridSpec(1, 2, figure=fig, width_ratios=[1, 4], hspace=0, wspace=0.02)
    
    # Left: feature importance bar plot
    ax_bar = fig.add_subplot(gs[0])
    n_features = len(feature_names_sorted)
    y_positions = np.arange(n_features)
    
    # Draw horizontal bar plot (from left y-axis, extending to the right）
    bar_color = '#6699CC'  # User-specified blue
    ax_bar.barh(y_positions, mean_abs_shap_sorted, 
                left=0.0, color=bar_color, height=0.8, alpha=0.8)
    
    # Set y-axis (feature names）- display on the left
    ax_bar.set_yticks(y_positions)
    ax_bar.set_yticklabels(feature_names_sorted)
    ax_bar.invert_yaxis()  # Invert y-axis, so the most important features are at the top
    ax_bar.set_ylabel('Features', fontsize=12, fontweight='bold')
    ax_bar.tick_params(axis='y', labelsize=10, left=True)
    
    # Set x-axis (importance values）- move to the top
    max_importance = np.max(mean_abs_shap_sorted) if len(mean_abs_shap_sorted) > 0 else 1.0
    ax_bar.set_xlim(0, max_importance * 1.15)
    ax_bar.xaxis.set_ticks_position('top')  # x-axis ticks at the top
    ax_bar.xaxis.set_label_position('top')  # x-axis labels at the top
    ax_bar.set_xlabel('Mean(|SHAP value|)', fontsize=12, fontweight='bold')
    ax_bar.tick_params(axis='x', labelsize=10, top=True, bottom=False)
    
    # Hide the bottom and right spines
    ax_bar.spines['bottom'].set_visible(False)
    ax_bar.spines['right'].set_visible(False)
    ax_bar.spines['top'].set_visible(True)
    ax_bar.spines['left'].set_visible(True)
    
    # Right: SHAP summary plot (shared y-axis）
    ax_summary = fig.add_subplot(gs[1], sharey=ax_bar)
    # Set the current axes to the right axes
    plt.sca(ax_summary)
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, show=False)
    
    # Ensure y-axis alignment (hide y-axis labels and ticks on the right plot, because the left plot already displays them）
    ax_summary.set_yticklabels([])
    ax_summary.tick_params(axis='y', left=False, right=False)
    
    # Ensure the x-axis of the right plot is at the bottom (summary plot is default at the bottom）
    ax_summary.tick_params(axis='x', labelsize=10, bottom=True, top=False)
    
    # Adjust layout
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_summary_combined.png'), dpi=300, bbox_inches='tight')
    plt.close()
    print(f"✓ Combined SHAP plot saved: shap_summary_combined.png")
    
    # 3. Waterfall plot (display SHAP values for a single sample）
    # Note: This is an analysis of a single sample, showing how each feature affects the prediction result
    # Select the first 3 samples as an example (if the data is enough）
    n_samples_waterfall = min(3, len(shap_values))
    
    for sample_idx in range(n_samples_waterfall):
        shap_explanation = shap.Explanation(
            values=shap_values[sample_idx:sample_idx+1],
            base_values=explainer.expected_value,
            data=X_explain[sample_idx:sample_idx+1],
            feature_names=feature_names
        )
        
        # Ensure negative sign is displayed normally - use DejaVu Sans font (supports negative sign）
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'SimHei', 'Microsoft YaHei']
        
        plt.figure(figsize=(12, 8))
        shap.plots.waterfall(shap_explanation[0], show=False)
        
        # Get current axes and ensure negative sign is displayed
        ax = plt.gca()
        # Force negative sign display and font
        ax.tick_params(which='major', labelsize=10)
        # Ensure all text uses a font that supports negative sign
        for text in ax.texts:
            text.set_fontfamily('DejaVu Sans')
        
        plt.tight_layout()
        if n_samples_waterfall == 1:
            plt.savefig(os.path.join(save_dir, 'shap_waterfall.png'), dpi=300, bbox_inches='tight')
        else:
            plt.savefig(os.path.join(save_dir, f'shap_waterfall_sample_{sample_idx+1}.png'), dpi=300, bbox_inches='tight')
        plt.close()
    
    print(f"✓ Generated {n_samples_waterfall} Waterfall plots")
    
    # 4. Force plot (display the prediction process for a single sample, long bar force plot）
    # Create an explanation containing multiple samples for force plot
    n_samples_force = min(3, len(shap_values))
    shap_explanation_force = shap.Explanation(
        values=shap_values[:n_samples_force],
        base_values=explainer.expected_value,
        data=X_explain[:n_samples_force],
        feature_names=feature_names
    )
    # Draw force plot for the first 3 samples
    for i in range(n_samples_force):
        # Ensure negative sign is displayed normally
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'SimHei', 'Microsoft YaHei']
        
        plt.figure(figsize=(16, 4))  # Long bar plot
        shap.plots.force(shap_explanation_force[i], matplotlib=True, show=False)
        
        # Ensure negative sign is displayed
        ax = plt.gca()
        for text in ax.texts:
            text.set_fontfamily('DejaVu Sans')
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, f'shap_force_sample_{i}.png'), dpi=300, bbox_inches='tight')
        # plt.show()
        plt.close()
    
    # 5. Calculate the average absolute SHAP value as feature importance
    shap_importance = pd.DataFrame({
        'feature': feature_names,
        'shap_importance': np.abs(shap_values).mean(axis=0),
        'shap_mean': shap_values.mean(axis=0),  # Average SHAP value (considering positive and negative）
        'shap_std': shap_values.std(axis=0),    # SHAP value standard deviation
        'shap_max': np.abs(shap_values).max(axis=0),  # Maximum absolute SHAP value
        'shap_min': np.abs(shap_values).min(axis=0)   # Minimum absolute SHAP value
    }).sort_values('shap_importance', ascending=False)
    
    # Save SHAP importance (will be saved in the main function）
    # Return DataFrame for main function to save
    
    # 6. Analyze the reasons why features have a small impact (print detailed analysis）
    print("\n" + "="*60)
    print("SHAP feature importance analysis")
    print("="*60)
    print(f"\nThe 5 most important features:")
    for idx, row in shap_importance.head(5).iterrows():
        print(f"  {row['feature']:10s}: Average impact = {row['shap_importance']:.4f}, "
              f"Average SHAP value = {row['shap_mean']:+.4f}, Standard deviation = {row['shap_std']:.4f}")
    
    print(f"\nThe 5 least important features:")
    for idx, row in shap_importance.tail(5).iterrows():
        print(f"  {row['feature']:10s}: Average impact = {row['shap_importance']:.4f}, "
              f"Average SHAP value = {row['shap_mean']:+.4f}, Standard deviation = {row['shap_std']:.4f}")
    
    # 分析影响小的可能原因
    low_importance_features = shap_importance[shap_importance['shap_importance'] < 0.1]
    if len(low_importance_features) > 0:
        print(f"\n⚠ Found {len(low_importance_features)} features with a small impact (average SHAP value < 0.1):")
        for idx, row in low_importance_features.iterrows():
            print(f"  - {row['feature']}: Average impact = {row['shap_importance']:.4f}")
        
        print("\nPossible reasons:")
        print("  1. The feature value has a small range (small variance), resulting in limited impact on prediction")
        print("  2. The feature is highly correlated with other features, replaced by other features")
        print("  3. The feature value is evenly distributed in the data set, lacking distinction")
        print("  4. The feature actually has a small impact on the target variable with physical/engineering laws）")
        print("  5. The sample size is small (94 samples), the impact of some features may be masked by noise")
        print("\nSuggestions:")
        print("  - Check the statistical distribution of these features (mean, standard deviation, range）")
        print("  - Analyze the correlation between features")
        print("  - View PDP plot to understand the marginal effect of these features")
        print("  - Consider feature engineering (such as feature combination, transformation）")
    
    print("="*60)
    
    # Restore original font settings
    plt.rcParams['font.sans-serif'] = original_font
    plt.rcParams['axes.unicode_minus'] = False
    
    return shap_importance

def analyze_feature_statistics(X, y, feature_names, shap_importance, save_dir):
    """Analyze feature statistics, explain why some features have a small impact"""
    print("\n" + "="*60)
    print("Feature statistics analysis (explain why some features have a small impact）")
    print("="*60)
    
    import pandas as pd
    
    # Create feature index mapping (feature_name -> column_index）
    feature_to_idx = {name: idx for idx, name in enumerate(feature_names)}
    
    # Create feature_stats according to the order of shap_importance (consistent with the SHAP summary plot）
    # shap_importance is already sorted by importance in descending order
    feature_stats_list = []
    
    for idx, row in shap_importance.iterrows():
        feat_name = row['feature']
        feat_idx = feature_to_idx[feat_name]
        
        # Calculate the statistical information for this feature
        feat_data = X[:, feat_idx]
        q25 = np.percentile(feat_data, 25)
        q50 = np.percentile(feat_data, 50)
        q75 = np.percentile(feat_data, 75)
        corr = np.corrcoef(feat_data, y)[0, 1]
        
        feat_range = np.max(feat_data) - np.min(feat_data)
        feat_std = np.std(feat_data)
        
        # Special check for w/c feature
        if feat_name == 'w/c':
            print(f"\n🔍 Detailed check for w/c feature:")
            print(f"  Data shape: {feat_data.shape}")
            print(f"  Unique value count: {len(np.unique(feat_data))}")
            print(f"  Unique values: {np.unique(feat_data)[:10]}")  # Display the first 10 unique values
            print(f"  Minimum value: {np.min(feat_data):.10f}")
            print(f"  Maximum value: {np.max(feat_data):.10f}")
            print(f"  Range: {feat_range:.10f}")
            print(f"  Mean: {np.mean(feat_data):.10f}")
            print(f"  Standard deviation: {feat_std:.10f}")
            print(f"  SHAP importance: {row['shap_importance']:.10f}")
            if feat_range < 1e-6:
                print(f"  ⚠ Warning: w/c feature value range is close to 0！")
                print(f"  Possible reasons:")
                print(f"    1. All values in the w/c column are the same in the data")
                print(f"    2. All values in the w/c column are the same after standardization during data preprocessing")
                print(f"    3. The w/c column is a constant in the data file")
                print(f"  Suggestions: Check the w/c column in the original data file")
        
        feature_stats_list.append({
            'feature': feat_name,
            'mean': np.mean(feat_data),
            'std': feat_std,
            'min': np.min(feat_data),
            'q25': q25,
            'median': q50,
            'q75': q75,
            'max': np.max(feat_data),
            'range': feat_range,
            'iqr': q75 - q25,
            'cv': feat_std / (np.mean(feat_data) + 1e-10),
            'shap_importance': row['shap_importance'],
            'correlation_with_target': corr
        })
    
    feature_stats = pd.DataFrame(feature_stats_list)
    
    # Rearrange the column order, make it more logical: basic statistics -> quartiles -> derived statistics -> importance
    column_order = ['feature', 'mean', 'std', 'min', 'q25', 'median', 'q75', 'max', 
                    'range', 'iqr', 'cv', 'shap_importance', 'correlation_with_target']
    feature_stats = feature_stats[column_order]
    
    # feature_stats is already sorted by shap_importance (consistent with the SHAP summary plot）
    
    # Save the statistical information (will be saved in the main function）
    # Return DataFrame for main function to save
    
    # Analyze the features with a small impact
    low_importance_threshold = 0.1
    low_importance = feature_stats[feature_stats['shap_importance'] < low_importance_threshold]
    
    print(f"\nAnalysis of features with a small impact (SHAP importance < {low_importance_threshold}）:")
    if len(low_importance) > 0:
        print(f"\nTotal {len(low_importance)} features with a small impact:")
        for idx, row in low_importance.iterrows():
            print(f"\n  {row['feature']}:")
            print(f"    - SHAP importance: {row['shap_importance']:.4f}")
            print(f"    - Correlation with the target variable: {row['correlation_with_target']:.4f}")
            print(f"    - Mean: {row['mean']:.4f}, Standard deviation: {row['std']:.4f}")
            print(f"    - Range: [{row['min']:.4f}, {row['max']:.4f}], CV: {row['cv']:.4f}")
            print(f"    - Quartiles: Q1={row['q25']:.4f}, Median={row['median']:.4f}, Q3={row['q75']:.4f}, IQR={row['iqr']:.4f}")
            
            # Determine possible reasons
            reasons = []
            if abs(row['correlation_with_target']) < 0.1:
                reasons.append("Low correlation with the target variable")
            if row['cv'] < 0.1:
                reasons.append("Small CV (small data range）")
            if row['range'] < row['mean'] * 0.1:
                reasons.append("Small range relative to the mean")
            
            if reasons:
                print(f"    - Possible reasons: {', '.join(reasons)}")
    
    # Analyze the features with a large impact
    high_importance = feature_stats.head(5)
    print(f"\nThe 5 most important features:")
    for idx, row in high_importance.iterrows():
        print(f"  {row['feature']:10s}: SHAP={row['shap_importance']:.4f}, "
              f"Correlation={row['correlation_with_target']:+.4f}, CV={row['cv']:.4f}")
    
    # Plot the feature statistics
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    
    # 1. SHAP importance vs correlation
    ax1 = axes[0, 0]
    scatter = ax1.scatter(feature_stats['correlation_with_target'], 
                         feature_stats['shap_importance'], 
                         s=100, alpha=0.6, c=feature_stats['cv'], cmap='viridis')
    for idx, row in feature_stats.iterrows():
        ax1.annotate(row['feature'], 
                    (row['correlation_with_target'], row['shap_importance']),
                    fontsize=9, alpha=0.7)
    ax1.set_xlabel('Correlation with the target variable', fontsize=12)
    ax1.set_ylabel('SHAP importance', fontsize=12)
    ax1.set_title('SHAP importance vs correlation', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3)
    plt.colorbar(scatter, ax=ax1, label='CV')
    
    # 2. SHAP importance vs CV
    ax2 = axes[0, 1]
    ax2.scatter(feature_stats['cv'], feature_stats['shap_importance'], s=100, alpha=0.6)
    for idx, row in feature_stats.iterrows():
        ax2.annotate(row['feature'], 
                    (row['cv'], row['shap_importance']),
                    fontsize=9, alpha=0.7)
    ax2.set_xlabel('CV (CV)', fontsize=12)
    ax2.set_ylabel('SHAP importance', fontsize=12)
    ax2.set_title('SHAP importance vs CV', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3)
    
    # 3. Feature value range (sorted by SHAP importance, consistent with the SHAP summary plot）
    ax3 = axes[1, 0]
    # feature_stats is already sorted by SHAP importance in descending order, take the first 10
    # Note: SHAP summary plot from top to bottom, so the most important is at the top
    top_features = feature_stats.head(10).copy()
    
    # Check and print detailed information for the w/c feature
    wc_feature = feature_stats[feature_stats['feature'] == 'w/c']
    if len(wc_feature) > 0:
        wc_row = wc_feature.iloc[0]
        print(f"\n⚠ Warning: w/c feature statistics:")
        print(f"  SHAP importance: {wc_row['shap_importance']:.6f}")
        print(f"  Minimum value: {wc_row['min']:.6f}")
        print(f"  Maximum value: {wc_row['max']:.6f}")
        print(f"  Range: {wc_row['range']:.6f}")
        print(f"  Mean: {wc_row['mean']:.6f}")
        print(f"  Standard deviation: {wc_row['std']:.6f}")
        print(f"  CV: {wc_row['cv']:.6f}")
        if wc_row['range'] < 1e-6:
            print(f"  ⚠ Warning: w/c feature value range is close to 0, all sample values are the same！")
            print(f"  This will cause the SHAP importance to be 0, because the feature value cannot affect the prediction when it is constant.")
            print(f"  Please check the actual values of the w/c column in the data.")
    
    # Filter out features with range 0 or close to 0 (these features should not appear in the importance plot）
    # But keep them in top_features for the user to see the problem
    valid_range_features = top_features[top_features['range'] > 1e-6]
    
    if len(valid_range_features) < len(top_features):
        print(f"\n⚠ Warning: Found {len(top_features) - len(valid_range_features)} features with range 0 or close to 0")
        print(f"  These features: {top_features[top_features['range'] <= 1e-6]['feature'].tolist()}")
        print(f"  All features will be plotted (including features with range 0) to find the problem")
    
    y_pos = np.arange(len(top_features))
    # Plot the horizontal bar chart, the most important feature is at the top (index 0）
    # For features with range 0, use a very small value (such as 1e-6） to display on the graph
    bar_values = top_features['range'].values.copy()
    bar_values[bar_values < 1e-6] = 1e-6  # Replace 0 values with a very small value to display
    
    bars = ax3.barh(y_pos, bar_values, alpha=0.7)
    # Use different colors for features with range 0
    for i, (idx, row) in enumerate(top_features.iterrows()):
        if row['range'] < 1e-6:
            bars[i].set_color('red')
            bars[i].set_alpha(0.5)
            bars[i].set_label('range=0')
    
    ax3.set_yticks(y_pos)
    ax3.set_yticklabels(top_features['feature'])
    ax3.set_xlabel('Feature value range (max - min)', fontsize=12)
    ax3.set_title('Value range of the 10 most important features (sorted by SHAP importance）', fontsize=14, fontweight='bold')
    ax3.grid(True, alpha=0.3, axis='x')
    # Invert the y axis, so the most important feature is at the top (consistent with the SHAP summary plot: from top to bottom, importance decreases）
    ax3.invert_yaxis()
    
    # If there are features with range 0, add a legend说明
    if (top_features['range'] < 1e-6).any():
        ax3.text(0.98, 0.02, 'Red: range=0', transform=ax3.transAxes, 
                fontsize=9, color='red', ha='right', va='bottom',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.7))
    
    # 4. Correlation sorting (sorted by absolute value）
    ax4 = axes[1, 1]
    # Calculate absolute correlation and sort
    feature_stats['abs_correlation'] = feature_stats['correlation_with_target'].abs()
    feature_stats_sorted = feature_stats.sort_values('abs_correlation', ascending=False)
    top_corr = feature_stats_sorted.head(10)
    colors = ['red' if x > 0 else 'blue' for x in top_corr['correlation_with_target']]
    y_pos = np.arange(len(top_corr))
    ax4.barh(y_pos, top_corr['correlation_with_target'], alpha=0.7, color=colors)
    ax4.set_yticks(y_pos)
    ax4.set_yticklabels(top_corr['feature'])
    ax4.set_xlabel('Correlation with the target variable', fontsize=12)
    ax4.set_title('Correlation between features and the target variable (top 10 by absolute value）', fontsize=14, fontweight='bold')
    ax4.axvline(x=0, color='black', linestyle='--', linewidth=1)
    ax4.grid(True, alpha=0.3, axis='x')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'feature_statistics_analysis.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    print("\n✓ Feature statistics analysis completed, results saved")
    print("="*60)
    
    return feature_stats

def plot_pdp_analysis(model, X_train, feature_names, save_dir, n_top_features=5):
    """PDP（部分依赖图）分析 - 使用scikit-learn"""
    print("\n=== 进行PDP分析 ===")
    
    # Ensure X_train is a DataFrame
    if isinstance(X_train, np.ndarray):
        X_train_df = pd.DataFrame(X_train, columns=feature_names)
    else:
        X_train_df = X_train
    
    # Get the n most important features
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': model.feature_importances_
    }).sort_values('importance', ascending=False)
    
    top_features = importance_df.head(n_top_features)['feature'].tolist()
    top_indices = [feature_names.index(f) for f in top_features]
    
    # Dictionary for saving Excel data
    excel_data_dict = {}

    def sanitize_filename(name):
        """Remove illegal characters from the file name"""
        return ''.join(c if c.isalnum() or c in ('_', '-', '.', ' ') else '_' for c in str(name))

    # Single variable PDP single plot output directory
    single_pdp_dir = os.path.join(save_dir, 'pdp_univariate_single')
    os.makedirs(single_pdp_dir, exist_ok=True)
    
    # 1. Single variable PDP analysis
    n_cols = 3  # Change to 3 columns
    n_rows = (n_top_features + 2) // 3  # Adjust the number of rows calculation
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, 6*n_rows))
    if n_top_features == 1:
        axes = [axes]
    else:
        axes = axes.flatten()
    
    for idx, (feature, ax) in enumerate(zip(top_features, axes)):
        try:
            # Get the feature index
            feature_idx = feature_names.index(feature)
            
            # Calculate PDP (average) and ICE (individual)
            # Increase grid_resolution to get more dense grid points, so the curve is smoother
            pd_result = partial_dependence(
                model, 
                X_train_df, 
                features=[feature_idx],
                kind='average',
                grid_resolution=200  # Originally 100, now encrypted to 200
            )
            
            ice_result = partial_dependence(
                model,
                X_train_df,
                features=[feature_idx],
                kind='individual',
                grid_resolution=200  # Consistent with PDP
            )
            
            # Calculate the 95% confidence interval for the ICE curve
            ice_curves = ice_result['individual'][0]  # shape: (n_samples, n_grid_points)
            ice_mean = np.mean(ice_curves, axis=0)
            ice_std = np.std(ice_curves, axis=0)
            ice_upper_ci = ice_mean + 1.96 * ice_std  # 95% CI
            ice_lower_ci = ice_mean - 1.96 * ice_std
            
            grid_values = pd_result['grid_values'][0]
            pdp_values = pd_result['average'][0]

            # Use spline interpolation to smooth the curve
            # Create more dense interpolation points (further smoothing）
            grid_dense = np.linspace(grid_values.min(), grid_values.max(), 400)
            
            # Use spline interpolation for PDP (using UnivariateSpline, s parameter controls the smoothness）
            try:
                # Use spline interpolation, the smaller the s parameter, the closer to the original data, the larger the smoother
                # Here we use a moderate smoothness
                spline_pdp = UnivariateSpline(grid_values, pdp_values, s=0.1*len(grid_values))
                pdp_smooth = spline_pdp(grid_dense)
            except:
                # If spline interpolation fails, use linear interpolation
                interp_pdp = interp1d(grid_values, pdp_values, kind='cubic', 
                                     bounds_error=False, fill_value='extrapolate')
                pdp_smooth = interp_pdp(grid_dense)
            
            # Use spline interpolation for ICE mean
            try:
                spline_ice = UnivariateSpline(grid_values, ice_mean, s=0.1*len(grid_values))
                ice_mean_smooth = spline_ice(grid_dense)
            except:
                interp_ice = interp1d(grid_values, ice_mean, kind='cubic',
                                      bounds_error=False, fill_value='extrapolate')
                ice_mean_smooth = interp_ice(grid_dense)
            
            # Use interpolation for the confidence interval
            try:
                interp_upper = interp1d(grid_values, ice_upper_ci, kind='cubic',
                                       bounds_error=False, fill_value='extrapolate')
                interp_lower = interp1d(grid_values, ice_lower_ci, kind='cubic',
                                       bounds_error=False, fill_value='extrapolate')
                ice_upper_smooth = interp_upper(grid_dense)
                ice_lower_smooth = interp_lower(grid_dense)
            except:
                ice_upper_smooth = ice_upper_ci
                ice_lower_smooth = ice_lower_ci
                grid_dense = grid_values
            
            # Calculate ALE (real ALE, not approximate）
            # ALE is obtained by calculating the local effect and accumulating them
            n_grid = len(grid_values)
            n_samples = len(X_train_df)
            
            # Create grid intervals
            ale_values = np.zeros(n_grid)
            
            # For each grid point, calculate the local effect
            for i in range(n_grid - 1):
                # Get the samples in the current interval
                if i == 0:
                    mask = (X_train_df.iloc[:, feature_idx].values <= grid_values[i+1])
                elif i == n_grid - 2:
                    mask = (X_train_df.iloc[:, feature_idx].values > grid_values[i])
                else:
                    mask = ((X_train_df.iloc[:, feature_idx].values > grid_values[i]) & 
                           (X_train_df.iloc[:, feature_idx].values <= grid_values[i+1]))
                
                # If there are samples in the current interval
                if np.sum(mask) > 0:
                    # Calculate the local effect: change the feature value, other features remain unchanged
                    X_low = X_train_df.copy()
                    X_high = X_train_df.copy()
                    
                    # Set the feature value to the ends of the interval
                    X_low.loc[mask, X_train_df.columns[feature_idx]] = grid_values[i]
                    X_high.loc[mask, X_train_df.columns[feature_idx]] = grid_values[i+1]
                    
                    # Calculate the prediction difference
                    pred_low = model.predict(X_low.iloc[mask, :].values)
                    pred_high = model.predict(X_high.iloc[mask, :].values)
                    
                    # Local effect
                    local_effect = np.mean(pred_high - pred_low) / (grid_values[i+1] - grid_values[i])
                    
                    # Accumulate to ALE
                    if i == 0:
                        ale_values[i+1] = local_effect * (grid_values[i+1] - grid_values[i])
                    else:
                        ale_values[i+1] = ale_values[i] + local_effect * (grid_values[i+1] - grid_values[i])
            
            # Center the ALE (so the mean is 0）
            ale_values = ale_values - np.mean(ale_values) + np.mean(pdp_values)
            
            # Use spline interpolation for ALE
            try:
                spline_ale = UnivariateSpline(grid_values, ale_values, s=0.1*len(grid_values))
                ale_smooth = spline_ale(grid_dense)
            except:
                interp_ale = interp1d(grid_values, ale_values, kind='cubic',
                                     bounds_error=False, fill_value='extrapolate')
                ale_smooth = interp_ale(grid_dense)
            
            # Plot the smoothed PDP line (blue solid line, no marker, smoother）
            ax.plot(grid_dense, pdp_smooth, 
                   linewidth=2.5, color='blue', 
                   label='PDP', zorder=4)
            
            # Plot the smoothed Mean c-ICE line (coral dashed line, no marker）
            ax.plot(grid_dense, ice_mean_smooth, 
                   linewidth=2.5, color='coral', linestyle='--', 
                   label='Mean c-ICE', zorder=4)
            
            # Plot the smoothed ALE line (green dotted line, no marker）
            ax.plot(grid_dense, ale_smooth, 
                   linewidth=2.5, color='green', linestyle='-.', 
                   label='ALE', zorder=4)
            
            # Remove the 95% CI shadow region (user requirement: the range is too wide, affecting the trend observation）
            # ax.fill_between(grid_dense, ice_lower_smooth, ice_upper_smooth, 
            #                color='coral', alpha=0.3, label='95% CI of c-ICE', zorder=1)
            
            # Save data to the dictionary (for subsequent saving to Excel）
            # Simplified to a few columns, so it is convenient to draw a clean PDP curve directly in Excel
            # Only keep: grid points + smoothed PDP / c-ICE / ALE
            excel_data_dict[feature] = pd.DataFrame({
                'Grid': grid_dense,
                'PDP_Smooth': pdp_smooth,
                'Mean_cICE_Smooth': ice_mean_smooth,
                'ALE_Smooth': ale_smooth,
            })
            
            ax.set_xlabel(feature, fontsize=10)
            ax.set_ylabel('Partial Effect', fontsize=10)
            ax.set_title(f'PDP, ALE & c-ICE for {feature}', fontsize=12, fontweight='bold')
            ax.legend(loc='best', fontsize=9)
            ax.grid(True, alpha=0.3)

            # Save the single variable PDP plot
            single_fig, single_ax = plt.subplots(figsize=(7, 5))
            single_ax.plot(grid_dense, pdp_smooth, linewidth=2.5, color='blue', label='PDP', zorder=4)
            single_ax.plot(grid_dense, ice_mean_smooth, linewidth=2.5, color='coral',
                           linestyle='--', label='Mean c-ICE', zorder=4)
            single_ax.plot(grid_dense, ale_smooth, linewidth=2.5, color='green',
                           linestyle='-.', label='ALE', zorder=4)
            single_ax.set_xlabel(feature, fontsize=11, fontweight='bold')
            single_ax.set_ylabel('Partial Effect', fontsize=11, fontweight='bold')
            single_ax.set_title(f'PDP / c-ICE / ALE for {feature}', fontsize=13, fontweight='bold')
            single_ax.legend(loc='best', fontsize=10)
            single_ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.6)
            single_ax.set_axisbelow(True)
            safe_feature_name = sanitize_filename(feature).strip().replace(' ', '_')
            single_path = os.path.join(single_pdp_dir, f'pdp_univariate_{safe_feature_name}.png')
            single_fig.tight_layout()
            single_fig.savefig(single_path, dpi=300, bbox_inches='tight')
            plt.close(single_fig)
        except Exception as e:
            print(f"Warning: Could not create PDP for {feature}: {e}")
            ax.text(0.5, 0.5, f'Error creating PDP', 
                   ha='center', va='center', transform=ax.transAxes)
            ax.axis('off')
    
    # Hide the extra subplots
    for idx in range(n_top_features, len(axes)):
        axes[idx].axis('off')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'pdp_analysis.png'), dpi=300, bbox_inches='tight')
    # plt.show()  # Show the single variable PDP plot
    plt.close()
    
    # Save PDP data to the existing model_analysis_results.xlsx file (append mode）
    print(f"\nPreparing to save single variable PDP data to Excel...")
    print(f"  Number of successfully calculated features: {len(excel_data_dict)}")
    if len(excel_data_dict) > 0:
        print(f"  Feature list: {list(excel_data_dict.keys())}")
    
    if excel_data_dict:
        excel_path = os.path.join(save_dir, 'model_analysis_results.xlsx')
        try:
            # Check if the file exists, if it exists, use the append mode, otherwise create a new file
            file_exists = os.path.exists(excel_path)
            mode = 'a' if file_exists else 'w'
            
            print(f"  Excel file path: {excel_path}")
            print(f"  File exists: {file_exists}")
            print(f"  Save mode: {mode}")
            
            saved_sheets = []
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
                # Create a worksheet for each feature
                for feature_name, data_df in excel_data_dict.items():
                    try:
                        # The worksheet name cannot exceed 31 characters, and cannot contain special characters
                        # Add the prefix "PDP_" to avoid conflicts with existing worksheets
                        sheet_name = f'PDP_{feature_name}'
                        if len(sheet_name) > 31:
                            sheet_name = f'PDP_{feature_name[:27]}'
                        # Replace the characters Excel does not support
                        sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                        
                        print(f"  Saving worksheet: {sheet_name} (feature: {feature_name})")
                        print(f"    Data shape: {data_df.shape}")
                        print(f"    Column names: {list(data_df.columns)}")
                        
                        data_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        saved_sheets.append(sheet_name)
                        print(f"    ✓ Successfully saved: {sheet_name}")
                    except Exception as e:
                        print(f"    ⚠ Error saving feature {feature_name}: {e}")
                        import traceback
                        traceback.print_exc()
            
            if file_exists:
                print(f"\n✓ Single variable PDP data has been appended to the existing Excel file: {excel_path}")
            else:
                print(f"\n✓ Single variable PDP data has been saved to the new Excel file: {excel_path}")
            print(f"  Successfully saved {len(saved_sheets)} worksheets: {saved_sheets}")
            print(f"  Each worksheet contains the following columns:")
            print(f"    - Grid_Values_Original: Original grid points")
            print(f"    - Grid_Values_Dense: Dense interpolation points (for smoothing the curve)")
            print(f"    - PDP_Original: Original PDP value")
            print(f"    - PDP_Smooth: Smoothed PDP value")
            print(f"    - Mean_cICE_Original: Original Mean c-ICE value")
            print(f"    - Mean_cICE_Smooth: Smoothed Mean c-ICE value")
            print(f"    - ALE_Original: Original ALE value")
            print(f"    - ALE_Smooth: Smoothed ALE value")
            print(f"    - ICE_Upper_CI: ICE upper confidence interval (95%)")
            print(f"    - ICE_Lower_CI: ICE lower confidence interval (95%)")
        except Exception as e:
            print(f"\n⚠ Error saving single variable PDP data to Excel: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"\n⚠ Warning: No single variable PDP data has been successfully calculated, cannot save to Excel")
        print(f"  Possible reasons:")
        print(f"    1. All PDP calculations for features have failed")
        print(f"    2. n_top_features is 0 or negative")
        print(f"    3. Data format problem causing calculation failure")
    
    print(f"\nSingle variable PDP analysis completed, the dependency plots of the top {n_top_features} important features have been drawn")
    
    # 2. Double variable PDP analysis (feature interaction）
    if n_top_features >= 2:
        print("\n=== Double variable PDP analysis ===")
        
        # Select the top 4 most important features for double variable interaction analysis
        n_interaction_features = min(4, n_top_features)
        interaction_features = top_features[:n_interaction_features]
        
        # Create interaction feature pairs
        interaction_pairs = []
        for i in range(len(interaction_features)):
            for j in range(i+1, len(interaction_features)):
                interaction_pairs.append((interaction_features[i], interaction_features[j]))
        
        # Dictionary for saving 2D PDP data
        excel_2d_data_dict = {}
        
        # Plot double variable PDP (using 3D surface plot）
        from mpl_toolkits.mplot3d import Axes3D
        
        n_pairs = len(interaction_pairs)
        n_cols = 3  # 3 columns
        n_rows = (n_pairs + 2) // 3  # Adjust the number of rows calculation
        fig = plt.figure(figsize=(18, 6*n_rows))

        pdp2d_dir = os.path.join(save_dir, 'pdp_2d_single')
        os.makedirs(pdp2d_dir, exist_ok=True)
        
        for idx, (feat1, feat2) in enumerate(interaction_pairs):
            try:
                # Get the feature index
                feat1_idx = feature_names.index(feat1)
                feat2_idx = feature_names.index(feat2)
                
                # Calculate double variable partial dependence - increase grid_resolution to make the grid denser
                pd_result = partial_dependence(
                    model,
                    X_train_df,
                    features=[feat1_idx, feat2_idx],
                    kind='average',
                    grid_resolution=80  # Further encrypt the grid points, improve the surface smoothness
                )
                
                # Get the grid values and average predicted values
                grid_values_1 = pd_result['grid_values'][0]
                grid_values_2 = pd_result['grid_values'][1]
                average = pd_result['average'][0]  # Take the first element (if it is a 3D array）
                
                # Ensure average is 2D
                if average.ndim == 3:
                    average = average[0]
                
                # Create the original grid
                X1, X2 = np.meshgrid(grid_values_1, grid_values_2)
                
                # Ensure average matches the grid dimensions
                # meshgrid returns the shape (len(grid_values_2), len(grid_values_1))
                if average.shape != X1.shape and average.shape == (len(grid_values_1), len(grid_values_2)):
                        average = average.T
                
                # Use a more dense interpolation grid, significantly improve the visualization density
                dense_resolution = 150
                grid_dense_1 = np.linspace(grid_values_1.min(), grid_values_1.max(), dense_resolution)
                grid_dense_2 = np.linspace(grid_values_2.min(), grid_values_2.max(), dense_resolution)
                try:
                    interpolator = RegularGridInterpolator(
                        (grid_values_2, grid_values_1), average,
                        bounds_error=False, fill_value=None
                    )
                    X1_dense, X2_dense = np.meshgrid(grid_dense_1, grid_dense_2)
                    dense_points = np.stack([X2_dense.ravel(), X1_dense.ravel()], axis=-1)
                    average_dense = interpolator(dense_points).reshape(X1_dense.shape)
                except Exception:
                    # RegularGridInterpolator may fail (for example, a dimension is approximately constant), try using griddata for 2D interpolation
                    try:
                        X1_dense, X2_dense = np.meshgrid(grid_dense_1, grid_dense_2)
                        original_points = np.column_stack([X1.flatten(), X2.flatten()])
                        target_points = np.column_stack([X1_dense.flatten(), X2_dense.flatten()])
                        average_dense = griddata(original_points, average.flatten(), target_points, method='cubic')
                        # If cubic produces NaN, fall back to linear; if not, use nearest
                        if np.isnan(average_dense).any():
                            average_dense = griddata(original_points, average.flatten(), target_points, method='linear')
                        if np.isnan(average_dense).any():
                            average_dense = griddata(original_points, average.flatten(), target_points, method='nearest')
                        average_dense = average_dense.reshape(X1_dense.shape)
                    except Exception:
                        # Still failed, fall back to the original grid
                        X1_dense, X2_dense, average_dense = X1, X2, average
                
                # Save 2D PDP data to the dictionary (using the encrypted grid, if available）
                grid1_flat = X1_dense.flatten()
                grid2_flat = X2_dense.flatten()
                average_flat = average_dense.flatten()
                
                # Create DataFrame
                pair_name = f'{feat1}_vs_{feat2}'
                data_df = pd.DataFrame({
                    f'{feat1}': grid1_flat,
                    f'{feat2}': grid2_flat,
                    'Partial_Dependence': average_flat
                })
                
                # Ensure the data length is consistent
                if len(grid1_flat) == len(grid2_flat) == len(average_flat):
                    excel_2d_data_dict[pair_name] = data_df
                    print(f"  ✓ Successfully calculated and saved the data for the feature pair {pair_name} (shape: {data_df.shape})")
                else:
                    print(f"  ⚠ Warning: The data length for the feature pair {pair_name} is inconsistent, skipping save")
                    print(f"    grid1_flat length: {len(grid1_flat)}, grid2_flat length: {len(grid2_flat)}, average_flat length: {len(average_flat)}")
                
                # Create a 3D subplot
                ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                
                # Plot the 3D surface plot (using the encrypted grid）
                surf = ax.plot_surface(X1_dense, X2_dense, average_dense, cmap='viridis', alpha=0.9, 
                                       linewidth=0, antialiased=True, edgecolor='none')
                
                # Set the labels and title
                ax.set_xlabel(feat1, fontsize=10, labelpad=8)
                ax.set_ylabel(feat2, fontsize=10, labelpad=8)
                ax.set_zlabel('Partial Dependence', fontsize=10, labelpad=8)
                ax.set_title(f'Interaction: {feat1} vs {feat2}', fontsize=12, fontweight='bold', pad=15)
                
                # Add the color bar
                fig.colorbar(surf, ax=ax, shrink=0.6, aspect=20, pad=0.1)

                # Save the single 3D interaction plot
                single_fig = plt.figure(figsize=(8, 6))
                single_ax = single_fig.add_subplot(111, projection='3d')
                single_surf = single_ax.plot_surface(X1_dense, X2_dense, average_dense, cmap='viridis', alpha=0.9,
                                                     linewidth=0, antialiased=True, edgecolor='none')
                single_ax.set_xlabel(feat1, fontsize=11, labelpad=10)
                single_ax.set_ylabel(feat2, fontsize=11, labelpad=10)
                single_ax.set_zlabel('Partial Dependence', fontsize=11, labelpad=10)
                single_ax.set_title(f'Interaction: {feat1} vs {feat2}', fontsize=13, fontweight='bold', pad=18)
                single_ax.view_init(elev=25, azim=135)
                single_fig.colorbar(single_surf, shrink=0.6, aspect=20, pad=0.12)
                safe_pair_name = sanitize_filename(pair_name).replace(' ', '_')
                single_path = os.path.join(pdp2d_dir, f'pdp_2d_{safe_pair_name}.png')
                single_fig.tight_layout()
                single_fig.savefig(single_path, dpi=300, bbox_inches='tight')
                plt.close(single_fig)
                
            except Exception as e:
                print(f"Warning: Could not create 2D PDP for {feat1} vs {feat2}: {e}")
                import traceback
                traceback.print_exc()
                # Create an empty 3D subplot to display the error information
                ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                ax.text(0.5, 0.5, 0.5, f'Error creating 2D PDP', 
                       ha='center', va='center', transform=ax.transAxes)
                ax.set_axis_off()
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, 'pdp_2d_interaction.png'), dpi=300, bbox_inches='tight')
        # plt.show()  # Show the 3D plot
        plt.close()
        
        # Save 2D PDP data to the existing model_analysis_results.xlsx file
        print(f"\nPreparing to save 2D PDP data...")
        print(f"  Number of successfully calculated feature pairs: {len(excel_2d_data_dict)}")
        if len(excel_2d_data_dict) > 0:
            print(f"  Feature pair list: {list(excel_2d_data_dict.keys())}")
            for pair_name, data_df in excel_2d_data_dict.items():
                print(f"    - {pair_name}: shape {data_df.shape}, columns {list(data_df.columns)}")
        else:
            print(f"  ⚠ Warning: excel_2d_data_dict is empty, no data to save")
        
        if excel_2d_data_dict and len(excel_2d_data_dict) > 0:
            excel_path = os.path.join(save_dir, 'model_analysis_results.xlsx')
            try:
                # Check if the file exists, if it exists, use the append mode, otherwise create a new file
                file_exists = os.path.exists(excel_path)
                mode = 'a' if file_exists else 'w'
                
                print(f"  Excel file path: {excel_path}")
                print(f"  File exists: {file_exists}")
                print(f"  Save mode: {mode}")
                
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
                    # Create a worksheet for each feature pair
                    saved_sheets = []
                    failed_sheets = []
                    
                    print(f"  Starting to save the data for {len(excel_2d_data_dict)} feature pairs...")
                    for idx, (pair_name, data_df) in enumerate(excel_2d_data_dict.items(), 1):
                        try:
                            # The worksheet name cannot exceed 31 characters, and cannot contain special characters
                            # Add the prefix "PDP2D_" to avoid conflicts with existing worksheets
                            sheet_name = f'PDP2D_{pair_name}'
                            if len(sheet_name) > 31:
                                sheet_name = f'PDP2D_{pair_name[:25]}'
                            # Replace the characters Excel does not support
                            sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                            
                            print(f"  [{idx}/{len(excel_2d_data_dict)}] Saving worksheet: {sheet_name} (feature pair: {pair_name})")
                            print(f"    Data shape: {data_df.shape}")
                            print(f"    Column names: {list(data_df.columns)}")
                            
                            # Check if the data is empty
                            if data_df.empty:
                                print(f"    ⚠ Warning: The data for the feature pair {pair_name} is empty, skipping save")
                                failed_sheets.append((pair_name, "Empty data"))
                                continue
                            
                            # Save to Excel
                            data_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            saved_sheets.append(sheet_name)
                            print(f"    ✓ Successfully saved: {sheet_name}")
                        except Exception as e:
                            error_msg = str(e)
                            print(f"    ⚠ Error saving the feature pair {pair_name}: {error_msg}")
                            failed_sheets.append((pair_name, error_msg))
                            import traceback
                            traceback.print_exc()
                            # Continue to process the next feature pair, do not interrupt the loop
                            continue
                    
                    # Print the save result summary
                    print(f"\n  Save result summary:")
                    print(f"    ✓ Successfully saved: {len(saved_sheets)} worksheets")
                    if saved_sheets:
                        print(f"      Successfully saved worksheets: {saved_sheets}")
                    if failed_sheets:
                        print(f"    ✗ Failed to save: {len(failed_sheets)} worksheets")
                        for pair_name, error in failed_sheets:
                            print(f"      - {pair_name}: {error}")
                
                if file_exists:
                    print(f"\n✓ 2D PDP data has been appended to the existing Excel file: {excel_path}")
                else:
                    print(f"\n✓ 2D PDP data has been saved to the new Excel file: {excel_path}")
                print(f"  Successfully saved {len(saved_sheets)} worksheets: {saved_sheets}")
                print(f"  Each worksheet contains the following columns:")
                print(f"    - First feature name: The grid values of the first feature")
                print(f"    - Second feature name: The grid values of the second feature")
                print(f"    - Partial_Dependence: The corresponding partial dependence values")
            except Exception as e:
                print(f"\n⚠ Error saving 2D PDP data to Excel: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"\n⚠ Warning: No 2D PDP data has been successfully calculated, cannot save to Excel")
            print(f"  Possible reasons:")
            print(f"    1. All feature pairs calculations have failed")
            print(f"    2. The number of feature pairs is 0 (n_interaction_features < 2）")
            print(f"    3. Data format problem causing calculation failure")
        
        print(f"\nDouble variable PDP analysis completed, the interaction plots of the top {n_interaction_features} important features have been drawn")

def analyze_noise_robustness(model, X_test, y_test, feature_names, save_dir):
    """Analyze the robustness of the model to noise and outliers"""
    print("\n=== Analyze the robustness of the model ===")
    
    # 1. Add Gaussian noise test
    print("Testing the robustness of the model to Gaussian noise...")
    # Use 0, 2, 4, 6, 8, 10 (%) as noise levels
    noise_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Noise levels (percentage of feature standard deviation)
    noise_rmses = []
    noise_stds = []
    
    for noise_level in noise_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times to get the standard deviation
            # Add Gaussian noise
            X_test_noisy = X_test.copy()
            for i in range(X_test.shape[1]):
                feature_std = np.std(X_test[:, i])
                noise = np.random.normal(0, noise_level * feature_std, size=len(X_test))
                X_test_noisy[:, i] = X_test[:, i] + noise
            
            # Predict
            y_pred_noisy = model.predict(X_test_noisy)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred_noisy))
            rmses.append(rmse)
        
        noise_rmses.append(np.mean(rmses))
        noise_stds.append(np.std(rmses))
    
    # 2. Add outlier test
    print("Testing the robustness of the model to outliers...")
    # Use 0, 2, 4, 6, 8, 10 (%) as outlier比例
    outlier_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Outlier比例
    outlier_severity = 5.0  # Outlier severity (multiple of standard deviation)
    outlier_rmses = []
    outlier_stds = []
    
    for outlier_pct in outlier_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times
            # Add outliers
            X_test_outlier = X_test.copy()
            n_outliers = int(len(X_test) * outlier_pct)
            if n_outliers > 0:
                outlier_indices = np.random.choice(len(X_test), n_outliers, replace=False)
                for idx in outlier_indices:
                    # Randomly select a feature to add outliers
                    feature_idx = np.random.randint(X_test.shape[1])
                    feature_mean = np.mean(X_test[:, feature_idx])
                    feature_std = np.std(X_test[:, feature_idx])
                    X_test_outlier[idx, feature_idx] = feature_mean + outlier_severity * feature_std
            
            # Predict
            y_pred_outlier = model.predict(X_test_outlier)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred_outlier))
            rmses.append(rmse)
        
        outlier_rmses.append(np.mean(rmses))
        outlier_stds.append(np.std(rmses))
    
    # 3. Plot the robustness analysis graph
    fig, axes = plt.subplots(1, 2, figsize=(15, 6))
    
    # Gaussian noise robustness
    axes[0].errorbar([l * 100 for l in noise_levels], noise_rmses, yerr=noise_stds, 
                     marker='o', capsize=5, capthick=2, linewidth=2, color='blue')
    axes[0].set_xlabel('Noise level (% of feature std)', fontsize=12)
    axes[0].set_ylabel('RMSE (MPa)', fontsize=12)
    axes[0].set_title('Robustness to Additive Gaussian Noise', fontsize=14, fontweight='bold')
    axes[0].grid(True, alpha=0.3)
    
    # Outlier robustness
    axes[1].errorbar([l * 100 for l in outlier_levels], outlier_rmses, yerr=outlier_stds, 
                     marker='o', capsize=5, capthick=2, linewidth=2, color='orange')
    axes[1].set_xlabel('Injected outliers (% of samples)', fontsize=12)
    axes[1].set_ylabel('RMSE (MPa)', fontsize=12)
    axes[1].set_title(f'Robustness to Outliers (severity={outlier_severity}×std)', fontsize=14, fontweight='bold')
    axes[1].grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'robustness_analysis.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # Save results (will be saved in the main function)
    robustness_df = pd.DataFrame({
        'Noise Level (%)': [l * 100 for l in noise_levels],
        'RMSE (MPa)': noise_rmses,
        'Std': noise_stds
    })
    
    outlier_df = pd.DataFrame({
        'Outlier Percentage (%)': [l * 100 for l in outlier_levels],
        'RMSE (MPa)': outlier_rmses,
        'Std': outlier_stds
    })
    
    # 打印结果摘要
    print("\n=== Robustness analysis result summary ===")
    print("\n1. Gaussian noise robustness:")
    print(f"   - 0% noise RMSE: {noise_rmses[0]:.3f} MPa")
    print(f"   - 2% noise RMSE: {noise_rmses[1]:.3f} MPa (increase {noise_rmses[1]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 4% noise RMSE: {noise_rmses[2]:.3f} MPa (increase {noise_rmses[2]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 6% noise RMSE: {noise_rmses[3]:.3f} MPa (increase {noise_rmses[3]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 8% noise RMSE: {noise_rmses[4]:.3f} MPa (increase {noise_rmses[4]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 10% noise RMSE: {noise_rmses[5]:.3f} MPa (increase {noise_rmses[5]-noise_rmses[0]:.3f} MPa)")
    
    print("\n2. Outlier robustness:")
    print(f"   - 0% outlier RMSE: {outlier_rmses[0]:.3f} MPa")
    print(f"   - 2% outlier RMSE: {outlier_rmses[1]:.3f} MPa (change {outlier_rmses[1]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 4% outlier RMSE: {outlier_rmses[2]:.3f} MPa (change {outlier_rmses[2]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 6% outlier RMSE: {outlier_rmses[3]:.3f} MPa (change {outlier_rmses[3]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 8% outlier RMSE: {outlier_rmses[4]:.3f} MPa (change {outlier_rmses[4]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 10% outlier RMSE: {outlier_rmses[5]:.3f} MPa (change {outlier_rmses[5]-outlier_rmses[0]:.3f} MPa)")
    
    # Calculate performance degradation rate
    noise_degradation = (noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100
    outlier_degradation = (outlier_rmses[-1] - outlier_rmses[0]) / outlier_rmses[0] * 100
    
    print("\n3. Performance degradation rate (10% noise/outlier vs no perturbation):")
    print(f"   - Gaussian noise performance degradation: {noise_degradation:.1f}%")
    print(f"   - Outlier performance degradation: {outlier_degradation:.1f}%")
    
    # Determine robustness
    if noise_degradation < 20:
        noise_robust = "Excellent"
    elif noise_degradation < 40:
        noise_robust = "Good"
    elif noise_degradation < 60:
        noise_robust = "Average"
    else:
        noise_robust = "Poor"
    
    if abs(outlier_degradation) < 5:
        outlier_robust = "Excellent"
    elif abs(outlier_degradation) < 10:
        outlier_robust = "Good"
    elif abs(outlier_degradation) < 20:
        outlier_robust = "Average"
    else:
        outlier_robust = "Poor"
    
    print("\n4. Robustness evaluation:")
    print(f"   - Robustness to Gaussian noise: {noise_robust} (10% noise level时 performance degradation{noise_degradation:.1f}%)")
    print(f"   - Robustness to outliers: {outlier_robust} (10% outlier percentage时 performance change{outlier_degradation:.1f}%)")
    
    print(f"\nRobustness analysis completed, detailed results have been saved")
    return noise_rmses, outlier_rmses, robustness_df, outlier_df

def plot_radar_chart(y_true, y_pred, noise_rmses, outlier_rmses, bootstrap_coverage, 
                     quantile_coverage, save_dir, normalize=True, training_metrics_path=None):
    """Plot radar chart to show model performance indicators, and save data to Excel
    
    Parameters:
        y_true: True values
        y_pred: Predicted values
        noise_rmses: Noise robustness RMSE list
        outlier_rmses: Outlier robustness RMSE list
        bootstrap_coverage: Bootstrap prediction interval coverage rate
        quantile_coverage: Quantile Regression prediction interval coverage rate
        save_dir: Save directory
        normalize: Whether to normalize indicators to 0-1 range (for radar chart display)
        training_metrics_path: Training metrics file path (for reading three-fold average and final model metrics)
    
    Returns:
        radar_data_df: DataFrame containing radar chart data
        metrics_table_df: DataFrame containing three-fold average and final model metrics
    """
    print("\n=== Plot radar chart ===")
    
    os.makedirs(save_dir, exist_ok=True)
    
    # Try to read three-fold average and final model metrics from training metrics file
    mean_train_metrics = None
    mean_test_metrics = None
    final_train_metrics = None
    final_test_metrics = None
    
    if training_metrics_path is None:
        training_metrics_path = os.path.join(save_dir, 'training_metrics.xlsx')
    
    if os.path.exists(training_metrics_path):
        try:
            print(f"Reading training metrics from {training_metrics_path}...")
            # Read each Fold metrics
            df_folds = pd.read_excel(training_metrics_path, sheet_name='Each Fold metrics')
            # Read final model metrics
            df_final = pd.read_excel(training_metrics_path, sheet_name='Final model metrics')
            
            # Try to read three-fold average metrics (if new worksheet is saved)
            try:
                df_avg = pd.read_excel(training_metrics_path, sheet_name='Three-fold average metrics')
                if len(df_avg) > 0:
                    train_row = df_avg[df_avg['Dataset'].str.contains('Training set', na=False)]
                    test_row = df_avg[df_avg['Dataset'].str.contains('Test set', na=False)]
                    
                    if len(train_row) > 0 and len(test_row) > 0:
                        mean_train_metrics = {
                            'R²': train_row['R²'].values[0] if 'R²' in train_row.columns else np.nan,
                            'RMSE': train_row['RMSE (MPa)'].values[0] if 'RMSE (MPa)' in train_row.columns else np.nan,
                            'MSE': train_row['MSE (MPa²)'].values[0] if 'MSE (MPa²)' in train_row.columns else np.nan,
                            'MAE': train_row['MAE (MPa)'].values[0] if 'MAE (MPa)' in train_row.columns else np.nan,
                            'MAPE': train_row['MAPE (%)'].values[0] if 'MAPE (%)' in train_row.columns else np.nan
                        }
                        mean_test_metrics = {
                            'R²': test_row['R²'].values[0] if 'R²' in test_row.columns else np.nan,
                            'RMSE': test_row['RMSE (MPa)'].values[0] if 'RMSE (MPa)' in test_row.columns else np.nan,
                            'MSE': test_row['MSE (MPa²)'].values[0] if 'MSE (MPa²)' in test_row.columns else np.nan,
                            'MAE': test_row['MAE (MPa)'].values[0] if 'MAE (MPa)' in test_row.columns else np.nan,
                            'MAPE': test_row['MAPE (%)'].values[0] if 'MAPE (%)' in test_row.columns else np.nan
                        }
                        print("✓ Three-fold average metrics (R², RMSE, MSE, MAE, MAPE) have been read")
                        print(f"  Training set: R²={mean_train_metrics['R²']:.5f}, RMSE={mean_train_metrics['RMSE']:.5f}, "
                              f"MSE={mean_train_metrics['MSE']:.5f}, MAE={mean_train_metrics['MAE']:.5f}, "
                              f"MAPE={mean_train_metrics['MAPE']:.5f}")
                        print(f"  Test set: R²={mean_test_metrics['R²']:.5f}, RMSE={mean_test_metrics['RMSE']:.5f}, "
                              f"MSE={mean_test_metrics['MSE']:.5f}, MAE={mean_test_metrics['MAE']:.5f}, "
                              f"MAPE={mean_test_metrics['MAPE']:.5f}")
                    else:
                        print("⚠ Warning: Three-fold average metrics worksheet did not find training set or test set row, will be calculated from each Fold metrics")
                        raise ValueError("Training set or test set row not found")
            except Exception as e:
                print(f"⚠ Unable to read from 'Three-fold average metrics' worksheet: {e}")
                print("  Will try to calculate three-fold average metrics from 'Each Fold metrics' worksheet...")
            
            # If three-fold average metrics are still not read successfully, calculate from each Fold metrics
            if mean_train_metrics is None or mean_test_metrics is None:
                if len(df_folds) > 0:
                    # Print all column names, for debugging
                    print(f"  Each Fold metrics worksheet column names: {list(df_folds.columns)}")
                    
                    # Extract training set metrics (using more flexible matching)
                    train_r2_col = [col for col in df_folds.columns if 'Training set' in str(col) and ('R²' in str(col) or 'R2' in str(col))]
                    train_rmse_col = [col for col in df_folds.columns if 'Training set' in str(col) and 'RMSE' in str(col)]
                    train_mse_col = [col for col in df_folds.columns if 'Training set' in str(col) and 'MSE' in str(col) and 'RMSE' not in str(col)]
                    train_mae_col = [col for col in df_folds.columns if 'Training set' in str(col) and 'MAE' in str(col)]
                    train_mape_col = [col for col in df_folds.columns if 'Training set' in str
                    # Extract test set metrics (using more flexible matching)
                    test_r2_col = [col for col in df_folds.columns if 'Test set' in str(col) and ('R²' in str(col) or 'R2' in str(col))]
                    test_rmse_col = [col for col in df_folds.columns if 'Test set' in str(col) and 'RMSE' in str(col)]
                    test_mse_col = [col for col in df_folds.columns if 'Test set' in str(col) and 'MSE' in str(col) and 'RMSE' not in str(col)]
                    test_mae_col = [col for col in df_folds.columns if 'Test set' in str(col) and 'MAE' in str(col)]
                    test_mape_col = [col for col in df_folds.columns if 'Test set' in str(col) and 'MAPE' in str(col)]
                    
                    # Ensure all columns are found, if not found set to NaN
                    mean_train_metrics = {
                        'R²': df_folds[train_r2_col[0]].mean() if train_r2_col and len(train_r2_col) > 0 else np.nan,
                        'RMSE': df_folds[train_rmse_col[0]].mean() if train_rmse_col and len(train_rmse_col) > 0 else np.nan,
                        'MSE': df_folds[train_mse_col[0]].mean() if train_mse_col and len(train_mse_col) > 0 else np.nan,
                        'MAE': df_folds[train_mae_col[0]].mean() if train_mae_col and len(train_mae_col) > 0 else np.nan,
                        'MAPE': df_folds[train_mape_col[0]].mean() if train_mape_col and len(train_mape_col) > 0 else np.nan
                    }
                    mean_test_metrics = {
                        'R²': df_folds[test_r2_col[0]].mean() if test_r2_col and len(test_r2_col) > 0 else np.nan,
                        'RMSE': df_folds[test_rmse_col[0]].mean() if test_rmse_col and len(test_rmse_col) > 0 else np.nan,
                        'MSE': df_folds[test_mse_col[0]].mean() if test_mse_col and len(test_mse_col) > 0 else np.nan,
                        'MAE': df_folds[test_mae_col[0]].mean() if test_mae_col and len(test_mae_col) > 0 else np.nan,
                        'MAPE': df_folds[test_mape_col[0]].mean() if test_mape_col and len(test_mape_col) > 0 else np.nan
                    }
                    
                    # 打印找到的列名，用于调试
                    print(f"  Found training set columns: R²={train_r2_col[0] if train_r2_col else 'not found'}, "
                          f"RMSE={train_rmse_col[0] if train_rmse_col else 'not found'}, "
                          f"MSE={train_mse_col[0] if train_mse_col else 'not found'}, "
                          f"MAE={train_mae_col[0] if train_mae_col else 'not found'}, "
                          f"MAPE={train_mape_col[0] if train_mape_col else 'not found'}")
                    print(f"  Found test set columns: R²={test_r2_col[0] if test_r2_col else 'not found'}, "
                          f"RMSE={test_rmse_col[0] if test_rmse_col else 'not found'}, "
                          f"MSE={test_mse_col[0] if test_mse_col else 'not found'}, "
                          f"MAE={test_mae_col[0] if test_mae_col else 'not found'}, "
                          f"MAPE={test_mape_col[0] if test_mape_col else 'not found'}")
                    
                    print("✓ Three-fold average metrics have been calculated from each Fold metrics")
            
            # Read final model metrics
            if len(df_final) > 0:
                # Find training set and test set rows
                train_row = df_final[df_final['Dataset'].str.contains('Training set', na=False)]
                test_row = df_final[df_final['Dataset'].str.contains('Test set', na=False)]
                
                if len(train_row) > 0 and len(test_row) > 0:
                    final_train_metrics = {
                        'R²': train_row['R²'].values[0] if 'R²' in train_row.columns else np.nan,
                        'RMSE': train_row['RMSE (MPa)'].values[0] if 'RMSE (MPa)' in train_row.columns else np.nan,
                        'MSE': train_row['MSE (MPa²)'].values[0] if 'MSE (MPa²)' in train_row.columns else np.nan,
                        'MAE': train_row['MAE (MPa)'].values[0] if 'MAE (MPa)' in train_row.columns else np.nan,
                        'MAPE': train_row['MAPE (%)'].values[0] if 'MAPE (%)' in train_row.columns else np.nan
                    }
                    final_test_metrics = {
                        'R²': test_row['R²'].values[0] if 'R²' in test_row.columns else np.nan,
                        'RMSE': test_row['RMSE (MPa)'].values[0] if 'RMSE (MPa)' in test_row.columns else np.nan,
                        'MSE': test_row['MSE (MPa²)'].values[0] if 'MSE (MPa²)' in test_row.columns else np.nan,
                        'MAE': test_row['MAE (MPa)'].values[0] if 'MAE (MPa)' in test_row.columns else np.nan,
                        'MAPE': test_row['MAPE (%)'].values[0] if 'MAPE (%)' in test_row.columns else np.nan
                    }
                    print("✓ Final model metrics (R², RMSE, MSE, MAE, MAPE) have been read")
                    
                    # Check if there are missing metrics
                    missing_train = [k for k, v in final_train_metrics.items() if np.isnan(v)]
                    missing_test = [k for k, v in final_test_metrics.items() if np.isnan(v)]
                    if missing_train or missing_test:
                        print(f"  ⚠ Warning: Final model metrics have missing values!")
                        if missing_train:
                            print(f"    Training set missing metrics: {', '.join(missing_train)}")
                        if missing_test:
                            print(f"    Test set missing metrics: {', '.join(missing_test)}")
                        print(f"    Suggestion: Re-run XGBoost_train.py to generate a complete file with all metrics")
        except Exception as e:
            print(f"⚠ Error reading training metrics file: {e}")
            import traceback
            traceback.print_exc()
    
    # After reading all metrics, check and give a prompt
    if training_metrics_path and os.path.exists(training_metrics_path):
        has_missing = False
        if mean_train_metrics:
            missing = [k for k, v in mean_train_metrics.items() if np.isnan(v)]
            if missing:
                has_missing = True
                print(f"  ⚠ Three-fold average training set metrics missing: {', '.join(missing)}")
        if mean_test_metrics:
            missing = [k for k, v in mean_test_metrics.items() if np.isnan(v)]
            if missing:
                has_missing = True
                print(f"  ⚠ Three-fold average test set metrics missing: {', '.join(missing)}")
        if final_train_metrics:
            missing = [k for k, v in final_train_metrics.items() if np.isnan(v)]
            if missing:
                has_missing = True
                print(f"  ⚠ Final model training set metrics missing: {', '.join(missing)}")
        if final_test_metrics:
            missing = [k for k, v in final_test_metrics.items() if np.isnan(v)]
            if missing:
                has_missing = True
                print(f"  ⚠ Final model test set metrics missing: {', '.join(missing)}")
        
        if has_missing:
            print(f"\n{'='*70}")
            print("Important提示：Detected missing performance indicators in training_metrics.xlsx file!")
            print("This may be because the file is an old version generated, missing some columns (such as MSE, MAPE).")
            print("Solution:")
            print(f"  1. Run XGBoost_train.py to re-train the model and generate a complete metrics file")
            print(f"  2. The new file will contain all five indicators: R², RMSE, MSE, MAE, MAPE")
            print(f"  3. Then run this noise analysis script again")
            print(f"{'='*70}\n")
    
    # Calculate current test set performance indicators (as a supplement)
    r2 = r2_score(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    mse = mean_squared_error(y_true, y_pred)
    rmse = np.sqrt(mse)
    mape = mean_absolute_percentage_error(y_true, y_pred)
    
    # If final model metrics are incomplete, use the current calculated metrics as a supplement
    if final_test_metrics is None:
        final_test_metrics = {
            'R²': r2,
            'RMSE': rmse,
            'MAE': mae,
            'MSE': mse,
            'MAPE': mape
        }
    else:
        # Supplement missing metrics (if some metrics are NaN, try to calculate or use the current value)
        if np.isnan(final_test_metrics['MSE']):
            final_test_metrics['MSE'] = final_test_metrics['RMSE'] ** 2 if not np.isnan(final_test_metrics['RMSE']) else mse
        if np.isnan(final_test_metrics['MAPE']):
            final_test_metrics['MAPE'] = mape
    
    if final_train_metrics is None:
        final_train_metrics = {
            'R²': np.nan,
            'RMSE': np.nan,
            'MAE': np.nan,
            'MSE': np.nan,
            'MAPE': np.nan
        }
    else:
        # Supplement missing metrics
        if np.isnan(final_train_metrics['MSE']):
            final_train_metrics['MSE'] = final_train_metrics['RMSE'] ** 2 if not np.isnan(final_train_metrics['RMSE']) else np.nan
    
    # If three-fold average metrics are incomplete, set to NaN
    if mean_train_metrics is None:
        mean_train_metrics = {
            'R²': np.nan,
            'RMSE': np.nan,
            'MAE': np.nan,
            'MSE': np.nan,
            'MAPE': np.nan
        }
    else:
        # Supplement missing metrics
        if np.isnan(mean_train_metrics['MSE']):
            mean_train_metrics['MSE'] = mean_train_metrics['RMSE'] ** 2 if not np.isnan(mean_train_metrics['RMSE']) else np.nan
    
    if mean_test_metrics is None:
        mean_test_metrics = {
            'R²': np.nan,
            'RMSE': np.nan,
            'MAE': np.nan,
            'MSE': np.nan,
            'MAPE': np.nan
        }
    else:
        # Supplement missing metrics
        if np.isnan(mean_test_metrics['MSE']):
            mean_test_metrics['MSE'] = mean_test_metrics['RMSE'] ** 2 if not np.isnan(mean_test_metrics['RMSE']) else np.nan
    
    # Calculate noise robustness score (based on performance degradation rate)
    if noise_rmses is not None and len(noise_rmses) > 0:
        noise_degradation = (noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100 if noise_rmses[0] > 0 else 0
        # Robustness score: Performance degradation rate越小越好，转换为0-1分数（退化率20%以下为优秀=1.0，100%以上为0）
        noise_robustness_score = max(0, min(1, 1 - noise_degradation / 100))
    else:
        noise_degradation = np.nan
        noise_robustness_score = np.nan
    
    # Calculate outlier robustness score
    if outlier_rmses is not None and len(outlier_rmses) > 0:
        outlier_degradation = (outlier_rmses[-1] - outlier_rmses[0]) / outlier_rmses[0] * 100 if outlier_rmses[0] > 0 else 0
        outlier_robustness_score = max(0, min(1, 1 - abs(outlier_degradation) / 100))
    else:
        outlier_degradation = np.nan
        outlier_robustness_score = np.nan
    
    # Prediction interval coverage (directly used, already in 0-1 range)
    avg_coverage = (bootstrap_coverage + quantile_coverage) / 2 if (bootstrap_coverage is not None and quantile_coverage is not None) else np.nan
    
    # Calculate normalized scores (for radar chart)
    # For RMSE, MAE, MAPE, normalization is needed (value is better smaller, convert to score)
    # Use relative minimum as reference (here use the maximum value of the test set as reference)
    y_max = np.max(np.abs(y_true))
    
    # RMSE normalized score (RMSE is better smaller, convert to 0-1 score)
    rmse_max_ref = y_max * 0.5  # Assume maximum RMSE is 50% of the maximum true value
    rmse_normalized = max(0, min(1, 1 - rmse / rmse_max_ref))
    
    # MAE normalized score
    mae_max_ref = y_max * 0.4
    mae_normalized = max(0, min(1, 1 - mae / mae_max_ref))
    
    # MAPE normalized score (MAPE is better smaller, assume maximum MAPE is 50%)
    mape_max_ref = 50.0
    mape_normalized = max(0, min(1, 1 - mape / mape_max_ref)) if not np.isnan(mape) else np.nan
    
    # Prepare radar chart data
    categories = ['R²', 'RMSE\nScore', 'MAE\nScore', 'MAPE\nScore', 
                  'Noise\nRobustness', 'Outlier\nRobustness', 'Prediction\nCoverage']
    
    # Original values
    values_raw = [r2, rmse, mae, mape, noise_degradation, outlier_degradation, avg_coverage * 100 if not np.isnan(avg_coverage) else np.nan]
    
    # Normalized values (for radar chart)
    values_normalized = [r2, rmse_normalized, mae_normalized, mape_normalized,
                        noise_robustness_score, outlier_robustness_score, avg_coverage]
    
    # Create DataFrame to save data
    radar_data_df = pd.DataFrame({
        'Indicator category': categories,
        'R²': [r2, np.nan, np.nan, np.nan, np.nan, np.nan, np.nan],
        'RMSE (MPa)': [np.nan, rmse, np.nan, np.nan, np.nan, np.nan, np.nan],
        'MAE (MPa)': [np.nan, np.nan, mae, np.nan, np.nan, np.nan, np.nan],
        'MAPE (%)': [np.nan, np.nan, np.nan, mape, np.nan, np.nan, np.nan],
        'Noise robustness degradation rate (%)': [np.nan, np.nan, np.nan, np.nan, noise_degradation, np.nan, np.nan],
        'Outlier robustness degradation rate (%)': [np.nan, np.nan, np.nan, np.nan, np.nan, outlier_degradation, np.nan],
        'Prediction interval coverage (%)': [np.nan, np.nan, np.nan, np.nan, np.nan, np.nan, avg_coverage * 100 if not np.isnan(avg_coverage) else np.nan],
        'Normalized score (0-1)': values_normalized,
        'Original value': values_raw
    })
    
    # No longer plot radar chart, only save data
    print("✓ Radar chart data has been calculated (no plotting, only saving data)")
    
    # Create a table containing three-fold average and final model metrics (five indicators: R², RMSE, MSE, MAE, MAPE)
    # Ensure all metrics are correctly read, if None then initialize as an empty dictionary
    if mean_train_metrics is None:
        mean_train_metrics = {'R²': np.nan, 'RMSE': np.nan, 'MSE': np.nan, 'MAE': np.nan, 'MAPE': np.nan}
    if mean_test_metrics is None:
        mean_test_metrics = {'R²': np.nan, 'RMSE': np.nan, 'MSE': np.nan, 'MAE': np.nan, 'MAPE': np.nan}
    if final_train_metrics is None:
        final_train_metrics = {'R²': np.nan, 'RMSE': np.nan, 'MSE': np.nan, 'MAE': np.nan, 'MAPE': np.nan}
    if final_test_metrics is None:
        final_test_metrics = {'R²': np.nan, 'RMSE': np.nan, 'MSE': np.nan, 'MAE': np.nan, 'MAPE': np.nan}
    
    # Ensure all keys exist, if not then set to NaN
    for metrics_dict in [mean_train_metrics, mean_test_metrics, final_train_metrics, final_test_metrics]:
        for key in ['R²', 'RMSE', 'MSE', 'MAE', 'MAPE']:
            if key not in metrics_dict:
                metrics_dict[key] = np.nan
    
    metrics_table_data = {
        'Indicator': ['R²', 'RMSE (MPa)', 'MSE (MPa²)', 'MAE (MPa)', 'MAPE (%)'],
        'Mean_Train': [
            mean_train_metrics.get('R²', np.nan),
            mean_train_metrics.get('RMSE', np.nan),
            mean_train_metrics.get('MSE', np.nan),
            mean_train_metrics.get('MAE', np.nan),
            mean_train_metrics.get('MAPE', np.nan)
        ],
        'Mean_Test': [
            mean_test_metrics.get('R²', np.nan),
            mean_test_metrics.get('RMSE', np.nan),
            mean_test_metrics.get('MSE', np.nan),
            mean_test_metrics.get('MAE', np.nan),
            mean_test_metrics.get('MAPE', np.nan)
        ],
        'Final_Train': [
            final_train_metrics.get('R²', np.nan),
            final_train_metrics.get('RMSE', np.nan),
            final_train_metrics.get('MSE', np.nan),
            final_train_metrics.get('MAE', np.nan),
            final_train_metrics.get('MAPE', np.nan)
        ],
        'Final_Test': [
            final_test_metrics.get('R²', np.nan),
            final_test_metrics.get('RMSE', np.nan),
            final_test_metrics.get('MSE', np.nan),
            final_test_metrics.get('MAE', np.nan),
            final_test_metrics.get('MAPE', np.nan)
        ]
    }
    metrics_table_df = pd.DataFrame(metrics_table_data)
    
    # Print debug information
    print("\n" + "="*60)
    print("Performance metrics table data check:")
    print("="*60)
    print(f"Mean_Train - R²: {mean_train_metrics.get('R²', 'N/A')}, RMSE: {mean_train_metrics.get('RMSE', 'N/A')}, MSE: {mean_train_metrics.get('MSE', 'N/A')}, MAE: {mean_train_metrics.get('MAE', 'N/A')}, MAPE: {mean_train_metrics.get('MAPE', 'N/A')}")
    print(f"Mean_Test - R²: {mean_test_metrics.get('R²', 'N/A')}, RMSE: {mean_test_metrics.get('RMSE', 'N/A')}, MSE: {mean_test_metrics.get('MSE', 'N/A')}, MAE: {mean_test_metrics.get('MAE', 'N/A')}, MAPE: {mean_test_metrics.get('MAPE', 'N/A')}")
    print(f"Final_Train - R²: {final_train_metrics.get('R²', 'N/A')}, RMSE: {final_train_metrics.get('RMSE', 'N/A')}, MSE: {final_train_metrics.get('MSE', 'N/A')}, MAE: {final_train_metrics.get('MAE', 'N/A')}, MAPE: {final_train_metrics.get('MAPE', 'N/A')}")
    print(f"Final_Test - R²: {final_test_metrics.get('R²', 'N/A')}, RMSE: {final_test_metrics.get('RMSE', 'N/A')}, MSE: {final_test_metrics.get('MSE', 'N/A')}, MAE: {final_test_metrics.get('MAE', 'N/A')}, MAPE: {final_test_metrics.get('MAPE', 'N/A')}")
    print("="*60)
    
    # Add detailed description data
    detailed_data = pd.DataFrame({
        'Indicator': ['R²', 'RMSE (MPa)', 'MAE (MPa)', 'MAPE (%)',
                'Noise robustness degradation rate (%)', 'Outlier robustness degradation rate (%)',
                'Bootstrap coverage (%)', 'Quantile coverage (%)',
                'Average prediction interval coverage (%)'],
        'Value': [r2, rmse, mae, mape,
                noise_degradation, outlier_degradation,
                bootstrap_coverage * 100 if bootstrap_coverage is not None else np.nan,
                quantile_coverage * 100 if quantile_coverage is not None else np.nan,
                avg_coverage * 100 if not np.isnan(avg_coverage) else np.nan],
        'Normalized score': [r2, rmse_normalized, mae_normalized, mape_normalized,
                      noise_robustness_score, outlier_robustness_score,
                      bootstrap_coverage if bootstrap_coverage is not None else np.nan,
                      quantile_coverage if quantile_coverage is not None else np.nan,
                      avg_coverage],
        'Description': ['R² is better closer to 1',
                'RMSE is better smaller',
                'MAE is better smaller',
                'MAPE is better smaller',
                'Noise robustness degradation rate is better smaller',
                'Outlier robustness degradation rate is better smaller',
                'Bootstrap method prediction interval coverage is better closer to target value (80%)',
                'Quantile Regression method prediction interval coverage is better closer to target value (80%)',
                'Average prediction interval coverage is better closer to target value (80%)']
    })
    
    print("\nRadar chart data summary:")
    print(f"  Current test set indicators - R²: {r2:.4f}, RMSE: {rmse:.3f} MPa, MAE: {mae:.3f} MPa, MAPE: {mape:.2f}%")
    if mean_train_metrics and not np.isnan(mean_train_metrics.get('R²', np.nan)):
        print(f"  Three-fold average training set - R²: {mean_train_metrics['R²']:.4f}, RMSE: {mean_train_metrics['RMSE']:.3f} MPa")
    if mean_test_metrics and not np.isnan(mean_test_metrics.get('R²', np.nan)):
        print(f"  Three-fold average test set - R²: {mean_test_metrics['R²']:.4f}, RMSE: {mean_test_metrics['RMSE']:.3f} MPa")
    if final_train_metrics and not np.isnan(final_train_metrics.get('R²', np.nan)):
        print(f"  Final model training set - R²: {final_train_metrics['R²']:.4f}, RMSE: {final_train_metrics['RMSE']:.3f} MPa")
    if final_test_metrics and not np.isnan(final_test_metrics.get('R²', np.nan)):
        print(f"  Final model test set - R²: {final_test_metrics['R²']:.4f}, RMSE: {final_test_metrics['RMSE']:.3f} MPa")
    
    print("\nPerformance metrics table (three-fold average vs final model):")
    print(metrics_table_df.to_string(index=False))
    
    # Ensure numerical format is correct (keep appropriate number of decimal places)
    metrics_table_df_formatted = metrics_table_df.copy()
    for col in ['Mean_Train', 'Mean_Test', 'Final_Train', 'Final_Test']:
        if col in metrics_table_df_formatted.columns:
            # For R², keep 5 decimal places; for other indicators, keep 2-5 decimal places
            for idx, metric in enumerate(metrics_table_df_formatted['Indicator']):
                value = metrics_table_df_formatted.loc[idx, col]
                if not pd.isna(value):
                    if 'R²' in metric:
                        metrics_table_df_formatted.loc[idx, col] = round(value, 5)
                    elif 'MAPE' in metric:
                        metrics_table_df_formatted.loc[idx, col] = round(value, 5)
                    else:
                        metrics_table_df_formatted.loc[idx, col] = round(value, 5)
    
    print("\nFormatted performance metrics table:")
    print(metrics_table_df_formatted.to_string(index=False))
    
    return radar_data_df, detailed_data, metrics_table_df

def analyze_prediction_intervals(model, X_test, y_test, save_dir, n_bootstrap=100, bootstrap_noise_level=0.01):
    """Analyze the confidence of the prediction intervals
    
    Parameters:
        model: Trained model
        X_test: Test features
        y_test: Test labels
        save_dir: Save directory
        n_bootstrap: Bootstrap resampling次数
        bootstrap_noise_level: Bootstrap noise level (percentage of feature standard deviation)
    """
    Method description:
    1. Bootstrap: Add
    
    # 1. Bootstrap prediction intervals - improved version
    print("Generating Bootstrap prediction intervals...")
    bootstrap_predictions = []
    for _ in range(n_bootstrap):
        # Add noise to simulate model uncertainty (similar to the idea of Monte Carlo Dropout)
        # Add small noise to the test data, reflecting the model's response to input uncertainty
        noise_std = bootstrap_noise_level  # Noise standard deviation (percentage of feature standard deviation)
        X_test_noisy = X_test.copy()
        for i in range(X_test.shape[1]):
            feature_std = np.std(X_test[:, i])
            noise = np.random.normal(0, noise_std * feature_std, size=len(X_test))
            X_test_noisy[:, i] = X_test[:, i] + noise
        
        # Predict
        y_pred = model.predict(X_test_noisy)
        bootstrap_predictions.append(y_pred)
    
    bootstrap_predictions = np.array(bootstrap_predictions)
    bootstrap_median = np.median(bootstrap_predictions, axis=0)
    bootstrap_lower = np.percentile(bootstrap_predictions, 10, axis=0)  # 80% confidence interval
    bootstrap_upper = np.percentile(bootstrap_predictions, 90, axis=0)
    
    # 2. Quantile Regression prediction intervals (using the residual distribution of the training data)
    print("Generating Quantile Regression prediction intervals...")
    print(f"Debug information:")
    print(f"  y_test range: {np.min(y_test):.4f} ~ {np.max(y_test):.4f}, mean: {np.mean(y_test):.4f}")
    print(f"  X_test shape: {X_test.shape}, mean range: [{np.mean(X_test, axis=0).min():.4f}, {np.mean(X_test, axis=0).max():.4f}]")
    y_pred_test = model.predict(X_test)
    print(f"  y_pred_test range: {np.min(y_pred_test):.4f} ~ {np.max(y_pred_test):.4f}, mean: {np.mean(y_pred_test):.4f}")
    residuals = y_test - y_pred_test
    print(f"  residuals range: {np.min(residuals):.4f} ~ {np.max(residuals):.4f}, mean: {np.mean(residuals):.4f}")
    residual_std = np.std(residuals)
    print(f"  residuals standard deviation: {residual_std:.4f}")
    
    # Use the standard deviation of the residuals to build the prediction intervals
    quantile_lower = y_pred_test - 1.28 * residual_std  # 10th percentile for 80% interval
    quantile_upper = y_pred_test + 1.28 * residual_std  # 90th percentile for 80% interval
    
    # Sort for visualization
    sort_indices = np.argsort(y_test)
    y_test_sorted = y_test[sort_indices]
    bootstrap_median_sorted = bootstrap_median[sort_indices]
    bootstrap_lower_sorted = bootstrap_lower[sort_indices]
    bootstrap_upper_sorted = bootstrap_upper[sort_indices]
    quantile_lower_sorted = quantile_lower[sort_indices]
    quantile_upper_sorted = quantile_upper[sort_indices]
    
    # Calculate coverage (before plotting)
    bootstrap_coverage = np.mean((y_test_sorted >= bootstrap_lower_sorted) & (y_test_sorted <= bootstrap_upper_sorted))
    quantile_coverage = np.mean((y_test_sorted >= quantile_lower_sorted) & (y_test_sorted <= quantile_upper_sorted))
    
    # Calculate average interval width
    bootstrap_width = np.mean(bootstrap_upper_sorted - bootstrap_lower_sorted)
    quantile_width = np.mean(quantile_upper_sorted - quantile_lower_sorted)
    
    # Debug: print interval width information
    print(f"\nInterval width calculation:")
    print(f"Bootstrap interval width: {bootstrap_width:.4f} MPa")
    print(f"Quantile interval width: {quantile_width:.4f} MPa")
    print(f"Bootstrap coverage: {bootstrap_coverage*100:.1f}%")
    print(f"Quantile coverage: {quantile_coverage*100:.1f}%")
    print(f"Note: Bootstrap coverage is smaller ({bootstrap_coverage*100:.1f}%), interval should be narrower ({bootstrap_width:.2f} MPa)")
    print(f"     Quantile coverage is larger ({quantile_coverage*100:.1f}%), interval should be wider ({quantile_width:.2f} MPa)")
    
    # 3. Plot prediction intervals
    # Keep the original "long bar" width, but split the top and bottom subplots into two independent images
    # Therefore, here we use figsize=(16, 5), keeping the original 16×10 total size in visual consistency
    
    # Create x-axis position (from 1 to n_samples, corresponding to sample number 1 to n_samples)
    n_samples = len(y_test_sorted)
    x_positions = np.arange(1, n_samples + 1)  # 1, 2, 3, ..., n_samples
    
    # Color scheme
    color_actual = '#000000'  # Pure black
    color_predicted = '#2563EB'  # Modern blue
    color_bootstrap_median = '#DC2626'  # Modern red
    
    # Blue gradient (Quantile confidence band)
    color_blue_start = '#2172B4'  # Deepest blue
    color_blue_end = '#FFFFFF'  # White (end color, enhance gradient effect)
    
    # Red gradient (Bootstrap confidence band)
    color_red_start = '#C41E3A'  # Deep red
    color_red_end = '#FFFFFF'  # White (end color, enhance gradient effect)
    
    # Number of gradient layers
    n_gradient_layers = 50
    
    # Helper function: convert RGB color to tuple
    def hex_to_rgb(hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    # Helper function: RGB color interpolation
    def interpolate_color(color1, color2, ratio):
        rgb1 = hex_to_rgb(color1)
        rgb2 = hex_to_rgb(color2)
        return tuple(int(rgb1[i] + (rgb2[i] - rgb1[i]) * ratio) for i in range(3))
    
    # Helper function: RGB to hexadecimal
    def rgb_to_hex(rgb):
        return '#{:02x}{:02x}{:02x}'.format(int(rgb[0]), int(rgb[1]), int(rgb[2]))
    
    # ==========================================================
    # Figure 1: Bootstrap prediction intervals (separate long bar)
    # ==========================================================
    fig_bootstrap, ax1 = plt.subplots(figsize=(16, 5))
    fig_bootstrap.patch.set_facecolor('white')
    
    # Plot Bootstrap prediction intervals
    bootstrap_baseline = bootstrap_median_sorted  # Bootstrap基线：Bootstrap中位数（Bootstrap分布的中心估计）
    bootstrap_range = bootstrap_upper_sorted - bootstrap_lower_sorted
    
    # 50 level gradient effect (from center dark to edge white)
    for i in range(n_gradient_layers):
        # layer_ratio from 0 to 1, 0=center (dark), 1=edge (white)
        layer_ratio = (i + 1) / n_gradient_layers
        # Calculate the boundary of the current layer
        layer_lower = bootstrap_baseline - (bootstrap_baseline - bootstrap_lower_sorted) * layer_ratio
        layer_upper = bootstrap_baseline + (bootstrap_upper_sorted - bootstrap_baseline) * layer_ratio
        
        # Color interpolation: layer_ratio=0 is deep red, layer_ratio=1 is white
        # Reverse layer_ratio for color: inner (near 0) dark, outer (near 1) white
        color_ratio = layer_ratio  # Use directly, because layer_ratio is from 0 to 1
        current_color = rgb_to_hex(interpolate_color(color_red_start, color_red_end, color_ratio))
        
        # Transparency: inner layer is less transparent, outer layer is more transparent (but not completely transparent)
        # From 0.6 (inner layer) to 0.1 (outer layer), ensure the outermost layer can also be seen
        alpha_layer = 0.6 - 0.5 * layer_ratio
        
        if i == 0:
            ax1.fill_between(x_positions, layer_lower, layer_upper, 
                           alpha=alpha_layer, color=current_color, 
                           label='Bootstrap 80% interval', zorder=1,
                           edgecolor='none', linewidth=0)
        else:
            ax1.fill_between(x_positions, layer_lower, layer_upper, 
                           alpha=alpha_layer, color=current_color, 
                           zorder=1, edgecolor='none', linewidth=0)
    
    # Plot Bootstrap median line
    ax1.plot(x_positions, bootstrap_median_sorted, '--', linewidth=2.5, 
             color=color_bootstrap_median, label='Bootstrap median', zorder=4, 
             alpha=0.80, dashes=(10, 5))
    
    # Plot predicted and actual values
    ax1.plot(x_positions, y_pred_test[sort_indices], '-', linewidth=3.5, 
             color=color_predicted, label='Predicted values', zorder=6, alpha=0.95)
    ax1.scatter(x_positions, y_test_sorted, s=120, c=color_actual, alpha=1.0, 
                label='Actual values', zorder=7, edgecolors='none', marker='o')
    
    # Set labels and styles
    ax1.set_xlabel('Test samples (sorted by actual values)',
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    ax1.set_ylabel('Peak Stress fc (MPa)', 
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    ax1.tick_params(axis='y', which='major', labelsize=18, width=2.0, length=6, colors='#000000')
    for label in ax1.get_yticklabels():
        label.set_fontweight('bold')
        label.set_fontfamily('Arial')
        label.set_color('#000000')
    for spine in ax1.spines.values():
        spine.set_linewidth(2.0)
        spine.set_color('#000000')
    ax1.grid(False)
    
    # Add Bootstrap performance metrics
    bootstrap_text = f'Coverage: {bootstrap_coverage*100:.1f}%\nAvg Width: {bootstrap_width:.2f} MPa'
    ax1.text(0.02, 0.98, bootstrap_text, transform=ax1.transAxes, fontsize=18, 
             fontweight='bold', fontfamily='Arial', verticalalignment='top', color='#000000')
    
    # Set legend
    ax1.legend(loc='lower right', fontsize=18, 
               prop={'family': 'Arial', 'weight': 'bold'},
               frameon=False, fancybox=False, shadow=False)
    
    # Unify y-axis range (ensure consistency with Quantile figure)
    y_min = min(np.min(y_test_sorted), np.min(y_pred_test[sort_indices]), 
                np.min(bootstrap_lower_sorted), np.min(quantile_lower_sorted))
    y_max = max(np.max(y_test_sorted), np.max(y_pred_test[sort_indices]), 
                np.max(bootstrap_upper_sorted), np.max(quantile_upper_sorted))
    y_padding = (y_max - y_min) * 0.05  # 5% margin
    y_range = [y_min - y_padding, y_max + y_padding]
    ax1.set_ylim(y_range)
    
    # Set x-axis ticks (ensure consistency with Quantile figure)
    tick_interval = max(1, n_samples // 6)
    x_ticks = np.arange(1, n_samples + 1, tick_interval)
    if len(x_ticks) == 0 or x_ticks[-1] != n_samples:
        x_ticks = np.append(x_ticks, n_samples)
    ax1.set_xticks(x_ticks)
    ax1.tick_params(axis='x', which='major', labelsize=18, width=2.0, length=6, colors='#000000')
    for label in ax1.get_xticklabels():
        label.set_fontweight('bold')
        label.set_fontfamily('Arial')
        label.set_color('#000000')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'prediction_intervals_bootstrap.png'),
                dpi=300, bbox_inches='tight')
    plt.close(fig_bootstrap)
    
    # ==========================================================
    # Figure 2: Quantile Regression prediction intervals (separate long bar)
    # ==========================================================
    fig_quantile, ax2 = plt.subplots(figsize=(16, 5))
    fig_quantile.patch.set_facecolor('white')
    
    # Plot Quantile Regression prediction intervals
    quantile_baseline = y_pred_test[sort_indices]  # Quantile baseline: predicted values
    quantile_range = quantile_upper_sorted - quantile_lower_sorted
    
    # 50 level gradient effect (from center dark to edge white)
    for i in range(n_gradient_layers):
        # layer_ratio from 0 to 1, 0=center (dark), 1=edge (white)
        layer_ratio = (i + 1) / n_gradient_layers
        # Calculate the boundary of the current layer
        layer_lower = quantile_baseline - (quantile_baseline - quantile_lower_sorted) * layer_ratio
        layer_upper = quantile_baseline + (quantile_upper_sorted - quantile_baseline) * layer_ratio
        
        # Color interpolation: layer_ratio=0 is deep blue, layer_ratio=1 is white
        # Reverse layer_ratio for color: inner (near 0) dark, outer (near 1) white
        color_ratio = layer_ratio  # Use directly, because layer_ratio is from 0 to 1
        current_color = rgb_to_hex(interpolate_color(color_blue_start, color_blue_end, color_ratio))
        
        # Transparency: inner layer is less transparent, outer layer is more transparent (but not completely transparent)
        # From 0.6 (inner layer) to 0.1 (outer layer), ensure the outermost layer can also be seen
        alpha_layer = 0.6 - 0.5 * layer_ratio
        
        if i == 0:
            ax2.fill_between(x_positions, layer_lower, layer_upper, 
                           alpha=alpha_layer, color=current_color, 
                           label='Quantile 80% interval', zorder=1,
                           edgecolor='none', linewidth=0)
        else:
            ax2.fill_between(x_positions, layer_lower, layer_upper, 
                           alpha=alpha_layer, color=current_color, 
                           zorder=1, edgecolor='none', linewidth=0)
    
    # Plot predicted and actual values
    ax2.plot(x_positions, y_pred_test[sort_indices], '-', linewidth=3.5, 
             color=color_predicted, label='Predicted values', zorder=6, alpha=0.95)
    ax2.scatter(x_positions, y_test_sorted, s=120, c=color_actual, alpha=1.0, 
                label='Actual values', zorder=7, edgecolors='none', marker='o')
    
    # Set subplot labels and styles
    ax2.set_xlabel('Test samples (sorted by actual values)', 
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    ax2.set_ylabel('Peak Stress fc (MPa)', 
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    
    # Set x-axis ticks (ensure consistency with Figure 1)
    ax2.set_xticks(x_ticks)
    
    # Set axis styles
    ax2.tick_params(axis='both', which='major', labelsize=18, width=2.0, length=6, colors='#000000')
    for label in ax2.get_xticklabels() + ax2.get_yticklabels():
        label.set_fontweight('bold')
        label.set_fontfamily('Arial')
        label.set_color('#000000')
    for spine in ax2.spines.values():
        spine.set_linewidth(2.0)
        spine.set_color('#000000')
    ax2.grid(False)
    
    # Add Quantile performance metrics
    quantile_text = f'Coverage: {quantile_coverage*100:.1f}%\nAvg Width: {quantile_width:.2f} MPa'
    ax2.text(0.02, 0.98, quantile_text, transform=ax2.transAxes, fontsize=18, 
             fontweight='bold', fontfamily='Arial', verticalalignment='top', color='#000000')
    
    # Set legend
    ax2.legend(loc='lower right', fontsize=18, 
               prop={'family': 'Arial', 'weight': 'bold'},
               frameon=False, fancybox=False, shadow=False)
    
    # y-axis range consistent with Bootstrap figure
    ax2.set_ylim(y_range)
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'prediction_intervals_quantile.png'),
                dpi=300, bbox_inches='tight')
    plt.close(fig_quantile)
    
    # Save prediction intervals results
    # Note: The figure is drawn in the order of actual values, so to match the figure,
    # Here we also save the results in the same order as the actual values (Sample_ID is the index of the original test sample in y_test).
    intervals_df = pd.DataFrame({
        'Sample_ID': sort_indices,
        'Actual': y_test_sorted,
        'Predicted': y_pred_test[sort_indices],
        'Bootstrap_Lower': bootstrap_lower_sorted,
        'Bootstrap_Upper': bootstrap_upper_sorted,
        'Quantile_Lower': quantile_lower_sorted,
        'Quantile_Upper': quantile_upper_sorted
    })
    # Save to Excel in the main function
    
    print(f"Prediction intervals analysis completed, 80% confidence interval coverage: Bootstrap={bootstrap_coverage*100:.1f}%, Quantile={quantile_coverage*100:.1f}%")
    print(f"Bootstrap average interval width: {bootstrap_width:.4f} MPa, Quantile average interval width: {quantile_width:.4f} MPa")
    return bootstrap_predictions, bootstrap_coverage, quantile_coverage, intervals_df

def load_trained_model(model_path):
    """Load the trained XGBoost model (supports .joblib and .pkl formats)"""
    print(f"\n=== Load the trained model ===")
    print(f"Model path: {model_path}")
    
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model file not found: {model_path}")
    
    loaded_data = joblib.load(model_path)
    
    # Determine the type of loaded data
    if isinstance(loaded_data, dict):
        # If it is a dictionary, try to extract the model and feature names
        print(f"  Loaded data type: dictionary")
        print(f"  Dictionary keys: {list(loaded_data.keys())}")
        model = loaded_data.get('model', None)
        if model is None:
            # If the dictionary does not have the 'model' key, try other possible keys
            for key in ['final_model', 'xgboost_model', 'model']:
                if key in loaded_data:
                    model = loaded_data[key]
                    print(f"  Extracted model from key '{key}'")
                    break
            if model is None:
                raise ValueError(f"Model object not found in dictionary, available keys: {list(loaded_data.keys())}")
        feature_names = loaded_data.get('feature_names', None)
        model_data = loaded_data
    else:
        # If it is a model object directly
        print(f"  Loaded data type: model object")
        model = loaded_data
        feature_names = None
        model_data = {'model': model}
    
    # Verify the model type
    if not hasattr(model, 'predict'):
        raise ValueError(f"Loaded object is not a valid model (missing predict method), type: {type(model)}")
    
    print(f"✓ Model loaded successfully")
    print(f"✓ Model type: {type(model).__name__}")
    if hasattr(model, 'n_features_in_'):
        print(f"✓ Expected number of features: {model.n_features_in_}")
    if feature_names:
        print(f"✓ Number of features: {len(feature_names)}")
        print(f"✓ Feature names: {feature_names}")
    else:
        print("⚠ Warning: Feature name information not found, will use feature names from data loading")
    
    return model, model_data

# Deleted: cross_subset_validation and cross_random_kfold_validation functions
# These are cross-validation functions related to training, not used in the main function, removed

def main():
    """Main function - load the trained model and perform SHAP analysis, noise analysis and uncertainty analysis"""
    print("=== XGBoost model SHAP analysis, noise analysis and uncertainty analysis ===")
    
    # 1. Load data
    data = load_data()
    if data is None:
        return
    
    X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df_processed = data
    
    # 2. Load the trained model - directly use xgboost_final_model.joblib
    model_path = os.path.join(SAVE_ROOT, 'xgboost_final_model.joblib')
    
    if not os.path.exists(model_path):
        print(f"Error: Model file not found: {model_path}")
        print("Ensure the model file exists in: peak_stress/XGBoost/save/xgboost_final_model.joblib")
        return
    
    model, model_data = load_trained_model(model_path)
    
    # Check the expected number of features
    if hasattr(model, 'n_features_in_'):
        expected_n_features = model.n_features_in_
        print(f"✓ Expected number of features: {expected_n_features}")
        if X.shape[1] != expected_n_features:
            print(f"❌ Error: Data feature number ({X.shape[1]}) does not match the expected number of features ({expected_n_features})!")
            print(f"  This will cause the prediction results to be completely incorrect!")
            return
    else:
        print("⚠ Warning: Model does not have n_features_in_ attribute, cannot verify the number of features")
    
    # If the model data has feature names, use the model's (ensure feature order consistency)
    if 'feature_names' in model_data and model_data['feature_names'] is not None:
        model_feature_names = model_data['feature_names']
        print(f"\nUse the model's saved feature names ({len(model_feature_names)} features)")
        # Ensure feature order consistency with the model
        if len(model_feature_names) == len(feature_names):
            # Check if the feature names match
            if set(model_feature_names) == set(feature_names):
                # Reorder features to match the model
                feature_order = [feature_names.index(f) for f in model_feature_names]
                X = X[:, feature_order]
                feature_names = model_feature_names
                print("✓ Feature order adjusted to match the model")
            else:
                print("⚠ Warning: Model feature names do not match the data feature names completely")
                print(f"  Model features: {model_feature_names}")
                print(f"  Data features: {feature_names}")
        else:
            print(f"⚠ Warning: Feature number mismatch - model: {len(model_feature_names)}, data: {len(feature_names)}")
    
    # Test model prediction (use the first 5 samples for quick validation)
    print(f"\n=== Model prediction test ===")
    print(f"Test data range: y_min={np.min(y):.4f}, y_max={np.max(y):.4f}, y_mean={np.mean(y):.4f}")
    print(f"Test data shape: X={X.shape}, y={y.shape}")
    print(f"Statistics of the first 5 samples of X: mean={np.mean(X[:5], axis=0)}, range=[{np.min(X[:5], axis=0)}, {np.max(X[:5], axis=0)}]")
    try:
        test_pred = model.predict(X[:5])
        print(f"First 5 samples prediction values: {test_pred}")
        print(f"First 5 samples actual values: {y[:5]}")
        print(f"Prediction value range: {np.min(test_pred):.4f} ~ {np.max(test_pred):.4f}")
        print(f"Actual value range: {np.min(y[:5]):.4f} ~ {np.max(y[:5]):.4f}")
        
        pred_mean = np.mean(test_pred)
        actual_mean = np.mean(y[:5])
        diff = abs(pred_mean - actual_mean)
        
        if diff > 10:
            print(f"\n❌ Severe error: Prediction values differ significantly from actual values!")
            print(f"  Average prediction value: {pred_mean:.4f}, average actual value: {actual_mean:.4f}")
            print(f"  Difference: {diff:.4f}")
            print(f"\nPossible reasons:")
            print(f"  1. Model file corrupted or loading error")
            print(f"  2. Data preprocessing inconsistency (standardization/normalization used during training, but not during prediction)")
            print(f"  3. Feature order mismatch")
            print(f"  4. Data format mismatch between model training and current data")
            print(f"\nSuggestions:")
            print(f"  1. Check if the model file is saved correctly")
            print(f"  2. Check if there are data preprocessing steps (e.g., StandardScaler) in the training script")
            print(f"  3. Retrain the model and ensure feature name information is included when saving")
            print(f"  4. Check if the data file is the same as the one used during training")
            return
        else:
            print(f"✓ Prediction test passed: average difference = {diff:.4f}")
    except Exception as e:
        print(f"❌ Model prediction error: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # 3. Data splitting (get training set and test set, for SHAP analysis and noise analysis)
    X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids = split_data_by_divisions(
        X, y, sample_divisions, sample_ids, test_ratio=0.2, val_ratio=0.2, random_state=42
    )
    
    if len(X_test) == 0:
        print("Warning: No test set found, will use validation set for analysis")
        X_test = X_val
        y_test = y_val
        test_ids = val_ids
    
    # Prepare data for SHAP analysis
    # Use all data for SHAP analysis (data set is small, total 94 samples)
    
    # Merge all datasets: training set + validation set + test set
    datasets_to_merge = []
    y_datasets_to_merge = []
    
    if len(X_train) > 0:
        datasets_to_merge.append(X_train)
        y_datasets_to_merge.append(y_train)
    if len(X_val) > 0:
        datasets_to_merge.append(X_val)
        y_datasets_to_merge.append(y_val)
    if len(X_test) > 0:
        datasets_to_merge.append(X_test)
        y_datasets_to_merge.append(y_test)
    
    if len(datasets_to_merge) > 0:
        X_all = np.vstack(datasets_to_merge)
        y_all = np.hstack(y_datasets_to_merge)
    else:
        X_all = X
        y_all = y
    
    # Use all data as background and explanation data set
    X_background = X_all
    y_background = y_all
    X_explain = X_all
    y_explain = y_all
    
    print(f"\nData preparation completed:")
    print(f"  Training set: {len(X_train)} samples")
    print(f"  Validation set: {len(X_val)} samples")
    print(f"  Test set: {len(X_test)} samples")
    print(f"\nSHAP analysis data set configuration (using all data):")
    print(f"  All data: {len(X_all)} samples (training set + validation set + test set)")
    print(f"  Background data set (for explainer initialization): {len(X_background)} samples")
    print(f"  Explanation data set (for calculating SHAP values): {len(X_explain)} samples")
    
    # 4. Create save directory
    analysis_save_dir = os.path.join(SAVE_ROOT)
    os.makedirs(analysis_save_dir, exist_ok=True)
    
    # 5. SHAP analysis (using all data)
    print(f"\n{'='*60}")
    print("Start SHAP analysis (using all data)")
    print(f"{'='*60}")
    shap_importance = None
    feature_stats = None
    try:
        shap_importance = plot_shap_analysis(
            model, X_background, X_explain, feature_names, analysis_save_dir
        )
        print("✓ SHAP analysis completed")
        
        # 5.1 Feature statistics analysis (explain why some features have a small impact)
        print(f"\n{'='*60}")
        print("Start feature statistics analysis")
        print(f"{'='*60}")
        try:
            feature_stats = analyze_feature_statistics(
                X_all, y_all, feature_names, shap_importance, analysis_save_dir
            )
            print("✓ Feature statistics analysis completed")
        except Exception as e:
            print(f"⚠ Feature statistics analysis error: {e}")
            import traceback
            traceback.print_exc()
            feature_stats = None
    except Exception as e:
        print(f"⚠ SHAP analysis error: {e}")
        import traceback
        traceback.print_exc()
        shap_importance = None
        feature_stats = None
    
    # 6. Feature importance analysis
    print(f"\n{'='*60}")
    print("Start feature importance analysis")
    print(f"{'='*60}")
    importance_df = None
    try:
        importance_df = plot_feature_importance(
            model, feature_names, analysis_save_dir, X=X_all
        )
        print("✓ Feature importance analysis completed")
    except Exception as e:
        print(f"⚠ Feature importance analysis error: {e}")
        import traceback
        traceback.print_exc()
        importance_df = None
    
    # 7. Noise robustness analysis
    print(f"\n{'='*60}")
    print("Start noise robustness analysis")
    print(f"{'='*60}")
    noise_rmses, outlier_rmses = None, None
    robustness_df, outlier_df = None, None
    try:
        noise_rmses, outlier_rmses, robustness_df, outlier_df = analyze_noise_robustness(
            model, X_test, y_test, feature_names, analysis_save_dir
        )
        print("✓ Noise robustness analysis completed")
    except Exception as e:
        print(f"⚠ Noise robustness analysis error: {e}")
        import traceback
        traceback.print_exc()
        noise_rmses, outlier_rmses = None, None
        robustness_df, outlier_df = None, None
    
    # 8. Uncertainty analysis (prediction intervals)
    print(f"\n{'='*60}")
    print("Start uncertainty analysis (prediction intervals)")
    print(f"{'='*60}")
    bootstrap_predictions, bootstrap_coverage, quantile_coverage = None, None, None
    intervals_df = None
    try:
        bootstrap_predictions, bootstrap_coverage, quantile_coverage, intervals_df = analyze_prediction_intervals(
            model, X_test, y_test, analysis_save_dir, n_bootstrap=100, bootstrap_noise_level=0.01
        )
        print("✓ Uncertainty analysis completed")
    except Exception as e:
        print(f"⚠ Uncertainty analysis error: {e}")
        import traceback
        traceback.print_exc()
        bootstrap_predictions, bootstrap_coverage, quantile_coverage = None, None, None
        intervals_df = None
    
    # 9. Plot prediction result comparison chart
    print(f"\n{'='*60}")
    print("Start plotting prediction result comparison chart")
    print(f"{'='*60}")
    prediction_df = None
    y_test_pred = None
    try:
        y_test_pred = model.predict(X_test)
        plot_results(y_test, y_test_pred, analysis_save_dir)
        # 使用训练集+验证集作为训练集，测试集作为测试集进行对比
        # 注意：X_background包含了全部数据（训练集+验证集+测试集），需要排除测试集
        # 合并训练集和验证集（排除测试集）
        X_train_val = np.vstack([X_train, X_val]) if len(X_val) > 0 else X_train
        y_train_val = np.hstack([y_train, y_val]) if len(y_val) > 0 else y_train
        prediction_df = plot_train_test_comparison(model, X_train_val, y_train_val, X_test, y_test, analysis_save_dir)
        print("✓ Prediction result comparison chart plotted")
    except Exception as e:
        print(f"⚠ Plot prediction result chart error: {e}")
        import traceback
        traceback.print_exc()
        prediction_df = None
        # If error occurs, at least ensure y_test_pred is calculated (for subsequent radar chart)
        if y_test_pred is None:
            try:
                y_test_pred = model.predict(X_test)
            except:
                pass
    
    # 10. PDP analysis (using training set + validation set): one-dimensional PDP and bivariate PDP of the top 4 most important features
    print(f"\n{'='*60}")
    print("Start PDP analysis (partial dependence plots, using training set + validation set)")
    print(f"{'='*60}")
    try:
        # n_top_features=4: generate 4 one-dimensional PDP plots + several bivariate PDP plots of the top 4 most important features
        plot_pdp_analysis(model, X_background, feature_names, analysis_save_dir, n_top_features=4)
        print("✓ PDP analysis completed")
    except Exception as e:
        print(f"⚠ PDP analysis error: {e}")
        import traceback
        traceback.print_exc()
    
    # 10.1. Radar chart analysis (plot performance indicator radar chart and save data)
    print(f"\n{'='*60}")
    print("Start plotting radar chart and saving data")
    print(f"{'='*60}")
    radar_data_df = None
    radar_detailed_df = None
    metrics_table_df = None
    try:
        # Ensure y_test_pred exists, if not, calculate it
        if y_test_pred is None:
            y_test_pred = model.predict(X_test)
        
        # Pass training metrics file path
        training_metrics_path = os.path.join(SAVE_ROOT, 'training_metrics.xlsx')
        radar_data_df, radar_detailed_df, metrics_table_df = plot_radar_chart(
            y_test, y_test_pred, noise_rmses, outlier_rmses,
            bootstrap_coverage, quantile_coverage, analysis_save_dir,
            training_metrics_path=training_metrics_path
        )
        print("✓ Radar chart plotted and data prepared")
    except Exception as e:
        print(f"⚠ Radar chart plotting error: {e}")
        import traceback
        traceback.print_exc()
        radar_data_df = None
        radar_detailed_df = None
        metrics_table_df = None
    
    # 11. Print summary
    print(f"\n{'='*60}")
    print("Analysis completed summary")
    print(f"{'='*60}")
    
    if noise_rmses is not None:
        print(f"\nNoise robustness analysis results:")
        print(f"  - No noise RMSE: {noise_rmses[0]:.3f} MPa")
        print(f"  - 10% noise RMSE: {noise_rmses[-1]:.3f} MPa")
        print(f"  - Performance degradation rate: {(noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100:.1f}%")
    
    if bootstrap_coverage is not None:
        print(f"\nUncertainty analysis results:")
        print(f"  - Bootstrap coverage: {bootstrap_coverage*100:.1f}%")
        print(f"  - Quantile Regression coverage: {quantile_coverage*100:.1f}%")
    
    # 11. Save all data to an Excel file with different sheets
    print(f"\n{'='*60}")
    print("Save all analysis results to an Excel file with different sheets")
    print(f"{'='*60}")
    try:
        excel_path = os.path.join(analysis_save_dir, 'model_analysis_results.xlsx')
        
        # Check if the file exists, if it exists, use append mode, otherwise create a new file
        # Note: PDP analysis has already added data in the previous sheets, so here should use append mode
        file_exists = os.path.exists(excel_path)
        mode = 'a' if file_exists else 'w'
        
        print(f"  Excel file path: {excel_path}")
        print(f"  File exists: {file_exists}")
        print(f"  Save mode: {mode} (append mode, keep PDP data)")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            # 1. Feature importance (including coefficient of variation)
            if importance_df is not None:
                importance_df.to_excel(writer, sheet_name='Feature importance', index=False)
                print("✓ Feature importance saved")
            
            # 2. SHAP importance
            if shap_importance is not None:
                shap_importance.to_excel(writer, sheet_name='SHAP importance', index=False)
                print("✓ SHAP importance saved")
            
            # 3. Feature statistics information
            if 'feature_stats' in locals() and feature_stats is not None:
                feature_stats.to_excel(writer, sheet_name='Feature statistics information', index=False)
                print("✓ Feature statistics information saved")
            
            # 4. Noise robustness
            if 'robustness_df' in locals() and robustness_df is not None:
                robustness_df.to_excel(writer, sheet_name='Noise robustness', index=False)
                print("✓ Noise robustness saved")
            
            # 5. Outlier robustness
            if 'outlier_df' in locals() and outlier_df is not None:
                outlier_df.to_excel(writer, sheet_name='Outlier robustness', index=False)
                print("✓ Outlier robustness saved")
            
            # 6. Prediction intervals
            if 'intervals_df' in locals() and intervals_df is not None:
                intervals_df.to_excel(writer, sheet_name='Prediction intervals', index=False)
                print("✓ Prediction intervals saved")
            
            # 7. Training set and test set prediction results
            if 'prediction_df' in locals() and prediction_df is not None:
                prediction_df.to_excel(writer, sheet_name='Prediction result comparison', index=False)
                print("✓ Prediction result comparison saved")
            
            # 8. Radar chart data (summary table)
            if 'radar_data_df' in locals() and radar_data_df is not None:
                radar_data_df.to_excel(writer, sheet_name='Radar chart data summary', index=False)
                print("✓ Radar chart data summary saved")
            
            # 9. Radar chart detailed data
            if 'radar_detailed_df' in locals() and radar_detailed_df is not None:
                radar_detailed_df.to_excel(writer, sheet_name='Radar chart detailed data', index=False)
                print("✓ Radar chart detailed data saved")
            
            # 10. Performance indicator table (three-fold average vs final model)
            if 'metrics_table_df' in locals() and metrics_table_df is not None:
                # Format numbers (keep appropriate decimal places)
                metrics_table_df_save = metrics_table_df.copy()
                for col in ['Mean_Train', 'Mean_Test', 'Final_Train', 'Final_Test']:
                    if col in metrics_table_df_save.columns:
                        for idx, metric in enumerate(metrics_table_df_save['Indicator']):
                            value = metrics_table_df_save.loc[idx, col]
                            if not pd.isna(value):
                                if 'R²' in metric:
                                    metrics_table_df_save.loc[idx, col] = round(value, 5)
                                elif 'MAPE' in metric:
                                    metrics_table_df_save.loc[idx, col] = round(value, 5)
                                else:
                                    metrics_table_df_save.loc[idx, col] = round(value, 5)
                
                metrics_table_df_save.to_excel(writer, sheet_name='Performance indicator table', index=False)
                print("✓ Performance indicator table saved (three-fold average and final model five indicators: R², RMSE, MSE, MAE, MAPE)")

        print(f"\n✓ All analysis results have been saved to: {excel_path}")
        # Count the number of sheets
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet_count = len(wb.sheetnames)
            print(f"  Contains {sheet_count} sheets: {', '.join(wb.sheetnames)}")
        except:
            print(f"  Contains multiple sheets")
    except Exception as e:
        print(f"⚠ Error saving Excel: {e}")
        import traceback
        traceback.print_exc()
    
    # 12. Predict all data and write to dataset_with_XGB_fc.xlsx
    print(f"\n{'='*60}")
    print("Predict all data and write to dataset_with_XGB_fc.xlsx")
    print(f"{'='*60}")
    try:
        # Predict all data using the final model
        print(f"Predicting all {len(X)} samples...")
        y_all_pred = model.predict(X)
        
        # Create a prediction result dictionary (using sample_ids as keys)
        pred_dict = {}
        for idx, sample_id in enumerate(sample_ids):
            pred_dict[sample_id] = float(y_all_pred[idx])
        
        # Read the original Excel file
        data_file = os.path.join(PROJECT_ROOT, "dataset", "dataset_final.xlsx")
        if not os.path.exists(data_file):
            print(f"⚠ Warning: Original data file not found: {data_file}")
            print("  Will try to use the processed data file...")
            # 如果原始文件不存在，尝试使用df_processed
            if 'df_processed' in locals() and df_processed is not None:
                df_original = df_processed.copy()
            else:
                print("  ✗ Unable to find data file, skip saving prediction results")
                raise FileNotFoundError(f"Unable to find data file: {data_file}")
        else:
            df_original = pd.read_excel(data_file, sheet_name=0)
        
        # Use map method to match prediction results
        if 'No_Customized' in df_original.columns:
            df_original['XGB_fc'] = df_original['No_Customized'].map(pred_dict)
            print(f"✓ Using No_Customized column to match prediction results")
        else:
            print(f"⚠ Warning: No_Customized column not found, will match prediction results by row order")
            # If there is no No_Customized column, match prediction results by row order (ensure the order is consistent)
            if len(y_all_pred) == len(df_original):
                df_original['XGB_fc'] = y_all_pred
                print(f"✓ Matching prediction results by row order")
            else:
                print(f"  ✗ Prediction sample number ({len(y_all_pred)}) does not match the number of rows in the original data ({len(df_original)})")
                print(f"  Will try to use original_df_indices to match...")
                # Try to use original_df_indices to match
                if 'original_df_indices' in locals() and original_df_indices is not None:
                    if len(original_df_indices) == len(y_all_pred):
                        # Create a Series, index is original_df_indices, value is prediction result
                        pred_series = pd.Series(y_all_pred, index=original_df_indices)
                        # Map the prediction results to the original DataFrame
                        df_original['XGB_fc'] = df_original.index.map(pred_series)
                        print(f"✓ Using original_df_indices to match prediction results")
                    else:
                        raise ValueError(f"Length of original_df_indices ({len(original_df_indices)}) does not match the length of prediction results ({len(y_all_pred)})")
                else:
                    raise ValueError("Sample number does not match, cannot write prediction results")
        
        # Save results
        output_file = os.path.join(PROJECT_ROOT, "dataset", "dataset_with_XGB_fc.xlsx")
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        df_original.to_excel(output_file, index=False)
        print(f"✓ Prediction results have been written to: {output_file}")
        print(f"  - New column: XGB_fc (final model prediction value)")
        print(f"  - Total prediction sample number: {len(y_all_pred)}")
        print(f"  - Number of non-empty prediction values: {df_original['XGB_fc'].notna().sum()}")
        
        # Backup to save directory
        backup_file = os.path.join(analysis_save_dir, 'dataset_with_XGB_fc_backup.xlsx')
        df_original.to_excel(backup_file, index=False)
        print(f"✓ Backup file has been saved: {backup_file}")
        
    except Exception as e:
        print(f"⚠ Warning: Error writing to dataset_with_XGB_fc.xlsx: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"\nAll results have been saved to: {analysis_save_dir}")
    print(f"{'='*60}")
    
    return {
        'model': model,
        'model_data': model_data,
        'noise_rmses': noise_rmses,
        'outlier_rmses': outlier_rmses,
        'bootstrap_coverage': bootstrap_coverage,
        'quantile_coverage': quantile_coverage,
        'shap_importance': shap_importance,
        'importance_df': importance_df
    }

if __name__ == "__main__":
    results = main()
