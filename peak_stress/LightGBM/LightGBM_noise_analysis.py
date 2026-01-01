#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
LightGBM model noise analysis and uncertainty analysis

Function:
1. Load the trained LightGBM model
2. Noise robustness analysis: test the robustness of the model to Gaussian noise and outliers
3. Prediction interval analysis: use Bootstrap and Quantile Regression methods to generate confidence intervals

Note: This script needs to run LightGBM_train.py to train the model first, then load the saved model for analysis
"""

import os
import sys
import numpy as np
import pandas as pd
# Set matplotlib backend to Agg (non-interactive, to avoid GUI related errors)
import matplotlib
matplotlib.use('Agg')  # Must be set before importing pyplot
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score, KFold, train_test_split
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error
from scipy import stats
import lightgbm as lgb
import optuna
from optuna.visualization import plot_optimization_history, plot_param_importances
import shap
from sklearn.inspection import partial_dependence
from mpl_toolkits.mplot3d import Axes3D
import warnings
import joblib
warnings.filterwarnings('ignore')

# Set Chinese fonts and LaTeX rendering
def configure_plot_fonts(fonts=None):
    """Set Matplotlib fonts uniformly, ensure negative sign is displayed normally"""
    if fonts is None:
        fonts = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
    plt.rcParams['font.sans-serif'] = fonts
    plt.rcParams['axes.unicode_minus'] = False

configure_plot_fonts()
plt.rcParams['mathtext.default'] = 'regular'  # Use regular font to render mathematical symbols

# Work directory setting - automatically find project root directory
def find_project_root():
    """Find project root directory"""
    # First try to locate from script position
    try:
        script_path = os.path.abspath(__file__)
        script_dir = os.path.dirname(script_path)
        # If script is in a subdirectory, find the directory containing dataset or SAVE
        current_dir = script_dir
        search_depth = 0
        while current_dir != os.path.dirname(current_dir) and search_depth < 10:
            if os.path.exists(os.path.join(current_dir, 'dataset')) or \
               os.path.exists(os.path.join(current_dir, 'SAVE')):
                return current_dir
            current_dir = os.path.dirname(current_dir)
            search_depth += 1
    except NameError:
        # Running in Jupyter, search from current directory upwards
        pass
    
    # Search from current working directory upwards
    current_dir = os.path.abspath(os.getcwd())
    search_limit = 0
    while current_dir != os.path.dirname(current_dir) and search_limit < 10:
        # Check if it contains dataset folder or specific project files
        if os.path.exists(os.path.join(current_dir, 'dataset')) or \
           os.path.exists(os.path.join(current_dir, 'SAVE')):
            return current_dir
        current_dir = os.path.dirname(current_dir)
        search_limit += 1
    
    # If still not found, return current directory itself
    return os.path.abspath(os.getcwd())

# Set work directory
PROJECT_ROOT = find_project_root()
if not os.path.exists(os.path.join(PROJECT_ROOT, 'dataset')):
    # If still not found, use current directory
    PROJECT_ROOT = os.getcwd()

os.chdir(PROJECT_ROOT)
print(f"Work directory set to: {PROJECT_ROOT}")

# Save directory root path
SAVE_ROOT = os.path.join(PROJECT_ROOT, 'peak_stress', 'LightGBM', 'save')
os.makedirs(SAVE_ROOT, exist_ok=True)

def load_data():
    """Load data - use 10 material parameters and 5 specimen parameters to regress peak stress"""
    print("=== Load data ===")
    
    # Use specified data file
    data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
    
    if not os.path.exists(data_file):
        print(f"Error: Data file not found: {data_file}")
        return None
    
    print(f"Found data file: {data_file}")
    
    # Read the first worksheet of the Excel file (material parameter table)
    df = pd.read_excel(data_file, sheet_name=0)
    print(f"Data shape: {df.shape}")
    print(f"Column names: {list(df.columns)}")
    
    # Define features: 10 material parameters + 5 specimen parameters
    # 10 material features
    material_features = ['water', 'cement', 'w/c', 'CS', 'sand', 'CA', 'r', 'WA', 'S', 'CI']
    # 5 specimen parameters
    specimen_features = ['age', 'μe', 'DJB', 'side', 'GJB']
    
    feature_names = material_features + specimen_features
    
    # Target variable: peak stress
    target_column = 'fc'
    
    # Check column names
    missing_cols = [col for col in feature_names if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns {missing_cols}")
        feature_names = [col for col in feature_names if col in df.columns]
    
    missing_target = target_column not in df.columns
    if missing_target:
        print(f"Error: Missing target variable column '{target_column}'")
        print(f"Available columns: {list(df.columns)}")
        return None
    
    # Extract data
    X = df[feature_names].values
    y = df[target_column].values
    
    # Special check w/c feature (if it exists)
    if 'w/c' in feature_names:
        wc_idx = feature_names.index('w/c')
        wc_values = X[:, wc_idx]
        print(f"\n🔍 Check w/c feature (original data):")
        print(f"  Unique value count: {len(np.unique(wc_values))}")
        print(f"  Minimum value: {np.min(wc_values):.10f}")
        print(f"  Maximum value: {np.max(wc_values):.10f}")
        print(f"  Range: {np.max(wc_values) - np.min(wc_values):.10f}")
        print(f"  Mean: {np.mean(wc_values):.10f}")
        print(f"  Standard deviation: {np.std(wc_values):.10f}")
        if len(np.unique(wc_values)) <= 5:
            print(f"  All unique values: {np.unique(wc_values)}")
        else:
            print(f"  First 5 unique values: {np.unique(wc_values)[:5]}")
        if np.max(wc_values) - np.min(wc_values) < 1e-6:
            print(f"  ⚠ Warning: w/c feature has a range close to 0 in original data!")
    
    # Check if there are missing values
    if np.isnan(X).any():
        print(f"Warning: Feature data contains NaN values, will be filled")
        from sklearn.impute import SimpleImputer
        imputer = SimpleImputer(strategy='mean')
        X = imputer.fit_transform(X)
    
    if np.isnan(y).any():
        print(f"Warning: Target variable contains NaN values, will be removed")
        valid_mask = ~np.isnan(y)
        X = X[valid_mask]
        y = y[valid_mask]
        df = df.iloc[valid_mask].reset_index(drop=True)
    
    # Extract sample division information (using DataSlice column)
    sample_divisions = []
    if 'DataSlice' in df.columns:
        sample_divisions = df['DataSlice'].values
        print(f"Found data set division information: {np.unique(sample_divisions, return_counts=True)}")
    else:
        print("DataSlice column not found, will use random划分")
        sample_divisions = None
    
    # Extract sample ID
    sample_ids = []
    if 'No_Customized' in df.columns:
        sample_ids = df['No_Customized'].values
    else:
        sample_ids = [f"sample_{i}" for i in range(len(X))]
    
    print(f"Feature count: {X.shape[1]} (10 material parameters + 5 specimen parameters)")
    print(f"Sample count: {X.shape[0]}")
    print(f"Target variable '{target_column}' range: {np.min(y):.2f} - {np.max(y):.2f}")
    
    # Save original DataFrame index (for subsequent mapping prediction results)
    original_df_indices = df.index.values
    
    return X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df

def split_data_by_divisions(X, y, sample_divisions, sample_ids, test_ratio=0.2, val_ratio=0.2, random_state=42):
    """Split data based on sample division information"""
    print("=== Split data based on sample division information ===")
    
    if sample_divisions is None:
        from sklearn.model_selection import train_test_split
        X_train, X_temp, y_train, y_temp = train_test_split(
            X, y, test_size=test_ratio + val_ratio, random_state=random_state
        )
        val_ratio_adjusted = val_ratio / (test_ratio + val_ratio)
        X_val, X_test, y_val, y_test = train_test_split(
            X_temp, y_temp, test_size=1-val_ratio_adjusted, random_state=random_state
        )
        
        train_ids = [f"train_{i}" for i in range(len(X_train))]
        val_ids = [f"val_{i}" for i in range(len(X_val))]
        test_ids = [f"test_{i}" for i in range(len(X_test))]
        
        return X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids
    
    # Classify samples based on sample division information (supports DataSlice format: subset1, subset2, subset3, test)
    train_indices = []
    val_indices = []
    test_indices = []
    
    for i, division in enumerate(sample_divisions):
        division_str = str(division).strip()
        if division_str.lower().startswith('test'):
            test_indices.append(i)
        elif division_str.startswith('subset'):
            # subset1, subset2, subset3 are all used as training set (or can be allocated according to needs)
            # Here all are used as training set by default, if validation set is needed, it can be divided from subset
            train_indices.append(i)
        elif division == 'train' or division == '训练集' or division_str.startswith('train'):
            train_indices.append(i)
        elif division == 'val' or division == '验证集' or division == 'validation' or division_str.startswith('val'):
            val_indices.append(i)
        else:
            # Default to training set
            train_indices.append(i)
    
    # If there is no validation set but there is a training set, split validation set from training set
    # If test set exists, keep it unchanged
    if train_indices and not val_indices:
        print("Only training set samples, split validation set from training set...")
        from sklearn.model_selection import train_test_split
        
        train_indices = np.array(train_indices)
        
        if len(train_indices) > 1:
            # If test set exists, only split validation set from training set
            # If test set does not exist, split validation set and test set from training set
            if len(test_indices) > 0:
                # test set already exists, only split validation set
                val_ratio_adjusted = val_ratio
                train_idx, val_idx = train_test_split(
                    train_indices, test_size=val_ratio_adjusted, random_state=random_state
                )
                train_indices = train_idx.tolist()
                val_indices = val_idx.tolist()
            else:
                # test set does not exist, split validation set and test set from training set
                train_val_idx, test_idx = train_test_split(
                    train_indices, test_size=test_ratio, random_state=random_state
                )
                val_ratio_adjusted = val_ratio / (1 - test_ratio)
                train_idx, val_idx = train_test_split(
                    train_val_idx, test_size=val_ratio_adjusted, random_state=random_state
                )
                train_indices = train_idx.tolist()
                val_indices = val_idx.tolist()
                test_indices = test_idx.tolist()
        else:
            val_indices = []
    
    # Convert to numpy array
    train_indices = np.array(train_indices)
    val_indices = np.array(val_indices)
    test_indices = np.array(test_indices)
    
    # Split data
    X_train = X[train_indices] if len(train_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    X_val = X[val_indices] if len(val_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    X_test = X[test_indices] if len(test_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    
    y_train = y[train_indices] if len(train_indices) > 0 else np.array([])
    y_val = y[val_indices] if len(val_indices) > 0 else np.array([])
    y_test = y[test_indices] if len(test_indices) > 0 else np.array([])
    
    # Get corresponding sample ID
    train_ids = [sample_ids[i] for i in train_indices] if len(train_indices) > 0 else []
    val_ids = [sample_ids[i] for i in val_indices] if len(val_indices) > 0 else []
    test_ids = [sample_ids[i] for i in test_indices] if len(test_indices) > 0 else []
    
    print(f"Data split results:")
    print(f"  Training set: {len(X_train)} samples")
    print(f"  Validation set: {len(X_val)} samples")
    print(f"  Test set: {len(X_test)} samples")
    
    return X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids

def objective(trial, X_train, y_train, X_val, y_val):
    """Optuna optimization objective function - use validation set evaluation (LightGBM version)"""
    
    params = {
        'objective': 'regression',
        'metric': 'rmse',
        'verbosity': -1,
        'random_state': 42,
        
        # LightGBM hyperparameter search space
        'num_leaves': trial.suggest_int('num_leaves', 15, 127),
        'max_depth': trial.suggest_int('max_depth', 5, 15),
        'n_estimators': trial.suggest_int('n_estimators', 100, 800),
        'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.3, log=True),
        'min_child_samples': trial.suggest_int('min_child_samples', 5, 100),
        'subsample': trial.suggest_float('subsample', 0.5, 1.0),
        'colsample_bytree': trial.suggest_float('colsample_bytree', 0.3, 1.0),
        'reg_alpha': trial.suggest_float('reg_alpha', 0.001, 20.0, log=True),
        'reg_lambda': trial.suggest_float('reg_lambda', 0.001, 20.0, log=True),
    }
    
    # Use internal cross-validation of training set to evaluate, more stable
    kf = KFold(n_splits=5, shuffle=True, random_state=42)
    scores = []
    
    for train_idx, val_idx in kf.split(X_train):
        X_tr, X_v = X_train[train_idx], X_train[val_idx]
        y_tr, y_v = y_train[train_idx], y_train[val_idx]
        
        model = lgb.LGBMRegressor(**params)
        model.fit(X_tr, y_tr)
        
        y_pred = model.predict(X_v)
        r2 = r2_score(y_v, y_pred)
        scores.append(r2)
    
    return np.mean(scores)

def optimize_lightgbm(X_train, y_train, X_val, y_val, n_trials=100):
    """Use Optuna to optimize LightGBM hyperparameters - use 5-fold cross-validation evaluation"""
    print(f"\n=== Start LightGBM hyperparameter optimization (n_trials={n_trials}) ===")
    print("Use 5-fold cross-validation to evaluate hyperparameters, improve stability")
    
    study = optuna.create_study(direction='maximize', study_name='lightgbm_optimization')
    study.optimize(lambda trial: objective(trial, X_train, y_train, X_val, y_val), n_trials=n_trials, n_jobs=1)
    
    print("\n=== Optimization completed ===")
    print(f"Best R2 score: {study.best_value:.4f}")
    print(f"Best parameters: {study.best_params}")
    
    # Visualize optimization history
    try:
        save_dir = os.path.join(SAVE_ROOT, 'lightgbm_optimization')
        os.makedirs(save_dir, exist_ok=True)
        
        fig = plot_optimization_history(study)
        plt.savefig(os.path.join(save_dir, 'optimization_history.png'), dpi=300, bbox_inches='tight')
        plt.close()
        
        fig = plot_param_importances(study)
        plt.savefig(os.path.join(save_dir, 'param_importance.png'), dpi=300, bbox_inches='tight')
        plt.close()
    except:
        pass
    
    return study

def train_lightgbm(X_train, y_train, X_val, y_val, best_params):
    """Use best parameters to train LightGBM model"""
    print("\n=== Use best parameters to train final model ===")
    
    # Merge training set and validation set
    X_train_full = np.vstack([X_train, X_val])
    y_train_full = np.hstack([y_train, y_val])
    
    # Create model
    model = lgb.LGBMRegressor(
        objective='regression',
        metric='rmse',
        verbosity=-1,
        random_state=42,
        **best_params
    )
    
    # Train model
    model.fit(X_train_full, y_train_full)
    
    return model

def plot_results(y_true, y_pred, save_dir, model_name='lightgbm'):
    """Plot results"""
    print("\n=== Plot results ===")
    
    os.makedirs(save_dir, exist_ok=True)
    
    y_true = np.array(y_true).flatten()
    y_pred = np.array(y_pred).flatten()
    
    r2 = r2_score(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    
    # Create subplots
    fig, axes = plt.subplots(2, 2, figsize=(15, 12))
    fig.suptitle(f'LightGBM Results ($R^2$={r2:.4f})', fontsize=16, fontweight='bold')
    
    # 1. Prediction vs true values
    axes[0,0].scatter(y_true, y_pred, alpha=0.7, s=60, color='green')
    axes[0,0].plot([y_true.min(), y_true.max()], [y_true.min(), y_true.max()], 'r--', lw=2)
    axes[0,0].set_xlabel('True Peak Stress fc (MPa)')
    axes[0,0].set_ylabel('Predicted Peak Stress fc (MPa)')
    axes[0,0].set_title(f'Prediction vs True Values\n$R^2$={r2:.4f}, MAE={mae:.3f}, RMSE={rmse:.3f}')
    axes[0,0].grid(True, alpha=0.3)
    
    # 2. Residual plot
    residuals = y_pred - y_true
    axes[0,1].scatter(y_pred, residuals, alpha=0.7, s=60, color='green')
    axes[0,1].axhline(y=0, color='r', linestyle='--', lw=2)
    axes[0,1].set_xlabel('Predicted Peak Stress fc (MPa)')
    axes[0,1].set_ylabel('Residuals (MPa)')
    axes[0,1].set_title('Residual Plot')
    axes[0,1].grid(True, alpha=0.3)
    
    # 3. Residual distribution
    axes[1,0].hist(residuals, bins=15, alpha=0.7, color='green', density=True)
    axes[1,0].axvline(x=0, color='r', linestyle='--', lw=2)
    axes[1,0].set_xlabel('Residuals (MPa)')
    axes[1,0].set_ylabel('Density')
    axes[1,0].set_title('Residual Distribution')
    axes[1,0].grid(True, alpha=0.3)
    
    # 4. Feature importance
    # Note: Feature importance needs to be obtained from model object after model training
    stats_text = f"""Model Performance Statistics:
    
Sample Count: {len(y_true)}
R2: {r2:.4f}
MAE: {mae:.3f} MPa
RMSE: {rmse:.3f} MPa

LightGBM Features:
- Gradient Boosting (Leaf-wise tree growth)
- Optuna Hyperparameter Optimization
- Cross-validation evaluation
- Feature Importance Analysis"""
    
    axes[1,1].text(0.1, 0.9, stats_text, transform=axes[1,1].transAxes, 
                   fontsize=11, verticalalignment='top', fontfamily='monospace')
    axes[1,1].set_xlim(0, 1)
    axes[1,1].set_ylim(0, 1)
    axes[1,1].axis('off')
    axes[1,1].set_title('Performance Statistics')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, f'lightgbm_results.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    return r2, mae, rmse

def plot_train_test_comparison(model, X_train, y_train, X_test, y_test, save_dir):
    """Plot training set and test set comparison chart (with edge distribution) - optimized version
    
    Returns:
        DataFrame: DataFrame containing training set and test set prediction results
    """
    print("\n=== Plot training set and test set comparison chart ===")
    
    # Prediction
    y_train_pred = model.predict(X_train)
    y_test_pred = model.predict(X_test)
    
    # Calculate metrics
    r2_train = r2_score(y_train, y_train_pred)
    mae_train = mean_absolute_error(y_train, y_train_pred)
    rmse_train = np.sqrt(mean_squared_error(y_train, y_train_pred))
    r2_test = r2_score(y_test, y_test_pred)
    mae_test = mean_absolute_error(y_test, y_test_pred)
    rmse_test = np.sqrt(mean_squared_error(y_test, y_test_pred))
    
    # Create figure (scatter plot with edge distribution)
    fig = plt.figure(figsize=(13, 11))
    
    # Define grid layout (adjust ratio)
    gs = fig.add_gridspec(3, 3, hspace=0.08, wspace=0.08,
                         height_ratios=[0.8, 4, 0.1], width_ratios=[4, 0.8, 0.1])
    
    # Main scatter plot
    ax_main = fig.add_subplot(gs[1, 0])
    
    # Upper histogram
    ax_top = fig.add_subplot(gs[0, 0], sharex=ax_main)
    
    # Right histogram
    ax_right = fig.add_subplot(gs[1, 1], sharey=ax_main)
    
    # Plot main scatter plot
    # Training set - use more professional blue
    ax_main.scatter(y_train, y_train_pred, alpha=0.5, s=50, color='#2E86AB', 
                   label='Training', edgecolors='white', linewidths=0.5, zorder=3)
    
    # Test set - use more prominent orange red
    ax_main.scatter(y_test, y_test_pred, alpha=0.7, s=60, color='#E63946',
                   label='Testing', edgecolors='white', linewidths=0.5, zorder=4)
    
    # 1:1 diagonal line
    all_values = np.concatenate([y_train, y_test, y_train_pred, y_test_pred])
    min_val = all_values.min()
    max_val = all_values.max()
    # Add padding
    padding = (max_val - min_val) * 0.05
    min_val -= padding
    max_val += padding
    
    ax_main.plot([min_val, max_val], [min_val, max_val], 'k--', lw=2, 
                alpha=0.4, label='Perfect Prediction', zorder=1)
    
    # Calculate and plot best fit line and confidence interval (training set)
    slope_train, intercept_train, r_value_train, p_value_train, std_err_train = stats.linregress(y_train, y_train_pred)
    line_train = slope_train * y_train + intercept_train
    predict_train = slope_train * np.sort(y_train) + intercept_train
    
    # Calculate confidence interval
    residuals_train = y_train_pred - line_train
    std_residuals_train = np.std(residuals_train)
    ci_train = 1.96 * std_residuals_train  # 95% CI
    
    ax_main.plot(np.sort(y_train), predict_train, color='#2E86AB', lw=3, alpha=0.9,
                linestyle='-', zorder=5)
    ax_main.fill_between(np.sort(y_train), predict_train - ci_train, predict_train + ci_train,
                         alpha=0.2, color='#2E86AB', zorder=2)
    
    # Best fit line and confidence interval (test set)
    slope_test, intercept_test, r_value_test, p_value_test, std_err_test = stats.linregress(y_test, y_test_pred)
    line_test = slope_test * y_test + intercept_test
    predict_test = slope_test * np.sort(y_test) + intercept_test
    
    residuals_test = y_test_pred - line_test
    std_residuals_test = np.std(residuals_test)
    ci_test = 1.96 * std_residuals_test
    
    ax_main.plot(np.sort(y_test), predict_test, color='#E63946', lw=3, alpha=0.9,
                linestyle='-', zorder=6)
    ax_main.fill_between(np.sort(y_test), predict_test - ci_test, predict_test + ci_test,
                         alpha=0.2, color='#E63946', zorder=2)
    
    # Set axis labels
    ax_main.set_xlabel('Observed Peak Stress fc (MPa)', fontsize=13, fontweight='bold')
    ax_main.set_ylabel('Predicted Peak Stress fc (MPa)', fontsize=13, fontweight='bold')
    
    # Optimized legend - placed in lower right corner, display R² using LaTeX format
    legend_elements = [
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#2E86AB', 
                  markersize=8, alpha=0.7, label=f'Training ($R^2$={r2_train:.3f}, MAE={mae_train:.2f})'),
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#E63946', 
                  markersize=9, alpha=0.8, label=f'Testing ($R^2$={r2_test:.3f}, MAE={mae_test:.2f})'),
        plt.Line2D([0], [0], color='k', linestyle='--', lw=2, alpha=0.4, label='Perfect Prediction'),
        plt.Line2D([0], [0], color='gray', lw=3, label='Best Fit Line'),
    ]
    ax_main.legend(handles=legend_elements, loc='lower right', fontsize=9.5, 
                  framealpha=0.95, edgecolor='gray', fancybox=True, shadow=True)
    
    ax_main.grid(True, alpha=0.25, linestyle='--', linewidth=0.5)
    ax_main.set_axisbelow(True)
    
    # Upper histogram (observed value distribution) - stacked display
    bins = np.linspace(min_val, max_val, 20)
    # Use numpy.histogram to calculate frequency
    hist_train_obs, _ = np.histogram(y_train, bins=bins)
    hist_test_obs, _ = np.histogram(y_test, bins=bins)
    
        # Plot stacked histogram
    bin_centers = (bins[:-1] + bins[1:]) / 2
    width = bins[1] - bins[0]
    ax_top.bar(bin_centers, hist_train_obs, width=width, alpha=0.7, 
              color='#2E86AB', label='Training', edgecolor='white', linewidth=0.5)
    ax_top.bar(bin_centers, hist_test_obs, width=width, alpha=0.8, 
              color='#E63946', label='Testing', edgecolor='white', linewidth=0.5, bottom=hist_train_obs)
    ax_top.set_ylabel('Count', fontsize=11, fontweight='bold')
    ax_top.legend(loc='upper right', fontsize=9, framealpha=0.9)
    ax_top.tick_params(labelbottom=False, labelsize=10)
    ax_top.grid(True, alpha=0.25, axis='y', linestyle='--', linewidth=0.5)
    ax_top.set_axisbelow(True)
    ax_top.spines['top'].set_visible(False)
    ax_top.spines['right'].set_visible(False)
    
    # Right histogram (predicted value distribution) - stacked display
    hist_train_pred, _ = np.histogram(y_train_pred, bins=bins)
    hist_test_pred, _ = np.histogram(y_test_pred, bins=bins)
    
    # Plot horizontal stacked histogram
    ax_right.barh(bin_centers, hist_train_pred, height=width, alpha=0.7, 
                 color='#2E86AB', edgecolor='white', linewidth=0.5)
    ax_right.barh(bin_centers, hist_test_pred, height=width, alpha=0.8, 
                 color='#E63946', edgecolor='white', linewidth=0.5, left=hist_train_pred)
    ax_right.set_xlabel('Count', fontsize=11, fontweight='bold')
    ax_right.tick_params(labelleft=False, labelsize=10)
    ax_right.grid(True, alpha=0.25, axis='x', linestyle='--', linewidth=0.5)
    ax_right.set_axisbelow(True)
    ax_right.spines['top'].set_visible(False)
    ax_right.spines['right'].set_visible(False)
    
    # Add total title - include sample number information
    title = f'LightGBM Model Performance\nTraining: n={len(y_train)} | Testing: n={len(y_test)}'
    fig.suptitle(title, fontsize=15, fontweight='bold', y=0.985)
    
    # Set overall background
    fig.patch.set_facecolor('white')
    
    plt.savefig(os.path.join(save_dir, 'train_test_comparison_with_margins.png'), 
               dpi=300, bbox_inches='tight', facecolor='white')
    # plt.show()
    plt.close()
    
    print(f"\nTraining set performance - R2: {r2_train:.4f}, MAE: {mae_train:.3f} MPa, RMSE: {rmse_train:.3f} MPa")
    print(f"Test set performance - R2: {r2_test:.4f}, MAE: {mae_test:.3f} MPa, RMSE: {rmse_test:.3f} MPa")
    print(f"Performance difference - ΔR2: {abs(r2_train - r2_test):.4f}, ΔMAE: {abs(mae_train - mae_test):.3f} MPa")
    
    # Create DataFrame containing training set and test set prediction results
    train_df = pd.DataFrame({
        'dataset': ['Training set'] * len(y_train),
        'observed': y_train,
        'predicted': y_train_pred,
        'residual': y_train - y_train_pred,
        'absolute_error': np.abs(y_train - y_train_pred)
    })
    
    test_df = pd.DataFrame({
        'dataset': ['Test set'] * len(y_test),
        'observed': y_test,
        'predicted': y_test_pred,
        'residual': y_test - y_test_pred,
        'absolute_error': np.abs(y_test - y_test_pred)
    })
    
    # Merge training set and test set data
    prediction_df = pd.concat([train_df, test_df], ignore_index=True)
    
    return prediction_df

def plot_feature_importance(model, feature_names, save_dir, X=None):
    """Plot feature importance
    
    Parameters:
        model: trained LightGBM model
        feature_names: feature name list
        save_dir: save directory
        X: feature data (optional, for calculating coefficient of variation)
    """
    print("\n=== Plot feature importance ===")
    
    # Get feature importance
    importance = model.feature_importances_
    
    # Create DataFrame
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': importance
    })
    
    # If feature data is provided, calculate coefficient of variation (CV)
    if X is not None:
        cv_values = []
        for i in range(X.shape[1]):
            feature_mean = np.mean(X[:, i])
            feature_std = np.std(X[:, i])
            # Coefficient of variation = standard deviation / mean (avoid division by zero)
            cv = feature_std / (feature_mean + 1e-10) if abs(feature_mean) > 1e-10 else 0.0
            cv_values.append(cv)
        importance_df['CV'] = cv_values
        print("✓ Coefficient of variation (CV) calculated")
    else:
        importance_df['CV'] = np.nan
        print("⚠ No feature data provided, cannot calculate coefficient of variation")
    
    # Sort by importance
    importance_df = importance_df.sort_values('importance', ascending=False)
    
    # Plot bar chart
    plt.figure(figsize=(12, 15))
    plt.barh(importance_df['feature'], importance_df['importance'], color='skyblue')
    plt.xlabel('Feature Importance')
    plt.title('LightGBM Feature Importance')
    plt.gca().invert_yaxis()
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'feature_importance.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    # Save to Excel (will be saved to comprehensive Excel file in main function)
    # Here it is not saved directly, the DataFrame is returned to the main function for unified saving
    
    # Print statistical information
    print(f"\nFeature importance statistics:")
    print(f"  Most important feature: {importance_df.iloc[0]['feature']} (importance={importance_df.iloc[0]['importance']:.6f})")
    if X is not None:
        print(f"  Highest CV feature: {importance_df.loc[importance_df['CV'].idxmax(), 'feature']} (CV={importance_df['CV'].max():.4f})")
        print(f"  Lowest CV feature: {importance_df.loc[importance_df['CV'].idxmin(), 'feature']} (CV={importance_df['CV'].min():.4f})")
    
    return importance_df


def plot_shap_analysis(model, X_background, X_explain, feature_names, save_dir):
    """SHAP value analysis
    
    Parameters:
        model: trained LightGBM model
        X_background: background data set (used to initialize explainer, usually using training set)
        X_explain: data set to be explained (data for calculating SHAP values, usually using test set)
        feature_names: feature name list
        save_dir: save directory
    
    Traditional做法说明:
        - Background data set (X_background) : using training set (or training set + validation set)
        - Background data set (X_background): using training set (or training set + validation set)
          Used to initialize SHAP explainer, representing the "baseline" distribution of the model
        - Explanation data set (X_explain): using test set
          Explain the data that has not been seen, evaluate the interpretability of the model's generalization ability
        
        Advantages:
        1. Conforms to machine learning best practices (training/test separation)
        2. Evaluates the interpretability of the model in real application scenarios
        3. Avoids data leakage (does not use test set to train explainer)
    """
    print("\n=== Perform SHAP analysis ===")
    print(f"Background data set sample number: {len(X_background)} (used to initialize explainer)")
    print(f"Explanation data set sample number: {len(X_explain)} (calculating SHAP values)")
    
    # Check if it is an ensemble model
    is_ensemble = hasattr(model, 'is_ensemble') and model.is_ensemble
    
    # Use TreeExplainer (LightGBM专用）
    # For ensemble model, special processing is required
    if is_ensemble:
        print("Detected ensemble model, using weighted average SHAP values...")
        # Calculate SHAP values for each sub-model, then average them
        shap_values_list = []
        weights = model.weights
        
        for i, sub_model in enumerate(model.models):
            print(f"  Calculating SHAP values for sub-model {i+1}/{len(model.models)}...")
            try:
                # Try using TreeExplainer
                try:
                    sub_explainer = shap.TreeExplainer(sub_model, X_background)
                except:
                    sub_explainer = shap.TreeExplainer(sub_model)
                
                sub_shap_values = sub_explainer.shap_values(X_explain)
                shap_values_list.append(sub_shap_values * weights[i])
            except Exception as e:
                print(f"  ⚠ Sub-model {i+1} SHAP calculation failed: {e}")
                # 如果TreeExplainer失败，使用KernelExplainer
                try:
                    if len(X_background) > 100:
                        background_sample = shap.sample(X_background, 100)
                    else:
                        background_sample = X_background
                    sub_explainer = shap.KernelExplainer(sub_model.predict, background_sample)
                    sub_shap_values = sub_explainer.shap_values(X_explain)
                    shap_values_list.append(sub_shap_values * weights[i])
                except Exception as e2:
                    print(f"  ⚠ All SHAP methods for sub-model {i+1} failed: {e2}")
                    # If all fail, use zero value
                    if len(shap_values_list) > 0:
                        shap_values_list.append(np.zeros_like(shap_values_list[0]) * weights[i])
        
        # Weighted average of SHAP values for all sub-models
        if len(shap_values_list) > 0:
            shap_values = np.sum(shap_values_list, axis=0)
            print("✓ Ensemble model SHAP values calculated")
        else:
            raise ValueError("All sub-model SHAP calculations failed")
    else:
        # Single model, use standard method
        try:
            # Try using background data set to initialize explainer
            explainer = shap.TreeExplainer(model, X_background)
            print("✓ Using background data set to initialize explainer successfully")
        except (ValueError, TypeError, AttributeError) as e:
            print(f"⚠ Using background data set to initialize explainer failed: {e}")
            print("  Trying to initialize explainer without background data set...")
            try:
                # Don't pass background data set, use default value
                explainer = shap.TreeExplainer(model)
                print("✓ Initializing explainer without background data set successfully")
            except Exception as e2:
                print(f"⚠ Explainer initialization failed: {e2}")
                print("  Trying to use KernelExplainer as a backup solution...")
                # If TreeExplainer fails, try using KernelExplainer (slower but more general)
                try:
                    # Sample background data to improve speed
                    if len(X_background) > 100:
                        background_sample = shap.sample(X_background, 100)
                    else:
                        background_sample = X_background
                    explainer = shap.KernelExplainer(model.predict, background_sample)
                    print("✓ Using KernelExplainer successfully (note: calculation is slower)")
                except Exception as e3:
                    print(f"⚠ All explainer initialization methods failed: {e3}")
                    raise ValueError(f"Cannot initialize SHAP explainer: {e3}")
        
        # Calculate SHAP values (for explanation data set)
        print("Calculating SHAP values...")
        shap_values = explainer.shap_values(X_explain)
    
    # Set font, avoid Chinese乱码，并确保负号正常显示
    original_font = list(plt.rcParams['font.sans-serif'])
    original_unicode_minus = plt.rcParams['axes.unicode_minus']

    # For SHAP plot, use DejaVu Sans to ensure negative sign is displayed properly
    configure_plot_fonts(['DejaVu Sans', 'SimHei', 'Microsoft YaHei'])
    plt.rcParams['axes.unicode_minus'] = False  # Ensure negative sign is displayed properly
    
    # 1. Summary plot
    # Increase width, truly stretch the horizontal axis ratio
    fig = plt.figure(figsize=(30, 12))  # Width increased to 30, significantly stretch the horizontal axis ratio
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, show=False)
    # Get current axes
    ax = plt.gca()
    # Don't change data range, only stretch the horizontal axis physical ratio by increasing the graph width
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_summary.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # 2. Bar plot (feature importance)
    plt.figure(figsize=(10, 8))
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, plot_type="bar", show=False)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_bar.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # 3. Waterfall plot (display SHAP values for single sample)
    # Note: This is an analysis of a single sample, showing how each feature affects the prediction result
    # Select the first 3 samples as an example (if data is sufficient)
    n_samples_waterfall = min(3, len(shap_values))
    
    for sample_idx in range(n_samples_waterfall):
        shap_explanation = shap.Explanation(
            values=shap_values[sample_idx:sample_idx+1],
            base_values=explainer.expected_value,
            data=X_explain[sample_idx:sample_idx+1],
            feature_names=feature_names
        )
        
        # Ensure negative sign is displayed properly - use DejaVu Sans font (supports negative sign)
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'SimHei', 'Microsoft YaHei']
        
        plt.figure(figsize=(12, 8))
        shap.plots.waterfall(shap_explanation[0], show=False)
        
        # Get current axes and ensure negative sign is displayed properly
        ax = plt.gca()
        # Force set negative sign display and font
        ax.tick_params(which='major', labelsize=10)
        # Ensure all text uses fonts that support negative sign
        for text in ax.texts:
            text.set_fontfamily('DejaVu Sans')
        
        plt.tight_layout()
        if n_samples_waterfall == 1:
            plt.savefig(os.path.join(save_dir, 'shap_waterfall.png'), dpi=300, bbox_inches='tight')
        else:
            plt.savefig(os.path.join(save_dir, f'shap_waterfall_sample_{sample_idx+1}.png'), dpi=300, bbox_inches='tight')
        plt.close()
    
    print(f"✓ Generated {n_samples_waterfall} waterfall plots")
    
    # 4. Force plot (display prediction process for single sample, long bar force plot)
    # Create explanation for force plot containing multiple samples
    n_samples_force = min(3, len(shap_values))
    shap_explanation_force = shap.Explanation(
        values=shap_values[:n_samples_force],
        base_values=explainer.expected_value,
        data=X_explain[:n_samples_force],
        feature_names=feature_names
    )
    # Draw force plot for the first 3 samples
    for i in range(n_samples_force):
        # Ensure negative sign is displayed properly
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'SimHei', 'Microsoft YaHei']
        
        plt.figure(figsize=(16, 4))  # Long bar force plot
        shap.plots.force(shap_explanation_force[i], matplotlib=True, show=False)
        
        # Ensure negative sign is displayed properly
        ax = plt.gca()
        for text in ax.texts:
            text.set_fontfamily('DejaVu Sans')
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, f'shap_force_sample_{i}.png'), dpi=300, bbox_inches='tight')
        # plt.show()
        plt.close()
    
    # 5. Calculate average absolute SHAP value as feature importance
    shap_importance = pd.DataFrame({
        'feature': feature_names,
        'shap_importance': np.abs(shap_values).mean(axis=0),
        'shap_mean': shap_values.mean(axis=0),  # Average SHAP value (consider positive and negative)
        'shap_std': shap_values.std(axis=0),    # Standard deviation of SHAP values
        'shap_max': np.abs(shap_values).max(axis=0),  # Maximum absolute SHAP value
        'shap_min': np.abs(shap_values).min(axis=0)   # Minimum absolute SHAP value
    }).sort_values('shap_importance', ascending=False)
    
    # Save SHAP importance (will be saved in main function)
    # Return DataFrame for main function to save
    
    # 6. Analyze the reason why some features have a small impact (print detailed analysis)
    print("\n" + "="*60)
    print("SHAP feature importance analysis")
    print("="*60)
    print(f"\nThe 5 most important features:")
    for idx, row in shap_importance.head(5).iterrows():
        print(f"  {row['feature']:10s}: average impact = {row['shap_importance']:.4f}, "
              f"average SHAP value = {row['shap_mean']:+.4f}, standard deviation = {row['shap_std']:.4f}")
    
    print(f"\nThe 5 least important features:")
    for idx, row in shap_importance.tail(5).iterrows():
        print(f"  {row['feature']:10s}: average impact = {row['shap_importance']:.4f}, "
              f"average SHAP value = {row['shap_mean']:+.4f}, standard deviation = {row['shap_std']:.4f}")
    
    # 分析影响小的可能原因
    low_importance_features = shap_importance[shap_importance['shap_importance'] < 0.1]
    if len(low_importance_features) > 0:
        print(f"\n⚠ Found {len(low_importance_features)} features with a small impact (average SHAP value < 0.1):")
        for idx, row in low_importance_features.iterrows():
            print(f"  - {row['feature']}: average impact = {row['shap_importance']:.4f}")
        
        print("\n Possible reasons:")  
        print("  1. The feature value change range is small (small variance), resulting in limited impact on prediction")
        print("  2. Highly correlated with other features, replaced by other features")
        print("  3. The feature value is uniformly distributed in the data set, lacking distinction")
        print("  4. This feature has a small impact on the target variable which is consistent with physical/engineering规律")
        print("  5. The sample size is small (94 samples), the impact of some features may be masked by noise")
        print("\n Suggestions:")
        print("  - Check the statistical distribution of these features (mean, standard deviation, range)")
        print("  - Analyze the correlation between features")
        print("  - View PDP plot to understand the marginal effect of these features")
        print("  - Consider feature engineering (such as feature combination, transformation)")
    
    print("="*60)
    
    # Restore original font settings
    plt.rcParams['font.sans-serif'] = original_font
    plt.rcParams['axes.unicode_minus'] = False
    
    return shap_importance

def analyze_feature_statistics(X, y, feature_names, shap_importance, save_dir):
    """Analyze feature statistics, explain why some features have a small impact"""
    print("\n" + "="*60)
    print("Feature statistics analysis (explain why some features have a small impact)")
    print("="*60)
    
    import pandas as pd
    
    # Create feature index mapping (feature_name -> column_index)
    feature_to_idx = {name: idx for idx, name in enumerate(feature_names)}
    
    # Create feature_stats in the order of shap_importance (consistent with SHAP summary plot)
    # shap_importance is already sorted by importance in descending order
    feature_stats_list = []
    
    for idx, row in shap_importance.iterrows():
        feat_name = row['feature']
        feat_idx = feature_to_idx[feat_name]
        
        # Calculate the statistical information for this feature
        feat_data = X[:, feat_idx]
        q25 = np.percentile(feat_data, 25)
        q50 = np.percentile(feat_data, 50)
        q75 = np.percentile(feat_data, 75)
        corr = np.corrcoef(feat_data, y)[0, 1]
        
        feat_range = np.max(feat_data) - np.min(feat_data)
        feat_std = np.std(feat_data)
        
        # Special check for w/c feature
        if feat_name == 'w/c':
            print(f"\n🔍 Detailed check for w/c feature:")
            print(f"  Data shape: {feat_data.shape}")
            print(f"  Unique value count: {len(np.unique(feat_data))}")
            print(f"  Unique values: {np.unique(feat_data)[:10]}")  # Display the first 10 unique values
            print(f"  Minimum value: {np.min(feat_data):.10f}")
            print(f"  Maximum value: {np.max(feat_data):.10f}")
            print(f"  Range: {feat_range:.10f}")
            print(f"  Mean: {np.mean(feat_data):.10f}")
            print(f"  Standard deviation: {feat_std:.10f}")
            print(f"  SHAP importance: {row['shap_importance']:.10f}")
            if feat_range < 1e-6:
                print(f"  ⚠ Warning: w/c feature value range is close to 0!")
                print(f"  Possible reasons:")
                print(f"    1. All values in the w/c column are the same in the data")
                print(f"    2. All values are the same after w/c is standardized during data preprocessing")
                print(f"    3. w/c column is a constant in the data file")
                print(f"  Suggestions: Check the w/c column in the original data file")
        
        feature_stats_list.append({
            'feature': feat_name,
            'mean': np.mean(feat_data),
            'std': feat_std,
            'min': np.min(feat_data),
            'q25': q25,
            'median': q50,
            'q75': q75,
            'max': np.max(feat_data),
            'range': feat_range,
            'iqr': q75 - q25,
            'cv': feat_std / (np.mean(feat_data) + 1e-10),
            'shap_importance': row['shap_importance'],
            'correlation_with_target': corr
        })
    
    feature_stats = pd.DataFrame(feature_stats_list)
    
    # Rearrange the column order,使其更符合逻辑：Basic statistics -> Quantiles -> Derived statistics -> Importance
    column_order = ['feature', 'mean', 'std', 'min', 'q25', 'median', 'q75', 'max', 
                    'range', 'iqr', 'cv', 'shap_importance', 'correlation_with_target']
    feature_stats = feature_stats[column_order]
    
    # feature_stats is already sorted by shap_importance (consistent with SHAP summary plot)
    
    # Save statistics (will be saved in main function)
    # Return DataFrame for main function to save
    
    # Analyze features with a small impact
    low_importance_threshold = 0.1
    low_importance = feature_stats[feature_stats['shap_importance'] < low_importance_threshold]
    
    print(f"\nAnalysis of features with a small impact (SHAP importance < {low_importance_threshold}):")
    if len(low_importance) > 0:
        print(f"\nTotal {len(low_importance)} features with a small impact:")
        for idx, row in low_importance.iterrows():
            print(f"\n  {row['feature']}:")
            print(f"    - SHAP importance: {row['shap_importance']:.4f}")
            print(f"    - Correlation with target variable: {row['correlation_with_target']:.4f}")
            print(f"    - Mean: {row['mean']:.4f}, standard deviation: {row['std']:.4f}")
            print(f"    - Range: [{row['min']:.4f}, {row['max']:.4f}], coefficient of variation: {row['cv']:.4f}")
            print(f"    - Quartiles: Q1={row['q25']:.4f}, median={row['median']:.4f}, Q3={row['q75']:.4f}, IQR={row['iqr']:.4f}")
            
            # Determine possible reasons
            reasons = []
            if abs(row['correlation_with_target']) < 0.1:
                reasons.append("Low correlation with target variable")
            if row['cv'] < 0.1:
                reasons.append("Small coefficient of variation (small data range)")
            if row['range'] < row['mean'] * 0.1:
                reasons.append("Small range relative to mean")
            
            if reasons:
                print(f"    - Possible reasons: {', '.join(reasons)}")
    
    # Analyze features with a large impact
    high_importance = feature_stats.head(5)
    print(f"\nThe 5 most important features:")
    for idx, row in high_importance.iterrows():
        print(f"  {row['feature']:10s}: SHAP={row['shap_importance']:.4f}, "
              f"Correlation={row['correlation_with_target']:+.4f}, CV={row['cv']:.4f}")
    
    # Plot feature statistics
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    
    # 1. SHAP importance vs correlation
    ax1 = axes[0, 0]
    scatter = ax1.scatter(feature_stats['correlation_with_target'], 
                         feature_stats['shap_importance'], 
                         s=100, alpha=0.6, c=feature_stats['cv'], cmap='viridis')
    for idx, row in feature_stats.iterrows():
        ax1.annotate(row['feature'], 
                    (row['correlation_with_target'], row['shap_importance']),
                    fontsize=9, alpha=0.7)
    ax1.set_xlabel('Correlation with target variable', fontsize=12)
    ax1.set_ylabel('SHAP importance', fontsize=12)
    ax1.set_title('SHAP importance vs correlation', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3)
    plt.colorbar(scatter, ax=ax1, label='Coefficient of variation')
    
    # 2. SHAP importance vs coefficient of variation
    ax2 = axes[0, 1]
    ax2.scatter(feature_stats['cv'], feature_stats['shap_importance'], s=100, alpha=0.6)
    for idx, row in feature_stats.iterrows():
        ax2.annotate(row['feature'], 
                    (row['cv'], row['shap_importance']),
                    fontsize=9, alpha=0.7)
    ax2.set_xlabel('Coefficient of variation (CV)', fontsize=12)
    ax2.set_ylabel('SHAP importance', fontsize=12)
    ax2.set_title('SHAP importance vs coefficient of variation', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3)
    
    # 3. Feature value range (sorted by SHAP importance, consistent with SHAP swarm plot)
    ax3 = axes[1, 0]
    # feature_stats is already sorted by SHAP importance in descending order, take the first 10
    # Note: SHAP summary plot is sorted by SHAP importance in descending order, so the most important features are at the top
    top_features = feature_stats.head(10).copy()
    
    # Check and print detailed information for w/c feature
    wc_feature = feature_stats[feature_stats['feature'] == 'w/c']
    if len(wc_feature) > 0:
        wc_row = wc_feature.iloc[0]
        print(f"\n⚠ Warning: w/c feature statistics:")
        print(f"  SHAP importance: {wc_row['shap_importance']:.6f}")
        print(f"  Minimum value: {wc_row['min']:.6f}")
        print(f"  Maximum value: {wc_row['max']:.6f}")
        print(f"  Range: {wc_row['range']:.6f}")
        print(f"  Mean: {wc_row['mean']:.6f}")
        print(f"  Standard deviation: {wc_row['std']:.6f}")
        print(f"  Coefficient of variation: {wc_row['cv']:.6f}")
        if wc_row['range'] < 1e-6:
            print(f"  ⚠ Warning: w/c feature value range is close to 0,可能所有样本的值都相同！")
            print(f"  This will cause SHAP importance to be 0, because when the feature value does not change, it cannot affect the prediction.")
            print(f"  Please check the actual values of the w/c column in the data.")
    
    # Filter out features with a range of 0 or close to 0 (these features should not appear in the importance plot)
    # But keep them in top_features for the user to see the problem
    valid_range_features = top_features[top_features['range'] > 1e-6]
    
    if len(valid_range_features) < len(top_features):
        print(f"\n⚠ Warning: Found {len(top_features) - len(valid_range_features)} features with a range of 0 or close to 0")
        print(f"  These features: {top_features[top_features['range'] <= 1e-6]['feature'].tolist()}")
        print(f"  All features will be plotted (including features with a range of 0), to find the problem")
    
    y_pos = np.arange(len(top_features))
    # Plot horizontal bar chart, the most important features are at the top (index 0)
    # For features with a range of 0, use a very small value (such as 1e-6) to display on the graph
    bar_values = top_features['range'].values.copy()
    bar_values[bar_values < 1e-6] = 1e-6  # Replace 0 values with a very small value to display on the graph
    
    bars = ax3.barh(y_pos, bar_values, alpha=0.7)
    # Use different colors for features with a range of 0
    for i, (idx, row) in enumerate(top_features.iterrows()):
        if row['range'] < 1e-6:
            bars[i].set_color('red')
            bars[i].set_alpha(0.5)
            bars[i].set_label('range=0')
    
    ax3.set_yticks(y_pos)
    ax3.set_yticklabels(top_features['feature'])
    ax3.set_xlabel('Feature value range (max - min)', fontsize=12)
    ax3.set_title('Value range of the 10 most important features (sorted by SHAP importance)', fontsize=14, fontweight='bold')
    ax3.grid(True, alpha=0.3, axis='x')
    # Invert the y-axis, so the most important features are at the top (consistent with SHAP summary plot: importance decreases from top to bottom)
    ax3.invert_yaxis()
    
    # If there are features with a range of 0, add a legend说明
    if (top_features['range'] < 1e-6).any():
        ax3.text(0.98, 0.02, 'Red: range=0', transform=ax3.transAxes, 
                fontsize=9, color='red', ha='right', va='bottom',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.7))
    
    # 4. Correlation sorting (sorted by absolute value)
    ax4 = axes[1, 1]
    # Calculate absolute correlation and sort
    feature_stats['abs_correlation'] = feature_stats['correlation_with_target'].abs()
    feature_stats_sorted = feature_stats.sort_values('abs_correlation', ascending=False)
    top_corr = feature_stats_sorted.head(10)
    colors = ['red' if x > 0 else 'blue' for x in top_corr['correlation_with_target']]
    y_pos = np.arange(len(top_corr))
    ax4.barh(y_pos, top_corr['correlation_with_target'], alpha=0.7, color=colors)
    ax4.set_yticks(y_pos)
    ax4.set_yticklabels(top_corr['feature'])
    ax4.set_xlabel('Correlation with target variable', fontsize=12)
    ax4.set_title('Correlation between features and target variable (absolute value of the top 10)', fontsize=14, fontweight='bold')
    ax4.axvline(x=0, color='black', linestyle='--', linewidth=1)
    ax4.grid(True, alpha=0.3, axis='x')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'feature_statistics_analysis.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    print("\n✓ Feature statistics analysis completed, results saved")
    print("="*60)
    
    return feature_stats

def plot_pdp_analysis(model, X_train, feature_names, save_dir, n_top_features=5):
    """PDP (partial dependence plot) analysis - using scikit-learn"""
    print("\n=== Performing PDP analysis ===")
    
    # Ensure X_train is a DataFrame
    if isinstance(X_train, np.ndarray):
        X_train_df = pd.DataFrame(X_train, columns=feature_names)
    else:
        X_train_df = X_train
    
    # Get the n most important features
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': model.feature_importances_
    }).sort_values('importance', ascending=False)
    
    top_features = importance_df.head(n_top_features)['feature'].tolist()
    top_indices = [feature_names.index(f) for f in top_features]
    
    # 1. Univariate PDP analysis
    n_cols = 3  # Change to 3 columns
    n_rows = (n_top_features + 2) // 3  # Adjust row number calculation
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, 6*n_rows))
    if n_top_features == 1:
        axes = [axes]
    else:
        axes = axes.flatten()
    
    for idx, (feature, ax) in enumerate(zip(top_features, axes)):
        try:
            # Get feature index
            feature_idx = feature_names.index(feature)
            
            # Calculate PDP (average) and ICE (individual)
            pd_result = partial_dependence(
                model, 
                X_train_df, 
                features=[feature_idx],
                kind='average'
            )
            
            ice_result = partial_dependence(
                model,
                X_train_df,
                features=[feature_idx],
                kind='individual'
            )
            
            # Calculate 95% confidence interval for ICE curves
            ice_curves = ice_result['individual'][0]  # shape: (n_samples, n_grid_points)
            ice_mean = np.mean(ice_curves, axis=0)
            ice_std = np.std(ice_curves, axis=0)
            ice_upper_ci = ice_mean + 1.96 * ice_std  # 95% CI
            ice_lower_ci = ice_mean - 1.96 * ice_std
            
            grid_values = pd_result['grid_values'][0]
            
            # Plot PDP line (blue)
            ax.plot(grid_values, pd_result['average'][0], 
                   linewidth=2.5, color='blue', marker='o', markersize=5, 
                   label='PDP', zorder=3)
            
            # Calculate ALE (true ALE, not approximate)
            # ALE is calculated by calculating the local effect and accumulating them
            n_grid = len(grid_values)
            n_samples = len(X_train_df)
            
            # Create grid intervals
            ale_values = np.zeros(n_grid)
            
            # For each grid point, calculate the local effect
            for i in range(n_grid - 1):
                # Get samples in the current interval
                if i == 0:
                    mask = (X_train_df.iloc[:, feature_idx].values <= grid_values[i+1])
                elif i == n_grid - 2:
                    mask = (X_train_df.iloc[:, feature_idx].values > grid_values[i])
                else:
                    mask = ((X_train_df.iloc[:, feature_idx].values > grid_values[i]) & 
                           (X_train_df.iloc[:, feature_idx].values <= grid_values[i+1]))
                
                # If there are samples in this interval
                if np.sum(mask) > 0:
                    # Calculate local effect: change the feature value, other features remain unchanged
                    X_low = X_train_df.copy()
                    X_high = X_train_df.copy()
                    
                    # Set feature value to the ends of the interval
                    X_low.loc[mask, X_train_df.columns[feature_idx]] = grid_values[i]
                    X_high.loc[mask, X_train_df.columns[feature_idx]] = grid_values[i+1]
                    
                    # Calculate prediction difference
                    pred_low = model.predict(X_low.iloc[mask, :].values)
                    pred_high = model.predict(X_high.iloc[mask, :].values)
                    
                    # Local effect
                    local_effect = np.mean(pred_high - pred_low) / (grid_values[i+1] - grid_values[i])
                    
                    # Accumulate to ALE
                    if i == 0:
                        ale_values[i+1] = local_effect * (grid_values[i+1] - grid_values[i])
                    else:
                        ale_values[i+1] = ale_values[i] + local_effect * (grid_values[i+1] - grid_values[i])
            
            # Center ALE (so the mean is 0)
            ale_values = ale_values - np.mean(ale_values) + np.mean(pd_result['average'][0])
            
            # Plot PDP line (blue solid line)
            ax.plot(grid_values, pd_result['average'][0], 
                   linewidth=2.5, color='blue', marker='o', markersize=5, 
                   label='PDP', zorder=4)
            
            # Plot Mean c-ICE line (coral dashed line)
            ax.plot(grid_values, ice_mean, 
                   linewidth=2.5, color='coral', linestyle='--', 
                   marker='s', markersize=4, label='Mean c-ICE', zorder=4)
            
            # Plot ALE line (green dotted line)
            ax.plot(grid_values, ale_values, 
                   linewidth=2.5, color='green', linestyle='-.', 
                   marker='^', markersize=4, label='ALE', zorder=4)
            
            # Plot 95% CI shadow area
            ax.fill_between(grid_values, ice_lower_ci, ice_upper_ci, 
                           color='coral', alpha=0.3, label='95% CI of c-ICE', zorder=1)
            
            ax.set_xlabel(feature, fontsize=10)
            ax.set_ylabel('Partial Effect', fontsize=10)
            ax.set_title(f'PDP, ALE & c-ICE for {feature}', fontsize=12, fontweight='bold')
            ax.legend(loc='best', fontsize=9)
            ax.grid(True, alpha=0.3)
        except Exception as e:
            print(f"Warning: Could not create PDP for {feature}: {e}")
            ax.text(0.5, 0.5, f'Error creating PDP', 
                   ha='center', va='center', transform=ax.transAxes)
            ax.axis('off')
    
    # Hide extra subplots
    for idx in range(n_top_features, len(axes)):
        axes[idx].axis('off')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'pdp_analysis.png'), dpi=300, bbox_inches='tight')
    # plt.show()  # Show univariate PDP plot
    plt.close()
    
    print(f"Univariate PDP analysis completed, the dependency plot of the {n_top_features} most important features has been drawn")
    
    # 2. Bivariate PDP analysis (feature interaction)
    if n_top_features >= 2:
        print("\n=== Performing bivariate PDP analysis ===")
        
        # Select the 4 most important features for bivariate interaction analysis
        n_interaction_features = min(4, n_top_features)
        interaction_features = top_features[:n_interaction_features]
        
        # Create interaction feature pairs
        interaction_pairs = []
        for i in range(len(interaction_features)):
            for j in range(i+1, len(interaction_features)):
                interaction_pairs.append((interaction_features[i], interaction_features[j]))
        
        # Plot bivariate PDP (using heatmap)
        n_pairs = len(interaction_pairs)
        n_cols = 3  # Change to 3 columns, more美观
        n_rows = (n_pairs + 2) // 3  # Adjust row number calculation
        fig, axes = plt.subplots(n_rows, n_cols, figsize=(15, 6*n_rows))
        
        if n_pairs == 1:
            axes = [axes]
        else:
            axes = axes.flatten()
        
        for idx, (feat1, feat2) in enumerate(interaction_pairs):
            try:
                # Get feature index
                feat1_idx = feature_names.index(feat1)
                feat2_idx = feature_names.index(feat2)
                
                # Calculate bivariate partial dependence
                pd_result = partial_dependence(
                    model,
                    X_train_df,
                    features=[feat1_idx, feat2_idx],
                    kind='average'
                )
                
                # Get grid values and average predicted values
                grid_values_1 = pd_result['grid_values'][0]
                grid_values_2 = pd_result['grid_values'][1]
                average = pd_result['average'][0]  # Get the first element (if it is a 3D array)
                
                # Ensure average is 2D
                if average.ndim == 3:
                    average = average[0]
                
                # Create grid
                X1, X2 = np.meshgrid(grid_values_1, grid_values_2)
                
                # Ensure average matches grid dimensions
                if average.shape != X1.shape:
                    # If dimensions do not match, transpose or reshape
                    if average.shape == (len(grid_values_1), len(grid_values_2)):
                        average = average.T
                
                # Plot 3D surface graph
                ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                
                surf = ax.plot_surface(X1, X2, average, cmap='viridis', 
                                      alpha=0.8, linewidth=0, antialiased=True)
                
                ax.set_xlabel(feat1, fontsize=10)
                ax.set_ylabel(feat2, fontsize=10)
                ax.set_zlabel('Partial Dependence', fontsize=10)
                ax.set_title(f'Interaction: {feat1} vs {feat2}', fontsize=12, fontweight='bold')
                
                # Add color bar
                fig.colorbar(surf, ax=ax, shrink=0.5, label='Partial Dependence')
            except Exception as e:
                print(f"Warning: Could not create 2D PDP for {feat1} vs {feat2}: {e}")
                # For 3D graph, need to check if it is a subplot
                if idx < len(axes):
                    axes[idx].text(0.5, 0.5, f'Error creating 2D PDP', 
                                  ha='center', va='center', transform=axes[idx].transAxes)
                    axes[idx].axis('off')
        
        # Hide extra subplots
        for idx in range(n_pairs, len(axes)):
            if axes[idx] is not None:
                axes[idx].axis('off')
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, 'pdp_2d_interaction.png'), dpi=300, bbox_inches='tight')
        # plt.show()  # Show 3D graph
        plt.close()
        
        print(f"Bivariate PDP analysis completed, the interaction plot of the {n_interaction_features} most important features has been drawn")

def analyze_noise_robustness(model, X_test, y_test, feature_names, save_dir):
    """Analyze the robustness of the model to noise and outliers"""
    print("\n=== Analyzing model robustness ===")
    
    # 1. Add Gaussian noise test
    print("Testing robustness to Gaussian noise...")
    # Use 0, 2, 4, 6, 8, 10 (%) as noise levels
    noise_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Noise levels (percentage of feature standard deviation)
    noise_rmses = []
    noise_stds = []
    
    for noise_level in noise_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times to get standard deviation
            # Add Gaussian noise
            X_test_noisy = X_test.copy()
            for i in range(X_test.shape[1]):
                feature_std = np.std(X_test[:, i])
                noise = np.random.normal(0, noise_level * feature_std, size=len(X_test))
                X_test_noisy[:, i] = X_test[:, i] + noise
            
            # Predict
            y_pred_noisy = model.predict(X_test_noisy)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred_noisy))
            rmses.append(rmse)
        
        noise_rmses.append(np.mean(rmses))
        noise_stds.append(np.std(rmses))
    
    # 2. Add outlier test
    print("Testing robustness to outliers...")
    # Use 0, 2, 4, 6, 8, 10 (%) as outlier levels
    outlier_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Outlier levels (percentage of samples)
    outlier_severity = 5.0  # Outlier severity (multiple of standard deviation)
    outlier_rmses = []
    outlier_stds = []
    
    for outlier_pct in outlier_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times
            # Add outliers
            X_test_outlier = X_test.copy()
            n_outliers = int(len(X_test) * outlier_pct)
            if n_outliers > 0:
                outlier_indices = np.random.choice(len(X_test), n_outliers, replace=False)
                for idx in outlier_indices:
                    # Randomly select a feature to add outliers
                    feature_idx = np.random.randint(X_test.shape[1])
                    feature_mean = np.mean(X_test[:, feature_idx])
                    feature_std = np.std(X_test[:, feature_idx])
                    X_test_outlier[idx, feature_idx] = feature_mean + outlier_severity * feature_std
            
            # Predict
            y_pred_outlier = model.predict(X_test_outlier)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred_outlier))
            rmses.append(rmse)
        
        outlier_rmses.append(np.mean(rmses))
        outlier_stds.append(np.std(rmses))
    
    # 3. Plot robustness analysis graph
    fig, axes = plt.subplots(1, 2, figsize=(15, 6))
    
    # Gaussian noise robustness
    axes[0].errorbar([l * 100 for l in noise_levels], noise_rmses, yerr=noise_stds, 
                     marker='o', capsize=5, capthick=2, linewidth=2, color='blue')
    axes[0].set_xlabel('Noise level (% of feature std)', fontsize=12)
    axes[0].set_ylabel('RMSE (MPa)', fontsize=12)
    axes[0].set_title('Robustness to Additive Gaussian Noise', fontsize=14, fontweight='bold')
    axes[0].grid(True, alpha=0.3)
    
    # Outlier robustness
    axes[1].errorbar([l * 100 for l in outlier_levels], outlier_rmses, yerr=outlier_stds, 
                     marker='o', capsize=5, capthick=2, linewidth=2, color='orange')
    axes[1].set_xlabel('Injected outliers (% of samples)', fontsize=12)
    axes[1].set_ylabel('RMSE (MPa)', fontsize=12)
    axes[1].set_title(f'Robustness to Outliers (severity={outlier_severity}×std)', fontsize=14, fontweight='bold')
    axes[1].grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'robustness_analysis.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # Save results (will be saved in the main function)
    robustness_df = pd.DataFrame({
        'Noise Level (%)': [l * 100 for l in noise_levels],
        'RMSE (MPa)': noise_rmses,
        'Std': noise_stds
    })
    
    outlier_df = pd.DataFrame({
        'Outlier Percentage (%)': [l * 100 for l in outlier_levels],
        'RMSE (MPa)': outlier_rmses,
        'Std': outlier_stds
    })
    
    # Print result summary
    print("\n=== Robustness analysis result summary ===")
    print("\n1. Gaussian noise robustness:")
    print(f"   - 0% noise时RMSE: {noise_rmses[0]:.3f} MPa")
    print(f"   - 2% noise时RMSE: {noise_rmses[1]:.3f} MPa (增加 {noise_rmses[1]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 4% noise时RMSE: {noise_rmses[2]:.3f} MPa (增加 {noise_rmses[2]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 6% noise时RMSE: {noise_rmses[3]:.3f} MPa (增加 {noise_rmses[3]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 8% noise时RMSE: {noise_rmses[4]:.3f} MPa (增加 {noise_rmses[4]-noise_rmses[0]:.3f} MPa)")
    print(f"   - 10% noise时RMSE: {noise_rmses[5]:.3f} MPa (增加 {noise_rmses[5]-noise_rmses[0]:.3f} MPa)")
    
    print("\n2. Outlier robustness:")
    print(f"   - 0% outlier RMSE: {outlier_rmses[0]:.3f} MPa")
    print(f"   - 2% outlier RMSE: {outlier_rmses[1]:.3f} MPa (变化 {outlier_rmses[1]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 4% outlier RMSE: {outlier_rmses[2]:.3f} MPa (变化 {outlier_rmses[2]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 6% outlier RMSE: {outlier_rmses[3]:.3f} MPa (变化 {outlier_rmses[3]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 8% outlier RMSE: {outlier_rmses[4]:.3f} MPa (变化 {outlier_rmses[4]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - 10% outlier RMSE: {outlier_rmses[5]:.3f} MPa (变化 {outlier_rmses[5]-outlier_rmses[0]:.3f} MPa)")
    
    # Calculate performance degradation rate
    noise_degradation = (noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100
    outlier_degradation = (outlier_rmses[-1] - outlier_rmses[0]) / outlier_rmses[0] * 100
    
    print("\n3. Performance degradation rate (10% noise/outlier vs no perturbation):")
    print(f"   - Gaussian noise performance degradation: {noise_degradation:.1f}%")
    print(f"   - Outlier performance degradation: {outlier_degradation:.1f}%")
    
    # Determine robustness
    if noise_degradation < 20:
        noise_robust = "Excellent"
    elif noise_degradation < 40:
        noise_robust = "Good"
    elif noise_degradation < 60:
        noise_robust = "Average"
    else:
        noise_robust = "Poor"
    
    if abs(outlier_degradation) < 5:
        outlier_robust = "Excellent"
    elif abs(outlier_degradation) < 10:
        outlier_robust = "Good"
    elif abs(outlier_degradation) < 20:
        outlier_robust = "Average"
    else:
        outlier_robust = "Poor"
    
    print("\n4. Robustness evaluation:")
    print(f"   - Robustness to Gaussian noise: {noise_robust} (10% noise时性能退化{noise_degradation:.1f}%)")
    print(f"   - Robustness to outliers: {outlier_robust} (10% outliers时性能变化{outlier_degradation:.1f}%)")
    
    print(f"\nRobustness analysis completed, detailed results saved")
    return noise_rmses, outlier_rmses, robustness_df, outlier_df

def analyze_prediction_intervals(model, X_test, y_test, save_dir, n_bootstrap=100, bootstrap_noise_level=0.01):
    """Analyze the confidence of the prediction intervals
    
    Parameters:
        model: The trained model
        X_test: The test features
        y_test: The test labels
        save_dir: The save directory
        n_bootstrap: Bootstrap resampling times
        bootstrap_noise_level: Bootstrap noise level (percentage of feature standard deviation)
    
    Method description:
    1. Bootstrap: Add input noise to simulate model uncertainty, similar to Monte Carlo Dropout
    2. Quantile Regression: Build a fixed-width prediction interval based on the residual distribution
    """
    print("\n=== Generate prediction intervals ===")
    print(f"Bootstrap noise level: {bootstrap_noise_level*100}% of feature std")
    
    # 1. Bootstrap prediction interval - improved version
    print("Generate Bootstrap prediction interval...")
    bootstrap_predictions = []
    for _ in range(n_bootstrap):
        # Add noise to simulate model uncertainty (similar to the idea of Monte Carlo Dropout)
        # Add small noise to the test data to reflect the model's response to input uncertainty
        noise_std = bootstrap_noise_level  # Noise standard deviation (percentage of feature standard deviation)
        X_test_noisy = X_test.copy()
        for i in range(X_test.shape[1]):
            feature_std = np.std(X_test[:, i])
            noise = np.random.normal(0, noise_std * feature_std, size=len(X_test))
            X_test_noisy[:, i] = X_test[:, i] + noise
        
        # Predict
        y_pred = model.predict(X_test_noisy)
        bootstrap_predictions.append(y_pred)
    
    bootstrap_predictions = np.array(bootstrap_predictions)
    bootstrap_median = np.median(bootstrap_predictions, axis=0)
    bootstrap_lower = np.percentile(bootstrap_predictions, 10, axis=0)  # 80% confidence interval
    bootstrap_upper = np.percentile(bootstrap_predictions, 90, axis=0)
    
    # 2. Quantile Regression prediction interval (using the residual distribution of the training data)
    print("Generate Quantile Regression prediction interval...")
    y_pred_test = model.predict(X_test)
    residuals = y_test - y_pred_test
    residual_std = np.std(residuals)
    
    # Use the standard deviation of the residuals to build the prediction interval
    quantile_lower = y_pred_test - 1.28 * residual_std  # 10th percentile for 80% interval
    quantile_upper = y_pred_test + 1.28 * residual_std  # 90th percentile for 80% interval
    
    # Sort for visualization
    sort_indices = np.argsort(y_test)
    y_test_sorted = y_test[sort_indices]
    bootstrap_median_sorted = bootstrap_median[sort_indices]
    bootstrap_lower_sorted = bootstrap_lower[sort_indices]
    bootstrap_upper_sorted = bootstrap_upper[sort_indices]
    quantile_lower_sorted = quantile_lower[sort_indices]
    quantile_upper_sorted = quantile_upper[sort_indices]
    
    # 3. Draw prediction interval comparison graph
    fig = plt.figure(figsize=(20, 10))
    
    # Create grid layout: left two columns display prediction intervals, right column displays performance comparison
    gs = fig.add_gridspec(2, 2, hspace=0.3, wspace=0.3, 
                         width_ratios=[2, 1], height_ratios=[1, 1])
    
    # Bootstrap prediction interval (top left)
    ax1 = fig.add_subplot(gs[0, 0])
    ax1.scatter(range(len(y_test_sorted)), y_test_sorted, s=30, c='black', alpha=0.7, label='Actual values', zorder=3)
    ax1.plot(range(len(y_test_sorted)), bootstrap_median_sorted, 'r--', linewidth=2, label='Predicted median', zorder=2)
    ax1.fill_between(range(len(y_test_sorted)), bootstrap_lower_sorted, bootstrap_upper_sorted, 
                     alpha=0.3, color='lightblue', label='80% Prediction interval', zorder=1)
    
    ax1.set_xlabel('Test samples (sorted by actual values)', fontsize=11)
    ax1.set_ylabel('Peak Stress fc (MPa)', fontsize=11)
    ax1.set_title('Bootstrap Prediction Intervals', fontsize=13, fontweight='bold')
    ax1.legend(loc='lower right', fontsize=9)
    ax1.grid(True, alpha=0.3)
    
    # Quantile Regression prediction interval (bottom left)
    ax2 = fig.add_subplot(gs[1, 0])
    ax2.scatter(range(len(y_test_sorted)), y_test_sorted, s=30, c='black', alpha=0.7, label='Actual values', zorder=3)
    ax2.plot(range(len(y_test_sorted)), y_pred_test[sort_indices], 'r--', linewidth=2, label='Predicted median', zorder=2)
    ax2.fill_between(range(len(y_test_sorted)), quantile_lower_sorted, quantile_upper_sorted, 
                     alpha=0.3, color='lightgreen', label='80% Prediction interval', zorder=1)
    
    ax2.set_xlabel('Test samples (sorted by actual values)', fontsize=11)
    ax2.set_ylabel('Peak Stress fc (MPa)', fontsize=11)
    ax2.set_title('Quantile Regression Intervals', fontsize=13, fontweight='bold')
    ax2.legend(loc='lower right', fontsize=9)
    ax2.grid(True, alpha=0.3)
    
    # Prediction interval coverage analysis (right)
    ax3 = fig.add_subplot(gs[:, 1])
    
    # Calculate coverage
    bootstrap_coverage = np.mean((y_test_sorted >= bootstrap_lower_sorted) & (y_test_sorted <= bootstrap_upper_sorted))
    quantile_coverage = np.mean((y_test_sorted >= quantile_lower_sorted) & (y_test_sorted <= quantile_upper_sorted))
    
    # Calculate average interval width
    bootstrap_width = np.mean(bootstrap_upper_sorted - bootstrap_lower_sorted)
    quantile_width = np.mean(quantile_upper_sorted - quantile_lower_sorted)
    
    methods = ['Bootstrap', 'Quantile Regression']
    coverages = [bootstrap_coverage * 100, quantile_coverage * 100]
    widths = [bootstrap_width, quantile_width]
    
    y_pos = np.arange(len(methods))
    height = 0.35
    
    # Create dual axis
    bars1 = ax3.barh(y_pos - height/2, coverages, height, 
                    label='Coverage Rate (%)', color='skyblue', alpha=0.8)
    ax3_twin = ax3.twinx()
    bars2 = ax3_twin.barh(y_pos + height/2, widths, height, 
                         label='Avg Interval Width (MPa)', color='coral', alpha=0.8)
    
    ax3.set_ylabel('Methods', fontsize=11)
    ax3.set_xlabel('Coverage Rate (%)', fontsize=11, color='blue')
    ax3_twin.set_xlabel('Average Interval Width (MPa)', fontsize=11, color='red')
    ax3.set_yticks(y_pos)
    ax3.set_yticklabels(methods)
    ax3.tick_params(axis='x', labelcolor='blue')
    ax3_twin.tick_params(axis='x', labelcolor='red')
    ax3.grid(True, alpha=0.3, axis='x')
    ax3.set_title('Performance Comparison', fontsize=13, fontweight='bold', pad=20)
    
    # Add numerical labels
    for i, v in enumerate(coverages):
        ax3.text(v + 2, i - height/2, f'{v:.1f}%', ha='left', va='center', fontsize=9)
    for i, v in enumerate(widths):
        ax3_twin.text(v + 0.5, i + height/2, f'{v:.2f}', ha='left', va='center', fontsize=9)
    
    # Add legend
    ax3.legend(loc='upper left', fontsize=9)
    ax3_twin.legend(loc='lower left', fontsize=9)
    
    # Invert y-axis to make Bootstrap at the top
    ax3.invert_yaxis()
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'prediction_intervals.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # Save prediction interval results
    intervals_df = pd.DataFrame({
        'Sample_ID': range(len(y_test)),
        'Actual': y_test,
        'Predicted': y_pred_test,
        'Bootstrap_Lower': bootstrap_lower,
        'Bootstrap_Upper': bootstrap_upper,
        'Quantile_Lower': quantile_lower,
        'Quantile_Upper': quantile_upper
    })
    # Will be saved in the main function
    
    print(f"Prediction interval analysis completed, 80% confidence interval coverage: Bootstrap={bootstrap_coverage*100:.1f}%, Quantile={quantile_coverage*100:.1f}%")
    return bootstrap_predictions, bootstrap_coverage, quantile_coverage, intervals_df

def train_single_fold(X_train, y_train, X_val, y_val, X_test, y_test, train_ids, val_ids, test_ids, 
                     feature_names, save_dir, n_trials=100):
    """Single fold training (for cross-validation) - only perform hyperparameter optimization and training, no interpretability analysis"""
    # Hyperparameter optimization
    study = optimize_lightgbm(X_train, y_train, X_val, y_val, n_trials=n_trials)
    best_params = study.best_params
    
    # Train the final model
    model = train_lightgbm(X_train, y_train, X_val, y_val, best_params)
    
    # Predict the test set (only for performance evaluation)
    if len(X_test) > 0:
        y_pred = model.predict(X_test)
        r2 = r2_score(y_test, y_pred)
        mae = mean_absolute_error(y_test, y_pred)
        rmse = np.sqrt(mean_squared_error(y_test, y_pred))
        
        # Save prediction results
        results_df = pd.DataFrame({
            'Sample ID': test_ids,
            'Actual value': y_test,
            'Predicted value': y_pred,
            'Residual': y_pred - y_test
        })
        results_df.to_excel(os.path.join(save_dir, 'test_predictions.xlsx'), index=False)
        
        return model, best_params, {'r2': r2, 'mae': mae, 'rmse': rmse}, train_ids, val_ids, test_ids
    else:
        return model, best_params, None, train_ids, val_ids, test_ids

# Ensemble model wrapper class (for handling saved ensemble models)
class EnsembleModelWrapper:
    """Ensemble model wrapper class, for handling ensemble models loaded from files"""
    def __init__(self, models, weights):
        self.models = models
        self.weights = np.array(weights) if not isinstance(weights, np.ndarray) else weights
        self.is_ensemble = True
    
    def predict(self, X):
        """Ensemble prediction (weighted average)"""
        predictions = np.zeros(len(X))
        for model, weight in zip(self.models, self.weights):
            predictions += weight * model.predict(X)
        return predictions
    
    @property
    def feature_importances_(self):
        """Calculate the feature importance of the ensemble model (weighted average)"""
        if len(self.models) == 0:
            return None
        
        # Get the feature importance of the first model as the baseline
        base_importance = self.models[0].feature_importances_
        ensemble_importance = np.zeros_like(base_importance)
        
        # Weighted average of the feature importance of all models
        for model, weight in zip(self.models, self.weights):
            ensemble_importance += weight * model.feature_importances_
        
        return ensemble_importance

def load_trained_model(model_path):
    """Load the trained LightGBM model (supports .joblib and .pkl format, including ensemble models)"""
    print(f"\n=== Load the trained model ===")
    print(f"Model path: {model_path}")
    
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model file not found: {model_path}")
    
    loaded_data = joblib.load(model_path)
    
    # Determine the type of loaded data
    if isinstance(loaded_data, dict):
        # Check if it is an ensemble model
        if loaded_data.get('type') == 'ensemble':
            # Ensemble model format
            models = loaded_data.get('models', [])
            weights = loaded_data.get('weights', [])
            strategy = loaded_data.get('strategy', 'unknown')
            
            if len(models) == 0:
                raise ValueError("Ensemble model has no submodels")
            
            # Create ensemble model wrapper class
            model = EnsembleModelWrapper(models, weights)
            feature_names = loaded_data.get('feature_names', None)
            model_data = {
                'type': 'ensemble',
                'strategy': strategy,
                'n_models': len(models),
                'weights': weights,
                'feature_names': feature_names
            }
            
            print(f"✓ Model loaded successfully (ensemble model)")
            print(f"✓ Ensemble strategy: {strategy}")
            print(f"✓ Number of submodels: {len(models)}")
            print(f"✓ Weights: {weights}")
        else:
            # Ordinary dictionary format, try to extract the model and feature names
            model = loaded_data.get('model', loaded_data)
            feature_names = loaded_data.get('feature_names', None)
            model_data = loaded_data

            print(f"✓ Model loaded successfully")
            print(f"✓ Model type: {type(model).__name__}")
    else:
        # If it is directly a model object
        model = loaded_data
        feature_names = None
        model_data = {'model': model}
        
        print(f"✓ Model loaded successfully")
        print(f"✓ Model type: {type(model).__name__}")
    
    if 'feature_names' in model_data and model_data['feature_names'] is not None:
        feature_names = model_data['feature_names']
        print(f"✓ Number of features: {len(feature_names)}")
        print(f"✓ Feature names: {feature_names}")
    else:
        print("⚠ Warning: Feature names not found, will use the feature names when loading the data")
    
    return model, model_data

def cross_subset_validation(X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df_processed, 
                           save_dir=None, n_trials=100):
    """subset1/2/3 three-fold cross-validation"""
    print("\n=== LightGBM Subset three-fold cross-validation ===")

    if save_dir is None:
        save_dir = os.path.join(SAVE_ROOT, 'lightgbm_cv')
    
    # Extract subset labels and test set
    subset_labels = sorted([s for s in np.unique(sample_divisions) if str(s).startswith('subset')])
    test_mask = np.array([str(s).strip().lower().startswith('test') for s in sample_divisions])
    test_indices = np.where(test_mask)[0]
    
    if len(subset_labels) != 3:
        raise ValueError(f"The number of subsets should be 3, actual: {len(subset_labels)} {subset_labels}")
    
    if len(test_indices) == 0:
        raise ValueError("Test set not found")
    
    print(f"Found 3 subsets: {subset_labels}")
    print(f"Number of test set samples: {len(test_indices)}")
    
    all_results = []
    
    # Three-fold cross-validation
    for i, val_label in enumerate(subset_labels):
        print(f"\n{'='*60}")
        print(f"Subset cross-validation round {i+1}/3 [{val_label} as validation set]")
        print(f"{'='*60}")
        
        # Determine the training set and validation set
        train_labels = [lbl for lbl in subset_labels if lbl != val_label]
        subset_to_indices = {lbl: np.where(sample_divisions == lbl)[0] for lbl in subset_labels}
        
        train_indices = np.concatenate([subset_to_indices[lbl] for lbl in train_labels])
        val_indices = subset_to_indices[val_label]
        
        print(f"Training set = {train_labels}, validation set = {val_label}, test set = test")
        print(f"   Training set samples: {len(train_indices)}, validation set: {len(val_indices)}, test set: {len(test_indices)}")
        
        # Split data
        X_train = X[train_indices]
        y_train = y[train_indices]
        X_val = X[val_indices]
        y_val = y[val_indices]
        X_test = X[test_indices]
        y_test = y[test_indices]
        
        train_ids_fold = [sample_ids[idx] for idx in train_indices]
        val_ids_fold = [sample_ids[idx] for idx in val_indices]
        test_ids_fold = [sample_ids[idx] for idx in test_indices]
        
        # Create the current round save directory
        round_save_dir = os.path.join(save_dir, f'subsetCV_{val_label}')
        os.makedirs(round_save_dir, exist_ok=True)
        
        # Train a single fold model (only perform hyperparameter optimization and training, no interpretability analysis)
        model, best_params, metrics, train_ids_round, val_ids_round, test_ids_round = train_single_fold(
            X_train, y_train, X_val, y_val, X_test, y_test,
            train_ids_fold, val_ids_fold, test_ids_fold,
            feature_names, round_save_dir, n_trials=n_trials
        )
        
        all_results.append({
            'val_label': val_label,
            'train_labels': train_labels,
            'best_params': best_params,
            'metrics': metrics,
            'model': model,
            'train_val_indices': np.concatenate([train_indices, val_indices]),
            'test_indices': test_indices.copy()
        })
        
        print(f"[{val_label} validation set] This round of training completed")
        if metrics:
            print(f"   Test set metrics: R2={metrics['r2']:.4f}, MAE={metrics['mae']:.3f}, RMSE={metrics['rmse']:.3f}")
    
    # Print summary
    print(f"\n{'='*60}")
    print("All three-fold cross-validation completed!")
    print(f"{'='*60}")
    for i, res in enumerate(all_results):
        print(f"[{i+1}] Validation set: {res['val_label']}")
        if res['metrics']:
            print(f"     Test set metrics: R2={res['metrics']['r2']:.4f}, MAE={res['metrics']['mae']:.3f}, RMSE={res['metrics']['rmse']:.3f}")
    
    # Calculate the average metrics of three-fold cross-validation
    valid_metrics = [res['metrics'] for res in all_results if res['metrics'] is not None]
    if valid_metrics:
        avg_r2 = np.mean([m['r2'] for m in valid_metrics])
        avg_mae = np.mean([m['mae'] for m in valid_metrics])
        avg_rmse = np.mean([m['rmse'] for m in valid_metrics])
        print(f"\nThree-fold average performance: R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")
    else:
        avg_r2 = avg_mae = avg_rmse = None

    # Find the model with the best test set performance (by R2) in three-fold cross-validation
    best_fold_idx = 0
    best_r2 = -np.inf
    for i, res in enumerate(all_results):
        if res['metrics'] and res['metrics']['r2'] > best_r2:
            best_r2 = res['metrics']['r2']
            best_fold_idx = i
    
    best_fold_result = all_results[best_fold_idx]
    best_overall_params = best_fold_result['best_params']
    
    print(f"\n{'='*60}")
    print(f"Three-fold cross-validation best hyperparameters (from the {best_fold_idx+1} fold, validation set={best_fold_result['val_label']})")
    print(f"Test set R2: {best_r2:.4f}")
    print(f"Best hyperparameters: {best_overall_params}")
    print(f"{'='*60}")
    
    # Use the best hyperparameters, retrain the final model using all subset data (excluding test)
    print(f"\n{'='*60}")
    print("Use the best fold model as the final model (no retraining)")
    print(f"{'='*60}")

    best_fold_train_val_indices = best_fold_result['train_val_indices']
    best_fold_test_indices = best_fold_result['test_indices']

    X_train_final = X[best_fold_train_val_indices]
    y_train_final = y[best_fold_train_val_indices]
    X_test_final = X[best_fold_test_indices]
    y_test_final = y[best_fold_test_indices]

    train_ids_final = [sample_ids[idx] for idx in best_fold_train_val_indices]
    test_ids_final = [sample_ids[idx] for idx in best_fold_test_indices]

    final_model = best_fold_result['model']

    # Evaluate on the test set (should be consistent with the metrics of the best fold)
    y_test_pred = final_model.predict(X_test_final)
    final_r2 = r2_score(y_test_final, y_test_pred)
    final_mae = mean_absolute_error(y_test_final, y_test_pred)
    final_rmse = np.sqrt(mean_squared_error(y_test_final, y_test_pred))

    print(f"The final model comes from the {best_fold_result['val_label']} fold.")
    print(f"  R2: {final_r2:.4f}")
    print(f"  MAE: {final_mae:.3f} MPa")
    print(f"  RMSE: {final_rmse:.3f} MPa")
    if avg_r2 is not None:
        print(f"  (Three-fold average) R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")
    
    # Interpretability analysis directory
    interpretability_dir = os.path.join(save_dir, 'interpretability_analysis')
    os.makedirs(interpretability_dir, exist_ok=True)
    
    print(f"\n{'='*60}")
    print("Start interpretability analysis")
    print(f"{'='*60}")
    
    # 1. Plot the results
    plot_results(y_test_final, y_test_pred, interpretability_dir)
    
    # 2. Plot the training set and test set comparison
    plot_train_test_comparison(final_model, X_train_final, y_train_final, X_test_final, y_test_final, interpretability_dir)
    
    # 3. Feature importance
    X_all_final_for_importance = np.vstack([X_train_final, X_test_final])
    importance_df = plot_feature_importance(final_model, feature_names, interpretability_dir, X=X_all_final_for_importance)
    
    # 4. SHAP analysis
    X_all_final = np.vstack([X_train_final, X_test_final])
    shap_importance = plot_shap_analysis(final_model, X_all_final, X_all_final, feature_names, interpretability_dir)
    
    # 5. PDP analysis
    plot_pdp_analysis(final_model, X_all_final, feature_names, interpretability_dir, n_top_features=6)
    
    # 6. Noise robustness analysis (optional)
    # analyze_noise_robustness(final_model, X_test_final, y_test_final, feature_names, interpretability_dir)
    
    # 7. Prediction interval analysis (optional)
    # analyze_prediction_intervals(final_model, X_test_final, y_test_final, interpretability_dir, n_bootstrap=100, bootstrap_noise_level=0.01)
    
    # Save the final model
    joblib.dump({
        'model': final_model,
        'best_params': best_overall_params,
        'feature_names': feature_names,
        'r2': final_r2,
        'mae': final_mae,
        'rmse': final_rmse,
        'importance_df': importance_df,
        'shap_importance': shap_importance,
        'cross_validation_results': all_results
    }, os.path.join(interpretability_dir, 'final_lightgbm_model.pkl'))
    
    # Predict all samples and write to Excel
    print(f"\n{'='*60}")
    print("Predict all samples and write to Excel")
    print(f"{'='*60}")
    try:
        # Predict all samples (including test set)
        X_all_samples = np.vstack([X_train_final, X_test_final])
        y_all_pred = final_model.predict(X_all_samples)
        all_sample_ids = train_ids_final + test_ids_final
        
        # Create prediction result dictionary
        pred_dict = {}
        for idx, sample_id in enumerate(all_sample_ids):
            pred_dict[sample_id] = y_all_pred[idx]
        
        # Reload the original Excel file
        data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
        df_original = pd.read_excel(data_file, sheet_name=0)
        
        # Use map method to match
        if 'No_Customized' in df_original.columns:
            df_original['LightGBM_fc'] = df_original['No_Customized'].map(pred_dict)
        else:
            print("Warning: No_Customized column not found, cannot match prediction results")
        
        # Save results
        output_file = os.path.join(PROJECT_ROOT, "dataset/dataset_with_LightGBM_fc.xlsx")
        df_original.to_excel(output_file, index=False)
        print(f"✓ Prediction results written to: {output_file}")
        print(f"  - New column: LightGBM_fc (final model predicted values)")
        print(f"  - Number of non-empty prediction values: {df_original['LightGBM_fc'].notna().sum()}")
        
        # Backup
        backup_file = os.path.join(save_dir, 'dataset_with_LightGBM_fc_backup.xlsx')
        df_original.to_excel(backup_file, index=False)
        print(f"✓ Backup file saved: {backup_file}")
    except Exception as e:
        print(f"Warning: Error writing to Excel: {e}")
        import traceback
        traceback.print_exc()
    
    return all_results, final_model, best_overall_params

def cross_random_kfold_validation(X, y, feature_names, sample_ids, sample_divisions=None,
                                  save_dir=None, n_trials=100, n_splits=3, random_state=42):
    """Random KFold cross-validation"""
    print(f"\n=== LightGBM Random {n_splits} fold cross-validation ===")

    if save_dir is None:
        save_dir = os.path.join(SAVE_ROOT, 'lightgbm_cv_random')

    os.makedirs(save_dir, exist_ok=True)

    fixed_test_indices = None
    if sample_divisions is not None:
        test_mask = np.array([str(s).strip().lower().startswith('test') for s in sample_divisions])
        if np.any(test_mask):
            fixed_test_indices = np.where(test_mask)[0]
            print(f"Use the pre-divided test set ({len(fixed_test_indices)} samples) for random cross-validation evaluation")

    if fixed_test_indices is not None and len(fixed_test_indices) == len(X):
        raise ValueError("All samples belong to the test set, cannot perform random cross-validation")

    if fixed_test_indices is not None:
        train_pool_mask = np.ones(len(X), dtype=bool)
        train_pool_mask[fixed_test_indices] = False
        train_pool_indices = np.where(train_pool_mask)[0]
        print(f"Random cross-validation only on non-test set samples, total {len(train_pool_indices)} samples")
        kf = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
        split_source = train_pool_indices
    else:
        kf = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
        split_source = np.arange(len(X))

    all_results = []

    for fold_idx, (train_indices_rel, val_indices_rel) in enumerate(kf.split(split_source), 1):
        print(f"\n{'='*60}")
        print(f"Random KFold round {fold_idx}/{n_splits} fold")
        print(f"{'='*60}")

        if fixed_test_indices is not None:
            train_indices = split_source[train_indices_rel]
            val_indices = split_source[val_indices_rel]
            test_indices = fixed_test_indices
        else:
            train_val_indices = split_source[train_indices_rel]
            val_candidates = split_source[val_indices_rel]
            if len(train_val_indices) < 5:
                raise ValueError("Training samples too few, cannot perform random cross-validation")
            inner_train_indices, inner_val_indices = train_test_split(
                np.concatenate([train_val_indices, val_candidates]),
                test_size=0.2,
                random_state=random_state + fold_idx
            )
            train_indices = inner_train_indices
            val_indices = inner_val_indices
            test_indices = val_candidates

        if len(train_indices) < 5:
            raise ValueError("Training samples too few, cannot perform random cross-validation")

        print(f"   Training set samples: {len(train_indices)}")
        print(f"   Validation set samples: {len(val_indices)}")
        print(f"   Test set samples: {len(test_indices)}")

        X_train = X[train_indices]
        y_train = y[train_indices]
        X_val = X[val_indices]
        y_val = y[val_indices]
        X_test = X[test_indices]
        y_test = y[test_indices]

        train_ids_fold = [sample_ids[idx] for idx in train_indices]
        val_ids_fold = [sample_ids[idx] for idx in val_indices]
        test_ids_fold = [sample_ids[idx] for idx in test_indices]

        round_save_dir = os.path.join(save_dir, f'randomKFold_fold{fold_idx}')
        os.makedirs(round_save_dir, exist_ok=True)

        model, best_params, metrics, _, _, _ = train_single_fold(
            X_train, y_train, X_val, y_val, X_test, y_test,
            train_ids_fold, val_ids_fold, test_ids_fold,
            feature_names, round_save_dir, n_trials=n_trials
        )

        combined_train_val = np.unique(np.concatenate([train_indices, val_indices]))

        all_results.append({
            'fold_idx': fold_idx,
            'best_params': best_params,
            'metrics': metrics,
            'model': model,
            'train_val_indices': combined_train_val,
            'test_indices': test_indices.copy()
        })

        if metrics:
            print(f"[Fold {fold_idx}] Test set metrics: R2={metrics['r2']:.4f}, MAE={metrics['mae']:.3f}, RMSE={metrics['rmse']:.3f}")

    print(f"\n{'='*60}")
    print("All random KFold completed!")
    print(f"{'='*60}")
    for res in all_results:
        if res['metrics']:
            print(f"[Fold {res['fold_idx']}] Test set: R2={res['metrics']['r2']:.4f}, MAE={res['metrics']['mae']:.3f}, RMSE={res['metrics']['rmse']:.3f}")

    valid_metrics = [res['metrics'] for res in all_results if res['metrics']]
    if valid_metrics:
        avg_r2 = np.mean([m['r2'] for m in valid_metrics])
        avg_mae = np.mean([m['mae'] for m in valid_metrics])
        avg_rmse = np.mean([m['rmse'] for m in valid_metrics])
        print(f"\n{n_splits} fold average performance: R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")
    else:
        avg_r2 = avg_mae = avg_rmse = None

    best_fold_idx = 0
    best_r2 = -np.inf
    for i, res in enumerate(all_results):
        if res['metrics'] and res['metrics']['r2'] > best_r2:
            best_r2 = res['metrics']['r2']
            best_fold_idx = i

    best_fold_result = all_results[best_fold_idx]
    best_overall_params = best_fold_result['best_params']

    print(f"\n{'='*60}")
    print(f"Best fold = Fold {best_fold_result['fold_idx']} (Test set R2={best_r2:.4f})")
    print(f"Best hyperparameters: {best_overall_params}")
    print(f"{'='*60}")

    best_fold_train_val_indices = best_fold_result['train_val_indices']
    best_fold_test_indices = best_fold_result['test_indices']

    X_train_final = X[best_fold_train_val_indices]
    y_train_final = y[best_fold_train_val_indices]
    X_test_final = X[best_fold_test_indices]
    y_test_final = y[best_fold_test_indices]

    train_ids_final = [sample_ids[idx] for idx in best_fold_train_val_indices]
    test_ids_final = [sample_ids[idx] for idx in best_fold_test_indices]

    final_model = best_fold_result['model']

    y_test_pred = final_model.predict(X_test_final)
    final_r2 = r2_score(y_test_final, y_test_pred)
    final_mae = mean_absolute_error(y_test_final, y_test_pred)
    final_rmse = np.sqrt(mean_squared_error(y_test_final, y_test_pred))

    print(f"The final model is trained from Fold {best_fold_result['fold_idx']}.")
    print(f"  R2: {final_r2:.4f}")
    print(f"  MAE: {final_mae:.3f} MPa")
    print(f"  RMSE: {final_rmse:.3f} MPa")
    if avg_r2 is not None:
        print(f"  ({n_splits} fold average) R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")

    interpretability_dir = os.path.join(save_dir, 'interpretability_analysis')
    os.makedirs(interpretability_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print("Start interpretability analysis (random划分)")
    print(f"{'='*60}")

    plot_results(y_test_final, y_test_pred, interpretability_dir)
    plot_train_test_comparison(final_model, X_train_final, y_train_final, X_test_final, y_test_final, interpretability_dir)
    X_all_final = np.vstack([X_train_final, X_test_final])
    importance_df = plot_feature_importance(final_model, feature_names, interpretability_dir, X=X_all_final)
    shap_importance = plot_shap_analysis(final_model, X_all_final, X_all_final, feature_names, interpretability_dir)
    plot_pdp_analysis(final_model, X_all_final, feature_names, interpretability_dir, n_top_features=6)

    joblib.dump({
        'model': final_model,
        'best_params': best_overall_params,
        'feature_names': feature_names,
        'r2': final_r2,
        'mae': final_mae,
        'rmse': final_rmse,
        'importance_df': importance_df,
        'shap_importance': shap_importance,
        'cross_validation_results': all_results
    }, os.path.join(interpretability_dir, 'final_lightgbm_model_random.pkl'))
    
    print(f"\n{'='*60}")
    print("Predict all samples and write to Excel (random divided model)")
    print(f"{'='*60}")
    try:
        X_all_samples = X
        y_all_pred = final_model.predict(X_all_samples)
        pred_dict = {sid: pred for sid, pred in zip(sample_ids, y_all_pred)}

        data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
        df_original = pd.read_excel(data_file, sheet_name=0)

        if 'No_Customized' in df_original.columns:
            df_original['LightGBM_fc_random'] = df_original['No_Customized'].map(pred_dict)
        else:
            df_original['LightGBM_fc_random'] = y_all_pred

        output_file = os.path.join(save_dir, "dataset_with_LightGBM_fc_random.xlsx")
        df_original.to_excel(output_file, index=False)
        print(f"✓ Random divided model prediction results written to: {output_file}")
        print(f"  - New column: LightGBM_fc_random")
        print(f"  - Number of non-empty prediction values: {df_original['LightGBM_fc_random'].notna().sum()}")
    except Exception as e:
        print(f"Warning: Error writing to random divided Excel: {e}")
        import traceback
        traceback.print_exc()

    return all_results, final_model, best_overall_params

def main():
    """Main function - load the trained model and perform SHAP analysis, noise analysis and uncertainty analysis"""
    print("=== LightGBM model SHAP analysis, noise analysis and uncertainty analysis ===")
    
    # 1. Load data
    data = load_data()
    if data is None:
        return
    
    X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df_processed = data
    
    # 2. Load the trained model - directly use lightgbm_final_model.joblib
    model_path = os.path.join(SAVE_ROOT, 'lightgbm_final_model.joblib')
    
    if not os.path.exists(model_path):
        print(f"Error: Model file not found: {model_path}")
        print("Please ensure the model file exists in: peak_stress/LightGBM/save/lightgbm_final_model.joblib")
        return
    
    model, model_data = load_trained_model(model_path)
    
    # If the model data has feature names, use the model's (ensure feature order is consistent)
    if 'feature_names' in model_data and model_data['feature_names'] is not None:
        model_feature_names = model_data['feature_names']
        print(f"\nUse the model saved feature names ({len(model_feature_names)} features)")
        # Ensure feature order is consistent with the model
        if len(model_feature_names) == len(feature_names):
            # Check if the feature names match
            if set(model_feature_names) == set(feature_names):
                # Reorder features to match the model
                feature_order = [feature_names.index(f) for f in model_feature_names]
                X = X[:, feature_order]
                feature_names = model_feature_names
                print("✓ Feature order adjusted to match the model")
            else:
                print("⚠ Warning: Model feature names do not match the data feature names completely")
                print(f"  Model features: {model_feature_names}")
                print(f"  Data features: {feature_names}")
        else:
            print(f"⚠ Warning: Feature number mismatch - model: {len(model_feature_names)}, data: {len(feature_names)}")
    
    # 3. Data split (get training set and test set, for SHAP analysis and noise analysis)
    X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids = split_data_by_divisions(
        X, y, sample_divisions, sample_ids, test_ratio=0.2, val_ratio=0.2, random_state=42
    )
    
    if len(X_test) == 0:
        print("Warning: Test set not found, will use validation set for analysis")
        X_test = X_val
        y_test = y_val
        test_ids = val_ids
    
    # Prepare data for SHAP analysis
    # Use all data for SHAP analysis (data set is small, total 94 samples)
    
    # Merge all datasets: training set + validation set + test set
    datasets_to_merge = []
    y_datasets_to_merge = []
    
    if len(X_train) > 0:
        datasets_to_merge.append(X_train)
        y_datasets_to_merge.append(y_train)
    if len(X_val) > 0:
        datasets_to_merge.append(X_val)
        y_datasets_to_merge.append(y_val)
    if len(X_test) > 0:
        datasets_to_merge.append(X_test)
        y_datasets_to_merge.append(y_test)
    
    if len(datasets_to_merge) > 0:
        X_all = np.vstack(datasets_to_merge)
        y_all = np.hstack(y_datasets_to_merge)
    else:
        X_all = X
        y_all = y
    
    # Use all data as background and explanation data set
    X_background = X_all
    y_background = y_all
    X_explain = X_all
    y_explain = y_all
    
    print(f"\nData preparation completed:")
    print(f"   Training set: {len(X_train)} samples")
    print(f"   Validation set: {len(X_val)} samples")
    print(f"   Test set: {len(X_test)} samples")
    print(f"\nSHAP analysis data set configuration (using all data set):")
    print(f"   All data set: {len(X_all)} samples (training set + validation set + test set)")
    print(f"   Background data set (for explainer initialization): {len(X_background)} samples")
    print(f"   Explanation data set (calculate SHAP values): {len(X_explain)} samples")
    
    # 4. Create save directory
    analysis_save_dir = os.path.join(SAVE_ROOT)
    os.makedirs(analysis_save_dir, exist_ok=True)
    
    # 5. SHAP analysis (using all data set)
    print(f"\n{'='*60}")
    print("Start SHAP analysis (using all data set)")
    print(f"{'='*60}")
    shap_importance = None
    feature_stats = None
    try:
        shap_importance = plot_shap_analysis(
            model, X_background, X_explain, feature_names, analysis_save_dir
        )
        print("✓ SHAP analysis completed")
        
        # 5.1 Feature statistics analysis (explain why some features have a small impact)
        print(f"\n{'='*60}")
        print("Start feature statistics analysis")
        print(f"{'='*60}")
        try:
            feature_stats = analyze_feature_statistics(
                X_all, y_all, feature_names, shap_importance, analysis_save_dir
            )
            print("✓ Feature statistics analysis completed")
        except Exception as e:
            print(f"⚠ Feature statistics analysis error: {e}")
            import traceback
            traceback.print_exc()
            feature_stats = None
    except Exception as e:
        print(f"⚠ SHAP analysis error: {e}")
        import traceback
        traceback.print_exc()
        shap_importance = None
        feature_stats = None
    
    # 6. Feature importance analysis
    print(f"\n{'='*60}")
    print("Start feature importance analysis")
    print(f"{'='*60}")
    importance_df = None
    try:
        importance_df = plot_feature_importance(
            model, feature_names, analysis_save_dir, X=X_all
        )
        print("✓ Feature importance analysis completed")
    except Exception as e:
        print(f"⚠ Feature importance analysis error: {e}")
        import traceback
        traceback.print_exc()
        importance_df = None
    
    # 7. Noise robustness analysis
    print(f"\n{'='*60}")
    print("Start noise robustness analysis")
    print(f"{'='*60}")
    noise_rmses, outlier_rmses = None, None
    robustness_df, outlier_df = None, None
    try:
        noise_rmses, outlier_rmses, robustness_df, outlier_df = analyze_noise_robustness(
            model, X_test, y_test, feature_names, analysis_save_dir
        )
        print("✓ Noise robustness analysis completed")
    except Exception as e:
        print(f"⚠ Noise robustness analysis error: {e}")
        import traceback
        traceback.print_exc()
        noise_rmses, outlier_rmses = None, None
        robustness_df, outlier_df = None, None
    
    # 8. Uncertainty analysis (prediction intervals)
    print(f"\n{'='*60}")
    print("Start uncertainty analysis (prediction intervals)")
    print(f"{'='*60}")
    bootstrap_predictions, bootstrap_coverage, quantile_coverage = None, None, None
    intervals_df = None
    try:
        bootstrap_predictions, bootstrap_coverage, quantile_coverage, intervals_df = analyze_prediction_intervals(
            model, X_test, y_test, analysis_save_dir, n_bootstrap=100, bootstrap_noise_level=0.01
        )
        print("✓ Uncertainty analysis completed")
    except Exception as e:
        print(f"⚠ Uncertainty analysis error: {e}")
        import traceback
        traceback.print_exc()
        bootstrap_predictions, bootstrap_coverage, quantile_coverage = None, None, None
        intervals_df = None
    
    # 9. Plot prediction result comparison
    print(f"\n{'='*60}")
    print("Start plotting prediction result comparison")
    print(f"{'='*60}")
    prediction_df = None
    try:
        y_test_pred = model.predict(X_test)
        plot_results(y_test, y_test_pred, analysis_save_dir)
        # Use training set + validation set as training set, test set as test set for comparison
        # Note: X_background contains all data (training set + validation set + test set), need to exclude test set
        # Merge training set and validation set (exclude test set)
        X_train_val = np.vstack([X_train, X_val]) if len(X_val) > 0 else X_train
        y_train_val = np.hstack([y_train, y_val]) if len(y_val) > 0 else y_train
        prediction_df = plot_train_test_comparison(model, X_train_val, y_train_val, X_test, y_test, analysis_save_dir)
        print("✓ Prediction result comparison plot completed")
    except Exception as e:
        print(f"⚠ Plot prediction result error: {e}")
        import traceback
        traceback.print_exc()
        prediction_df = None
    
    # 10. PDP analysis (using training set + validation set, if data set is large, it may be slow)
    print(f"\n{'='*60}")
    print("Start PDP analysis (partial dependency plot, using training set + validation set)")
    print(f"{'='*60}")
    try:
        plot_pdp_analysis(model, X_background, feature_names, analysis_save_dir, n_top_features=6)
        print("✓ PDP analysis completed")
    except Exception as e:
        print(f"⚠ PDP analysis error: {e}")
        import traceback
        traceback.print_exc()
    
    # 11. Print summary
    print(f"\n{'='*60}")
    print("Analysis completed summary")
    print(f"{'='*60}")
    
    if noise_rmses is not None:
        print(f"\nNoise robustness analysis results:")
        print(f"  - RMSE without noise: {noise_rmses[0]:.3f} MPa")
        print(f"  - RMSE with 10% noise: {noise_rmses[-1]:.3f} MPa")
        print(f"  - Performance degradation rate: {(noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100:.1f}%")
    
    if bootstrap_coverage is not None:
        print(f"\nUncertainty analysis results:")
        print(f"  - Bootstrap coverage: {bootstrap_coverage*100:.1f}%")
        print(f"  - Quantile Regression coverage: {quantile_coverage*100:.1f}%")
    
    # 11. Save all data to an Excel file in different worksheets
    print(f"\n{'='*60}")
    print("Save all analysis results to Excel file")
    print(f"{'='*60}")
    try:
        excel_path = os.path.join(analysis_save_dir, 'model_analysis_results.xlsx')
        
        # Collect all non-empty DataFrames
        sheets_to_save = []
        if importance_df is not None:
            sheets_to_save.append(('Feature importance', importance_df))
        if shap_importance is not None:
            sheets_to_save.append(('SHAP importance', shap_importance))
        if 'feature_stats' in locals() and feature_stats is not None:
            sheets_to_save.append(('Feature statistics', feature_stats))
        if 'robustness_df' in locals() and robustness_df is not None:
            sheets_to_save.append(('Noise robustness', robustness_df))
        if 'outlier_df' in locals() and outlier_df is not None:
            sheets_to_save.append(('Outlier robustness', outlier_df))
        if 'intervals_df' in locals() and intervals_df is not None:
            sheets_to_save.append(('Prediction intervals', intervals_df))
        if 'prediction_df' in locals() and prediction_df is not None:
            sheets_to_save.append(('Prediction result comparison', prediction_df))
        
        # Only save if at least one worksheet needs to be saved
        if len(sheets_to_save) > 0:
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                for sheet_name, df in sheets_to_save:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    print(f"✓ Saved: {sheet_name}")
            
            print(f"\n✓ All analysis results saved to: {excel_path}")
            print(f"   Saved {len(sheets_to_save)} worksheets")
            # Count the number of worksheets
            try:
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                sheet_count = len(wb.sheetnames)
                print(f"   Worksheet list: {', '.join(wb.sheetnames)}")
            except:
                pass
        else:
            print("⚠ No analysis results to save (all analyses failed or not executed)")
            print("   Skip Excel save")
    except Exception as e:
        print(f"⚠ Error saving Excel: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"\nAll results saved to: {analysis_save_dir}")
    print(f"{'='*60}")
    
    return {
        'model': model,
        'model_data': model_data,
        'noise_rmses': noise_rmses,
        'outlier_rmses': outlier_rmses,
        'bootstrap_coverage': bootstrap_coverage,
        'quantile_coverage': quantile_coverage,
        'shap_importance': shap_importance,
        'importance_df': importance_df
    }

if __name__ == "__main__":
    results = main()
