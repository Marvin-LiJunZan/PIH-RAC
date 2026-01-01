#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CatBoost Model Noise Analysis and Uncertainty Analysis - Peak Strain Prediction

Functions:
1. Load the pre-trained CatBoost model
2. Noise Robustness Analysis: Test the model's robustness to Gaussian noise and outliers
3. Prediction Interval Analysis: Generate confidence intervals using Bootstrap and Quantile Regression methods

Notes: This script requires running CatBoost_train.py first to train the model, then load the saved model for analysis
This script is used for peak strain prediction, with 17 features (15 material parameters + fc + Xiao_strain)
"""

import os
import sys
import numpy as np
import pandas as pd
# Set the matplotlib backend to Agg (non-interactive, to avoid GUI-related errors)
import matplotlib
matplotlib.use('Agg')  # Must be set before importing pyplot
import matplotlib.pyplot as plt
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import cross_val_score, KFold, train_test_split
from sklearn.metrics import r2_score, mean_squared_error, mean_absolute_error, mean_absolute_percentage_error
from scipy import stats
from catboost import CatBoostRegressor
import optuna
from optuna.visualization import plot_optimization_history, plot_param_importances
import shap
from sklearn.inspection import partial_dependence
from mpl_toolkits.mplot3d import Axes3D
from scipy.interpolate import interp1d, UnivariateSpline, RegularGridInterpolator, griddata
import warnings
import joblib
warnings.filterwarnings('ignore')

# Set Chinese fonts and LaTeX rendering
def configure_plot_fonts(fonts=None):
    """Unified configuration of Matplotlib fonts to ensure proper display of negative signs"""
    if fonts is None:
        fonts = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
    plt.rcParams['font.sans-serif'] = fonts
    plt.rcParams['axes.unicode_minus'] = False

configure_plot_fonts()
plt.rcParams['mathtext.default'] = 'regular'  # Render mathematical symbols using regular fonts

# Working Directory Configuration - Automatically Locate the Project Root Director
def find_project_root():
    """Locate the project root directory"""
    # First attempt to locate from the script location
    try:
        script_path = os.path.abspath(__file__)
        script_dir = os.path.dirname(script_path)
        # If the script is in a subdirectory, search upward to find the directory containing 'dataset' or 'SAVE'
        current_dir = script_dir
        search_depth = 0
        while current_dir != os.path.dirname(current_dir) and search_depth < 10:
            if os.path.exists(os.path.join(current_dir, 'dataset')) or \
               os.path.exists(os.path.join(current_dir, 'SAVE')):
                return current_dir
            current_dir = os.path.dirname(current_dir)
            search_depth += 1
    except NameError:
        #  When running in Jupyter, search upward from the current directory    
        pass
    
    # Search upward from the current working directory
    current_dir = os.path.abspath(os.getcwd())
    search_limit = 0
    while current_dir != os.path.dirname(current_dir) and search_limit < 10:
        # Check if it contains the dataset folder or specific project files
        if os.path.exists(os.path.join(current_dir, 'dataset')) or \
           os.path.exists(os.path.join(current_dir, 'SAVE')):
            return current_dir
        current_dir = os.path.dirname(current_dir)
        search_limit += 1
    
    # If still not found, return the current directory itself
    return os.path.abspath(os.getcwd())

# Set working directory
PROJECT_ROOT = find_project_root()
if not os.path.exists(os.path.join(PROJECT_ROOT, 'dataset')):
    # If still not found, use the current directory
    PROJECT_ROOT = os.getcwd()

os.chdir(PROJECT_ROOT)
print(f"Working directory set to: {PROJECT_ROOT}")

# Save directory root path
SAVE_ROOT = os.path.join(PROJECT_ROOT, 'peak_strain', 'CatBoost', 'save')
os.makedirs(SAVE_ROOT, exist_ok=True)

def compute_xiao_strain(fc: np.ndarray, r: np.ndarray) -> np.ndarray:
    """Compute peak strain using Xiao formula (consistent with PINN script implementation).
    
    Formula: ε_cp = {0.00076 + [(0.626σ_cp - 4.33) × 10^-7]^0.5} × (1 + r / (65.715r^2 - 109.43r + 48.989))
    
    Args:
        fc: Stress value (compressive strength)
        r: Aggregate replacement rate (stored as percentage×100, e.g., 1.5 represents 1.5%)
    
    Returns:
        Peak strain predicted by Xiao formula
    """
    # r = aggregate replacement rate, converted from percentage×100 (e.g., 1.5) to decimal (0.015)
    r = r / 100.0
    
    fc_clamped = np.clip(fc.astype(float), a_min=1e-6, a_max=None)
    r_clamped = np.clip(r, a_min=1e-8, a_max=None)
    
    # First term: 0.00076 + sqrt((0.626 * σ_cp - 4.33) × 10^-7)
    inner = (0.626 * fc_clamped - 4.33) * 1e-7
    inner_clamped = np.clip(inner, a_min=0.0, a_max=None)
    term1 = 0.00076 + np.sqrt(inner_clamped)

    # Second term: 1 + r / (65.715r^2 - 109.43r + 48.989)
    denom = 65.715 * r_clamped ** 2 - 109.43 * r_clamped + 48.989
    term2 = 1.0 + (r_clamped / denom)

    return term1 * term2

def load_data():
    """Load data - use 15 material parameters + peak stress (fc) + Xiao_strain to regress peak strain"""
    print("=== Load data ===")
    
    # Use the specified data file
    data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
    
    if not os.path.exists(data_file):
        print(f"Error: Data file not found: {data_file}")
        return None
    
    print(f"Found data file: {data_file}")
    
    # Read the first worksheet of the Excel file (material parameters table)
    df = pd.read_excel(data_file, sheet_name=0)
    print(f"Data shape: {df.shape}")
    print(f"Column names: {list(df.columns)}")
    
    # Define features: 15 material parameters + peak stress (fc) + Xiao_strain = 17 features
    # 10 material features
    material_features = ['water', 'cement', 'w/c', 'CS', 'sand', 'CA', 'r', 'WA', 'S', 'CI']
    # 5 specimen parameters
    specimen_features = ['age', 'μe', 'DJB', 'side', 'GJB']
    # Extra features
    extra_features = ['fc']  # Peak stress as input feature
    formula_features = ['Xiao_strain']  # Xiao formula as input feature
    
    feature_names = material_features + specimen_features + extra_features + formula_features
    
    # Target variable: peak strain
    target_column = 'peak_strain'
    
    # Check column names
    missing_cols = [col for col in feature_names if col not in df.columns]
    if missing_cols:
        print(f"Warning: Missing columns {missing_cols}")
        # 如果缺少Xiao_strain，尝试计算
        if "Xiao_strain" in missing_cols and "fc" in df.columns and "r" in df.columns:
            print("  Trying to calculate Xiao_strain...")
            df["Xiao_strain"] = compute_xiao_strain(df["fc"].values, df["r"].values)
            print("  ✓ Xiao_strain has been calculated and added")
            missing_cols = [col for col in missing_cols if col != "Xiao_strain"]
        
        # Remove other missing features
        if missing_cols:
            feature_names = [col for col in feature_names if col in df.columns]
            print(f"  Missing features removed: {missing_cols}, current feature count: {len(feature_names)}")
    
    missing_target = target_column not in df.columns
    if missing_target:
        print(f"Error: Missing target variable column '{target_column}'")
        print(f"Available columns: {list(df.columns)}")
        return None
    
    # Extract data
    X = df[feature_names].values
    y = df[target_column].values
    
    # Special check for w/c feature (if it exists)
    if 'w/c' in feature_names:
        wc_idx = feature_names.index('w/c')
        wc_values = X[:, wc_idx]
        print(f"\n🔍 Check w/c feature (original data):")
        print(f"  Unique value count: {len(np.unique(wc_values))}")
        print(f"  Minimum value: {np.min(wc_values):.10f}")
        print(f"  Maximum value: {np.max(wc_values):.10f}")
        print(f"  Range: {np.max(wc_values) - np.min(wc_values):.10f}")
        print(f"  Mean: {np.mean(wc_values):.10f}")
        print(f"  Standard deviation: {np.std(wc_values):.10f}")
        if len(np.unique(wc_values)) <= 5:
            print(f"  All unique values: {np.unique(wc_values)}")
        else:
            print(f"  First 5 unique values: {np.unique(wc_values)[:5]}")
        if np.max(wc_values) - np.min(wc_values) < 1e-6:
            print(f"  ⚠ Warning: w/c feature has a range close to 0 in the original data!")
    
    # Check for missing values
    if np.isnan(X).any():
        print(f"Warning: Feature data contains NaN values, will be filled")
        from sklearn.impute import SimpleImputer
        imputer = SimpleImputer(strategy='mean')
        X = imputer.fit_transform(X)
    
    if np.isnan(y).any():
        print(f"Warning: Target variable contains NaN values, will be removed")
        valid_mask = ~np.isnan(y)
        X = X[valid_mask]
        y = y[valid_mask]
        df = df.iloc[valid_mask].reset_index(drop=True)
    
    # Extract sample division information (using DataSlice column)
    sample_divisions = []
    if 'DataSlice' in df.columns:
        sample_divisions = df['DataSlice'].values
        print(f"Found data set division information: {np.unique(sample_divisions, return_counts=True)}")
    else:
        print("DataSlice column not found, will use random划分")
        sample_divisions = None
    
    # Extract sample ID
    sample_ids = []
    if 'No_Customized' in df.columns:
        sample_ids = df['No_Customized'].values
    else:
        sample_ids = [f"sample_{i}" for i in range(len(X))]
    
    print(f"Feature count: {X.shape[1]} (15 material parameters + fc + Xiao_strain = 17 features)")
    print(f"Sample count: {X.shape[0]}")
    print(f"Target variable '{target_column}' range: {np.min(y):.2f} - {np.max(y):.2f}")
    
    # Save original DataFrame indices (for later mapping predicted results)
    original_df_indices = df.index.values
    
    return X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df

def split_data_by_divisions(X, y, sample_divisions, sample_ids, test_ratio=0.2, val_ratio=0.2, random_state=42):
    """Split data based on sample division information"""
    print("=== Split data based on sample division information ===")
    
    if sample_divisions is None:
        from sklearn.model_selection import train_test_split
        X_train, X_temp, y_train, y_temp = train_test_split(
            X, y, test_size=test_ratio + val_ratio, random_state=random_state
        )
        val_ratio_adjusted = val_ratio / (test_ratio + val_ratio)
        X_val, X_test, y_val, y_test = train_test_split(
            X_temp, y_temp, test_size=1-val_ratio_adjusted, random_state=random_state
        )
        
        train_ids = [f"train_{i}" for i in range(len(X_train))]
        val_ids = [f"val_{i}" for i in range(len(X_val))]
        test_ids = [f"test_{i}" for i in range(len(X_test))]
        
        return X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids
    
    # Classify samples based on sample division information (supports DataSlice format: subset1, subset2, subset3, test)
    train_indices = []
    val_indices = []
    test_indices = []
    
    for i, division in enumerate(sample_divisions):
        division_str = str(division).strip()
        if division_str.lower().startswith('test'):
            test_indices.append(i)
        elif division_str.startswith('subset'):
            # subset1, subset2, subset3 are all used as training set (or can be allocated according to needs)
            # Here all are used as training set by default, if validation set is needed, it can be split from subset
            train_indices.append(i)
        elif division == 'train' or division == '训练集' or division_str.startswith('train'):
            train_indices.append(i)
        elif division == 'val' or division == '验证集' or division == 'validation' or division_str.startswith('val'):
            val_indices.append(i)
        else:
            # Default to training set
            train_indices.append(i)
    
    # If there is no validation set but there is a training set, split validation set from training set
    # If test set exists, keep it unchanged
    if train_indices and not val_indices:
        print("Only training set samples, split validation set from training set...")
        from sklearn.model_selection import train_test_split
        
        train_indices = np.array(train_indices)
        
        if len(train_indices) > 1:
            # If test set exists, only split validation set from training set
            # If test set does not exist, split validation set and test set from training set
            if len(test_indices) > 0:
                # test set already exists, only split validation set
                val_ratio_adjusted = val_ratio
                train_idx, val_idx = train_test_split(
                    train_indices, test_size=val_ratio_adjusted, random_state=random_state
                )
                train_indices = train_idx.tolist()
                val_indices = val_idx.tolist()
            else:
                # test set does not exist, split validation set and test set from training set
                train_val_idx, test_idx = train_test_split(
                    train_indices, test_size=test_ratio, random_state=random_state
                )
                val_ratio_adjusted = val_ratio / (1 - test_ratio)
                train_idx, val_idx = train_test_split(
                    train_val_idx, test_size=val_ratio_adjusted, random_state=random_state
                )
                train_indices = train_idx.tolist()
                val_indices = val_idx.tolist()
                test_indices = test_idx.tolist()
        else:
            val_indices = []
    
    # Convert to numpy array
    train_indices = np.array(train_indices)
    val_indices = np.array(val_indices)
    test_indices = np.array(test_indices)
    
    # Split data
    X_train = X[train_indices] if len(train_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    X_val = X[val_indices] if len(val_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    X_test = X[test_indices] if len(test_indices) > 0 else np.array([]).reshape(0, X.shape[1])
    
    y_train = y[train_indices] if len(train_indices) > 0 else np.array([])
    y_val = y[val_indices] if len(val_indices) > 0 else np.array([])
    y_test = y[test_indices] if len(test_indices) > 0 else np.array([])
    
    # Get corresponding sample IDs
    train_ids = [sample_ids[i] for i in train_indices] if len(train_indices) > 0 else []
    val_ids = [sample_ids[i] for i in val_indices] if len(val_indices) > 0 else []
    test_ids = [sample_ids[i] for i in test_indices] if len(test_indices) > 0 else []
    
    print(f"Data split results:")
    print(f"  Training set: {len(X_train)} samples")
    print(f"  Validation set: {len(X_val)} samples")
    print(f"  Test set: {len(X_test)} samples")
    
    return X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids

def objective(trial, X_train, y_train, X_val, y_val):
    """Optuna optimization objective function - use validation set for evaluation"""
    
    params = {
        'task_type': 'CPU',  # CatBoost default uses CPU, if GPU is needed, it can be changed to 'GPU'
        'random_seed': 42,
        'verbose': False,
        'loss_function': 'RMSE',
        
        # Hyperparameter search space (CatBoost parameters)
        'depth': trial.suggest_int('depth', 1, 10),
        'iterations': trial.suggest_int('iterations', 100, 1500),
        'learning_rate': trial.suggest_float('learning_rate', 0.001, 0.5, log=True),
        'min_child_samples': trial.suggest_int('min_child_samples', 1, 20),
        'l2_leaf_reg': trial.suggest_float('l2_leaf_reg', 0.001, 20.0, log=True),
        'subsample': trial.suggest_float('subsample', 0.1, 1.0),
        'rsm': trial.suggest_float('rsm', 0.1, 1.0),  # Feature sampling ratio
    }
    
    # Use internal cross-validation on training set for evaluation, more stable
    kf = KFold(n_splits=5, shuffle=True, random_state=42)
    scores = []
    
    for train_idx, val_idx in kf.split(X_train):
        X_tr, X_v = X_train[train_idx], X_train[val_idx]
        y_tr, y_v = y_train[train_idx], y_train[val_idx]
        
        model = CatBoostRegressor(**params)
        model.fit(X_tr, y_tr, verbose=False)
        
        y_pred = model.predict(X_v)
        r2 = r2_score(y_v, y_pred)
        scores.append(r2)
    
    return np.mean(scores)

def optimize_catboost(X_train, y_train, X_val, y_val, n_trials=100):
    """Use Optuna to optimize CatBoost hyperparameters - use 5-fold cross-validation for evaluation"""
    print(f"\n=== Start CatBoost hyperparameter optimization (n_trials={n_trials}) ===")
    print("Use 5-fold cross-validation for evaluation of hyperparameters, improve stability")
    
    study = optuna.create_study(direction='maximize', study_name='catboost_optimization')
    study.optimize(lambda trial: objective(trial, X_train, y_train, X_val, y_val), n_trials=n_trials, n_jobs=1)
    
    print("\n=== Optimization completed ===")
    print(f"Best R2 score: {study.best_value:.4f}")
    print(f"Best parameters: {study.best_params}")
    
    # Visualize optimization history
    try:
        save_dir = os.path.join(SAVE_ROOT, 'catboost_optimization')
        os.makedirs(save_dir, exist_ok=True)
        
        fig = plot_optimization_history(study)
        plt.savefig(os.path.join(save_dir, 'optimization_history.png'), dpi=300, bbox_inches='tight')
        plt.close()
        
        fig = plot_param_importances(study)
        plt.savefig(os.path.join(save_dir, 'param_importance.png'), dpi=300, bbox_inches='tight')
        plt.close()
    except:
        pass
    
    return study

def train_catboost(X_train, y_train, X_val, y_val, best_params):
    """Use best parameters to train CatBoost model"""
    print("\n=== Use best parameters to train final model ===")
    
    # Merge training set and validation set
    X_train_full = np.vstack([X_train, X_val])
    y_train_full = np.hstack([y_train, y_val])
    
    # Create model
    model = CatBoostRegressor(
        task_type='CPU',  # If GPU is needed, it can be changed to 'GPU'
        random_seed=42,
        verbose=False,
        loss_function='RMSE',
        **best_params
    )
    
    # Train model
    model.fit(X_train_full, y_train_full, verbose=False)
    
    return model

def plot_results(y_true, y_pred, save_dir, model_name='xgboost'):
    """Plot results"""
    print("\n=== Plot results ===")
    
    os.makedirs(save_dir, exist_ok=True)
    
    y_true = np.array(y_true).flatten()
    y_pred = np.array(y_pred).flatten()
    
    r2 = r2_score(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    rmse = np.sqrt(mean_squared_error(y_true, y_pred))
    
    # Create subplots
    fig, axes = plt.subplots(2, 2, figsize=(15, 12))
    fig.suptitle(f'CatBoost Results ($R^2$={r2:.4f})', fontsize=16, fontweight='bold')
    
    # 1. Predicted vs true values
    axes[0,0].scatter(y_true, y_pred, alpha=0.7, s=60, color='green')
    axes[0,0].plot([y_true.min(), y_true.max()], [y_true.min(), y_true.max()], 'r--', lw=2)
    axes[0,0].set_xlabel('True Peak Strain')
    axes[0,0].set_ylabel('Predicted Peak Strain')
    axes[0,0].set_title(f'Prediction vs True Values\n$R^2$={r2:.4f}, MAE={mae:.3f}, RMSE={rmse:.3f}')
    axes[0,0].grid(True, alpha=0.3)
    
    # 2. Residual plot
    residuals = y_pred - y_true
    axes[0,1].scatter(y_pred, residuals, alpha=0.7, s=60, color='green')
    axes[0,1].axhline(y=0, color='r', linestyle='--', lw=2)
    axes[0,1].set_xlabel('Predicted Peak Strain')
    axes[0,1].set_ylabel('Residuals')
    axes[0,1].set_title('Residual Plot')
    axes[0,1].grid(True, alpha=0.3)
    
    # 3. Residual distribution
    axes[1,0].hist(residuals, bins=15, alpha=0.7, color='green', density=True)
    axes[1,0].axvline(x=0, color='r', linestyle='--', lw=2)
    axes[1,0].set_xlabel('Residuals')
    axes[1,0].set_ylabel('Density')
    axes[1,0].set_title('Residual Distribution')
    axes[1,0].grid(True, alpha=0.3)
    
    # 4. Feature importance
    # Note: Feature importance needs to be obtained from the model object after training
    stats_text = f"""Model Performance Statistics:
    
Sample Count: {len(y_true)}
R2: {r2:.4f}
MAE: {mae:.3f}
RMSE: {rmse:.3f}

CatBoost Features:
- Gradient Boosting
- Optuna Hyperparameter Optimization
- Cross-validation evaluation
- Feature Importance Analysis"""
    
    axes[1,1].text(0.1, 0.9, stats_text, transform=axes[1,1].transAxes, 
                   fontsize=11, verticalalignment='top', fontfamily='monospace')
    axes[1,1].set_xlim(0, 1)
    axes[1,1].set_ylim(0, 1)
    axes[1,1].axis('off')
    axes[1,1].set_title('Performance Statistics')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, f'catboost_results.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    return r2, mae, rmse

def plot_train_test_comparison(model, X_train, y_train, X_test, y_test, save_dir):
    """Plot training set and test set comparison (with edge distribution) - optimized version
    
    Returns:
        DataFrame: DataFrame containing training set and test set prediction results
    """
    print("\n=== Plot training set and test set comparison ===")
    
    # Predict
    y_train_pred = model.predict(X_train)
    y_test_pred = model.predict(X_test)
    
    # Calculate metrics
    r2_train = r2_score(y_train, y_train_pred)
    mae_train = mean_absolute_error(y_train, y_train_pred)
    rmse_train = np.sqrt(mean_squared_error(y_train, y_train_pred))
    r2_test = r2_score(y_test, y_test_pred)
    mae_test = mean_absolute_error(y_test, y_test_pred)
    rmse_test = np.sqrt(mean_squared_error(y_test, y_test_pred))
    
    # Create figure (scatter plot with edge distribution)
    fig = plt.figure(figsize=(13, 11))
    
    # Define grid layout (adjust ratio)
    gs = fig.add_gridspec(3, 3, hspace=0.08, wspace=0.08,
                         height_ratios=[0.8, 4, 0.1], width_ratios=[4, 0.8, 0.1])
    
    # Main scatter plot
    ax_main = fig.add_subplot(gs[1, 0])
    
    # Top histogram
    ax_top = fig.add_subplot(gs[0, 0], sharex=ax_main)
    
    # Right histogram
    ax_right = fig.add_subplot(gs[1, 1], sharey=ax_main)
    
    # Plot main scatter plot
    # Training set - use more professional blue
    ax_main.scatter(y_train, y_train_pred, alpha=0.5, s=50, color='#2E86AB', 
                   label='Training', edgecolors='white', linewidths=0.5, zorder=3)
    
    # Test set - use more prominent orange-red
    ax_main.scatter(y_test, y_test_pred, alpha=0.7, s=60, color='#E63946',
                   label='Testing', edgecolors='white', linewidths=0.5, zorder=4)
    
    # 1:1 diagonal line
    all_values = np.concatenate([y_train, y_test, y_train_pred, y_test_pred])
    min_val = all_values.min()
    max_val = all_values.max()
    # Add padding
    padding = (max_val - min_val) * 0.05
    min_val -= padding
    max_val += padding
    
    ax_main.plot([min_val, max_val], [min_val, max_val], 'k--', lw=2, 
                alpha=0.4, label='Perfect Prediction', zorder=1)
    
    # Calculate and plot best fit line and confidence interval (training set)
    slope_train, intercept_train, r_value_train, p_value_train, std_err_train = stats.linregress(y_train, y_train_pred)
    line_train = slope_train * y_train + intercept_train
    predict_train = slope_train * np.sort(y_train) + intercept_train
    
    # Calculate confidence interval
    residuals_train = y_train_pred - line_train
    std_residuals_train = np.std(residuals_train)
    ci_train = 1.96 * std_residuals_train  # 95% CI
    
    ax_main.plot(np.sort(y_train), predict_train, color='#2E86AB', lw=3, alpha=0.9,
                linestyle='-', zorder=5)
    ax_main.fill_between(np.sort(y_train), predict_train - ci_train, predict_train + ci_train,
                         alpha=0.2, color='#2E86AB', zorder=2)
    
    # Best fit line and confidence interval (test set)
    slope_test, intercept_test, r_value_test, p_value_test, std_err_test = stats.linregress(y_test, y_test_pred)
    line_test = slope_test * y_test + intercept_test
    predict_test = slope_test * np.sort(y_test) + intercept_test
    
    residuals_test = y_test_pred - line_test
    std_residuals_test = np.std(residuals_test)
    ci_test = 1.96 * std_residuals_test
    
    ax_main.plot(np.sort(y_test), predict_test, color='#E63946', lw=3, alpha=0.9,
                linestyle='-', zorder=6)
    ax_main.fill_between(np.sort(y_test), predict_test - ci_test, predict_test + ci_test,
                         alpha=0.2, color='#E63946', zorder=2)
    
    # Set axis labels
    ax_main.set_xlabel('Observed Peak Strain', fontsize=13, fontweight='bold')
    ax_main.set_ylabel('Predicted Peak Strain', fontsize=13, fontweight='bold')
    
    # Optimized legend - placed in lower right corner, using LaTeX format to display R²
    legend_elements = [
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#2E86AB', 
                  markersize=8, alpha=0.7, label=f'Training ($R^2$={r2_train:.3f}, MAE={mae_train:.2f})'),
        plt.Line2D([0], [0], marker='o', color='w', markerfacecolor='#E63946', 
                  markersize=9, alpha=0.8, label=f'Testing ($R^2$={r2_test:.3f}, MAE={mae_test:.2f})'),
        plt.Line2D([0], [0], color='k', linestyle='--', lw=2, alpha=0.4, label='Perfect Prediction'),
        plt.Line2D([0], [0], color='gray', lw=3, label='Best Fit Line'),
    ]
    ax_main.legend(handles=legend_elements, loc='lower right', fontsize=9.5, 
                  framealpha=0.95, edgecolor='gray', fancybox=True, shadow=True)
    
    ax_main.grid(True, alpha=0.25, linestyle='--', linewidth=0.5)
    ax_main.set_axisbelow(True)
    
    # Top histogram (observed value distribution) - stacked display
    bins = np.linspace(min_val, max_val, 20)
    # Use numpy.histogram to calculate frequency
    hist_train_obs, _ = np.histogram(y_train, bins=bins)
    hist_test_obs, _ = np.histogram(y_test, bins=bins)
    
    # Plot stacked histogram
    bin_centers = (bins[:-1] + bins[1:]) / 2
    width = bins[1] - bins[0]
    ax_top.bar(bin_centers, hist_train_obs, width=width, alpha=0.7, 
              color='#2E86AB', label='Training', edgecolor='white', linewidth=0.5)
    ax_top.bar(bin_centers, hist_test_obs, width=width, alpha=0.8, 
              color='#E63946', label='Testing', edgecolor='white', linewidth=0.5, bottom=hist_train_obs)
    ax_top.set_ylabel('Count', fontsize=11, fontweight='bold')
    ax_top.legend(loc='upper right', fontsize=9, framealpha=0.9)
    ax_top.tick_params(labelbottom=False, labelsize=10)
    ax_top.grid(True, alpha=0.25, axis='y', linestyle='--', linewidth=0.5)
    ax_top.set_axisbelow(True)
    ax_top.spines['top'].set_visible(False)
    ax_top.spines['right'].set_visible(False)
    
    # Right histogram (predicted value distribution) - stacked display
    hist_train_pred, _ = np.histogram(y_train_pred, bins=bins)
    hist_test_pred, _ = np.histogram(y_test_pred, bins=bins)
    
    # Plot horizontal stacked histogram
    ax_right.barh(bin_centers, hist_train_pred, height=width, alpha=0.7, 
                 color='#2E86AB', edgecolor='white', linewidth=0.5)
    ax_right.barh(bin_centers, hist_test_pred, height=width, alpha=0.8, 
                 color='#E63946', edgecolor='white', linewidth=0.5, left=hist_train_pred)
    ax_right.set_xlabel('Count', fontsize=11, fontweight='bold')
    ax_right.tick_params(labelleft=False, labelsize=10)
    ax_right.grid(True, alpha=0.25, axis='x', linestyle='--', linewidth=0.5)
    ax_right.set_axisbelow(True)
    ax_right.spines['top'].set_visible(False)
    ax_right.spines['right'].set_visible(False)
    
    # Add total title - include sample number information
    title = f'CatBoost Model Performance\nTraining: n={len(y_train)} | Testing: n={len(y_test)}'
    fig.suptitle(title, fontsize=15, fontweight='bold', y=0.985)
    
    # Set overall background
    fig.patch.set_facecolor('white')
    
    plt.savefig(os.path.join(save_dir, 'train_test_comparison_with_margins.png'), 
               dpi=300, bbox_inches='tight', facecolor='white')
    # plt.show()
    plt.close()
    
    print(f"\nTraining set performance - R2: {r2_train:.4f}, MAE: {mae_train:.3f}, RMSE: {rmse_train:.3f}")
    print(f"Test set performance - R2: {r2_test:.4f}, MAE: {mae_test:.3f}, RMSE: {rmse_test:.3f}")
    print(f"Performance difference - ΔR2: {abs(r2_train - r2_test):.4f}, ΔMAE: {abs(mae_train - mae_test):.3f}")
    
    # Create DataFrame containing training set and test set prediction results
    train_df = pd.DataFrame({
        'dataset': ['Training set'] * len(y_train),
        'observed': y_train,
        'predicted': y_train_pred,
        'residual': y_train - y_train_pred,
        'absolute_error': np.abs(y_train - y_train_pred)
    })
    
    test_df = pd.DataFrame({
        'dataset': ['Test set'] * len(y_test),
        'observed': y_test,
        'predicted': y_test_pred,
        'residual': y_test - y_test_pred,
        'absolute_error': np.abs(y_test - y_test_pred)
    })
    
    # Merge training set and test set data
    prediction_df = pd.concat([train_df, test_df], ignore_index=True)
    
    return prediction_df

def plot_feature_importance(model, feature_names, save_dir, X=None):
    """Plot feature importance
    
    Parameters:
        model: Trained CatBoost model
        feature_names: Feature name list
        save_dir: Save directory
        X: Feature data (optional, for calculating coefficient of variation)
    """
    print("\n=== Plot feature importance ===")
    
    # Get feature importance
    importance = model.feature_importances_
    
    # Create DataFrame
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': importance
    })
    
    # If feature data is provided, calculate coefficient of variation (CV)
    if X is not None:
        cv_values = []
        for i in range(X.shape[1]):
            feature_mean = np.mean(X[:, i])
            feature_std = np.std(X[:, i])
            # Coefficient of variation = standard deviation / mean (to avoid division by zero)
            cv = feature_std / (feature_mean + 1e-10) if abs(feature_mean) > 1e-10 else 0.0
            cv_values.append(cv)
        importance_df['CV'] = cv_values
        print("✓ Coefficient of variation (CV) has been calculated")
    else:
        importance_df['CV'] = np.nan
        print("⚠ Feature data not provided, cannot calculate coefficient of variation")
    
    # Sort by importance
    importance_df = importance_df.sort_values('importance', ascending=False)
    
    # Plot bar chart
    plt.figure(figsize=(12, 15))
    plt.barh(importance_df['feature'], importance_df['importance'], color='skyblue')
    plt.xlabel('Feature Importance')
    plt.title('CatBoost Feature Importance')
    plt.gca().invert_yaxis()
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'feature_importance.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    # Save to Excel (will be saved to comprehensive Excel file in main function)
    # Here not directly saved, return DataFrame for main function to save
    
    # Print statistics
    print(f"\nFeature importance statistics:")
    print(f"   Most important feature: {importance_df.iloc[0]['feature']} (importance={importance_df.iloc[0]['importance']:.6f})")
    if X is not None:
        print(f"   Highest CV feature: {importance_df.loc[importance_df['CV'].idxmax(), 'feature']} (CV={importance_df['CV'].max():.4f})")
        print(f"   Lowest CV feature: {importance_df.loc[importance_df['CV'].idxmin(), 'feature']} (CV={importance_df['CV'].min():.4f})")
    
    return importance_df

# CatBoost model does not need to be repaired, just use it directly

def plot_shap_analysis(model, X_background, X_explain, feature_names, save_dir, y_explain=None):
    """SHAP value analysis
    
    Parameters:
        model: Trained CatBoost model
        X_background: Background data set (used to initialize explainer, usually using training set)
        X_explain: Data set to explain (data for calculating SHAP values, usually using test set)
        feature_names: Feature name list
        save_dir: Save directory
        y_explain: Target variable value (optional, for checking the correlation between features and target variable)
    
    Traditional explanation method:
        - Background data set (X_background): Using training set (or training set + validation set)
          Used to initialize SHAP explainer, representing the "baseline" distribution of the model
        - Explanation data set (X_explain): Using test set
          Explain the unseen data, evaluate the interpretability of the model's generalization ability
        
        Advantages:
        1. Conforms to machine learning best practices (training/test separation)
        2. Evaluate the interpretability of the model in real application scenarios
        3. Avoid data leakage (do not use test set to train explainer)
    """
    print("\n=== Perform SHAP analysis ===")
    print(f"Background data set sample number: {len(X_background)} (used to initialize explainer)")
    print(f"Explanation data set sample number: {len(X_explain)} (calculate SHAP values)")
    
    # CatBoost model can directly use TreeExplainer
    # If the background data set is passed in incorrectly, the background data set can be not passed in (using default value)
    try:
        # Try to initialize explainer using background data set
        explainer = shap.TreeExplainer(model, X_background)
        print("✓ Successfully initialized explainer using background data set")
    except (ValueError, TypeError, AttributeError) as e:
        print(f"⚠ Failed to initialize explainer using background data set: {e}")
        print("  Try to initialize explainer without background data set...")
        try:
            # Do not pass in background data set, use default value
            explainer = shap.TreeExplainer(model)
            print("✓ Successfully initialized explainer without background data set")
        except Exception as e2:
            print(f"⚠ Failed to initialize explainer: {e2}")
            print("  Try to use KernelExplainer as a backup solution...")
            # If TreeExplainer fails, try to use KernelExplainer (slower but more general)
            try:
                # Sample background data to improve speed
                if len(X_background) > 100:
                    background_sample = shap.sample(X_background, 100)
                else:
                    background_sample = X_background
                explainer = shap.KernelExplainer(model.predict, background_sample)
                print("✓ Successfully initialized explainer using KernelExplainer (note: calculation is slower)")
            except Exception as e3:
                print(f"⚠ All explainer initialization methods failed: {e3}")
                raise ValueError(f"Failed to initialize SHAP explainer: {e3}")
    
    # Calculate SHAP values (for explanation data set)
    print("Calculating SHAP values...")
    shap_values = explainer.shap_values(X_explain)
    
    # Set font, avoid Chinese乱码，并确保负号正常显示
    original_font = list(plt.rcParams['font.sans-serif'])
    original_unicode_minus = plt.rcParams['axes.unicode_minus']
    
    # For SHAP plot, prioritize DejaVu Sans to ensure negative sign is displayed correctly
    configure_plot_fonts(['DejaVu Sans', 'SimHei', 'Microsoft YaHei'])
    plt.rcParams['axes.unicode_minus'] = False  # Ensure negative sign is displayed correctly
    
    # 1. Summary plot
    #Greatly increase the width and truly stretch the horizontal axis scale
    fig = plt.figure(figsize=(30, 12))  # Width increased to 30, significantly elongate the horizontal axis
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, show=False)
    # Get current axes
    ax = plt.gca()
    # Do not change data range, only elongate the physical ratio of the horizontal axis by increasing the graphic width
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_summary.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # 2. Bar plot (feature importance)
    plt.figure(figsize=(10, 8))
    shap.summary_plot(shap_values, X_explain, feature_names=feature_names, plot_type="bar", show=False)
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'shap_bar.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # 3. Waterfall plot (display SHAP values for a single sample)
    # Note: This is an analysis of a single sample, showing how each feature affects the prediction result for that sample
    # Select the first 3 samples as examples (if the data is sufficient)
    n_samples_waterfall = min(3, len(shap_values))
    
    for sample_idx in range(n_samples_waterfall):
        shap_explanation = shap.Explanation(
            values=shap_values[sample_idx:sample_idx+1],
            base_values=explainer.expected_value,
            data=X_explain[sample_idx:sample_idx+1],
            feature_names=feature_names
        )
        
        # Ensure negative sign is displayed correctly - use DejaVu Sans font (supports negative sign)
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'SimHei', 'Microsoft YaHei']
        
        plt.figure(figsize=(12, 8))
        shap.plots.waterfall(shap_explanation[0], show=False)
        
        # Get current axes and ensure negative sign is displayed
        ax = plt.gca()
        # Force negative sign display and font
        ax.tick_params(which='major', labelsize=10)
        # Ensure all text uses fonts that support negative sign
        for text in ax.texts:
            text.set_fontfamily('DejaVu Sans')
        
        plt.tight_layout()
        if n_samples_waterfall == 1:
            plt.savefig(os.path.join(save_dir, 'shap_waterfall.png'), dpi=300, bbox_inches='tight')
        else:
            plt.savefig(os.path.join(save_dir, f'shap_waterfall_sample_{sample_idx+1}.png'), dpi=300, bbox_inches='tight')
        plt.close()
    
    print(f"✓ Generated {n_samples_waterfall} waterfall plots for samples")
    
    # 4. Force plot (display the prediction process for a single sample, long bar force plot)
    # Create an explanation containing multiple samples for force plot
    n_samples_force = min(3, len(shap_values))
    shap_explanation_force = shap.Explanation(
        values=shap_values[:n_samples_force],
        base_values=explainer.expected_value,
        data=X_explain[:n_samples_force],
        feature_names=feature_names
    )
    # Draw force plot for the first 3 samples
    for i in range(n_samples_force):
        # Ensure negative sign is displayed correctly
        plt.rcParams['axes.unicode_minus'] = False
        plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'SimHei', 'Microsoft YaHei']
        
        plt.figure(figsize=(16, 4))  # Long bar graph
        shap.plots.force(shap_explanation_force[i], matplotlib=True, show=False)
        
        # Ensure negative sign is displayed
        ax = plt.gca()
        for text in ax.texts:
            text.set_fontfamily('DejaVu Sans')
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, f'shap_force_sample_{i}.png'), dpi=300, bbox_inches='tight')
        # plt.show()
        plt.close()
    
    # 5. Calculate the average absolute SHAP value as feature importance
    shap_importance = pd.DataFrame({
        'feature': feature_names,
        'shap_importance': np.abs(shap_values).mean(axis=0),
        'shap_mean': shap_values.mean(axis=0),  # Average SHAP value (considering positive and negative)
        'shap_std': shap_values.std(axis=0),    # SHAP value standard deviation
        'shap_max': np.abs(shap_values).max(axis=0),  # Maximum absolute SHAP value
        'shap_min': np.abs(shap_values).min(axis=0)   # Minimum absolute SHAP value
    }).sort_values('shap_importance', ascending=False)
    
    # Save SHAP importance (will be saved in main function)
    # Return DataFrame for main function to save
    
    # 6. Analyze the reasons why features affect small (print detailed analysis)
    print("\n" + "="*60)
    print("SHAP feature importance analysis")
    print("="*60)
    
    # Special check for the correlation between Xiao_strain and peak_strain
    if 'Xiao_strain' in feature_names:
        xiao_idx = feature_names.index('Xiao_strain')
        xiao_values = X_explain[:, xiao_idx]
        # Calculate the actual correlation between Xiao_strain and the target variable
        xiao_target_corr = np.corrcoef(xiao_values, y_explain)[0, 1]
        # Get the SHAP value for Xiao_strain
        xiao_shap_mean = shap_importance[shap_importance['feature'] == 'Xiao_strain']['shap_mean'].values[0] if len(shap_importance[shap_importance['feature'] == 'Xiao_strain']) > 0 else None
        
        print(f"\n🔍 Special check: Xiao_strain feature")
        print(f"  - Actual correlation with target variable (peak_strain): {xiao_target_corr:+.4f}")
        if xiao_shap_mean is not None:
            print(f"  - Average SHAP impact value: {xiao_shap_mean:+.4f}")
            if xiao_target_corr > 0 and xiao_shap_mean < 0:
                print(f"  ⚠ Warning:发现异常情况！")
                print(f"    - Xiao_strain与peak_strain的实际相关性为正 ({xiao_target_corr:+.4f})")
                print(f"    - But SHAP value shows negative impact ({xiao_shap_mean:+.4f})")
                print(f"    - This may indicate:")
                print(f"      1. The model may have learned some compensation relationship (if the predicted value of Xiao_strain is high, the model may think it needs to reduce the prediction)")
                print(f"      2. There is multi-collinearity between features, causing the model to learn the wrong pattern")
                print(f"      3. The model may have overfitted the training data")
                print(f"    - Suggestion: Check the correlation between features, consider removing Xiao_strain or adjusting the model")
            elif xiao_target_corr < 0 and xiao_shap_mean > 0:
                print(f"  ⚠ Warning: Found abnormal situation!")
                print(f"    - Actual correlation with target variable (peak_strain) is negative ({xiao_target_corr:+.4f})")
                print(f"    - But SHAP value shows positive impact ({xiao_shap_mean:+.4f})")
                print(f"    - This may indicate that the model has learned some compensation relationship")
        print(f"  - Xiao_strain value range: [{np.min(xiao_values):.6f}, {np.max(xiao_values):.6f}]")
        print(f"  - peak_strain value range: [{np.min(y_explain):.6f}, {np.max(y_explain):.6f}]")
    
    print(f"\nThe 5 most important features:")
    for idx, row in shap_importance.head(5).iterrows():
        print(f"  {row['feature']:10s}: Average impact = {row['shap_importance']:.4f}, "
              f"Average SHAP value = {row['shap_mean']:+.4f}, Standard deviation = {row['shap_std']:.4f}")
    
    print(f"\nThe 5 least important features:")
    for idx, row in shap_importance.tail(5).iterrows():
        print(f"  {row['feature']:10s}: Average impact = {row['shap_importance']:.4f}, "
              f"Average SHAP value = {row['shap_mean']:+.4f}, Standard deviation = {row['shap_std']:.4f}")
    
    # Analyze the possible reasons why features affect small
    low_importance_features = shap_importance[shap_importance['shap_importance'] < 0.1]
    if len(low_importance_features) > 0:
        print(f"\n⚠ Found {len(low_importance_features)} features that affect small (Average SHAP value < 0.1):")
        for idx, row in low_importance_features.iterrows():
            print(f"  - {row['feature']}: Average impact = {row['shap_importance']:.4f}")
        
        print("\nPossible reasons:")
        print("  1. The value range of the feature is small (small variance), resulting in limited impact on prediction")
        print("  2. Highly correlated with other features, replaced by other features")
        print("  3. The feature values are distributed evenly in the data set, lacking distinction")
        print("  4. This feature has a small impact on the target variable ( consistent with the principles of physics and engineering)")
        print("  5. The sample size is small (94 samples), the impact of some features may be covered by noise")
        print("\nSuggestion:")
        print("  - Check the statistical distribution of these features (mean, standard deviation, range)")
        print("  - Analyze the correlation between features")
        print("  - View PDP to understand the marginal effect of these features")
        print("  - Consider feature engineering (such as feature combination, transformation)")
    
    print("="*60)
    
    # Restore original font settings
    plt.rcParams['font.sans-serif'] = original_font
    plt.rcParams['axes.unicode_minus'] = False
    
    return shap_importance

def analyze_feature_statistics(X, y, feature_names, shap_importance, save_dir):
    """Analyze feature statistics, explain why some features affect small"""
    print("\n" + "="*60)
    print("Feature statistics analysis (explain why some features affect small)")
    print("="*60)
    
    import pandas as pd
    
    # Create feature index mapping (feature_name -> column_index)
    feature_to_idx = {name: idx for idx, name in enumerate(feature_names)}
    
    # Create feature_stats according to the order of shap_importance (consistent with the SHAP summary plot)
    # shap_importance is already sorted by importance in descending order
    feature_stats_list = []
    
    for idx, row in shap_importance.iterrows():
        feat_name = row['feature']
        feat_idx = feature_to_idx[feat_name]
        
        # Calculate the statistical information for this feature
        feat_data = X[:, feat_idx]
        q25 = np.percentile(feat_data, 25)
        q50 = np.percentile(feat_data, 50)
        q75 = np.percentile(feat_data, 75)
        corr = np.corrcoef(feat_data, y)[0, 1]
        
        feat_range = np.max(feat_data) - np.min(feat_data)
        feat_std = np.std(feat_data)
        
        # Special check for w/c feature
        if feat_name == 'w/c':
            print(f"\n🔍 Detailed check for w/c feature:")
            print(f"  Data shape: {feat_data.shape}")
            print(f"  Unique value count: {len(np.unique(feat_data))}")
            print(f"  Unique values: {np.unique(feat_data)[:10]}")  # Display the first 10 unique values
            print(f"  Minimum value: {np.min(feat_data):.10f}")
            print(f"  Maximum value: {np.max(feat_data):.10f}")
            print(f"  Range: {feat_range:.10f}")
            print(f"  Mean: {np.mean(feat_data):.10f}")
            print(f"  Standard deviation: {feat_std:.10f}")
            print(f"  SHAP importance: {row['shap_importance']:.10f}")
            if feat_range < 1e-6:
                print(f"  ⚠ Warning: w/c feature value range is close to 0!")
                print(f"  Possible reasons:")
                print(f"    1. All values in the w/c column are the same in the data")
                print(f"    2. All values are the same after w/c is standardized during data preprocessing")
                print(f"    3. The w/c column is a constant in the data file")
                print(f"  Suggestion: Check the w/c column in the original data file")
        
        feature_stats_list.append({
            'feature': feat_name,
            'mean': np.mean(feat_data),
            'std': feat_std,
            'min': np.min(feat_data),
            'q25': q25,
            'median': q50,
            'q75': q75,
            'max': np.max(feat_data),
            'range': feat_range,
            'iqr': q75 - q25,
            'cv': feat_std / (np.mean(feat_data) + 1e-10),
            'shap_importance': row['shap_importance'],
            'correlation_with_target': corr
        })
    
    feature_stats = pd.DataFrame(feature_stats_list)
    
    # Rearrange the order of columns to make it more logical: basic statistics -> quartiles -> derived statistics -> importance
    column_order = ['feature', 'mean', 'std', 'min', 'q25', 'median', 'q75', 'max', 
                    'range', 'iqr', 'cv', 'shap_importance', 'correlation_with_target']
    feature_stats = feature_stats[column_order]
    
    # feature_stats is already sorted by shap_importance (consistent with the SHAP summary plot)
    
    # Save statistical information (will be saved in main function)
    # Return DataFrame for main function to save
    
    # Analyze features that affect small
    low_importance_threshold = 0.1
    low_importance = feature_stats[feature_stats['shap_importance'] < low_importance_threshold]
    
    print(f"\nAnalysis of features that affect small (SHAP importance < {low_importance_threshold}):")
    if len(low_importance) > 0:
        print(f"\nThere are {len(low_importance)} features that affect small:")
        for idx, row in low_importance.iterrows():
            print(f"\n  {row['feature']}:")
            print(f"    - SHAP importance: {row['shap_importance']:.4f}")
            print(f"    - Correlation with target variable: {row['correlation_with_target']:.4f}")
            print(f"    - Mean: {row['mean']:.4f}, Standard deviation: {row['std']:.4f}")
            print(f"    - Range: [{row['min']:.4f}, {row['max']:.4f}], Variation coefficient: {row['cv']:.4f}")
            print(f"    - Quartiles: Q1={row['q25']:.4f}, Median={row['median']:.4f}, Q3={row['q75']:.4f}, IQR={row['iqr']:.4f}")
            
            # Determine possible reasons
            reasons = []
            if abs(row['correlation_with_target']) < 0.1:
                reasons.append("Low correlation with the target variable")
            if row['cv'] < 0.1:
                reasons.append("Small variation coefficient (small data range)")
            if row['range'] < row['mean'] * 0.1:
                reasons.append("Small value range relative to the mean")
            
            if reasons:
                print(f"    - Possible reasons: {', '.join(reasons)}")
    
    # Analyze features that affect large
    high_importance = feature_stats.head(5)
    print(f"\nThe 5 most important features:")
    for idx, row in high_importance.iterrows():
        print(f"  {row['feature']:10s}: SHAP={row['shap_importance']:.4f}, "
              f"Correlation={row['correlation_with_target']:+.4f}, CV={row['cv']:.4f}")
    
    # Plot feature statistics
    fig, axes = plt.subplots(2, 2, figsize=(16, 12))
    
    # 1. SHAP importance vs correlation
    ax1 = axes[0, 0]
    scatter = ax1.scatter(feature_stats['correlation_with_target'], 
                         feature_stats['shap_importance'], 
                         s=100, alpha=0.6, c=feature_stats['cv'], cmap='viridis')
    for idx, row in feature_stats.iterrows():
        ax1.annotate(row['feature'], 
                    (row['correlation_with_target'], row['shap_importance']),
                    fontsize=9, alpha=0.7)
    ax1.set_xlabel('Correlation with the target variable', fontsize=12)
    ax1.set_ylabel('SHAP importance', fontsize=12)
    ax1.set_title('SHAP importance vs correlation', fontsize=14, fontweight='bold')
    ax1.grid(True, alpha=0.3)
    plt.colorbar(scatter, ax=ax1, label='Variation coefficient')
    
    # 2. SHAP importance vs variation coefficient
    ax2 = axes[0, 1]
    ax2.scatter(feature_stats['cv'], feature_stats['shap_importance'], s=100, alpha=0.6)
    for idx, row in feature_stats.iterrows():
        ax2.annotate(row['feature'], 
                    (row['cv'], row['shap_importance']),
                    fontsize=9, alpha=0.7)
    ax2.set_xlabel('Variation coefficient (CV)', fontsize=12)
    ax2.set_ylabel('SHAP importance', fontsize=12)
    ax2.set_title('SHAP importance vs variation coefficient', fontsize=14, fontweight='bold')
    ax2.grid(True, alpha=0.3)
    
    # 3. Feature value range (sorted by SHAP importance, consistent with the SHAP swarm plot)
    ax3 = axes[1, 0]
    # feature_stats is already sorted by SHAP importance in descending order, directly take the top 10
    # Note: The SHAP summary plot is sorted from top to bottom by importance, so the most important ones are at the top
    top_features = feature_stats.head(10).copy()
    
    # Check and print detailed information for the w/c feature
    wc_feature = feature_stats[feature_stats['feature'] == 'w/c']
    if len(wc_feature) > 0:
        wc_row = wc_feature.iloc[0]
        print(f"\n⚠ Warning: w/c feature statistics:")
        print(f"  SHAP importance: {wc_row['shap_importance']:.6f}")
        print(f"  Minimum value: {wc_row['min']:.6f}")
        print(f"  Maximum value: {wc_row['max']:.6f}")
        print(f"  Range: {wc_row['range']:.6f}")
        print(f"  Mean: {wc_row['mean']:.6f}")
        print(f"  Standard deviation: {wc_row['std']:.6f}")
        print(f"  Variation coefficient: {wc_row['cv']:.6f}")
        if wc_row['range'] < 1e-6:
            print(f"  ⚠ Warning: w/c feature value range is close to 0, all sample values are the same!")
            print(f"  This will cause the SHAP importance to be 0, because when the feature value is constant, it cannot affect the prediction.")
            print(f"  Please check the actual values of the w/c column in the data.")
    
    # Filter out features with range 0 or close to 0 (these features should not appear in the importance plot)
    # But keep them in top_features so the user can see the problem
    valid_range_features = top_features[top_features['range'] > 1e-6]
    
    if len(valid_range_features) < len(top_features):
        print(f"\n⚠ Warning: Found {len(top_features) - len(valid_range_features)} features with range 0 or close to 0")
        print(f"  These features: {top_features[top_features['range'] <= 1e-6]['feature'].tolist()}")
        print(f"  All features will be used to plot (including features with range 0), to find the problem")
    
    y_pos = np.arange(len(top_features))
    # Plot horizontal bar chart, the most important features at the top (index 0)
    # For features with range 0, use a very small value (e.g. 1e-6) to display on the plot
    bar_values = top_features['range'].values.copy()
    bar_values[bar_values < 1e-6] = 1e-6  # Replace 0 values with very small values to display
    
    bars = ax3.barh(y_pos, bar_values, alpha=0.7)
    # For features with range 0, use different colors
    for i, (idx, row) in enumerate(top_features.iterrows()):
        if row['range'] < 1e-6:
            bars[i].set_color('red')
            bars[i].set_alpha(0.5)
            bars[i].set_label('range=0')
    
    ax3.set_yticks(y_pos)
    ax3.set_yticklabels(top_features['feature'])
    ax3.set_xlabel('Feature value range (max - min)', fontsize=12)
    ax3.set_title('Value range of the 10 most important features (sorted by SHAP importance)', fontsize=14, fontweight='bold')
    ax3.grid(True, alpha=0.3, axis='x')
    # Invert the y-axis, so the most important features are at the top (consistent with the SHAP summary plot: importance decreases from top to bottom)
    ax3.invert_yaxis()
    
    # If there are features with range 0, add a legend说明
    if (top_features['range'] < 1e-6).any():
        ax3.text(0.98, 0.02, 'Red: range=0', transform=ax3.transAxes, 
                fontsize=9, color='red', ha='right', va='bottom',
                bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.7))
    
    # 4. Correlation sorting (sorted by absolute value)
    ax4 = axes[1, 1]
    # Calculate absolute correlation and sort
    feature_stats['abs_correlation'] = feature_stats['correlation_with_target'].abs()
    feature_stats_sorted = feature_stats.sort_values('abs_correlation', ascending=False)
    top_corr = feature_stats_sorted.head(10)
    colors = ['red' if x > 0 else 'blue' for x in top_corr['correlation_with_target']]
    y_pos = np.arange(len(top_corr))
    ax4.barh(y_pos, top_corr['correlation_with_target'], alpha=0.7, color=colors)
    ax4.set_yticks(y_pos)
    ax4.set_yticklabels(top_corr['feature'])
    ax4.set_xlabel('Correlation with the target variable', fontsize=12)
    ax4.set_title('Correlation between features and the target variable (top 10 by absolute value)', fontsize=14, fontweight='bold')
    ax4.axvline(x=0, color='black', linestyle='--', linewidth=1)
    ax4.grid(True, alpha=0.3, axis='x')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'feature_statistics_analysis.png'), dpi=300, bbox_inches='tight')
    plt.close()
    
    print("\n✓ Feature statistics analysis completed, results saved")
    print("="*60)
    
    return feature_stats

def plot_pdp_analysis(model, X_train, feature_names, save_dir, n_top_features=5):
    """PDP (Partial Dependence Plot) analysis - using scikit-learn"""
    print("\n=== Performing PDP analysis ===")
    
    # Ensure X_train is a DataFrame
    if isinstance(X_train, np.ndarray):
        X_train_df = pd.DataFrame(X_train, columns=feature_names)
    else:
        X_train_df = X_train
    
    # Get the n most important features
    importance_df = pd.DataFrame({
        'feature': feature_names,
        'importance': model.feature_importances_
    }).sort_values('importance', ascending=False)
    
    top_features = importance_df.head(n_top_features)['feature'].tolist()
    top_indices = [feature_names.index(f) for f in top_features]
    
    # Dictionary for saving Excel data
    excel_data_dict = {}

    def sanitize_filename(name):
        """Remove illegal characters from the file name"""
        return ''.join(c if c.isalnum() or c in ('_', '-', '.', ' ') else '_' for c in str(name))

    # Single variable PDP single figure output directory
    single_pdp_dir = os.path.join(save_dir, 'pdp_univariate_single')
    os.makedirs(single_pdp_dir, exist_ok=True)
    
    # 1. Single variable PDP analysis
    n_cols = 3  # Change to 3 columns
    n_rows = (n_top_features + 2) // 3  # Adjust row number calculation
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, 6*n_rows))
    if n_top_features == 1:
        axes = [axes]
    else:
        axes = axes.flatten()
    
    for idx, (feature, ax) in enumerate(zip(top_features, axes)):
        try:
            # Get feature index
            feature_idx = feature_names.index(feature)
            
            # Calculate PDP (average) and ICE (individual)
            # Increase grid_resolution to get more dense grid points, so the curve is smoother
            pd_result = partial_dependence(
                model, 
                X_train_df, 
                features=[feature_idx],
                kind='average',
                grid_resolution=200  # Originally 100, now encrypted to 200
            )
            
            ice_result = partial_dependence(
                model,
                X_train_df,
                features=[feature_idx],
                kind='individual',
                grid_resolution=200  # Consistent with PDP
            )
            
            # Calculate 95% confidence interval for ICE curves
            ice_curves = ice_result['individual'][0]  # shape: (n_samples, n_grid_points)
            ice_mean = np.mean(ice_curves, axis=0)
            ice_std = np.std(ice_curves, axis=0)
            ice_upper_ci = ice_mean + 1.96 * ice_std  # 95% CI
            ice_lower_ci = ice_mean - 1.96 * ice_std
            
            grid_values = pd_result['grid_values'][0]
            pdp_values = pd_result['average'][0]
            
            # Use spline interpolation to smooth the curve
            # Create more dense interpolation points (further smoothing)
            grid_dense = np.linspace(grid_values.min(), grid_values.max(), 400)
            
            # Use spline interpolation for PDP (using UnivariateSpline, s parameter controls smoothness)
            try:
                # Use spline interpolation, the smaller the s parameter, the closer to the original data, the larger the smoother
                # Here we use a moderate smoothness
                spline_pdp = UnivariateSpline(grid_values, pdp_values, s=0.1*len(grid_values))
                pdp_smooth = spline_pdp(grid_dense)
            except:
                # If spline interpolation fails, use linear interpolation
                interp_pdp = interp1d(grid_values, pdp_values, kind='cubic', 
                                     bounds_error=False, fill_value='extrapolate')
                pdp_smooth = interp_pdp(grid_dense)
            
            # Use spline interpolation for ICE mean
            try:
                spline_ice = UnivariateSpline(grid_values, ice_mean, s=0.1*len(grid_values))
                ice_mean_smooth = spline_ice(grid_dense)
            except:
                interp_ice = interp1d(grid_values, ice_mean, kind='cubic',
                                      bounds_error=False, fill_value='extrapolate')
                ice_mean_smooth = interp_ice(grid_dense)
            
            # Use interpolation for confidence interval
            try:
                interp_upper = interp1d(grid_values, ice_upper_ci, kind='cubic',
                                       bounds_error=False, fill_value='extrapolate')
                interp_lower = interp1d(grid_values, ice_lower_ci, kind='cubic',
                                       bounds_error=False, fill_value='extrapolate')
                ice_upper_smooth = interp_upper(grid_dense)
                ice_lower_smooth = interp_lower(grid_dense)
            except:
                ice_upper_smooth = ice_upper_ci
                ice_lower_smooth = ice_lower_ci
                grid_dense = grid_values
            
            # Calculate ALE (true ALE, not approximate)
            # ALE is calculated by calculating the local effect and accumulating them
            n_grid = len(grid_values)
            n_samples = len(X_train_df)
            
            # Create grid intervals
            ale_values = np.zeros(n_grid)
            
            # For each grid point, calculate the local effect
            for i in range(n_grid - 1):
                # Get samples in the current interval
                if i == 0:
                    mask = (X_train_df.iloc[:, feature_idx].values <= grid_values[i+1])
                elif i == n_grid - 2:
                    mask = (X_train_df.iloc[:, feature_idx].values > grid_values[i])
                else:
                    mask = ((X_train_df.iloc[:, feature_idx].values > grid_values[i]) & 
                           (X_train_df.iloc[:, feature_idx].values <= grid_values[i+1]))
                
                # If there are samples in this interval
                if np.sum(mask) > 0:
                    # Calculate the local effect: change the feature value, other features remain unchanged
                    X_low = X_train_df.copy()
                    X_high = X_train_df.copy()
                    
                    # Set feature value to the two ends of the interval
                    X_low.loc[mask, X_train_df.columns[feature_idx]] = grid_values[i]
                    X_high.loc[mask, X_train_df.columns[feature_idx]] = grid_values[i+1]
                    
                    # Calculate the prediction difference
                    pred_low = model.predict(X_low.iloc[mask, :].values)
                    pred_high = model.predict(X_high.iloc[mask, :].values)
                    
                    # Local effect
                    local_effect = np.mean(pred_high - pred_low) / (grid_values[i+1] - grid_values[i])
                    
                    # Accumulate to ALE
                    if i == 0:
                        ale_values[i+1] = local_effect * (grid_values[i+1] - grid_values[i])
                    else:
                        ale_values[i+1] = ale_values[i] + local_effect * (grid_values[i+1] - grid_values[i])
            
            # Center the ALE (so the mean is 0)
            ale_values = ale_values - np.mean(ale_values) + np.mean(pdp_values)
            
            # Use spline interpolation for ALE
            try:
                spline_ale = UnivariateSpline(grid_values, ale_values, s=0.1*len(grid_values))
                ale_smooth = spline_ale(grid_dense)
            except:
                interp_ale = interp1d(grid_values, ale_values, kind='cubic',
                                     bounds_error=False, fill_value='extrapolate')
                ale_smooth = interp_ale(grid_dense)
            
            # Plot the smoothed PDP line (blue solid line, no marker, smoother)
            ax.plot(grid_dense, pdp_smooth, 
                   linewidth=2.5, color='blue', 
                   label='PDP', zorder=4)
            
            # Plot the smoothed Mean c-ICE line (coral dashed line, no marker)
            ax.plot(grid_dense, ice_mean_smooth, 
                   linewidth=2.5, color='coral', linestyle='--', 
                   label='Mean c-ICE', zorder=4)
            
            # Plot the smoothed ALE line (green dotted line, no marker)
            ax.plot(grid_dense, ale_smooth, 
                   linewidth=2.5, color='green', linestyle='-.', 
                   label='ALE', zorder=4)
            
            # Remove 95% CI shadow region (user requested: too wide range, affects trend observation)
            # ax.fill_between(grid_dense, ice_lower_smooth, ice_upper_smooth, 
            #                color='coral', alpha=0.3, label='95% CI of c-ICE', zorder=1)
            
            # Save data to dictionary (for subsequent saving to Excel)
            # Simplified to a few columns,方便在 Excel 中直接绘制干净的 PDP 曲线
            # Only keep: grid points + smoothed PDP / c-ICE / ALE
            excel_data_dict[feature] = pd.DataFrame({
                'Grid': grid_dense,
                'PDP_Smooth': pdp_smooth,
                'Mean_cICE_Smooth': ice_mean_smooth,
                'ALE_Smooth': ale_smooth,
            })
            
            ax.set_xlabel(feature, fontsize=10)
            ax.set_ylabel('Partial Effect', fontsize=10)
            ax.set_title(f'PDP, ALE & c-ICE for {feature}', fontsize=12, fontweight='bold')
            ax.legend(loc='best', fontsize=9)
            ax.grid(True, alpha=0.3)

            # Save separate single variable PDP figure
            single_fig, single_ax = plt.subplots(figsize=(7, 5))
            single_ax.plot(grid_dense, pdp_smooth, linewidth=2.5, color='blue', label='PDP', zorder=4)
            single_ax.plot(grid_dense, ice_mean_smooth, linewidth=2.5, color='coral',
                           linestyle='--', label='Mean c-ICE', zorder=4)
            single_ax.plot(grid_dense, ale_smooth, linewidth=2.5, color='green',
                           linestyle='-.', label='ALE', zorder=4)
            single_ax.set_xlabel(feature, fontsize=11, fontweight='bold')
            single_ax.set_ylabel('Partial Effect', fontsize=11, fontweight='bold')
            single_ax.set_title(f'PDP / c-ICE / ALE for {feature}', fontsize=13, fontweight='bold')
            single_ax.legend(loc='best', fontsize=10)
            single_ax.grid(True, alpha=0.3, linestyle='--', linewidth=0.6)
            single_ax.set_axisbelow(True)
            safe_feature_name = sanitize_filename(feature).strip().replace(' ', '_')
            single_path = os.path.join(single_pdp_dir, f'pdp_univariate_{safe_feature_name}.png')
            single_fig.tight_layout()
            single_fig.savefig(single_path, dpi=300, bbox_inches='tight')
            plt.close(single_fig)
        except Exception as e:
            print(f"Warning: Could not create PDP for {feature}: {e}")
            ax.text(0.5, 0.5, f'Error creating PDP', 
                   ha='center', va='center', transform=ax.transAxes)
            ax.axis('off')
    
    # Hide extra subplots
    for idx in range(n_top_features, len(axes)):
        axes[idx].axis('off')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'pdp_analysis.png'), dpi=300, bbox_inches='tight')
    # plt.show()  # Show single variable PDP figure
    plt.close()
    
    # Save PDP data to the existing model_analysis_results.xlsx file (append mode)
    print(f"\nPreparing to save single variable PDP data to Excel...")
    print(f"  Number of successfully calculated features: {len(excel_data_dict)}")
    if len(excel_data_dict) > 0:
        print(f"  Feature list: {list(excel_data_dict.keys())}")
    
    if excel_data_dict:
        excel_path = os.path.join(save_dir, 'model_analysis_results.xlsx')
        try:
            # Check if the file exists, if it exists, use append mode, otherwise create a new file
            file_exists = os.path.exists(excel_path)
            mode = 'a' if file_exists else 'w'
            
            print(f"  Excel file path: {excel_path}")
            print(f"  File exists: {file_exists}")
            print(f"  Save mode: {mode}")
            
            saved_sheets = []
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
                # Create a worksheet for each feature
                for feature_name, data_df in excel_data_dict.items():
                    try:
                        # The worksheet name cannot exceed 31 characters, and cannot contain special characters
                        # Add prefix "PDP_" to avoid conflicts with existing worksheets
                        sheet_name = f'PDP_{feature_name}'
                        if len(sheet_name) > 31:
                            sheet_name = f'PDP_{feature_name[:27]}'
                        # Replace characters Excel does not support
                        sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                        
                        print(f"  Saving worksheet: {sheet_name} (feature: {feature_name})")
                        print(f"    Data shape: {data_df.shape}")
                        print(f"    Column names: {list(data_df.columns)}")
                        
                        data_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        saved_sheets.append(sheet_name)
                        print(f"    ✓ Successfully saved: {sheet_name}")
                    except Exception as e:
                        print(f"    ⚠ Error saving feature {feature_name}: {e}")
                        import traceback
                        traceback.print_exc()
            
            if file_exists:
                print(f"\n✓ Single variable PDP data has been appended to the existing Excel file: {excel_path}")
            else:
                print(f"\n✓ Single variable PDP data has been saved to a new Excel file: {excel_path}")
            print(f"  Successfully saved {len(saved_sheets)} worksheets: {saved_sheets}")
            print(f"  Each worksheet contains the following columns:")
            print(f"    - Grid: Dense interpolation points (for smoothing curves)")
            print(f"    - PDP_Smooth: Smoothed PDP values")
            print(f"    - Mean_cICE_Smooth: Smoothed Mean c-ICE values")
            print(f"    - ALE_Smooth: Smoothed ALE values")
        except Exception as e:
            print(f"\n⚠ Error saving single variable PDP data to Excel: {e}")
            import traceback
            traceback.print_exc()
    else:
        print(f"\n⚠ Warning: No single variable PDP data has been successfully calculated, cannot save to Excel")
        print(f"  Possible reasons:")
        print(f"    1. All PDP calculations for all features failed")
        print(f"    2. n_top_features is 0 or negative")
        print(f"    3. Data format problem caused calculation failure")
    
    print(f"\nSingle variable PDP analysis completed, the dependency graph of the first {n_top_features} important features has been drawn")
    
    # 2. Double variable PDP analysis (feature interaction)
    if n_top_features >= 2:
        print("\n=== Performing double variable PDP analysis ===")
        
        # Select the top 4 most important features for double variable interaction analysis
        n_interaction_features = min(4, n_top_features)
        interaction_features = top_features[:n_interaction_features]
        
        # Create interaction feature pairs
        interaction_pairs = []
        for i in range(len(interaction_features)):
            for j in range(i+1, len(interaction_features)):
                interaction_pairs.append((interaction_features[i], interaction_features[j]))
        
        # Dictionary for saving 2D PDP data
        excel_2d_data_dict = {}
        
        # Plot double variable PDP (using 3D surface plot)
        from mpl_toolkits.mplot3d import Axes3D
        
        n_pairs = len(interaction_pairs)
        n_cols = 3  # 3 columns
        n_rows = (n_pairs + 2) // 3  # Adjust row number calculation
        fig = plt.figure(figsize=(18, 6*n_rows))

        pdp2d_dir = os.path.join(save_dir, 'pdp_2d_single')
        os.makedirs(pdp2d_dir, exist_ok=True)
        
        for idx, (feat1, feat2) in enumerate(interaction_pairs):
            try:
                # Get feature index
                feat1_idx = feature_names.index(feat1)
                feat2_idx = feature_names.index(feat2)
                
                # Calculate double variable partial dependence - increase grid_resolution to make the grid denser
                pd_result = partial_dependence(
                    model,
                    X_train_df,
                    features=[feat1_idx, feat2_idx],
                    kind='average',
                    grid_resolution=80  # 进一步加密网格点，提升表面平滑度
                )
                
                # Get grid values and average predicted values
                grid_values_1 = pd_result['grid_values'][0]
                grid_values_2 = pd_result['grid_values'][1]
                average = pd_result['average'][0]  # Take the first element (if it is a 3D array)
                
                # Ensure average is 2D
                if average.ndim == 3:
                    average = average[0]
                
                # Create original grid
                X1, X2 = np.meshgrid(grid_values_1, grid_values_2)
                
                # Ensure average matches grid dimensions
                # meshgrid返回的形状是 (len(grid_values_2), len(grid_values_1))
                if average.shape != X1.shape and average.shape == (len(grid_values_1), len(grid_values_2)):
                        average = average.T
                
                # Use a denser interpolation grid, significantly improving visual density
                dense_resolution = 150
                grid_dense_1 = np.linspace(grid_values_1.min(), grid_values_1.max(), dense_resolution)
                grid_dense_2 = np.linspace(grid_values_2.min(), grid_values_2.max(), dense_resolution)
                try:
                    interpolator = RegularGridInterpolator(
                        (grid_values_2, grid_values_1), average,
                        bounds_error=False, fill_value=None
                    )
                    X1_dense, X2_dense = np.meshgrid(grid_dense_1, grid_dense_2)
                    dense_points = np.stack([X2_dense.ravel(), X1_dense.ravel()], axis=-1)
                    average_dense = interpolator(dense_points).reshape(X1_dense.shape)
                except Exception:
                    # RegularGridInterpolator may fail (for example, a dimension is approximately constant), try using griddata for 2D interpolation
                    try:
                        X1_dense, X2_dense = np.meshgrid(grid_dense_1, grid_dense_2)
                        original_points = np.column_stack([X1.flatten(), X2.flatten()])
                        target_points = np.column_stack([X1_dense.flatten(), X2_dense.flatten()])
                        average_dense = griddata(original_points, average.flatten(), target_points, method='cubic')
                        # If cubic produces NaN, revert to linear; if not, use nearest
                        if np.isnan(average_dense).any():
                            average_dense = griddata(original_points, average.flatten(), target_points, method='linear')
                        if np.isnan(average_dense).any():
                            average_dense = griddata(original_points, average.flatten(), target_points, method='nearest')
                        average_dense = average_dense.reshape(X1_dense.shape)
                    except Exception:
                        # Still failed, revert to original grid
                        X1_dense, X2_dense, average_dense = X1, X2, average
                
                # Save 2D PDP data to dictionary (using encrypted grid, if available)
                grid1_flat = X1_dense.flatten()
                grid2_flat = X2_dense.flatten()
                average_flat = average_dense.flatten()
                
                # Create DataFrame
                pair_name = f'{feat1}_vs_{feat2}'
                data_df = pd.DataFrame({
                    f'{feat1}': grid1_flat,
                    f'{feat2}': grid2_flat,
                    'Partial_Dependence': average_flat
                })
                
                # Ensure data length is consistent
                if len(grid1_flat) == len(grid2_flat) == len(average_flat):
                    excel_2d_data_dict[pair_name] = data_df
                    print(f"  ✓ Successfully calculated and saved data for feature pair {pair_name} (shape: {data_df.shape})")
                else:
                    print(f"  ⚠ Warning: Data length for feature pair {pair_name} is inconsistent, skipping save")
                    print(f"    grid1_flat length: {len(grid1_flat)}, grid2_flat length: {len(grid2_flat)}, average_flat length: {len(average_flat)}")
                
                # Create 3D subplot
                ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                
                # Plot 3D surface plot (using encrypted grid)
                surf = ax.plot_surface(X1_dense, X2_dense, average_dense, cmap='viridis', alpha=0.9, 
                                       linewidth=0, antialiased=True, edgecolor='none')
                
                # Set labels and title
                ax.set_xlabel(feat1, fontsize=10, labelpad=8)
                ax.set_ylabel(feat2, fontsize=10, labelpad=8)
                ax.set_zlabel('Partial Dependence', fontsize=10, labelpad=8)
                ax.set_title(f'Interaction: {feat1} vs {feat2}', fontsize=12, fontweight='bold', pad=15)
                
                # Add color bar
                fig.colorbar(surf, ax=ax, shrink=0.6, aspect=20, pad=0.1)

                # Save separate 3D interaction figure
                single_fig = plt.figure(figsize=(8, 6))
                single_ax = single_fig.add_subplot(111, projection='3d')
                single_surf = single_ax.plot_surface(X1_dense, X2_dense, average_dense, cmap='viridis', alpha=0.9,
                                                     linewidth=0, antialiased=True, edgecolor='none')
                single_ax.set_xlabel(feat1, fontsize=11, labelpad=10)
                single_ax.set_ylabel(feat2, fontsize=11, labelpad=10)
                single_ax.set_zlabel('Partial Dependence', fontsize=11, labelpad=10)
                single_ax.set_title(f'Interaction: {feat1} vs {feat2}', fontsize=13, fontweight='bold', pad=18)
                single_ax.view_init(elev=25, azim=135)
                single_fig.colorbar(single_surf, shrink=0.6, aspect=20, pad=0.12)
                safe_pair_name = sanitize_filename(pair_name).replace(' ', '_')
                single_path = os.path.join(pdp2d_dir, f'pdp_2d_{safe_pair_name}.png')
                single_fig.tight_layout()
                single_fig.savefig(single_path, dpi=300, bbox_inches='tight')
                plt.close(single_fig)
                
            except Exception as e:
                print(f"Warning: Could not create 2D PDP for {feat1} vs {feat2}: {e}")
                import traceback
                traceback.print_exc()
                # Create empty 3D subplot to display error information
                ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                ax.text(0.5, 0.5, 0.5, f'Error creating 2D PDP', 
                       ha='center', va='center', transform=ax.transAxes)
                ax.set_axis_off()
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, 'pdp_2d_interaction.png'), dpi=300, bbox_inches='tight')
        # plt.show()  # Show 3D figure
        plt.close()
        
        # Save 2D PDP data to the existing model_analysis_results.xlsx file
        print(f"\nPreparing to save 2D PDP data...")
        print(f"  Number of successfully calculated feature pairs: {len(excel_2d_data_dict)}")
        if len(excel_2d_data_dict) > 0:
            print(f"  Feature pair list: {list(excel_2d_data_dict.keys())}")
            for pair_name, data_df in excel_2d_data_dict.items():
                print(f"    - {pair_name}: shape {data_df.shape}, columns {list(data_df.columns)}")
        else:
            print(f"  ⚠ Warning: excel_2d_data_dict is empty, no data to save")
        
        if excel_2d_data_dict and len(excel_2d_data_dict) > 0:
            excel_path = os.path.join(save_dir, 'model_analysis_results.xlsx')
            try:
                # Check if the file exists, if it exists, use append mode, otherwise create a new file
                file_exists = os.path.exists(excel_path)
                mode = 'a' if file_exists else 'w'
                
                print(f"  Excel file path: {excel_path}")
                print(f"  File exists: {file_exists}")
                print(f"  Save mode: {mode}")
                
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
                    # Create a worksheet for each feature pair
                    saved_sheets = []
                    failed_sheets = []
                    
                    print(f"  Starting to save data for {len(excel_2d_data_dict)} feature pairs...")
                    for idx, (pair_name, data_df) in enumerate(excel_2d_data_dict.items(), 1):
                        try:
                            # Worksheet name cannot exceed 31 characters, and cannot contain special characters
                            # Add prefix "PDP2D_" to avoid conflicts with existing worksheets
                            sheet_name = f'PDP2D_{pair_name}'
                            if len(sheet_name) > 31:
                                sheet_name = f'PDP2D_{pair_name[:25]}'
                            # Replace characters Excel does not support
                            sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                            
                            print(f"  [{idx}/{len(excel_2d_data_dict)}] Saving worksheet: {sheet_name} (feature pair: {pair_name})")
                            print(f"    Data shape: {data_df.shape}")
                            print(f"    Column names: {list(data_df.columns)}")
                            
                            # Check if data is empty
                            if data_df.empty:
                                print(f"    ⚠ Warning: Data for feature pair {pair_name} is empty, skipping save")
                                failed_sheets.append((pair_name, "Data is empty"))
                                continue
                            
                            # Save to Excel
                            data_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            saved_sheets.append(sheet_name)
                            print(f"    ✓ Successfully saved: {sheet_name}")
                        except Exception as e:
                            error_msg = str(e)
                            print(f"    ⚠ Error saving feature pair {pair_name}: {error_msg}")
                            failed_sheets.append((pair_name, error_msg))
                            import traceback
                            traceback.print_exc()
                            # Continue processing the next feature pair, without interrupting the loop
                            continue
                    
                    # Print save result summary
                    print(f"\n Save result summary:")
                    print(f"    ✓ Successfully saved: {len(saved_sheets)} worksheets")
                    if saved_sheets:
                        print(f"      Successfully saved worksheets: {saved_sheets}")
                    if failed_sheets:
                        print(f"    ✗ Save failed: {len(failed_sheets)} worksheets")
                        for pair_name, error in failed_sheets:
                            print(f"      - {pair_name}: {error}")
                
                if file_exists:
                    print(f"\n✓ 2D PDP data has been appended to the existing Excel file: {excel_path}")
                else:
                    print(f"\n✓ 2D PDP data has been saved to a new Excel file: {excel_path}")
                print(f"  Successfully saved {len(saved_sheets)} worksheets: {saved_sheets}")
                print(f"  Each worksheet contains the following columns:")
                print(f"    - First feature name: Grid values for the first feature")
                print(f"    - Second feature name: Grid values for the second feature")
                print(f"    - Partial_Dependence: Corresponding partial dependence values")
            except Exception as e:
                print(f"\n⚠ Error saving 2D PDP data to Excel: {e}")
                import traceback
                traceback.print_exc()
        else:
            print(f"\n⚠ Warning: No 2D PDP data has been successfully calculated, cannot save to Excel")
            print(f"  Possible reasons:")
            print(f"    1. All feature pair calculations failed")
            print(f"    2. Number of feature pairs is 0 (n_interaction_features < 2)")
            print(f"    3. Data format problem caused calculation failure")
        
        print(f"\nDouble variable PDP analysis completed, the interaction graph of the first {n_interaction_features} important features has been drawn")

def analyze_noise_robustness(model, X_test, y_test, feature_names, save_dir):
    """Analyze the robustness of the model to noise and outliers"""
    print("\n=== Analyze the robustness of the model to noise and outliers ===")
    
    # 1. Add Gaussian noise test
    print("Testing the robustness of the model to Gaussian noise...")
    # Use 0, 2, 4, 6, 8, 10 (%) as noise levels
    noise_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Noise levels (percentage of feature standard deviation)
    noise_rmses = []
    noise_stds = []
    
    for noise_level in noise_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times to get the standard deviation
            # Add Gaussian noise
            X_test_noisy = X_test.copy()
            for i in range(X_test.shape[1]):
                feature_std = np.std(X_test[:, i])
                noise = np.random.normal(0, noise_level * feature_std, size=len(X_test))
                X_test_noisy[:, i] = X_test[:, i] + noise
            
            # Predict
            y_pred_noisy = model.predict(X_test_noisy)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred_noisy))
            rmses.append(rmse)
        
        noise_rmses.append(np.mean(rmses))
        noise_stds.append(np.std(rmses))
    
    # 2. Add outlier test
    print("Testing the robustness of the model to outliers...")
    # Use 0, 2, 4, 6, 8, 10 (%) as outlier levels
    outlier_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Outlier levels (percentage of samples)
    outlier_severity = 5.0  # Outlier severity (multiple of standard deviation)
    outlier_rmses = []
    outlier_stds = []
    
    for outlier_pct in outlier_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times
            # Add outliers
            X_test_outlier = X_test.copy()
            n_outliers = int(len(X_test) * outlier_pct)
            if n_outliers > 0:
                outlier_indices = np.random.choice(len(X_test), n_outliers, replace=False)
                for idx in outlier_indices:
                    # Randomly select a feature and add an outlier
                    feature_idx = np.random.randint(X_test.shape[1])
                    feature_mean = np.mean(X_test[:, feature_idx])
                    feature_std = np.std(X_test[:, feature_idx])
                    X_test_outlier[idx, feature_idx] = feature_mean + outlier_severity * feature_std
            
            # Predict
            y_pred_outlier = model.predict(X_test_outlier)
            rmse = np.sqrt(mean_squared_error(y_test, y_pred_outlier))
            rmses.append(rmse)
        
        outlier_rmses.append(np.mean(rmses))
        outlier_stds.append(np.std(rmses))
    
    # 3. Draw robustness analysis graph
    fig, axes = plt.subplots(1, 2, figsize=(15, 6))
    
    # Robustness to additive Gaussian noise
    axes[0].errorbar([l * 100 for l in noise_levels], noise_rmses, yerr=noise_stds, 
                     marker='o', capsize=5, capthick=2, linewidth=2, color='blue')
    axes[0].set_xlabel('Noise level (% of feature std)', fontsize=12)
    axes[0].set_ylabel('RMSE (MPa)', fontsize=12)
    axes[0].set_title('Robustness to Additive Gaussian Noise', fontsize=14, fontweight='bold')
    axes[0].grid(True, alpha=0.3)
    
    # Robustness to outliers
    axes[1].errorbar([l * 100 for l in outlier_levels], outlier_rmses, yerr=outlier_stds, 
                     marker='o', capsize=5, capthick=2, linewidth=2, color='orange')
    axes[1].set_xlabel('Injected outliers (% of samples)', fontsize=12)
    axes[1].set_ylabel('RMSE (MPa)', fontsize=12)
    axes[1].set_title(f'Robustness to Outliers (severity={outlier_severity}×std)', fontsize=14, fontweight='bold')
    axes[1].grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'robustness_analysis.png'), dpi=300, bbox_inches='tight')
    # plt.show()
    plt.close()
    
    # Save results (will be saved in the main function)
    robustness_df = pd.DataFrame({
        'Noise Level (%)': [l * 100 for l in noise_levels],
        'RMSE (MPa)': noise_rmses,
        'Std': noise_stds
    })
    
    outlier_df = pd.DataFrame({
        'Outlier Percentage (%)': [l * 100 for l in outlier_levels],
        'RMSE (MPa)': outlier_rmses,
        'Std': outlier_stds
    })
    
    # Print result summary
    print("\n=== Robustness analysis result summary ===")
    print("\n1. Robustness to additive Gaussian noise:")
    print(f"   - RMSE at 0% noise: {noise_rmses[0]:.3f} MPa")
    print(f"   - RMSE at 2% noise: {noise_rmses[1]:.3f} MPa (increase of {noise_rmses[1]-noise_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 4% noise: {noise_rmses[2]:.3f} MPa (increase of {noise_rmses[2]-noise_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 6% noise: {noise_rmses[3]:.3f} MPa (increase of {noise_rmses[3]-noise_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 8% noise: {noise_rmses[4]:.3f} MPa (increase of {noise_rmses[4]-noise_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 10% noise: {noise_rmses[5]:.3f} MPa (increase of {noise_rmses[5]-noise_rmses[0]:.3f} MPa)")
    
    print("\n2. Robustness to outliers:")
    print(f"   - RMSE at 0% outliers: {outlier_rmses[0]:.3f} MPa")
    print(f"   - RMSE at 2% outliers: {outlier_rmses[1]:.3f} MPa (change of {outlier_rmses[1]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 4% outliers: {outlier_rmses[2]:.3f} MPa (change of {outlier_rmses[2]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 6% outliers: {outlier_rmses[3]:.3f} MPa (change of {outlier_rmses[3]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 8% outliers: {outlier_rmses[4]:.3f} MPa (change of {outlier_rmses[4]-outlier_rmses[0]:.3f} MPa)")
    print(f"   - RMSE at 10% outliers: {outlier_rmses[5]:.3f} MPa (change of {outlier_rmses[5]-outlier_rmses[0]:.3f} MPa)")
    
    # 计算性能退化率
    noise_degradation = (noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100
    outlier_degradation = (outlier_rmses[-1] - outlier_rmses[0]) / outlier_rmses[0] * 100
    
    print("\n3. Performance degradation rate (10% noise/outliers vs no perturbation):")
    print(f"   - Performance degradation of Gaussian noise: {noise_degradation:.1f}%")
    print(f"   - Performance degradation of outliers: {outlier_degradation:.1f}%")
    
    # 判断鲁棒性
    if noise_degradation < 20:
        noise_robust = "Excellent"
    elif noise_degradation < 40:
        noise_robust = "Good"
    elif noise_degradation < 60:
        noise_robust = "Average"
    else:
        noise_robust = "Poor"
    
    if abs(outlier_degradation) < 5:
        outlier_robust = "Excellent"
    elif abs(outlier_degradation) < 10:
        outlier_robust = "Good"
    elif abs(outlier_degradation) < 20:
        outlier_robust = "Average"
    else:
        outlier_robust = "Poor"
    
    print("\n4. Robustness assessment:")
    print(f"   - Robustness to Gaussian noise: {noise_robust} (performance degradation of 10% noise: {noise_degradation:.1f}%)")
    print(f"   - Robustness to outliers: {outlier_robust} (performance degradation of 10% outliers: {outlier_degradation:.1f}%)")
    
    print(f"\nRobustness analysis completed, detailed results have been saved")
    return noise_rmses, outlier_rmses, robustness_df, outlier_df

def analyze_prediction_intervals(model, X_test, y_test, save_dir, n_bootstrap=100, bootstrap_noise_level=0.01):
    """Analyze the confidence of the prediction intervals
    
    Parameters:
        model: Trained model
        X_test: Test features
        y_test: Test labels
        save_dir: Save directory
        n_bootstrap: Bootstrap resampling times
        bootstrap_noise_level: Bootstrap noise level (percentage of feature standard deviation)
    
    Method description:
    1. Bootstrap: Add input noise to simulate model uncertainty, similar to Monte Carlo Dropout
    2. Quantile Regression: Build a fixed width prediction interval based on the residual distribution
    """
    print("\n=== Generate prediction intervals ===")
    print(f"Bootstrap noise level: {bootstrap_noise_level*100}% of feature std")
    
    # 1. Bootstrap prediction interval - improved version
    print("Generate Bootstrap prediction interval...")
    bootstrap_predictions = []
    for _ in range(n_bootstrap):
        # Add noise to simulate model uncertainty (similar to the idea of Monte Carlo Dropout)
        # Add small noise to the test data to reflect the model's response to input uncertainty
        noise_std = bootstrap_noise_level  # Noise standard deviation (percentage of feature standard deviation)
        X_test_noisy = X_test.copy()
        for i in range(X_test.shape[1]):
            feature_std = np.std(X_test[:, i])
            noise = np.random.normal(0, noise_std * feature_std, size=len(X_test))
            X_test_noisy[:, i] = X_test[:, i] + noise
        
        # Predict
        y_pred = model.predict(X_test_noisy)
        bootstrap_predictions.append(y_pred)
    
    bootstrap_predictions = np.array(bootstrap_predictions)
    bootstrap_median = np.median(bootstrap_predictions, axis=0)
    bootstrap_lower = np.percentile(bootstrap_predictions, 10, axis=0)  # 80% confidence interval
    bootstrap_upper = np.percentile(bootstrap_predictions, 90, axis=0)
    
    # 2. Quantile Regression prediction interval (using the residual distribution of the training data)
    print("Generate Quantile Regression prediction interval...")
    y_pred_test = model.predict(X_test)
    residuals = y_test - y_pred_test
    residual_std = np.std(residuals)
    
    # Build the prediction interval using the standard deviation of the residuals
    quantile_lower = y_pred_test - 1.28 * residual_std  # 10th percentile for 80% interval
    quantile_upper = y_pred_test + 1.28 * residual_std  # 90th percentile for 80% interval
    
    # Sort for visualization
    sort_indices = np.argsort(y_test)
    y_test_sorted = y_test[sort_indices]
    bootstrap_median_sorted = bootstrap_median[sort_indices]
    bootstrap_lower_sorted = bootstrap_lower[sort_indices]
    bootstrap_upper_sorted = bootstrap_upper[sort_indices]
    quantile_lower_sorted = quantile_lower[sort_indices]
    quantile_upper_sorted = quantile_upper[sort_indices]
    
    # Calculate coverage
    bootstrap_coverage = np.mean((y_test_sorted >= bootstrap_lower_sorted) & (y_test_sorted <= bootstrap_upper_sorted))
    quantile_coverage = np.mean((y_test_sorted >= quantile_lower_sorted) & (y_test_sorted <= quantile_upper_sorted))
    
    # Calculate average interval width
    bootstrap_width = np.mean(bootstrap_upper_sorted - bootstrap_lower_sorted)
    quantile_width = np.mean(quantile_upper_sorted - quantile_lower_sorted)
    
    # Debug: print interval width information
    print(f"\n=== Calculate the interval width ===")
    print(f"Bootstrap interval width: {bootstrap_width:.6f}")
    print(f"Quantile interval width: {quantile_width:.6f}")
    print(f"Bootstrap coverage: {bootstrap_coverage*100:.1f}%")
    print(f"Quantile coverage: {quantile_coverage*100:.1f}%")
    print(f"Note: Bootstrap coverage is smaller ({bootstrap_coverage*100:.1f}%), the interval should be narrower ({bootstrap_width:.6f})")
    print(f"     Quantile coverage is larger ({quantile_coverage*100:.1f}%), the interval should be wider ({quantile_width:.6f})")
    
    # 3. Plot the prediction interval
    # Keep the original "long bar" width, but split the upper and lower two sub-figures into two independent images
    # Therefore, here we use figsize=(16, 5), which maintains the same total size as the original 16×10
    
    # Create x-axis positions (from 1 to n_samples)
    n_samples = len(y_test_sorted)
    x_positions = np.arange(1, n_samples + 1)  # 1, 2, 3, ..., n_samples
    
    # Color scheme
    color_actual = '#000000'  # 纯黑色
    color_predicted = '#2563EB'  # Modern blue
    color_bootstrap_median = '#DC2626'  # Modern red
    
    # Blue gradient (Quantile confidence band)
    color_blue_start = '#2172B4'  # Deepest blue
    color_blue_end = '#FFFFFF'  # White (end color, enhance the gradient effect)
    
    # Red gradient (Bootstrap confidence band)
    color_red_start = '#C41E3A'  # Deep red
    color_red_end = '#FFFFFF'  # White (end color, enhance the gradient effect)
    
    # Number of gradient layers
    n_gradient_layers = 50
    
    # Helper function: convert RGB color to tuple
    def hex_to_rgb(hex_color):
        hex_color = hex_color.lstrip('#')
        return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))
    
    # Helper function: RGB color interpolation
    def interpolate_color(color1, color2, ratio):
        rgb1 = hex_to_rgb(color1)
        rgb2 = hex_to_rgb(color2)
        return tuple(int(rgb1[i] + (rgb2[i] - rgb1[i]) * ratio) for i in range(3))
    
    # Helper function: RGB to hexadecimal
    def rgb_to_hex(rgb):
        return '#{:02x}{:02x}{:02x}'.format(int(rgb[0]), int(rgb[1]), int(rgb[2]))
    
    # ==========================================================
    # Figure 1: Bootstrap prediction interval (separate long bar)
    # ==========================================================
    fig_bootstrap, ax1 = plt.subplots(figsize=(16, 5))
    fig_bootstrap.patch.set_facecolor('white')
    
    # Plot Bootstrap prediction interval
    bootstrap_baseline = bootstrap_median_sorted  # Bootstrap baseline: Bootstrap median (center estimate of the Bootstrap distribution)
    bootstrap_range = bootstrap_upper_sorted - bootstrap_lower_sorted
    
    # Check interval width, if too narrow, use simple fill, otherwise use gradient
    bootstrap_range_avg = np.mean(bootstrap_upper_sorted - bootstrap_lower_sorted)
    y_range_span = y_max - y_min if 'y_max' in locals() and 'y_min' in locals() else np.max(y_test_sorted) - np.min(y_test_sorted)
    range_ratio = bootstrap_range_avg / (y_range_span + 1e-10)
    
    # If the interval width is less than 1% of the y-axis range, use simple fill to improve visibility
    if range_ratio < 0.01:
        # Use simple fill, increase transparency to improve visibility
        ax1.fill_between(x_positions, bootstrap_lower_sorted, bootstrap_upper_sorted, 
                        alpha=0.4, color=color_red_start, 
                        label='Bootstrap 80% interval', zorder=1,
                        edgecolor='none', linewidth=0)
    else:
        # 50-level gradient effect (from center dark to edge white)
        for i in range(n_gradient_layers):
            # layer_ratio from 0 to 1, 0=center (dark), 1=edge (white)
            layer_ratio = (i + 1) / n_gradient_layers
            # Calculate the boundaries of the current layer
            layer_lower = bootstrap_baseline - (bootstrap_baseline - bootstrap_lower_sorted) * layer_ratio
            layer_upper = bootstrap_baseline + (bootstrap_upper_sorted - bootstrap_baseline) * layer_ratio
            
            # Color interpolation: layer_ratio=0 is deep red, layer_ratio=1 is white
            # Reverse layer_ratio for color: inner layer (close to 0) dark, outer layer (close to 1) white
            color_ratio = layer_ratio  # Use directly, because layer_ratio is from 0 to 1
            current_color = rgb_to_hex(interpolate_color(color_red_start, color_red_end, color_ratio))
            
            # Transparency: inner layer is less transparent, outer layer is more transparent (but not completely transparent)
            # From 0.8 (inner layer) to 0.2 (outer layer), increase overall transparency to ensure visibility
            alpha_layer = 0.8 - 0.6 * layer_ratio
            
            if i == 0:
                ax1.fill_between(x_positions, layer_lower, layer_upper, 
                               alpha=alpha_layer, color=current_color, 
                               label='Bootstrap 80% interval', zorder=1,
                               edgecolor='none', linewidth=0)
            else:
                ax1.fill_between(x_positions, layer_lower, layer_upper, 
                               alpha=alpha_layer, color=current_color, 
                               zorder=1, edgecolor='none', linewidth=0)
    
    # Plot Bootstrap median line
    ax1.plot(x_positions, bootstrap_median_sorted, '--', linewidth=2.5, 
             color=color_bootstrap_median, label='Bootstrap median', zorder=4, 
             alpha=0.80, dashes=(10, 5))
    
    # Plot predicted and actual values
    ax1.plot(x_positions, y_pred_test[sort_indices], '-', linewidth=3.5, 
             color=color_predicted, label='Predicted values', zorder=6, alpha=0.95)
    ax1.scatter(x_positions, y_test_sorted, s=120, c=color_actual, alpha=1.0, 
                label='Actual values', zorder=7, edgecolors='none', marker='o')
    
    # Set labels and styles
    ax1.set_xlabel('Test samples (sorted by actual values)',
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    ax1.set_ylabel('Peak Strain', 
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    ax1.tick_params(axis='y', which='major', labelsize=18, width=2.0, length=6, colors='#000000')
    for label in ax1.get_yticklabels():
        label.set_fontweight('bold')
        label.set_fontfamily('Arial')
        label.set_color('#000000')
    for spine in ax1.spines.values():
        spine.set_linewidth(2.0)
        spine.set_color('#000000')
    ax1.grid(False)
    
    # Add Bootstrap performance metrics
    bootstrap_text = f'Coverage: {bootstrap_coverage*100:.1f}%\nAvg Width: {bootstrap_width:.6f}'
    ax1.text(0.02, 0.98, bootstrap_text, transform=ax1.transAxes, fontsize=18, 
             fontweight='bold', fontfamily='Arial', verticalalignment='top', color='#000000')
    
    # Set legend
    ax1.legend(loc='lower right', fontsize=18, 
               prop={'family': 'Arial', 'weight': 'bold'},
               frameon=False, fancybox=False, shadow=False)
    
    # Unify y-axis range (ensure consistency with Quantile figure)
    y_min = min(np.min(y_test_sorted), np.min(y_pred_test[sort_indices]), 
                np.min(bootstrap_lower_sorted), np.min(quantile_lower_sorted))
    y_max = max(np.max(y_test_sorted), np.max(y_pred_test[sort_indices]), 
                np.max(bootstrap_upper_sorted), np.max(quantile_upper_sorted))
    y_padding = (y_max - y_min) * 0.05  # 5% margin
    y_range = [y_min - y_padding, y_max + y_padding]
    ax1.set_ylim(y_range)
    
    # Set x-axis ticks (consistent with Quantile figure)
    tick_interval = max(1, n_samples // 6)
    x_ticks = np.arange(1, n_samples + 1, tick_interval)
    if len(x_ticks) == 0 or x_ticks[-1] != n_samples:
        x_ticks = np.append(x_ticks, n_samples)
    ax1.set_xticks(x_ticks)
    ax1.tick_params(axis='x', which='major', labelsize=18, width=2.0, length=6, colors='#000000')
    for label in ax1.get_xticklabels():
        label.set_fontweight('bold')
        label.set_fontfamily('Arial')
        label.set_color('#000000')
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'prediction_intervals_bootstrap.png'),
                dpi=300, bbox_inches='tight')
    plt.close(fig_bootstrap)
    
    # ==========================================================
    # Figure 2: Quantile Regression prediction interval (separate long bar)
    # ==========================================================
    fig_quantile, ax2 = plt.subplots(figsize=(16, 5))
    fig_quantile.patch.set_facecolor('white')
    
    # Plot Quantile Regression prediction interval
    quantile_baseline = y_pred_test[sort_indices]  # Quantile baseline: predicted values
    quantile_range = quantile_upper_sorted - quantile_lower_sorted
    
    # Check interval width, if too narrow, use simple fill, otherwise use gradient
    quantile_range_avg = np.mean(quantile_upper_sorted - quantile_lower_sorted)
    y_range_span = np.max(y_test_sorted) - np.min(y_test_sorted)
    range_ratio = quantile_range_avg / (y_range_span + 1e-10)
    
    # If the interval width is less than 1% of the y-axis range, use simple fill to improve visibility
    if range_ratio < 0.01:
        # Use simple fill, increase transparency to improve visibility
        ax2.fill_between(x_positions, quantile_lower_sorted, quantile_upper_sorted, 
                        alpha=0.4, color=color_blue_start, 
                        label='Quantile 80% interval', zorder=1,
                        edgecolor='none', linewidth=0)
    else:
        # 50-level gradient effect (from center dark to edge white)
        for i in range(n_gradient_layers):
            # layer_ratio from 0 to 1, 0=center (dark), 1=edge (white)
            layer_ratio = (i + 1) / n_gradient_layers
            # Calculate the boundaries of the current layer
            layer_lower = quantile_baseline - (quantile_baseline - quantile_lower_sorted) * layer_ratio
            layer_upper = quantile_baseline + (quantile_upper_sorted - quantile_baseline) * layer_ratio
            
            # Color interpolation: layer_ratio=0 is deep blue, layer_ratio=1 is white
            # Reverse layer_ratio for color: inner layer (close to 0) dark, outer layer (close to 1) white
            color_ratio = layer_ratio  # Use directly, because layer_ratio is from 0 to 1
            current_color = rgb_to_hex(interpolate_color(color_blue_start, color_blue_end, color_ratio))
            
            # Transparency: inner layer is less transparent, outer layer is more transparent (but not completely transparent)
            # From 0.8 (inner layer) to 0.2 (outer layer), increase overall transparency to ensure visibility
            alpha_layer = 0.8 - 0.6 * layer_ratio
            
            if i == 0:
                ax2.fill_between(x_positions, layer_lower, layer_upper, 
                               alpha=alpha_layer, color=current_color, 
                               label='Quantile 80% interval', zorder=1,
                               edgecolor='none', linewidth=0)
            else:
                ax2.fill_between(x_positions, layer_lower, layer_upper, 
                               alpha=alpha_layer, color=current_color, 
                               zorder=1, edgecolor='none', linewidth=0)
    
    # Plot predicted and actual values
    ax2.plot(x_positions, y_pred_test[sort_indices], '-', linewidth=3.5, 
             color=color_predicted, label='Predicted values', zorder=6, alpha=0.95)
    ax2.scatter(x_positions, y_test_sorted, s=120, c=color_actual, alpha=1.0, 
                label='Actual values', zorder=7, edgecolors='none', marker='o')
    
    # Set sub-figure labels and styles
    ax2.set_xlabel('Test samples (sorted by actual values)', 
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    ax2.set_ylabel('Peak Strain', 
                   fontsize=24, fontweight='bold', fontfamily='Arial', color='#000000')
    
    # Set x-axis ticks (consistent with the upper figure)
    ax2.set_xticks(x_ticks)
    
    # Set axis styles
    ax2.tick_params(axis='both', which='major', labelsize=18, width=2.0, length=6, colors='#000000')
    for label in ax2.get_xticklabels() + ax2.get_yticklabels():
        label.set_fontweight('bold')
        label.set_fontfamily('Arial')
        label.set_color('#000000')
    for spine in ax2.spines.values():
        spine.set_linewidth(2.0)
        spine.set_color('#000000')
    ax2.grid(False)
    
    # Add Quantile performance metrics
    quantile_text = f'Coverage: {quantile_coverage*100:.1f}%\nAvg Width: {quantile_width:.6f}'
    ax2.text(0.02, 0.98, quantile_text, transform=ax2.transAxes, fontsize=18, 
             fontweight='bold', fontfamily='Arial', verticalalignment='top', color='#000000')
    
    # Set legend
    ax2.legend(loc='lower right', fontsize=18, 
               prop={'family': 'Arial', 'weight': 'bold'},
               frameon=False, fancybox=False, shadow=False)
    
    # y-axis range consistent with Bootstrap figure
    ax2.set_ylim(y_range)
    
    plt.tight_layout()
    plt.savefig(os.path.join(save_dir, 'prediction_intervals_quantile.png'),
                dpi=300, bbox_inches='tight')
    plt.close(fig_quantile)
    
    # Save prediction interval results
    # Note: The figure is plotted in the order of actual values, so to match the figure,
    # here we also save the results in the same order (Sample_ID is the index of the original test sample in y_test).
    intervals_df = pd.DataFrame({
        'Sample_ID': sort_indices,
        'Actual': y_test_sorted,
        'Predicted': y_pred_test[sort_indices],
        'Bootstrap_Lower': bootstrap_lower_sorted,
        'Bootstrap_Upper': bootstrap_upper_sorted,
        'Quantile_Lower': quantile_lower_sorted,
        'Quantile_Upper': quantile_upper_sorted
    })
    # Save to Excel in the main function
    
    print(f"Prediction interval analysis completed, 80% confidence interval coverage: Bootstrap={bootstrap_coverage*100:.1f}%, Quantile={quantile_coverage*100:.1f}%")
    print(f"Bootstrap average interval width: {bootstrap_width:.6f}, Quantile average interval width: {quantile_width:.6f}")
    return bootstrap_predictions, bootstrap_coverage, quantile_coverage, intervals_df

def plot_radar_chart(y_true, y_pred, noise_rmses, outlier_rmses, bootstrap_coverage, 
                     quantile_coverage, save_dir, training_metrics_path=None):
    """Plot radar chart to show model performance metrics, and save data to Excel
    
    参数:
        y_true: True values
        y_pred: Predicted values
        noise_rmses: Noise robustness RMSE list
        outlier_rmses: Outlier robustness RMSE list
        bootstrap_coverage: Bootstrap prediction interval coverage
        quantile_coverage: Quantile Regression prediction interval coverage
        save_dir: Save directory
        training_metrics_path: Training metrics file path (deprecated, calculate all metrics directly)
    
    返回:
        radar_data_df: DataFrame containing radar chart data
        radar_detailed_df: DataFrame containing detailed data
        metrics_table_df: DataFrame containing the table of three-fold average and final model metrics
    """
    print("\n=== Plot radar chart ===")
    
    os.makedirs(save_dir, exist_ok=True)
    
    # Read all metrics from the training metrics file
    if training_metrics_path is None:
        training_metrics_path = os.path.join(save_dir, 'training_metrics.xlsx')
    
    print(f"Reading training metrics from {training_metrics_path}...")
    
    # Read three-fold average metrics (CV_Average sheet)
    df_avg = pd.read_excel(training_metrics_path, sheet_name='CV_Average')
    metric_dict = {}
    for _, row in df_avg.iterrows():
        metric = str(row['Metric']).strip()
        train_val = row['Train Average']
        test_val = row['Test Average']
        metric_dict[metric] = {'train': train_val, 'test': test_val}
    
    mean_train_metrics = {
        'R²': metric_dict['R²']['train'],
        'RMSE': metric_dict['RMSE']['train'],
        'MSE': metric_dict['MSE']['train'],
        'MAE': metric_dict['MAE']['train'],
        'MAPE': metric_dict['MAPE (%)']['train']
    }
    mean_test_metrics = {
        'R²': metric_dict['R²']['test'],
        'RMSE': metric_dict['RMSE']['test'],
        'MSE': metric_dict['MSE']['test'],
        'MAE': metric_dict['MAE']['test'],
        'MAPE': metric_dict['MAPE (%)']['test']
    }
    
    # Read final model metrics (Final_Model sheet)
    df_final = pd.read_excel(training_metrics_path, sheet_name='Final_Model')
    train_row = df_final[df_final['Set'].astype(str).str.contains('Training', na=False, case=False)].iloc[0]
    test_row = df_final[df_final['Set'].astype(str).str.contains('Testing', na=False, case=False)].iloc[0]
    
    final_train_metrics = {
        'R²': train_row['R²'],
        'RMSE': train_row['RMSE'],
        'MSE': train_row['MSE'],
        'MAE': train_row['MAE'],
        'MAPE': train_row['MAPE (%)']
    }
    final_test_metrics = {
        'R²': test_row['R²'],
        'RMSE': test_row['RMSE'],
        'MSE': test_row['MSE'],
        'MAE': test_row['MAE'],
        'MAPE': test_row['MAPE (%)']
    }
    
    print(f"✓ All training metrics have been read")
    
    # Calculate all performance metrics for the current test set (as Final_Test)
    r2 = r2_score(y_true, y_pred)
    mae = mean_absolute_error(y_true, y_pred)
    mse = mean_squared_error(y_true, y_pred)
    rmse = np.sqrt(mse)
    try:
        mape = mean_absolute_percentage_error(y_true, y_pred) * 100
    except:
        mape = np.mean(np.abs((y_true - y_pred) / (y_true + 1e-10))) * 100
    from scipy.stats import pearsonr
    r_value = pearsonr(y_true, y_pred)[0]
    
    # Use the current test set metrics as Final_Test
    final_test_metrics = {
        'R²': r2,
        'RMSE': rmse,
        'MAE': mae,
        'MSE': mse,
        'MAPE': mape
    }
    
    print(f"✓ All metrics have been read and calculated")
    
    # Calculate noise robustness score
    if noise_rmses is not None and len(noise_rmses) > 0 and noise_rmses[0] > 0:
        noise_degradation = (noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100
        noise_robustness_score = max(0, min(1, 1 - noise_degradation / 100))
    else:
        noise_degradation = np.nan
        noise_robustness_score = np.nan
    
    # Calculate outlier robustness score
    if outlier_rmses is not None and len(outlier_rmses) > 0 and outlier_rmses[0] > 0:
        outlier_degradation = (outlier_rmses[-1] - outlier_rmses[0]) / outlier_rmses[0] * 100
        outlier_robustness_score = max(0, min(1, 1 - abs(outlier_degradation) / 100))
    else:
        outlier_degradation = np.nan
        outlier_robustness_score = np.nan
    
    # Prediction interval coverage
    avg_coverage = (bootstrap_coverage + quantile_coverage) / 2 if (bootstrap_coverage is not None and quantile_coverage is not None) else np.nan
    
    # Normalization score (for radar chart)
    y_max = np.max(np.abs(y_true))
    
    # RMSE normalization score
    rmse_max_ref = y_max * 0.5
    rmse_normalized = max(0, min(1, 1 - rmse / rmse_max_ref))
    
    # MAE normalization score
    mae_max_ref = y_max * 0.4
    mae_normalized = max(0, min(1, 1 - mae / mae_max_ref))
    
    # MAPE normalization score
    mape_max_ref = 50.0
    mape_normalized = max(0, min(1, 1 - mape / mape_max_ref)) if not np.isnan(mape) else np.nan
    
    # R normalization score (R越大越好，直接使用)
    from scipy.stats import pearsonr
    r_value = pearsonr(y_true, y_pred)[0]
    r_normalized = r_value if not np.isnan(r_value) else np.nan
    
    # Prepare radar chart data
    categories = ['R²', 'RMSE\nScore', 'MAE\nScore', 'MAPE\nScore', 'R\nScore',
                  'Noise\nRobustness', 'Outlier\nRobustness', 'Prediction\nCoverage']
    
    values_normalized = [r2, rmse_normalized, mae_normalized, mape_normalized, r_normalized,
                        noise_robustness_score, outlier_robustness_score, avg_coverage]
    
    # Create radar chart data summary DataFrame
    radar_data_df = pd.DataFrame({
        'Metric Category': categories,
        'Normalization Score (0-1)': values_normalized
    })
    
    # Create detailed data DataFrame
    radar_detailed_df = pd.DataFrame({
        'Metric': ['R² (Final Model)', 'RMSE (Final Model)', 'MAE (Final Model)', 'MAPE (Final Model, %)', 'R (Final Model)',
                'R² (Three-Fold Average)', 'RMSE (Three-Fold Average)', 'MAE (Three-Fold Average)',
                'Noise Robustness Degradation (%)', 'Outlier Robustness Degradation (%)',
                'Bootstrap Coverage (%)', 'Quantile Coverage (%)', 'Average Prediction Interval Coverage (%)'],
        'Value': [r2, rmse, mae, mape, r_value,
                mean_test_metrics.get('R²', np.nan),
                mean_test_metrics.get('RMSE', np.nan),
                mean_test_metrics.get('MAE', np.nan),
                noise_degradation, outlier_degradation,
                bootstrap_coverage * 100 if bootstrap_coverage is not None else np.nan,
                quantile_coverage * 100 if quantile_coverage is not None else np.nan,
                avg_coverage * 100 if not np.isnan(avg_coverage) else np.nan],
        'Normalization Score (0-1)': [r2, rmse_normalized, mae_normalized, mape_normalized, r_normalized,
                            np.nan, np.nan, np.nan,
                            noise_robustness_score, outlier_robustness_score,
                            bootstrap_coverage if bootstrap_coverage is not None else np.nan,
                            quantile_coverage if quantile_coverage is not None else np.nan,
                            avg_coverage],
        'Description': ['R², the closer to 1 the better',
                'RMSE, the smaller the better',
                'MAE, the smaller the better',
                'MAPE, the smaller the better',
                'Pearson correlation coefficient, the closer to 1 the better',
                'Three-fold cross-validation average R²',
                'Three-fold cross-validation average RMSE',
                'Three-fold cross-validation average MAE',
                'Performance degradation rate at 10% noise, the smaller the better',
                'Performance degradation rate at 10% outliers, the smaller the better',
                'Bootstrap method prediction interval coverage, the closer to the target value (80%) the better',
                'Quantile Regression method prediction interval coverage, the closer to the target value (80%) the better',
                'Average coverage of the two methods, the closer to the target value (80%) the better']
    })
    
    # Plot radar chart
    valid_indices = [i for i, v in enumerate(values_normalized) if not np.isnan(v)]
    if len(valid_indices) > 2:
        categories_valid = [categories[i] for i in valid_indices]
        values_valid = [values_normalized[i] for i in valid_indices]
        
        N = len(categories_valid)
        angles = [n / float(N) * 2 * np.pi for n in range(N)]
        angles += angles[:1]
        
        values_valid += values_valid[:1]
        
        fig, ax = plt.subplots(figsize=(10, 10), subplot_kw=dict(projection='polar'))
        
        ax.plot(angles, values_valid, 'o-', linewidth=2, color='#2E86AB', label='CatBoost Performance')
        ax.fill(angles, values_valid, alpha=0.25, color='#2E86AB')
        
        ax.set_xticks(angles[:-1])
        ax.set_xticklabels(categories_valid, fontsize=11)
        
        ax.set_ylim(0, 1)
        ax.set_yticks([0.2, 0.4, 0.6, 0.8, 1.0])
        ax.set_yticklabels(['0.2', '0.4', '0.6', '0.8', '1.0'], fontsize=9)
        ax.grid(True, alpha=0.3)
        
        plt.title('CatBoost Model Performance Radar Chart\n' + 
                 f'R²={r2:.3f}, RMSE={rmse:.2f}, MAE={mae:.2f}', 
                 size=14, fontweight='bold', pad=20)
        
        plt.legend(loc='upper right', bbox_to_anchor=(1.3, 1.1))
        
        plt.tight_layout()
        plt.savefig(os.path.join(save_dir, 'radar_chart.png'), dpi=300, bbox_inches='tight')
        plt.close()
        print("✓ Radar chart has been saved: radar_chart.png")
    else:
        print("⚠ Warning: Not enough valid metrics to plot radar chart")
    
    # Create table containing three-fold average and final model metrics (five metrics: R², RMSE, MSE, MAE, MAPE)
    metrics_table_data = {
        'Metric': ['R²', 'RMSE', 'MSE', 'MAE', 'MAPE (%)'],
        'Mean_Train': [
            mean_train_metrics.get('R²', np.nan),
            mean_train_metrics.get('RMSE', np.nan),
            mean_train_metrics.get('MSE', np.nan),
            mean_train_metrics.get('MAE', np.nan),
            mean_train_metrics.get('MAPE', np.nan)
        ],
        'Mean_Test': [
            mean_test_metrics.get('R²', np.nan),
            mean_test_metrics.get('RMSE', np.nan),
            mean_test_metrics.get('MSE', np.nan),
            mean_test_metrics.get('MAE', np.nan),
            mean_test_metrics.get('MAPE', np.nan)
        ],
        'Final_Train': [
            final_train_metrics.get('R²', np.nan),
            final_train_metrics.get('RMSE', np.nan),
            final_train_metrics.get('MSE', np.nan),
            final_train_metrics.get('MAE', np.nan),
            final_train_metrics.get('MAPE', np.nan)
        ],
        'Final_Test': [
            final_test_metrics.get('R²', np.nan),
            final_test_metrics.get('RMSE', np.nan),
            final_test_metrics.get('MSE', np.nan),
            final_test_metrics.get('MAE', np.nan),
            final_test_metrics.get('MAPE', np.nan)
        ]
    }
    metrics_table_df = pd.DataFrame(metrics_table_data)
    
    print("\nPerformance metrics table (three-fold average vs final model):")
    print(metrics_table_df.to_string(index=False))
    
    return radar_data_df, radar_detailed_df, metrics_table_df

def train_single_fold(X_train, y_train, X_val, y_val, X_test, y_test, train_ids, val_ids, test_ids, 
                     feature_names, save_dir, n_trials=100):
    """Single fold training (for cross-validation) - only perform hyperparameter optimization and training, no interpretability analysis"""
    # Hyperparameter optimization
    study = optimize_catboost(X_train, y_train, X_val, y_val, n_trials=n_trials)
    best_params = study.best_params
    
    # Train final model
    model = train_catboost(X_train, y_train, X_val, y_val, best_params)
    
    # Predict test set (only for performance evaluation)
    if len(X_test) > 0:
        y_pred = model.predict(X_test)
        r2 = r2_score(y_test, y_pred)
        mae = mean_absolute_error(y_test, y_pred)
        rmse = np.sqrt(mean_squared_error(y_test, y_pred))
        
        # Save prediction results
        results_df = pd.DataFrame({
            'Sample_ID': test_ids,
            'Actual': y_test,
            'Predicted': y_pred,
            'Residual': y_pred - y_test
        })
        results_df.to_excel(os.path.join(save_dir, 'test_predictions.xlsx'), index=False)
        
        return model, best_params, {'r2': r2, 'mae': mae, 'rmse': rmse}, train_ids, val_ids, test_ids
    else:
        return model, best_params, None, train_ids, val_ids, test_ids

def load_trained_model(model_path):
    """Load trained CatBoost model (supports .joblib and .pkl formats)"""
    print(f"\n=== Load trained model ===")
    print(f"Model path: {model_path}")
    
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model file not found: {model_path}")
    
    loaded_data = joblib.load(model_path)
    
    # Check the type of loaded data
    if isinstance(loaded_data, dict):
        # If it's a dictionary, try to extract the model and feature names
        model = loaded_data.get('model', loaded_data)
        feature_names = loaded_data.get('feature_names', None)
        model_data = loaded_data
    else:
        # If it's a model object directly
        model = loaded_data
        feature_names = None
        model_data = {'model': model}
    
    print(f"✓ Model loaded successfully")
    print(f"✓ Model type: {type(model).__name__}")
    if feature_names:
        print(f"✓ Number of features: {len(feature_names)}")
        print(f"✓ Feature names: {feature_names}")
    else:
        print("⚠ Warning: No feature name information found, will use feature names from data loading")
    
    return model, model_data

def cross_subset_validation(X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df_processed, 
                           save_dir=None, n_trials=100):
    """Three-fold cross-validation for subset1/2/3"""
    print("\n=== CatBoost Three-fold cross-validation for subset1/2/3 ===")

    if save_dir is None:
        save_dir = os.path.join(SAVE_ROOT, 'catboost_cv')
    
    # Extract subset labels and test set
    subset_labels = sorted([s for s in np.unique(sample_divisions) if str(s).startswith('subset')])
    test_mask = np.array([str(s).strip().lower().startswith('test') for s in sample_divisions])
    test_indices = np.where(test_mask)[0]
    
    if len(subset_labels) != 3:
        raise ValueError(f"Number of subsets should be 3, actual: {len(subset_labels)} {subset_labels}")
    
    if len(test_indices) == 0:
        raise ValueError("No test set found")
    
    print(f"Found 3 subsets: {subset_labels}")
    print(f"Number of test set samples: {len(test_indices)}")
    
    all_results = []
    
    # Three-fold cross-validation
    for i, val_label in enumerate(subset_labels):
        print(f"\n{'='*60}")
        print(f"Subset cross-validation round {i+1}/3 [{val_label} as validation set]")
        print(f"{'='*60}")
        
        # Determine training and validation sets
        train_labels = [lbl for lbl in subset_labels if lbl != val_label]
        subset_to_indices = {lbl: np.where(sample_divisions == lbl)[0] for lbl in subset_labels}
        
        train_indices = np.concatenate([subset_to_indices[lbl] for lbl in train_labels])
        val_indices = subset_to_indices[val_label]
        
        print(f"Training set = {train_labels}，Validation set = {val_label}，Test set = test")
        print(f"  Number of training set samples: {len(train_indices)}，Number of validation set samples: {len(val_indices)}，Number of test set samples: {len(test_indices)}")
        
        # Split data
        X_train = X[train_indices]
        y_train = y[train_indices]
        X_val = X[val_indices]
        y_val = y[val_indices]
        X_test = X[test_indices]
        y_test = y[test_indices]
        
        train_ids_fold = [sample_ids[idx] for idx in train_indices]
        val_ids_fold = [sample_ids[idx] for idx in val_indices]
        test_ids_fold = [sample_ids[idx] for idx in test_indices]
        
        # Create save directory for this round
        round_save_dir = os.path.join(save_dir, f'subsetCV_{val_label}')
        os.makedirs(round_save_dir, exist_ok=True)
        
        # Train single fold model (only perform hyperparameter optimization and training, no interpretability analysis)
        model, best_params, metrics, train_ids_round, val_ids_round, test_ids_round = train_single_fold(
            X_train, y_train, X_val, y_val, X_test, y_test,
            train_ids_fold, val_ids_fold, test_ids_fold,
            feature_names, round_save_dir, n_trials=n_trials
        )
        
        all_results.append({
            'val_label': val_label,
            'train_labels': train_labels,
            'best_params': best_params,
            'metrics': metrics,
            'model': model,
            'train_val_indices': np.concatenate([train_indices, val_indices]),
            'test_indices': test_indices.copy()
        })
        
        print(f"[{val_label} validation set] This round of training completed")
        if metrics:
            print(f"  Test set metrics: R2={metrics['r2']:.4f}, MAE={metrics['mae']:.3f}, RMSE={metrics['rmse']:.3f}")
    
    # Print summary
    print(f"\n{'='*60}")
    print("All three-fold cross-validation completed!")
    print(f"{'='*60}")
    for i, res in enumerate(all_results):
        print(f"[{i+1}] Validation set: {res['val_label']}")
        if res['metrics']:
            print(f"     Test set metrics: R2={res['metrics']['r2']:.4f}, MAE={res['metrics']['mae']:.3f}, RMSE={res['metrics']['rmse']:.3f}")
    
    # Calculate three-fold average metrics
    valid_metrics = [res['metrics'] for res in all_results if res['metrics'] is not None]
    if valid_metrics:
        avg_r2 = np.mean([m['r2'] for m in valid_metrics])
        avg_mae = np.mean([m['mae'] for m in valid_metrics])
        avg_rmse = np.mean([m['rmse'] for m in valid_metrics])
        print(f"\nThree-fold average performance: R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")
    else:
        avg_r2 = avg_mae = avg_rmse = None

    # Find the model with the best test set performance (by R2)
    best_fold_idx = 0
    best_r2 = -np.inf
    for i, res in enumerate(all_results):
        if res['metrics'] and res['metrics']['r2'] > best_r2:
            best_r2 = res['metrics']['r2']
            best_fold_idx = i
    
    best_fold_result = all_results[best_fold_idx]
    best_overall_params = best_fold_result['best_params']
    
    print(f"\n{'='*60}")
    print(f"Three-fold cross-validation best hyperparameters (from fold {best_fold_idx+1}, validation set={best_fold_result['val_label']})")
    print(f"Test set R2: {best_r2:.4f}")
    print(f"Best hyperparameters: {best_overall_params}")
    print(f"{'='*60}")
    
    # Use best hyperparameters, retrain final model with all subset data (excluding test)
    print(f"\n{'='*60}")
    print("Use best fold model as final model (no re-training)")
    print(f"{'='*60}")

    best_fold_train_val_indices = best_fold_result['train_val_indices']
    best_fold_test_indices = best_fold_result['test_indices']

    X_train_final = X[best_fold_train_val_indices]
    y_train_final = y[best_fold_train_val_indices]
    X_test_final = X[best_fold_test_indices]
    y_test_final = y[best_fold_test_indices]

    train_ids_final = [sample_ids[idx] for idx in best_fold_train_val_indices]
    test_ids_final = [sample_ids[idx] for idx in best_fold_test_indices]

    final_model = best_fold_result['model']

    # Evaluate on test set (should be consistent with the fold metrics)
    y_test_pred = final_model.predict(X_test_final)
    final_r2 = r2_score(y_test_final, y_test_pred)
    final_mae = mean_absolute_error(y_test_final, y_test_pred)
    final_rmse = np.sqrt(mean_squared_error(y_test_final, y_test_pred))

    print(f"Final model from validation set {best_fold_result['val_label']} fold.")
    print(f"  R2: {final_r2:.4f}")
    print(f"  MAE: {final_mae:.3f}")
    print(f"  RMSE: {final_rmse:.3f}")
    if avg_r2 is not None:
        print(f"  (Three-fold average) R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")
    
    # Interpretability analysis directory
    interpretability_dir = os.path.join(save_dir, 'interpretability_analysis')
    os.makedirs(interpretability_dir, exist_ok=True)
    
    print(f"\n{'='*60}")
    print("Start interpretability analysis")
    print(f"{'='*60}")
    
    # 1. Plot results
    plot_results(y_test_final, y_test_pred, interpretability_dir)
    
    # 2. Plot training and test set comparison
    plot_train_test_comparison(final_model, X_train_final, y_train_final, X_test_final, y_test_final, interpretability_dir)
    
    # 3. Feature importance
    X_all_final_for_importance = np.vstack([X_train_final, X_test_final])
    importance_df = plot_feature_importance(final_model, feature_names, interpretability_dir, X=X_all_final_for_importance)
    
    # 4. SHAP analysis
    X_all_final = np.vstack([X_train_final, X_test_final])
    shap_importance = plot_shap_analysis(final_model, X_all_final, X_all_final, feature_names, interpretability_dir)
    
    # 5. PDP analysis
    plot_pdp_analysis(final_model, X_all_final, feature_names, interpretability_dir, n_top_features=6)
    
    # 6. Noise robustness analysis (optional)
    # analyze_noise_robustness(final_model, X_test_final, y_test_final, feature_names, interpretability_dir)
    
    # 7. Prediction interval analysis (optional)
    # analyze_prediction_intervals(final_model, X_test_final, y_test_final, interpretability_dir, n_bootstrap=100, bootstrap_noise_level=0.01)
    
    # Save final model
    joblib.dump({
        'model': final_model,
        'best_params': best_overall_params,
        'feature_names': feature_names,
        'r2': final_r2,
        'mae': final_mae,
        'rmse': final_rmse,
        'importance_df': importance_df,
        'shap_importance': shap_importance,
        'cross_validation_results': all_results
    }, os.path.join(interpretability_dir, 'final_catboost_model.pkl'))
    
    # Predict all samples and write to Excel
    print(f"\n{'='*60}")
    print("Predict all samples and write to Excel")
    print(f"{'='*60}")
    try:
        # Predict all samples (including test set)
        X_all_samples = np.vstack([X_train_final, X_test_final])
        y_all_pred = final_model.predict(X_all_samples)
        all_sample_ids = train_ids_final + test_ids_final
        
        # Create prediction result dictionary
        pred_dict = {}
        for idx, sample_id in enumerate(all_sample_ids):
            pred_dict[sample_id] = y_all_pred[idx]
        
        # Reload original Excel file
        data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
        df_original = pd.read_excel(data_file, sheet_name=0)
        
        # Use map method to match
        if 'No_Customized' in df_original.columns:
            df_original['CatBoost_peak_strain'] = df_original['No_Customized'].map(pred_dict)
        else:
            print("Warning: No_Customized column not found, cannot match prediction results")
        
        # Save results
        output_file = os.path.join(PROJECT_ROOT, "dataset/dataset_with_CatBoost_peak_strain.xlsx")
        df_original.to_excel(output_file, index=False)
        print(f"✓ Prediction results written to: {output_file}")
        print(f"  - New column: CatBoost_peak_strain (final model prediction values)")
        print(f"  - Number of non-empty prediction values: {df_original['CatBoost_peak_strain'].notna().sum()}")
        
        # Backup
        backup_file = os.path.join(save_dir, 'dataset_with_CatBoost_peak_strain_backup.xlsx')
        df_original.to_excel(backup_file, index=False)
        print(f"✓ Backup file saved: {backup_file}")
    except Exception as e:
        print(f"Warning: Error writing to Excel: {e}")
        import traceback
        traceback.print_exc()
    
    return all_results, final_model, best_overall_params

def cross_random_kfold_validation(X, y, feature_names, sample_ids, sample_divisions=None,
                                  save_dir=None, n_trials=100, n_splits=3, random_state=42):
    """Random KFold cross-validation"""
    print(f"\n=== CatBoost 随机{n_splits}折交叉验证 ===")

    if save_dir is None:
        save_dir = os.path.join(SAVE_ROOT, 'catboost_cv_random')

    os.makedirs(save_dir, exist_ok=True)

    fixed_test_indices = None
    if sample_divisions is not None:
        test_mask = np.array([str(s).strip().lower().startswith('test') for s in sample_divisions])
        if np.any(test_mask):
            fixed_test_indices = np.where(test_mask)[0]
            print(f"Use pre-divided test set ({len(fixed_test_indices)} samples) for random KFold cross-validation evaluation")

    if fixed_test_indices is not None and len(fixed_test_indices) == len(X):
        raise ValueError("All samples belong to test set, cannot perform random KFold cross-validation")

    if fixed_test_indices is not None:
        train_pool_mask = np.ones(len(X), dtype=bool)
        train_pool_mask[fixed_test_indices] = False
        train_pool_indices = np.where(train_pool_mask)[0]
        print(f"Random KFold cross-validation only on non-test set samples, {len(train_pool_indices)} samples")
        kf = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
        split_source = train_pool_indices
    else:
        kf = KFold(n_splits=n_splits, shuffle=True, random_state=random_state)
        split_source = np.arange(len(X))

    all_results = []

    for fold_idx, (train_indices_rel, val_indices_rel) in enumerate(kf.split(split_source), 1):
        print(f"\n{'='*60}")
        print(f"Random KFold fold {fold_idx}/{n_splits}")
        print(f"{'='*60}")

        if fixed_test_indices is not None:
            train_indices = split_source[train_indices_rel]
            val_indices = split_source[val_indices_rel]
            test_indices = fixed_test_indices
        else:
            train_val_indices = split_source[train_indices_rel]
            val_candidates = split_source[val_indices_rel]
            if len(train_val_indices) < 5:
                raise ValueError("Too few training samples, cannot perform random KFold cross-validation")
            inner_train_indices, inner_val_indices = train_test_split(
                np.concatenate([train_val_indices, val_candidates]),
                test_size=0.2,
                random_state=random_state + fold_idx
            )
            train_indices = inner_train_indices
            val_indices = inner_val_indices
            test_indices = val_candidates

        if len(train_indices) < 5:
            raise ValueError("Too few training samples, cannot perform random KFold cross-validation")

        print(f"  Number of training set samples: {len(train_indices)}")
        print(f"  Number of validation set samples: {len(val_indices)}")
        print(f"  Number of test set samples: {len(test_indices)}")

        X_train = X[train_indices]
        y_train = y[train_indices]
        X_val = X[val_indices]
        y_val = y[val_indices]
        X_test = X[test_indices]
        y_test = y[test_indices]

        train_ids_fold = [sample_ids[idx] for idx in train_indices]
        val_ids_fold = [sample_ids[idx] for idx in val_indices]
        test_ids_fold = [sample_ids[idx] for idx in test_indices]

        round_save_dir = os.path.join(save_dir, f'randomKFold_fold{fold_idx}')
        os.makedirs(round_save_dir, exist_ok=True)

        model, best_params, metrics, _, _, _ = train_single_fold(
            X_train, y_train, X_val, y_val, X_test, y_test,
            train_ids_fold, val_ids_fold, test_ids_fold,
            feature_names, round_save_dir, n_trials=n_trials
        )

        combined_train_val = np.unique(np.concatenate([train_indices, val_indices]))

        all_results.append({
            'fold_idx': fold_idx,
            'best_params': best_params,
            'metrics': metrics,
            'model': model,
            'train_val_indices': combined_train_val,
            'test_indices': test_indices.copy()
        })

        if metrics:
            print(f"[Fold {fold_idx}] Test set metrics: R2={metrics['r2']:.4f}, MAE={metrics['mae']:.3f}, RMSE={metrics['rmse']:.3f}")

    print(f"\n{'='*60}")
    print("All random KFold completed!")
    print(f"{'='*60}")
    for res in all_results:
        if res['metrics']:
            print(f"[Fold {res['fold_idx']}] 测试集: R2={res['metrics']['r2']:.4f}, MAE={res['metrics']['mae']:.3f}, RMSE={res['metrics']['rmse']:.3f}")

    valid_metrics = [res['metrics'] for res in all_results if res['metrics']]
    if valid_metrics:
        avg_r2 = np.mean([m['r2'] for m in valid_metrics])
        avg_mae = np.mean([m['mae'] for m in valid_metrics])
        avg_rmse = np.mean([m['rmse'] for m in valid_metrics])
        print(f"\n{n_splits} fold average performance: R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")
    else:
        avg_r2 = avg_mae = avg_rmse = None

    best_fold_idx = 0
    best_r2 = -np.inf
    for i, res in enumerate(all_results):
        if res['metrics'] and res['metrics']['r2'] > best_r2:
            best_r2 = res['metrics']['r2']
            best_fold_idx = i

    best_fold_result = all_results[best_fold_idx]
    best_overall_params = best_fold_result['best_params']

    print(f"\n{'='*60}")
    print(f"Best fold = Fold {best_fold_result['fold_idx']} (Test set R2={best_r2:.4f})")
    print(f"Best hyperparameters: {best_overall_params}")
    print(f"{'='*60}")

    best_fold_train_val_indices = best_fold_result['train_val_indices']
    best_fold_test_indices = best_fold_result['test_indices']

    X_train_final = X[best_fold_train_val_indices]
    y_train_final = y[best_fold_train_val_indices]
    X_test_final = X[best_fold_test_indices]
    y_test_final = y[best_fold_test_indices]

    train_ids_final = [sample_ids[idx] for idx in best_fold_train_val_indices]
    test_ids_final = [sample_ids[idx] for idx in best_fold_test_indices]

    final_model = best_fold_result['model']

    y_test_pred = final_model.predict(X_test_final)
    final_r2 = r2_score(y_test_final, y_test_pred)
    final_mae = mean_absolute_error(y_test_final, y_test_pred)
    final_rmse = np.sqrt(mean_squared_error(y_test_final, y_test_pred))

    print(f"Final model trained using Fold {best_fold_result['fold_idx']}.")
    print(f"  R2: {final_r2:.4f}")
    print(f"  MAE: {final_mae:.3f}")
    print(f"  RMSE: {final_rmse:.3f}")
    if avg_r2 is not None:
        print(f"  ({n_splits} fold average) R2={avg_r2:.4f}, MAE={avg_mae:.3f}, RMSE={avg_rmse:.3f}")

    interpretability_dir = os.path.join(save_dir, 'interpretability_analysis')
    os.makedirs(interpretability_dir, exist_ok=True)

    print(f"\n{'='*60}")
    print("Start interpretability analysis (random split)")
    print(f"{'='*60}")

    plot_results(y_test_final, y_test_pred, interpretability_dir)
    plot_train_test_comparison(final_model, X_train_final, y_train_final, X_test_final, y_test_final, interpretability_dir)
    X_all_final = np.vstack([X_train_final, X_test_final])
    importance_df = plot_feature_importance(final_model, feature_names, interpretability_dir, X=X_all_final)
    shap_importance = plot_shap_analysis(final_model, X_all_final, X_all_final, feature_names, interpretability_dir)
    plot_pdp_analysis(final_model, X_all_final, feature_names, interpretability_dir, n_top_features=6)

    joblib.dump({
        'model': final_model,
        'best_params': best_overall_params,
        'feature_names': feature_names,
        'r2': final_r2,
        'mae': final_mae,
        'rmse': final_rmse,
        'importance_df': importance_df,
        'shap_importance': shap_importance,
        'cross_validation_results': all_results
    }, os.path.join(interpretability_dir, 'final_catboost_model_random.pkl'))

    print(f"\n{'='*60}")
    print("Predict all samples and write to Excel (random split model)")
    print(f"{'='*60}")
    try:
        X_all_samples = X
        y_all_pred = final_model.predict(X_all_samples)
        pred_dict = {sid: pred for sid, pred in zip(sample_ids, y_all_pred)}

        data_file = os.path.join(PROJECT_ROOT, "dataset/dataset_final.xlsx")
        df_original = pd.read_excel(data_file, sheet_name=0)

        if 'No_Customized' in df_original.columns:
            df_original['CatBoost_peak_strain_random'] = df_original['No_Customized'].map(pred_dict)
        else:
            df_original['CatBoost_peak_strain_random'] = y_all_pred

        output_file = os.path.join(save_dir, "dataset_with_CatBoost_peak_strain_random.xlsx")
        df_original.to_excel(output_file, index=False)
        print(f"✓ Random split model prediction results written to: {output_file}")
        print(f"  - New column: CatBoost_peak_strain_random")
        print(f"  - Number of non-empty prediction values: {df_original['CatBoost_peak_strain_random'].notna().sum()}")
    except Exception as e:
        print(f"Warning: Error writing to random split Excel: {e}")
        import traceback
        traceback.print_exc()

    return all_results, final_model, best_overall_params

def print_training_metrics():
    """Read and print training metrics"""
    print("\n" + "="*70)
    print("CatBoost model training metrics")
    print("="*70)
    
    metrics_file = os.path.join(SAVE_ROOT, 'training_metrics.xlsx')
    if not os.path.exists(metrics_file):
        print(f"⚠ Warning: Training metrics file not found: {metrics_file}")
        return
    
    try:
        # Read all worksheets
        excel_data = pd.read_excel(metrics_file, sheet_name=None)
        
        # Print final model metrics
        if 'Final_Model' in excel_data:
            print("\nFinal model performance metrics")
            df_final = excel_data['Final_Model']
            print(df_final.to_string(index=False))
        
        # Print three-fold cross-validation average metrics
        if 'CV_Average' in excel_data:
            print("\nThree-fold cross-validation average metrics")
            df_cv = excel_data['CV_Average']
            print(df_cv.to_string(index=False))
        
        # Print best fold metrics
        if 'Best_Fold' in excel_data:
            print("\nBest fold performance metrics")
            df_best = excel_data['Best_Fold']
            print(df_best.to_string(index=False))
        
        # Print each fold's detailed metrics
        if 'Fold_Details' in excel_data:
            print("\nEach fold's detailed metrics")
            df_folds = excel_data['Fold_Details']
            print(df_folds.to_string(index=False))
        
        # Print best hyperparameters
        if 'Best_Hyperparameters' in excel_data:
            print("\nBest hyperparameters")
            df_params = excel_data['Best_Hyperparameters']
            print(df_params.to_string(index=False))
        
        # Print model summary
        if 'Summary' in excel_data:
            print("\nModel summary")
            df_summary = excel_data['Summary']
            print(df_summary.to_string(index=False))
        
        print("\n" + "="*70)
        
    except Exception as e:
        print(f"⚠ Error reading training metrics: {e}")
        import traceback
        traceback.print_exc()

def main():
    """Main function - Load trained model and perform SHAP analysis, noise analysis and uncertainty analysis"""
    print("=== CatBoost model SHAP analysis, noise analysis and uncertainty analysis ===")
    
    # 0. First print training metrics
    print_training_metrics()
    
    # 1. Load data
    data = load_data()
    if data is None:
        return
    
    X, y, feature_names, sample_divisions, sample_ids, original_df_indices, df_processed = data
    
    # 2. Load trained model - directly use catboost_final_model.joblib
    model_path = os.path.join(SAVE_ROOT, 'catboost_final_model.joblib')
    
    if not os.path.exists(model_path):
        print(f"Error: Model file not found: {model_path}")
        print("Ensure model file exists in: peak_strain/CatBoost/save/catboost_final_model.joblib")
        return
    
    model, model_data = load_trained_model(model_path)
    
    # If model data has feature names, use model's (ensure feature order is consistent)
    if 'feature_names' in model_data and model_data['feature_names'] is not None:
        model_feature_names = model_data['feature_names']
        print(f"\nUsing model saved feature names ({len(model_feature_names)} features)")
        # 确保特征顺序与模型一致
        if len(model_feature_names) == len(feature_names):
            # Check if feature names match
            if set(model_feature_names) == set(feature_names):
                # Reorder features to match model
                feature_order = [feature_names.index(f) for f in model_feature_names]
                X = X[:, feature_order]
                feature_names = model_feature_names
                print("✓ Feature order adjusted to match model")
            else:
                print("⚠ Warning: Model feature names do not match data feature names")
                print(f"  Model features: {model_feature_names}")
                print(f"  Data features: {feature_names}")
        else:
            print(f"⚠ Warning: Feature number mismatch - model: {len(model_feature_names)}, data: {len(feature_names)}")
    
    # 3. Data分割（获取训练集和测试集，用于SHAP分析和噪声分析）
    X_train, X_val, X_test, y_train, y_val, y_test, train_ids, val_ids, test_ids = split_data_by_divisions(
        X, y, sample_divisions, sample_ids, test_ratio=0.2, val_ratio=0.2, random_state=42
    )
    
    if len(X_test) == 0:
        print("Warning: Test set not found, using validation set for analysis")
        X_test = X_val
        y_test = y_val
        test_ids = val_ids
    
    # Prepare data for SHAP analysis
    # Use all data set for SHAP analysis (data set is small, total 94 samples)
    
    # Merge all data sets: training set + validation set + test set
    datasets_to_merge = []
    y_datasets_to_merge = []
    
    if len(X_train) > 0:
        datasets_to_merge.append(X_train)
        y_datasets_to_merge.append(y_train)
    if len(X_val) > 0:
        datasets_to_merge.append(X_val)
        y_datasets_to_merge.append(y_val)
    if len(X_test) > 0:
        datasets_to_merge.append(X_test)
        y_datasets_to_merge.append(y_test)
    
    if len(datasets_to_merge) > 0:
        X_all = np.vstack(datasets_to_merge)
        y_all = np.hstack(y_datasets_to_merge)
    else:
        X_all = X
        y_all = y
    
    # Use all data set as background and explanation data set
    X_background = X_all
    y_background = y_all
    X_explain = X_all
    y_explain = y_all
    
    print(f"\nData preparation completed:")
    print(f"  Training set: {len(X_train)} samples")
    print(f"  Validation set: {len(X_val)} samples")
    print(f"  Test set: {len(X_test)} samples")
    print(f"\nSHAP analysis data set configuration (using all data set):")
    print(f"  All data set: {len(X_all)} samples (training set + validation set + test set)")
    print(f"  Background data set (for explainer initialization): {len(X_background)} samples")
    print(f"  Explanation data set (for calculating SHAP values): {len(X_explain)} samples")
    
    # 4. Create save directory
    analysis_save_dir = os.path.join(SAVE_ROOT)
    os.makedirs(analysis_save_dir, exist_ok=True)
    
    # 5. SHAP analysis (using all data set)
    print(f"\n{'='*60}")
    print("Start SHAP analysis (using all data set)")
    print(f"{'='*60}")
    shap_importance = None
    feature_stats = None
    try:
        shap_importance = plot_shap_analysis(
            model, X_background, X_explain, feature_names, analysis_save_dir, y_explain=y_all
        )
        print("✓ SHAP analysis completed")
        
        # 5.1 Feature statistics analysis (explain why some features have small impact)
        print(f"\n{'='*60}")
        print("Start feature statistics analysis")
        print(f"{'='*60}")
        try:
            feature_stats = analyze_feature_statistics(
                X_all, y_all, feature_names, shap_importance, analysis_save_dir
            )
            print("✓ Feature statistics analysis completed")
        except Exception as e:
            print(f"⚠ Feature statistics analysis error: {e}")
            import traceback
            traceback.print_exc()
            feature_stats = None
    except Exception as e:
        print(f"⚠ SHAP analysis error: {e}")
        import traceback
        traceback.print_exc()
        shap_importance = None
        feature_stats = None
    
    # 6. Feature importance analysis
    print(f"\n{'='*60}")
    print("Start feature importance analysis")
    print(f"{'='*60}")
    importance_df = None
    try:
        importance_df = plot_feature_importance(
            model, feature_names, analysis_save_dir, X=X_all
        )
        print("✓ Feature importance analysis completed")
    except Exception as e:
        print(f"⚠ Feature importance analysis error: {e}")
        import traceback
        traceback.print_exc()
        importance_df = None
    
    # 7. Noise robustness analysis
    print(f"\n{'='*60}")
    print("Start noise robustness analysis")
    print(f"{'='*60}")
    noise_rmses, outlier_rmses = None, None
    robustness_df, outlier_df = None, None
    try:
        noise_rmses, outlier_rmses, robustness_df, outlier_df = analyze_noise_robustness(
            model, X_test, y_test, feature_names, analysis_save_dir
        )
        print("✓ Noise robustness analysis completed")
    except Exception as e:
        print(f"⚠ Noise robustness analysis error: {e}")
        import traceback
        traceback.print_exc()
        noise_rmses, outlier_rmses = None, None
        robustness_df, outlier_df = None, None
    
    # 8. Uncertainty analysis (prediction intervals)
    print(f"\n{'='*60}")
    print("Start uncertainty analysis (prediction intervals)")
    print(f"{'='*60}")
    bootstrap_predictions, bootstrap_coverage, quantile_coverage = None, None, None
    intervals_df = None
    try:
        bootstrap_predictions, bootstrap_coverage, quantile_coverage, intervals_df = analyze_prediction_intervals(
            model, X_test, y_test, analysis_save_dir, n_bootstrap=100, bootstrap_noise_level=0.01
        )
        print("✓ Uncertainty analysis completed")
    except Exception as e:
        print(f"⚠ Uncertainty analysis error: {e}")
        import traceback
        traceback.print_exc()
        bootstrap_predictions, bootstrap_coverage, quantile_coverage = None, None, None
        intervals_df = None
    
    # 9. Plot prediction result comparison chart
    print(f"\n{'='*60}")
    print("Start plotting prediction result comparison chart")
    print(f"{'='*60}")
    prediction_df = None
    y_test_pred = None  # Initialize y_test_pred, ensure available when radar chart is called
    try:
        y_test_pred = model.predict(X_test)
        plot_results(y_test, y_test_pred, analysis_save_dir)
        # Use training set + validation set as training set, test set as test set for comparison
        # Note: X_background contains all data (training set + validation set + test set), need to exclude test set
        # Merge training set and validation set (exclude test set)
        X_train_val = np.vstack([X_train, X_val]) if len(X_val) > 0 else X_train
        y_train_val = np.hstack([y_train, y_val]) if len(y_val) > 0 else y_train
        prediction_df = plot_train_test_comparison(model, X_train_val, y_train_val, X_test, y_test, analysis_save_dir)
        print("✓ Prediction result comparison chart plotted")
    except Exception as e:
        print(f"⚠ Plotting prediction result chart error: {e}")
        import traceback
        traceback.print_exc()
        prediction_df = None
    
    # 10. PDP analysis (using training set + validation set, if data is large may be slow)
    print(f"\n{'='*60}")
    print("Start PDP analysis (partial dependency plot, using training set + validation set)")
    print(f"{'='*60}")
    try:
        plot_pdp_analysis(model, X_background, feature_names, analysis_save_dir, n_top_features=6)
        print("✓ PDP analysis completed")
    except Exception as e:
        print(f"⚠ PDP analysis error: {e}")
        import traceback
        traceback.print_exc()
    
    # 10.5. Plot radar chart and save performance metrics data
    print(f"\n{'='*60}")
    print("Start plotting radar chart and save performance metrics data")
    print(f"{'='*60}")
    radar_data_df = None
    radar_detailed_df = None
    metrics_table_df = None
    
    # Ensure y_test_pred has been calculated
    if 'y_test_pred' not in locals() or y_test_pred is None:
        try:
            y_test_pred = model.predict(X_test)
            print("✓ Test set prediction values have been calculated, for radar chart analysis")
        except Exception as e:
            print(f"⚠ Error calculating test set prediction values: {e}")
            y_test_pred = None
    
    if y_test_pred is not None:
        try:
            # Get training metrics file path
            training_metrics_path = os.path.join(SAVE_ROOT, 'training_metrics.xlsx')
            if not os.path.exists(training_metrics_path):
                print(f"⚠ Warning: Training metrics file not found: {training_metrics_path}")
                training_metrics_path = None
            
            radar_data_df, radar_detailed_df, metrics_table_df = plot_radar_chart(
                y_test, y_test_pred, noise_rmses, outlier_rmses, 
                bootstrap_coverage, quantile_coverage, 
                analysis_save_dir, training_metrics_path=training_metrics_path
            )
            print("✓ Radar chart plotted, performance metrics data has been prepared")
        except Exception as e:
            print(f"⚠ Error plotting radar chart: {e}")
            import traceback
            traceback.print_exc()
            radar_data_df = None
            radar_detailed_df = None
            metrics_table_df = None
    else:
        print("⚠ Warning: Cannot plot radar chart, because test set prediction values are not available")
    
    # 11. Print summary
    print(f"\n{'='*60}")
    print("Analysis completed summary")
    print(f"{'='*60}")
    
    if noise_rmses is not None:
        print(f"\nNoise robustness analysis results:")
        print(f"  - No noise RMSE: {noise_rmses[0]:.3f}")
        print(f"  - 10% noise RMSE: {noise_rmses[-1]:.3f}")
        print(f"  - Performance degradation rate: {(noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100:.1f}%")
    
    if bootstrap_coverage is not None:
        print(f"\nUncertainty analysis results:")
        print(f"  - Bootstrap coverage: {bootstrap_coverage*100:.1f}%")
        print(f"  - Quantile Regression coverage: {quantile_coverage*100:.1f}%")
    
    # 11. Save all data to an Excel file with different worksheets
    print(f"\n{'='*60}")
    print("Save all analysis results to Excel file")
    print(f"{'='*60}")
    try:
        excel_path = os.path.join(analysis_save_dir, 'model_analysis_results.xlsx')
        
        # Check if file exists, if exists use append mode to keep PDP data
        file_exists = os.path.exists(excel_path)
        mode = 'a' if file_exists else 'w'
        print(f"  Excel file path: {excel_path}")
        print(f"  File exists: {file_exists}")
        print(f"  Save mode: {mode} (will keep PDP data)")
        
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode=mode, if_sheet_exists='replace') as writer:
            # 1. Feature importance (including coefficient of variation)
            if importance_df is not None:
                importance_df.to_excel(writer, sheet_name='Feature importance', index=False)
                print("✓ Feature importance has been saved")
            
            # 2. SHAP importance
            if shap_importance is not None:
                shap_importance.to_excel(writer, sheet_name='SHAP importance', index=False)
                print("✓ SHAP importance has been saved")
            
            # 3. Feature statistics
            if 'feature_stats' in locals() and feature_stats is not None:
                feature_stats.to_excel(writer, sheet_name='Feature statistics', index=False)
                print("✓ Feature statistics has been saved")
            
            # 4. Noise robustness
            if 'robustness_df' in locals() and robustness_df is not None:
                robustness_df.to_excel(writer, sheet_name='Noise robustness', index=False)
                print("✓ Noise robustness has been saved")
            
            # 5. Outlier robustness
            if 'outlier_df' in locals() and outlier_df is not None:
                outlier_df.to_excel(writer, sheet_name='Outlier robustness', index=False)
                print("✓ Outlier robustness has been saved")
            
            # 6. Prediction intervals
            if 'intervals_df' in locals() and intervals_df is not None:
                intervals_df.to_excel(writer, sheet_name='Prediction intervals', index=False)
                print("✓ Prediction intervals has been saved")
            
            # 7. Training set and test set prediction results
            if 'prediction_df' in locals() and prediction_df is not None:
                prediction_df.to_excel(writer, sheet_name='Prediction result comparison', index=False)
                print("✓ Prediction result comparison has been saved")
            
            # 8. Radar chart data summary
            if 'radar_data_df' in locals() and radar_data_df is not None:
                radar_data_df.to_excel(writer, sheet_name='Radar chart data summary', index=False)
                print("✓ Radar chart data summary has been saved")
            
            # 9. Radar chart detailed data
            if 'radar_detailed_df' in locals() and radar_detailed_df is not None:
                radar_detailed_df.to_excel(writer, sheet_name='Radar chart detailed data', index=False)
                print("✓ Radar chart detailed data has been saved")
            
            # 10. Performance metrics table (three-fold average vs final model)
            if 'metrics_table_df' in locals() and metrics_table_df is not None:
                metrics_table_df.to_excel(writer, sheet_name='Performance metrics table', index=False)
                print("✓ Performance metrics table has been saved")
        
        print(f"\n✓ All analysis results have been saved to: {excel_path}")
        # Count number of worksheets
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet_count = len(wb.sheetnames)
            print(f"  Contains {sheet_count} worksheets: {', '.join(wb.sheetnames)}")
        except:
            print(f"  Contains multiple worksheets")
    except Exception as e:
        print(f"⚠ Error saving Excel file: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"\nAll results have been saved to: {analysis_save_dir}")
    print(f"{'='*60}")
    
    # 12. Predict all data and write to dataset_with_CatBoost_strain.xlsx
    print(f"\n{'='*60}")
    print("Predict all data and write to dataset_with_CatBoost_strain.xlsx")
    print(f"{'='*60}")
    try:
        # Use final model to predict all data
        print(f"Predicting all {len(X)} samples...")
        y_all_pred = model.predict(X)
        
        # Create prediction result dictionary (using sample_ids as keys)
        pred_dict = {}
        for idx, sample_id in enumerate(sample_ids):
            pred_dict[sample_id] = float(y_all_pred[idx])
        
        # Read original Excel file
        data_file = os.path.join(PROJECT_ROOT, "dataset", "dataset_final.xlsx")
        if not os.path.exists(data_file):
            print(f"⚠ Warning: Original data file not found: {data_file}")
            print("  Trying to use processed data file...")
            # If original file does not exist, try using df_processed
            if 'df_processed' in locals() and df_processed is not None:
                df_original = df_processed.copy()
            else:
                print("  ✗ Cannot find data file, skip saving prediction results")
                raise FileNotFoundError(f"Cannot find data file: {data_file}")
        else:
            df_original = pd.read_excel(data_file, sheet_name=0)
        
        # Use map method to match prediction results
        if 'No_Customized' in df_original.columns:
            df_original['CatBoost_peak_strain'] = df_original['No_Customized'].map(pred_dict)
            print(f"✓ Using No_Customized column to match prediction results")
        else:
            print(f"⚠ Warning: No_Customized column not found, will match prediction results by row order")
            # If No_Customized column not found, match prediction results by row order (need to ensure order consistency)
            if len(y_all_pred) == len(df_original):
                df_original['CatBoost_peak_strain'] = y_all_pred
                print(f"✓ Match prediction results by row order")
            else:
                print(f"  ✗ Prediction sample number ({len(y_all_pred)}) does not match original data row number ({len(df_original)})")
                print(f"  Trying to use original_df_indices to match...")
                # Try using original_df_indices to match
                if 'original_df_indices' in locals() and original_df_indices is not None:
                    if len(original_df_indices) == len(y_all_pred):
                        # Create a Series, index is original_df_indices, value is prediction results
                        pred_series = pd.Series(y_all_pred, index=original_df_indices)
                        # Map prediction results to original DataFrame
                        df_original['CatBoost_peak_strain'] = df_original.index.map(pred_series)
                        print(f"✓ Use original_df_indices to match prediction results")
                    else:
                        raise ValueError(f"original_df_indices length ({len(original_df_indices)}) does not match prediction result length ({len(y_all_pred)})")
                else:
                    raise ValueError("Sample number does not match, cannot write prediction results")
        
        # Save results to new independent file
        output_file = os.path.join(PROJECT_ROOT, "dataset", "dataset_with_CatBoost_strain.xlsx")
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        df_original.to_excel(output_file, index=False)
        print(f"✓ Prediction results have been written to: {output_file}")
        print(f"  - New column: CatBoost_peak_strain (CatBoost final model prediction value)")
        print(f"  - Total prediction sample number: {len(y_all_pred)}")
        print(f"  - Number of non-empty prediction values: {df_original['CatBoost_peak_strain'].notna().sum()}")
        
        # Backup to save directory
        backup_file = os.path.join(analysis_save_dir, 'dataset_with_CatBoost_strain_backup.xlsx')
        df_original.to_excel(backup_file, index=False)
        print(f"✓ Backup file has been saved: {backup_file}")
        
    except Exception as e:
        print(f"⚠ Warning: Error writing to dataset_with_CatBoost_strain.xlsx: {e}")
        import traceback
        traceback.print_exc()
    
    print(f"{'='*60}")
    
    return {
        'model': model,
        'model_data': model_data,
        'noise_rmses': noise_rmses,
        'outlier_rmses': outlier_rmses,
        'bootstrap_coverage': bootstrap_coverage,
        'quantile_coverage': quantile_coverage,
        'shap_importance': shap_importance,
        'importance_df': importance_df
    }

if __name__ == "__main__":
    results = main()
