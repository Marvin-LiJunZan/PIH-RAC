#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
PINN Peak Strain Prediction Model - Inference and Batch Comparison Analysis Script
=================================================================================

Functionality Overview:
-----------------------
This script integrates model inference, batch comparison analysis, interpretability analysis, and noise analysis capabilities:
1. Load trained PINN model for prediction
2. Calculate predicted values using Xiao's formula
3. Batch compare existing prediction columns in dataset (e.g., Xiao_strain, XGB_Xiao_strain, etc.)
4. SHAP interpretability analysis (feature importance, SHAP value analysis)
5. Noise analysis and uncertainty analysis (Bootstrap prediction intervals)

Core Features:
--------------
1. Model Inference (Optional):
   - Load trained PINN model (weights, configuration, preprocessors)
   - Perform batch prediction on dataset
   - Calculate predicted values using Xiao's formula

2. Batch Comparison Analysis:
   - Compare existing prediction columns in dataset with true values
   - Support custom comparison column lists

3. Interpretability Analysis (Optional):
   - SHAP value analysis (summary plot, bar plot, waterfall plot)
   - Feature importance analysis
   - Partial Dependence Plot (PDP) analysis

4. Noise Analysis and Uncertainty Analysis (Optional):
   - Bootstrap prediction intervals
   - Model robustness testing

5. Unified Performance Metrics:
   - R², EVS, MAE, MSE, RMSE, MedAE, MAPE, MPE, Max Error
   - Correlation coefficient, median absolute error

✅ Output Files:
   - peak_strain/SAVE/pinn_analysis_results.xlsx: All analysis results (multiple worksheets)
     * PINN performance metrics
     * Xiao's formula performance metrics
     * SHAP analysis results
     * Uncertainty analysis results
   - dataset/dataset_with_PINN_peak_strain.xlsx: Original data + PINN prediction column
   - peak_strain/SAVE/shap_*.png: SHAP analysis plots (if SHAP is available)

Usage:
------
# Run directly (automatically loads model and data, similar to XGBoost noise analysis.py)
python peak_strain_inference.py

The script will automatically:
- Load trained model from peak_strain/SAVE/
- Load data from dataset/dataset_final.xlsx
- Perform PINN and Xiao's formula predictions
- Conduct SHAP interpretability analysis
- Perform uncertainty analysis (Bootstrap prediction intervals)
- Save all results to multiple worksheets in Excel
- Save prediction results to dataset_with_PINN_peak_strain.xlsx

Input Requirements:
-------------------
1. Model directory (peak_strain/SAVE/) must contain:
   - pinn_peak_strain.pt (model weights)
   - training_summary.json (training configuration)
   - scalers.pkl (data preprocessors)
   - model_architecture.json (model architecture)

2. Dataset (dataset/dataset_final.xlsx) must contain:
   - peak_strain (true strain column, required)
   - fc, r (required for calculating xiao_formula)
   - Other feature columns (required)
   - Note: If xiao_formula column is missing, the script will calculate it automatically
"""

from __future__ import annotations
import argparse
import json
import logging
import os
import pickle
import sys
from pathlib import Path
from typing import Any, Dict, Optional

import numpy as np
import pandas as pd
from sklearn.metrics import (
    explained_variance_score,
    max_error,
    mean_absolute_error,
    mean_absolute_percentage_error,
    mean_squared_error,
    median_absolute_error,
    r2_score,
)
from sklearn.preprocessing import MinMaxScaler
from sklearn.inspection import partial_dependence
import torch
import torch.nn as nn
import matplotlib
matplotlib.use('Agg')   # Non-interactive backend
import matplotlib.pyplot as plt
import warnings
warnings.filterwarnings('ignore')

# Try importing SHAP (optional dependency)
try:
    import shap
    SHAP_AVAILABLE = True
except ImportError:
    SHAP_AVAILABLE = False
     logging.warning("SHAP is not installed, interpretability analysis features will be unavailable. Installation command: pip install shap")

# ---------------------------------------------------------------------------
# Configure Logging
# ---------------------------------------------------------------------------

def configure_logging() -> None:
    """Configure logging output"""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
        force=True,
    )


# ---------------------------------------------------------------------------
# Model Definition (consistent with training script)
# ---------------------------------------------------------------------------

class PINNRegressor(nn.Module):
    """PINN regression model (same definition as training script)."""

    def __init__(
        self,
        input_dim: int,
        hidden_dim: int,
        num_layers: int,
        dropout: float = 0.0,
    ) -> None:
        super().__init__()

        activation = nn.ReLU()
        layers: list[nn.Module] = []
        in_dim = input_dim

        for _ in range(num_layers):
            layers.append(nn.Linear(in_dim, hidden_dim))
            layers.append(activation)
            if dropout > 0:
                layers.append(nn.Dropout(dropout))
            in_dim = hidden_dim

        layers.append(nn.Linear(in_dim, 1))
        layers.append(nn.Sigmoid())
        self.model = nn.Sequential(*layers)

    def forward(self, x: torch.Tensor) -> torch.Tensor:
        return self.model(x)


# ---------------------------------------------------------------------------
# Xiao's Formula (consistent with training script)
# ---------------------------------------------------------------------------

def compute_xiao_formula(fc: np.ndarray, r: np.ndarray) -> np.ndarray:
    """Xiao's formula calculation (exact implementation as training script).
    
    Formula: ε_cp = {0.00076 + [(0.626σ_cp - 4.33) × 10^-7]^0.5} × (1 + r / (65.715r^2 - 109.43r + 48.989))
    
    Args:
        fc: Stress value (compressive strength)
        r: Aggregate replacement rate (stored as percentage×100 in data, e.g., 1.5 represents 1.5%)
    
    Returns:
        Peak strain predicted by Xiao's formula
    """
    # r = aggregate replacement rate, convert from percentage×100 (e.g., 1.5) to decimal (0.015)
    r = r / 100.0
    
    fc_clamped = np.clip(fc.astype(float), a_min=1e-6, a_max=None)
    r_clamped = np.clip(r, a_min=1e-8, a_max=None)
    
    # First term: 0.00076 + sqrt((0.626 * σ_cp - 4.33) × 10^-7)
    inner = (0.626 * fc_clamped - 4.33) * 1e-7
    inner_clamped = np.clip(inner, a_min=0.0, a_max=None)
    term1 = 0.00076 + np.sqrt(inner_clamped)

    # Second term: 1 + r / (65.715r^2 - 109.43r + 48.989)
    denom = 65.715 * r_clamped ** 2 - 109.43 * r_clamped + 48.989
    term2 = 1.0 + (r_clamped / denom)

    return term1 * term2


# ---------------------------------------------------------------------------
# Model Loading
# ---------------------------------------------------------------------------

def load_trained_model(model_dir: str, device: str = "cpu") -> Dict[str, Any]:
     """Load trained model and configuration.
    
    Returns:
        Dictionary containing:
        - model: Model with loaded weights
        - feature_scaler: Feature standardizer
        - target_scaler: Target standardizer
        - feature_cols: List of feature column names
        - train_config: Training configuration
        - base_config: Base configuration
    """
    model_dir = Path(model_dir)
    
    if not model_dir.exists():
        raise FileNotFoundError(f"Model directory does not exist: {model_dir}")
    
    # 1. Load training summary (contains configuration and hyperparameters)
    summary_path = model_dir / "training_summary.json"
    if not summary_path.exists():
        raise FileNotFoundError(f"Training summary file not found: {summary_path}")
    
    with open(summary_path, "r", encoding="utf-8") as f:
        summary = json.load(f)
    
    train_config = summary["train_config"]
    base_config = summary["base_config"]
    
    logging.info("✓ Loaded training configuration")
    logging.info("  - Best validation R² = %.4f", summary["validation_metrics"]["r2"])
    logging.info("  - Best epoch = %d", summary["best_epoch"])
    
    # 2. Load model architecture
    arch_path = model_dir / "model_architecture.json"
    if not arch_path.exists():
        raise FileNotFoundError(f"Model architecture file not found: {arch_path}")
    
    with open(arch_path, "r", encoding="utf-8") as f:
        model_arch = json.load(f)
    
    # 3. Load data preprocessors
    scaler_path = model_dir / "scalers.pkl"
    if not scaler_path.exists():
        raise FileNotFoundError(f"Data preprocessor file not found: {scaler_path}")
    
    with open(scaler_path, "rb") as f:
        scaler_data = pickle.load(f)
    
    feature_scaler = scaler_data["feature_scaler"]
    target_scaler = scaler_data["target_scaler"]
    feature_cols = scaler_data["feature_cols"]
    
    logging.info("✓ Loaded data preprocessors")
    logging.info("  - Number of features = %d", len(feature_cols))
    logging.info("  - Feature columns: %s", ", ".join(feature_cols))
    
    # 4. Reconstruct model and load weights
    # If input_dim is 0 or missing, infer from number of feature columns
    input_dim = model_arch.get("input_dim", 0)
    if input_dim == 0 or input_dim is None:
        input_dim = len(feature_cols)
        logging.warning(f"input_dim is 0 in model architecture, inferring from feature columns: {input_dim}")
    
    model = PINNRegressor(
        input_dim=input_dim,
        hidden_dim=model_arch["hidden_dim"],
        num_layers=model_arch["num_layers"],
        dropout=model_arch["dropout"],
    )
    
    model_path = model_dir / "pinn_peak_strain.pt"
    if not model_path.exists():
        raise FileNotFoundError(f"Model weight file not found: {model_path}")
    
    model.load_state_dict(torch.load(model_path, map_location=device))
    model = model.to(device)
    model.eval()
    
    logging.info("✓ Loaded model weights")
    logging.info("  - Network structure: %d layers × %d dimensions (dropout=%.2f)", 
                model_arch["num_layers"], model_arch["hidden_dim"], model_arch["dropout"])
    
    return {
        "model": model,
        "feature_scaler": feature_scaler,
        "target_scaler": target_scaler,
        "feature_cols": feature_cols,
        "train_config": train_config,
        "base_config": base_config,
        "summary": summary,
    }


# ---------------------------------------------------------------------------
# Data Loading and Preprocessing
# ---------------------------------------------------------------------------

def load_and_preprocess_data(
    dataset_path: str,
    feature_cols: list[str],
    feature_scaler: MinMaxScaler,
    target_col: str = "peak_strain",
    stress_col: str = "fc",
) -> Dict[str, Any]:
     """Load dataset and perform preprocessing.
    
    Returns:
        Dictionary containing:
        - df: Original DataFrame
        - features_scaled: Standardized features
        - features_raw: Raw features
        - target: Target values (true strain)
        - stress: Stress values (fc)
        - r_values: Aggregate replacement rates
    """
     if not os.path.exists(dataset_path):
        raise FileNotFoundError(f"Data file does not exist: {dataset_path}")
    
    df = pd.read_excel(dataset_path)
    logging.info("✓ Loaded dataset: %s (shape: %s)", Path(dataset_path).name, df.shape)
    
    # If xiao_formula was used as input feature during model training, it's also needed for inference
    # If xiao_formula column is missing in dataset, calculate automatically (consistent with training script)
    if "xiao_formula" in feature_cols and "xiao_formula" not in df.columns:
        logging.info("  Dataset missing xiao_formula column, calculating automatically...")
        if stress_col not in df.columns:
            raise ValueError(f"Dataset missing stress column: {stress_col} (cannot calculate xiao_formula)")
        if "r" not in df.columns:
            raise ValueError("Dataset missing aggregate replacement rate column: r (cannot calculate xiao_formula)")
        df["xiao_formula"] = compute_xiao_formula(df[stress_col].values, df["r"].values)
        logging.info("  ✓ xiao_formula column calculated and added")
    
    # Check required columns
    missing_cols = [col for col in feature_cols if col not in df.columns]
    if missing_cols:
        raise ValueError(f"Dataset missing feature columns: {missing_cols}")
    
    if target_col not in df.columns:
        raise ValueError(f"Dataset missing target column: {target_col}")
    
    if stress_col not in df.columns:
        raise ValueError(f"Dataset missing stress column: {stress_col}")
    
    if "r" not in df.columns:
        raise ValueError("Dataset missing aggregate replacement rate column: r")
    
    # Extract data
    features_raw = df[feature_cols].values.astype(np.float32)
    target = df[target_col].values.astype(np.float32)
    stress = df[stress_col].values.astype(np.float32)
    r_values = df["r"].values.astype(np.float32)
    
    # Standardize features
    features_scaled = feature_scaler.transform(features_raw).astype(np.float32)
    
    logging.info("✓ Data preprocessing completed")
    logging.info("  - Number of samples = %d", len(df))
    logging.info("  - Feature dimension = %d", features_raw.shape[1])
    
    return {
        "df": df,
        "features_scaled": features_scaled,
        "features_raw": features_raw,
        "target": target,
        "stress": stress,
        "r_values": r_values,
    }


# ---------------------------------------------------------------------------
# Model Inference
# ---------------------------------------------------------------------------

@torch.no_grad()
def predict_with_pinn(
    model: nn.Module,
    features_scaled: np.ndarray,
    target_scaler: MinMaxScaler,
    device: str = "cpu",
) -> np.ndarray:
     """Perform prediction using PINN model.
    
    Returns:
        Denormalized predictions (original scale)
    """
    model.eval()
    features_tensor = torch.from_numpy(features_scaled).to(device)
    
    preds_scaled = model(features_tensor).cpu().numpy().flatten()
    preds = target_scaler.inverse_transform(preds_scaled.reshape(-1, 1)).flatten()
    
    return preds


# ---------------------------------------------------------------------------
#  Performance Evaluation
# ---------------------------------------------------------------------------

def compute_metrics(y_true: np.ndarray, y_pred: np.ndarray, include_extended: bool = False) -> Dict[str, float]:
    """Calculate regression performance metrics.
    
    Args:
        y_true: True values
        y_pred: Predicted values
        include_extended: Whether to include extended metrics (MPE, correlation coefficient, etc.)
    
    Returns:
        Dictionary containing performance metrics
    """
    metrics = {
        "r2": float(r2_score(y_true, y_pred)),
        "evs": float(explained_variance_score(y_true, y_pred)),
        "mae": float(mean_absolute_error(y_true, y_pred)),
        "mse": float(mean_squared_error(y_true, y_pred)),
        "rmse": float(np.sqrt(mean_squared_error(y_true, y_pred))),
        "medae": float(median_absolute_error(y_true, y_pred)),
        "mape": float(mean_absolute_percentage_error(y_true, y_pred)) * 100,  # Convert to percentage
        "max_error": float(max_error(y_true, y_pred)),
    }
    
    if include_extended:
        # MPE (Mean Percentage Error) = mean((pred - true) / true) * 100
        metrics["mpe"] = float(np.mean((y_pred - y_true) / np.abs(y_true)) * 100)
        # Correlation coefficient
        metrics["correlation"] = float(np.corrcoef(y_true, y_pred)[0, 1])
    
    return metrics


def print_comparison_table(
    pinn_metrics: Dict[str, float],
    xiao_metrics: Dict[str, float],
) -> None:
    """Print performance comparison table for PINN and Xiao's formula."""
    logging.info("\n" + "="*100)
    logging.info("PINN vs Xiao's Formula - Performance Comparison")
    logging.info("="*100)
    logging.info("%-12s | %-8s | %-8s | %-10s | %-10s | %-10s | %-10s",
                "模型", "R²", "EVS", "MAE", "MSE", "RMSE", "MAPE(%)")
    logging.info("-" * 100)
    
    logging.info("%-12s | %-8.4f | %-8.4f | %-10.3e | %-10.3e | %-10.3e | %-10.2f",
                "PINN",
                pinn_metrics["r2"], pinn_metrics["evs"], pinn_metrics["mae"],
                pinn_metrics["mse"], pinn_metrics["rmse"], pinn_metrics["mape"])
    
    logging.info("%-12s | %-8.4f | %-8.4f | %-10.3e | %-10.3e | %-10.3e | %-10.2f",
                "Xiao公式",
                xiao_metrics["r2"], xiao_metrics["evs"], xiao_metrics["mae"],
                xiao_metrics["mse"], xiao_metrics["rmse"], xiao_metrics["mape"])
    
    logging.info("-" * 100)
    
     # Calculate improvement percentage
    r2_improve = (pinn_metrics["r2"] - xiao_metrics["r2"]) / abs(xiao_metrics["r2"]) * 100 if xiao_metrics["r2"] != 0 else 0
    mae_improve = (xiao_metrics["mae"] - pinn_metrics["mae"]) / xiao_metrics["mae"] * 100
    rmse_improve = (xiao_metrics["rmse"] - pinn_metrics["rmse"]) / xiao_metrics["rmse"] * 100
    
    logging.info("PINN Improvement: R² increase=%.1f%%, MAE decrease=%.1f%%, RMSE decrease=%.1f%%",
                r2_improve, mae_improve, rmse_improve)
    logging.info("="*100 + "\n")


# ---------------------------------------------------------------------------
# Result Export
# ---------------------------------------------------------------------------

def prepare_predictions_data(
    df: pd.DataFrame,
    pinn_preds: np.ndarray,
    xiao_preds: np.ndarray,
    target_col: str,
) -> pd.DataFrame:
    """Prepare prediction results DataFrame (without direct saving).
    
    Added columns:
    - peak_strain_pinn: PINN model predictions
    - peak_strain_xiao: Xiao's formula predictions
    - error_pinn: PINN prediction error
    - error_xiao: Xiao's prediction error
    - abs_error_pinn: PINN absolute error
    - abs_error_xiao: Xiao's absolute error
    
    Returns:
        DataFrame containing prediction results
    """
    df_out = df.copy()
    
    true_values = df[target_col].values
    
    # Add prediction columns
    df_out[f"{target_col}_pinn"] = pinn_preds
    df_out[f"{target_col}_xiao"] = xiao_preds
    
    # Add error columns
    df_out[f"error_pinn"] = pinn_preds - true_values
    df_out[f"error_xiao"] = xiao_preds - true_values
    
    df_out[f"abs_error_pinn"] = np.abs(pinn_preds - true_values)
    df_out[f"abs_error_xiao"] = np.abs(xiao_preds - true_values)
    
    return df_out


def prepare_comparison_data(
    df: pd.DataFrame,
    pinn_preds: np.ndarray,
    xiao_preds: np.ndarray,
    target_col: str,
) -> pd.DataFrame:
    """Prepare detailed comparison results DataFrame for PINN and Xiao's formula (without direct saving).
    
    Returns:
        DataFrame containing comparison results
    """
    true_values = df[target_col].values
    
    comparison_df = pd.DataFrame({
        "Sample_ID": df.get("No_Customized", range(len(df))),
        "True_Strain": true_values,
        "PINN_Prediction": pinn_preds,
        "Xiao_Prediction": xiao_preds,
        "PINN_Error": pinn_preds - true_values,
        "Xiao_Error": xiao_preds - true_values,
        "PINN_Absolute_Error": np.abs(pinn_preds - true_values),
        "Xiao_Absolute_Error": np.abs(xiao_preds - true_values),
        "PINN_Relative_Error(%)": np.abs((pinn_preds - true_values) / true_values) * 100,
        "Xiao_Relative_Error(%)": np.abs((xiao_preds - true_values) / true_values) * 100,
    })
    
    return comparison_df


def prepare_metrics_data(
    pinn_metrics: Optional[Dict[str, float]],
    xiao_metrics: Optional[Dict[str, float]],
) -> pd.DataFrame:
    """Prepare performance metrics DataFrame (without direct saving).
    
    Returns:
        DataFrame containing performance metrics
    """
    metrics_list = []
    
    if pinn_metrics is not None:
        metrics_list.append({
            "Model": "PINN",
            "R²": pinn_metrics["r2"],
            "EVS": pinn_metrics["evs"],
            "MAE": pinn_metrics["mae"],
            "MSE": pinn_metrics["mse"],
            "RMSE": pinn_metrics["rmse"],
            "MedAE": pinn_metrics["medae"],
            "MAPE (%)": pinn_metrics["mape"],
            "Max_Error": pinn_metrics["max_error"],
        })
    
    if xiao_metrics is not None:
        metrics_list.append({
            "Model": "Xiao's Formula",
            "R²": xiao_metrics["r2"],
            "EVS": xiao_metrics["evs"],
            "MAE": xiao_metrics["mae"],
            "MSE": xiao_metrics["mse"],
            "RMSE": xiao_metrics["rmse"],
            "MedAE": xiao_metrics["medae"],
            "MAPE (%)": xiao_metrics["mape"],
            "Max_Error": xiao_metrics["max_error"],
        })
    
    if pinn_metrics is not None and xiao_metrics is not None:
        metrics_list.append({
            "Model": "Improvement",
            "R²": pinn_metrics["r2"] - xiao_metrics["r2"],
            "EVS": pinn_metrics["evs"] - xiao_metrics["evs"],
            "MAE": xiao_metrics["mae"] - pinn_metrics["mae"],
            "MSE": xiao_metrics["mse"] - pinn_metrics["mse"],
            "RMSE": xiao_metrics["rmse"] - pinn_metrics["rmse"],
            "MedAE": xiao_metrics["medae"] - pinn_metrics["medae"],
            "MAPE (%)": xiao_metrics["mape"] - pinn_metrics["mape"],
            "Max_Error": xiao_metrics["max_error"] - pinn_metrics["max_error"],
        })
    
    return pd.DataFrame(metrics_list)


def save_metrics_json(
    pinn_metrics: Optional[Dict[str, float]],
    xiao_metrics: Optional[Dict[str, float]],
    output_path: str,
) -> None:
    """Save performance metrics to JSON file (preserve JSON format for programmatic reading)."""
    metrics_data = {}
    
    if pinn_metrics is not None:
        metrics_data["pinn"] = pinn_metrics
    if xiao_metrics is not None:
        metrics_data["xiao_formula"] = xiao_metrics
    
    if pinn_metrics is not None and xiao_metrics is not None:
        metrics_data["improvement"] = {
            "r2_delta": pinn_metrics["r2"] - xiao_metrics["r2"],
            "mae_delta": xiao_metrics["mae"] - pinn_metrics["mae"],
            "rmse_delta": xiao_metrics["rmse"] - pinn_metrics["rmse"],
        }
    
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(metrics_data, f, ensure_ascii=False, indent=2)
    
    logging.info("✓ Performance metrics JSON saved: %s", output_path)


# ---------------------------------------------------------------------------
# Batch Comparison Analysis
# ---------------------------------------------------------------------------

def batch_compare_columns(
    df: pd.DataFrame,
    true_col: str = "peak_strain",
    compare_cols: Optional[list[str]] = None,
    default_compare_cols: Optional[list[str]] = None,
) -> pd.DataFrame:
     """Batch compare performance metrics of specified columns with true values in dataset.
    
    Args:
        df: Dataset DataFrame
        true_col: True value column name
        compare_cols: List of column names to compare (use default_compare_cols if None)
        default_compare_cols: Default list of comparison columns
    
    Returns:
        DataFrame containing all comparison results, sorted by R² descending
    """
    if default_compare_cols is None:
        default_compare_cols = [
            "Xiao_strain",
            "XGB_Xiao_strain",
            "True_Yan_strain",
            "True_Belen_strain",
            "peak_strain_pinn_pred",
            "peak_strain_pinn_pred_xgbfc",
        ]
    
    if compare_cols is None:
        compare_cols = default_compare_cols
    
    # Check if the true value column exists
    if true_col not in df.columns:
        raise ValueError(f"数据集中不存在真实值列: {true_col}")
    
    # Get the true values
    true_values = df[true_col].values
    mask = ~np.isnan(true_values)
    true_values = true_values[mask]
    
    logging.info(f"✓ Valid sample number: {len(true_values)}")
    
    # Check which columns exist
    existing_cols = [col for col in compare_cols if col in df.columns]
    missing_cols = [col for col in compare_cols if col not in df.columns]
    
    if missing_cols:
        logging.warning(f"The following columns do not exist in the dataset: {missing_cols}")
    if not existing_cols:
        raise ValueError("No columns found to compare!")
    
    logging.info(f"找到 {len(existing_cols)} 个需要对比的列: {existing_cols}")
    
    # Store all metrics results
    results = []
    
    for col in existing_cols:
        pred_values = df[col].values[mask]
        
        # Check if there are valid values
        valid_mask = ~np.isnan(pred_values)
        if valid_mask.sum() == 0:
            logging.warning(f"{col} column has no valid values, skipping")
            continue
        
        true_vals = true_values[valid_mask]
        pred_vals = pred_values[valid_mask]
        
        # Avoid division by zero error
        if len(true_vals) == 0 or np.std(true_vals) == 0:
            logging.warning(f"{col} column data is invalid, skipping")
            continue
        
        try:
            # Calculate all metrics (including extended metrics)
            metrics = compute_metrics(true_vals, pred_vals, include_extended=True)
            
            results.append({
                "Column Name": col,
                "R²": metrics["r2"],
                "RMSE": metrics["rmse"],
                "MSE": metrics["mse"],
                "MAE": metrics["mae"],
                "MAPE (%)": metrics["mape"],
                "MPE (%)": metrics.get("mpe", np.nan),
                "EVS": metrics["evs"],
                "Max Error": metrics["max_error"],
                "MedAE": metrics["medae"],
                "Correlation": metrics.get("correlation", np.nan),
                "Valid Sample Number": len(true_vals),
            })
            
            logging.info(f"✓ {col}: R²={metrics['r2']:.4f}, RMSE={metrics['rmse']:.6f}, MAPE={metrics['mape']:.2f}%")
            
        except Exception as e:
            logging.error(f"✗ {col} calculation failed: {e}")
            continue
    
    # Create result DataFrame
    results_df = pd.DataFrame(results)
    
    if len(results_df) == 0:
        logging.warning("No metrics were successfully calculated for any column!")
        return pd.DataFrame()
    
    # Sort by R² descending
    results_df = results_df.sort_values("R²", ascending=False)
    
    return results_df


def print_batch_comparison_summary(results_df: pd.DataFrame) -> None:
    """Print batch comparison result summary."""
    if len(results_df) == 0:
        logging.warning("No results to display")
        return
    
    # Print top 10
    logging.info(f"\nTop 10 (sorted by R²):")
    display_cols = ["Column Name", "R²", "RMSE", "MAPE (%)", "Correlation"]
    available_cols = [col for col in display_cols if col in results_df.columns]
    logging.info(f"\n{results_df[available_cols].head(10).to_string(index=False)}")


# ---------------------------------------------------------------------------
# Interpretability Analysis (SHAP)
# ---------------------------------------------------------------------------

def predict_wrapper(model: nn.Module, target_scaler: MinMaxScaler, device: str):
    """Create a prediction wrapper function for SHAP."""
    def predict_fn(X: np.ndarray) -> np.ndarray:
        """SHAP compatible prediction function."""
        model.eval()
        with torch.no_grad():
            X_tensor = torch.from_numpy(X.astype(np.float32)).to(device)
            preds_scaled = model(X_tensor).cpu().numpy()
            preds = target_scaler.inverse_transform(preds_scaled).flatten()
        return preds
    return predict_fn


def plot_shap_analysis(
    model: nn.Module,
    X_background: np.ndarray,
    X_explain: np.ndarray,
    feature_names: list[str],
    target_scaler: MinMaxScaler,
    device: str,
    save_dir: Path,
    max_background_samples: int = 100,
) -> None:
    """SHAP value analysis (suitable for neural network models).
    
    Args:
        model: Trained PINN model
        X_background: Background dataset (used to initialize explainer, usually using training set)
        X_explain: Dataset to explain (data for calculating SHAP values, usually using test set)
        feature_names: List of feature names
        target_scaler: Target value scaler (used for inverse transformation)
        device: Running device
        save_dir: Save directory
        max_background_samples: Maximum number of background dataset samples (used for acceleration)
    """
    if not SHAP_AVAILABLE:
        logging.warning("SHAP is not installed, interpretability analysis will be skipped. Installation command: pip install shap")
        return
    
    # Set matplotlib font (ensure Chinese and negative sign are displayed correctly, set at the beginning of the function)
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    
    logging.info("\n" + "="*80)
    logging.info("Performing SHAP interpretability analysis")
    logging.info("="*80)
    logging.info(f"Background dataset sample number: {len(X_background)} (used to initialize explainer)")
    logging.info(f"Explanation dataset sample number: {len(X_explain)} (used to calculate SHAP values)")
    
    # Sample background data to improve speed
    if len(X_background) > max_background_samples:
        indices = np.random.choice(len(X_background), max_background_samples, replace=False)
        X_background_sample = X_background[indices]
        logging.info(f"Background dataset sampled to {max_background_samples} samples")
    else:
        X_background_sample = X_background
    
    # Create prediction wrapper function
    predict_fn = predict_wrapper(model, target_scaler, device)
    
    # Initialize SHAP explainer (using KernelExplainer, suitable for any model)
    try:
        logging.info("初始化SHAP KernelExplainer...")
        explainer = shap.KernelExplainer(predict_fn, X_background_sample)
        logging.info("✓ SHAP explainer initialized successfully")
    except Exception as e:
        logging.error(f"✗ SHAP explainer initialization failed: {e}")
        return
    
    # Calculate SHAP values (for explanation dataset)
    logging.info("Calculating SHAP values (this may take some time)...")
    try:
        # Limit explanation sample number to improve speed
        max_explain_samples = min(100, len(X_explain))
        if len(X_explain) > max_explain_samples:
            indices = np.random.choice(len(X_explain), max_explain_samples, replace=False)
            X_explain_sample = X_explain[indices]
            logging.info(f"Explanation dataset sampled to {max_explain_samples} samples")
        else:
            X_explain_sample = X_explain
        
        shap_values = explainer.shap_values(X_explain_sample, nsamples=100)
        logging.info("✓ SHAP values calculated successfully")
    except Exception as e:
        logging.error(f"✗ SHAP values calculation failed: {e}")
        return
    
    # Calculate feature importance (average absolute SHAP values)
    shap_importance_values = np.abs(shap_values).mean(axis=0)
    
    # Set importance threshold (filter out features with importance 0 or close to 0)
    importance_threshold = 1e-6
    important_feature_indices = np.where(shap_importance_values > importance_threshold)[0]
    
    logging.info(f"\nFeature importance statistics:")
    logging.info(f"  Total number of features: {len(feature_names)}")
    logging.info(f"  Number of features with importance > 0: {len(important_feature_indices)}")  # Fix negative sign display problem
    
    # 1. Summary plot (full version)
    try:
        logging.info("Generating SHAP summary plot...")
        fig = plt.figure(figsize=(30, 12))
        shap.summary_plot(shap_values, X_explain_sample, feature_names=feature_names, show=False)
        plt.tight_layout()
        plt.savefig(save_dir / 'shap_summary.png', dpi=300, bbox_inches='tight')
        plt.close()
        logging.info("✓ SHAP summary plot saved")
    except Exception as e:
        logging.warning(f"Summary plot generation failed: {e}")
    
    # 1.1 Summary plot (simplified version - only includes features with importance > 0)
    if len(important_feature_indices) < len(feature_names):
        try:
            logging.info("Generating simplified SHAP summary plot...")
            shap_values_filtered = shap_values[:, important_feature_indices]
            X_explain_filtered = X_explain_sample[:, important_feature_indices]
            feature_names_filtered = [feature_names[i] for i in important_feature_indices]
            
            fig = plt.figure(figsize=(30, 12))
            shap.summary_plot(shap_values_filtered, X_explain_filtered, 
                            feature_names=feature_names_filtered, show=False)
            plt.tight_layout()
            plt.savefig(save_dir / 'shap_summary_filtered.png', dpi=300, bbox_inches='tight')
            plt.close()
            logging.info("✓ Simplified summary plot saved")
        except Exception as e:
            logging.warning(f"Simplified summary plot generation failed: {e}")
    
    # 2. Bar plot (full version)
    try:
        logging.info("Generating SHAP bar plot...")
        plt.figure(figsize=(10, 8))
        shap.summary_plot(shap_values, X_explain_sample, feature_names=feature_names, 
                         plot_type="bar", show=False)
        plt.tight_layout()
        plt.savefig(save_dir / 'shap_bar.png', dpi=300, bbox_inches='tight')
        plt.close()
        logging.info("✓ SHAP bar plot saved")
    except Exception as e:
        logging.warning(f"Bar plot generation failed: {e}")
    
    # 2.1 Bar plot (simplified version - with value labels)
    if len(important_feature_indices) < len(feature_names):
        try:
            logging.info("Generating simplified SHAP bar plot (with value labels)...")
            shap_values_filtered = shap_values[:, important_feature_indices]
            X_explain_filtered = X_explain_sample[:, important_feature_indices]
            feature_names_filtered = [feature_names[i] for i in important_feature_indices]
            
            plt.figure(figsize=(10, 8))
            shap.summary_plot(shap_values_filtered, X_explain_filtered, 
                            feature_names=feature_names_filtered, plot_type="bar", show=False)
            
            # 添加数值标签
            ax = plt.gca()
            bars = ax.patches
            mean_abs_shap = np.abs(shap_values_filtered).mean(axis=0)
            sorted_indices = np.argsort(mean_abs_shap)[::-1]
            mean_abs_shap_sorted = mean_abs_shap[sorted_indices]
            
            for i, (bar, value) in enumerate(zip(bars, mean_abs_shap_sorted)):
                width = bar.get_width()
                y_pos = bar.get_y() + bar.get_height() / 2
                ax.text(width, y_pos, f'{value:.3f}', 
                       ha='left', va='center', fontsize=9, fontweight='bold')
            
            plt.tight_layout()
            plt.savefig(save_dir / 'shap_bar_filtered.png', dpi=300, bbox_inches='tight')
            plt.close()
            logging.info("✓ Simplified bar plot saved")
        except Exception as e:
            logging.warning(f"Simplified bar plot generation failed: {e}")
    
    # 3. Waterfall plot (display SHAP values for single sample)
    try:
        logging.info("Generating SHAP waterfall plot...")
        n_samples_waterfall = min(3, len(shap_values))
        
        # Check SHAP value range, used to set appropriate display format
        shap_min = np.min(shap_values)
        shap_max = np.max(shap_values)
        shap_abs_max = max(abs(shap_min), abs(shap_max))
        logging.info(f"  SHAP value range: [{shap_min:.6e}, {shap_max:.6e}], maximum absolute value: {shap_abs_max:.6e}")
        
        # Determine display format based on SHAP value range
        if shap_abs_max < 0.001:
            # Use scientific notation
            fmt_str = '{:.2e}'
            use_scientific = True
        elif shap_abs_max < 0.01:
            # Use 4 decimal places
            fmt_str = '{:.4f}'
            use_scientific = False
        else:
            # Use 3 decimal places
            fmt_str = '{:.3f}'
            use_scientific = False
        
        logging.info(f"  Using display format: {fmt_str}")
        
        for sample_idx in range(n_samples_waterfall):
            sample_shap = shap_values[sample_idx]
            shap_explanation = shap.Explanation(
                values=shap_values[sample_idx:sample_idx+1],
                base_values=explainer.expected_value,
                data=X_explain_sample[sample_idx:sample_idx+1],
                feature_names=feature_names
            )
            
            # Create figure and set font and negative sign display
            fig = plt.figure(figsize=(12, 8))
            
            try:
                # Generate waterfall plot
                shap.plots.waterfall(shap_explanation[0], show=False, max_display=20)
                
                # Get current axes
                ax = plt.gca()
                
                # Manually fix all text labels
                # Get SHAP values for this sample
                sample_shap_dict = {feature_names[i]: sample_shap[i] for i in range(len(feature_names))}
                
                # Iterate through all text objects and update
                for text_obj in ax.texts:
                    try:
                        text_str = text_obj.get_text().strip()
                        
                        # Check if it is a SHAP value label (usually "+0", "-0", "0", "+0.000", "-0.000" etc.)
                        if text_str in ['+0', '-0', '0', '+0.0', '-0.0', '0.0', '+0.00', '-0.00', '0.00']:
                            # Try to infer the corresponding feature from text position
                            # Since the text position of SHAP waterfall is complex, we use another method
                            # Directly delete these texts, and add them back later
                            text_obj.set_text('')
                        elif text_str.startswith('+') or text_str.startswith('-'):
                            # Try to parse existing value and reformat
                            try:
                                val = float(text_str)
                                if abs(val) < 1e-6:
                                    # Value too small, use scientific notation
                                    new_text = fmt_str.format(val)
                                    text_obj.set_text(new_text)
                            except:
                                pass
                    except Exception as e_text:
                        pass
                
                # Re-add SHAP value labels (more reliable method)
                # Get all feature positions and SHAP values
                # Note: The layout of SHAP waterfall may be complex, we try another method here
                # Use SHAP's force plot instead, or custom drawing
                
            except Exception as e_inner:
                logging.warning(f"  Sample {sample_idx+1} using standard waterfall failed: {e_inner}")
                plt.close(fig)
                
                # Use custom method to draw waterfall plot
                try:
                    fig, ax = plt.subplots(figsize=(12, 8))
                    
                    # Sort by SHAP value absolute value, take top 20
                    sorted_indices = np.argsort(np.abs(sample_shap))[::-1][:20]
                    sorted_shap = sample_shap[sorted_indices]
                    sorted_features = [feature_names[i] for i in sorted_indices]
                    sorted_data = X_explain_sample[sample_idx, sorted_indices]
                    
                    # Calculate cumulative value
                    cumulative = explainer.expected_value
                    y_positions = np.arange(len(sorted_indices))
                    
                    # Draw cumulative bar chart
                    for i, (shap_val, feat_name, feat_val) in enumerate(zip(sorted_shap, sorted_features, sorted_data)):
                        color = 'red' if shap_val > 0 else 'blue'
                        width = shap_val
                        left = cumulative
                        right = cumulative + width
                        
                        # Draw bar
                        ax.barh(i, width, left=left, color=color, alpha=0.7)
                        
                        # Add label (using scientific notation or high precision)
                        if abs(shap_val) > 1e-6:
                            label_x = (left + right) / 2
                            label_text = fmt_str.format(shap_val)
                            ax.text(label_x, i, label_text, ha='center', va='center', 
                                   fontsize=9, fontweight='bold', color='white')
                        
                        cumulative = right
                    
                    # Add baseline
                    ax.axvline(x=explainer.expected_value, color='black', linestyle='--', linewidth=2, label='基准值')
                    
                    # Add final prediction value
                    final_pred = explainer.expected_value + np.sum(sorted_shap)
                    ax.axvline(x=final_pred, color='green', linestyle='--', linewidth=2, label='最终预测')
                    
                    ax.set_yticks(y_positions)
                    ax.set_yticklabels([f'{feat_name}={feat_val:.3f}' for feat_name, feat_val in zip(sorted_features, sorted_data)])
                    ax.set_xlabel('SHAP value', fontsize=12)
                    ax.set_ylabel('Feature', fontsize=12)
                    ax.set_title(f'SHAP Waterfall Plot (Sample {sample_idx+1})', fontsize=14, fontweight='bold')
                    ax.legend()
                    ax.grid(True, alpha=0.3)
                    
                except Exception as e_custom:
                    logging.warning(f"  Custom waterfall also failed: {e_custom}")
                    plt.close(fig)
                    continue
            
            plt.tight_layout()
            
            if n_samples_waterfall == 1:
                plt.savefig(save_dir / 'shap_waterfall.png', dpi=300, bbox_inches='tight')
            else:
                plt.savefig(save_dir / f'shap_waterfall_sample_{sample_idx+1}.png', 
                           dpi=300, bbox_inches='tight')
            plt.close()
            
            # Print SHAP value summary for this sample (for debugging)
            top_positive = np.argsort(sample_shap)[-5:][::-1]
            top_negative = np.argsort(sample_shap)[:5]
            logging.info(f"  Sample {sample_idx+1} SHAP value summary:")
            logging.info(f"    Maximum positive contribution: {feature_names[top_positive[0]]} = {sample_shap[top_positive[0]]:.6e}")
            logging.info(f"    Maximum negative contribution: {feature_names[top_negative[0]]} = {sample_shap[top_negative[0]]:.6e}")
        
        logging.info(f"✓ Generated {n_samples_waterfall} waterfall plots")
    except Exception as e:
        logging.warning(f"Waterfall plot generation failed: {e}")
        import traceback
        traceback.print_exc()
    
    logging.info("="*80 + "\n")


# ---------------------------------------------------------------------------
# PDP (Partial Dependence Plot) analysis
# ---------------------------------------------------------------------------

def plot_pdp_analysis(
    model: nn.Module,
    X_background: np.ndarray,
    feature_names: list[str],
    target_scaler: MinMaxScaler,
    device: str,
    save_dir: Path,
    n_top_features: int = 6,
    shap_importance_values: Optional[np.ndarray] = None,
) -> None:
    """PDP (Partial Dependence Plot) analysis (suitable for neural network models).
    
    Args:
        model: Trained PINN model
        X_background: Background dataset (for PDP calculation)
        feature_names: Feature name list
        target_scaler: Target value scaler
        device: Running device
        save_dir: Save directory
        n_top_features: Number of top important features to analyze
        shap_importance_values: SHAP importance values (if calculated, used to select important features)
    """
    logging.info("\n" + "="*80)
    logging.info("Performing PDP (Partial Dependence Plot) analysis")
    logging.info("="*80)
    
    # Set matplotlib font (ensure Chinese and negative sign are displayed correctly)
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    
    # Ensure X_background is DataFrame format (for sklearn's partial_dependence)
    if isinstance(X_background, np.ndarray):
        X_background_df = pd.DataFrame(X_background, columns=feature_names)
    else:
        X_background_df = X_background
    
    # Get most important features (using SHAP importance or random selection)
    if shap_importance_values is not None and len(shap_importance_values) == len(feature_names):
        importance_df = pd.DataFrame({
            'feature': feature_names,
            'importance': shap_importance_values
        }).sort_values('importance', ascending=False)
    else:
        # If no SHAP values, use random order or all features
        logging.warning("No SHAP importance values provided, will analyze top N features")
        importance_df = pd.DataFrame({
            'feature': feature_names,
            'importance': np.ones(len(feature_names))
        })
    
    top_features = importance_df.head(n_top_features)['feature'].tolist()
    top_indices = [feature_names.index(f) for f in top_features]
    
    logging.info(f"Will analyze top {n_top_features} important features: {top_features}")
    
    # Create prediction wrapper function (for sklearn's partial_dependence)
    def predict_fn(X: np.ndarray) -> np.ndarray:
        """sklearn compatible prediction function."""
        model.eval()
        with torch.no_grad():
            X_tensor = torch.from_numpy(X.astype(np.float32)).to(device)
            preds_scaled = model(X_tensor).cpu().numpy()
            preds = target_scaler.inverse_transform(preds_scaled).flatten()
        return preds
    
    # Create sklearn compatible model wrapper
    # sklearn's partial_dependence needs estimator to implement fit and predict methods
    # Use sklearn.base.BaseEstimator as base class, ensure compatibility
    from sklearn.base import BaseEstimator, RegressorMixin
    from sklearn.utils.validation import check_is_fitted
    
    class ModelWrapper(BaseEstimator, RegressorMixin):
        def __init__(self, predict_fn):
            super().__init__()
            self.predict_fn = predict_fn
            # Mark as un-fitted during initialization, need to explicitly call fit
        
        def fit(self, X, y=None):
            """sklearn required fit method (mark as fitted for trained models)"""
            # Use sklearn's mechanism to mark as fitted
            self._is_fitted = True
            return self
        
        def predict(self, X):
            """Prediction method"""
            return self.predict_fn(X)
        
        def __sklearn_is_fitted__(self):
            """sklearn check if fitted method (priority call)"""
            return hasattr(self, '_is_fitted') and self._is_fitted
    
    wrapped_model = ModelWrapper(predict_fn)
    # Explicitly call fit method, let sklearn recognize as fitted model
    wrapped_model.fit(X_background_df)
    
    # Verify model passed sklearn's fit check
    try:
        check_is_fitted(wrapped_model)
        logging.info("✓ ModelWrapper passed sklearn fit check")
    except Exception as e:
        logging.warning(f"⚠ ModelWrapper fit check failed: {e}")
        # If standard check fails, try to force set all necessary attributes
        wrapped_model._is_fitted = True
        # Ensure __sklearn_is_fitted__ method exists and returns True
        if not hasattr(wrapped_model, '__sklearn_is_fitted__') or not wrapped_model.__sklearn_is_fitted__():
            # Redefine method to ensure returns True
            wrapped_model.__sklearn_is_fitted__ = lambda: True
        # Verify again
        try:
            check_is_fitted(wrapped_model)
            logging.info("✓ ModelWrapper passed sklearn fit check after repair")
        except Exception as e2:
            logging.error(f"✗ Repair still failed: {e2}")
            # Last attempt: directly modify partial_dependence call method
            # Use estimator=None, then manually implement PDP calculation
    
    # Dictionary for saving Excel data
    excel_data_dict = {}
    
    # 1. Single variable PDP analysis
    logging.info("Generating single variable PDP plot...")
    n_cols = 3
    n_rows = (n_top_features + 2) // 3
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(18, 6*n_rows))
    if n_top_features == 1:
        axes = [axes]
    else:
        axes = axes.flatten()
    
    for idx, (feature, ax) in enumerate(zip(top_features, axes)):
        try:
            feature_idx = feature_names.index(feature)
            
            # Manually calculate PDP (bypass sklearn's check mechanism)
            # Get feature value range
            feature_values = X_background_df.iloc[:, feature_idx].values
            feature_min, feature_max = feature_values.min(), feature_values.max()
            
            # Create grid points
            grid_resolution = 50
            grid_values = np.linspace(feature_min, feature_max, grid_resolution)
            
            # Calculate average prediction value for each grid point
            pdp_values = []
            for grid_val in grid_values:
                # Copy background data
                X_temp = X_background_df.copy()
                # Set current feature to grid value
                X_temp.iloc[:, feature_idx] = grid_val
                # Predict
                preds = wrapped_model.predict(X_temp.values)
                # Calculate average prediction value
                pdp_values.append(np.mean(preds))
            
            pdp_values = np.array(pdp_values)
            
            pd_result = {
                'grid_values': [grid_values],
                'average': [pdp_values]
            }
            
            grid_values = pd_result['grid_values'][0]
            pdp_values = pd_result['average'][0]
            
            # Plot PDP curve
            ax.plot(grid_values, pdp_values, linewidth=2.5, color='blue', label='PDP')
            ax.set_xlabel(feature, fontsize=10)
            ax.set_ylabel('Partial dependence value', fontsize=10)
            ax.set_title(f'PDP for {feature}', fontsize=12, fontweight='bold')
            ax.legend(loc='best', fontsize=9)
            ax.grid(True, alpha=0.3)
            
            # Save data
            excel_data_dict[feature] = pd.DataFrame({
                'Grid_Values': grid_values,
                'PDP_Values': pdp_values
            })
            
            logging.info(f"  ✓ {feature}: PDP calculation completed")
            
        except Exception as e:
            logging.warning(f"  ✗ {feature}: PDP calculation failed: {e}")
            ax.text(0.5, 0.5, f'PDP calculation failed', 
                   ha='center', va='center', transform=ax.transAxes)
            ax.axis('off')
    
    # Hide extra subplots
    for idx in range(n_top_features, len(axes)):
        axes[idx].axis('off')
    
    plt.tight_layout()
    plt.savefig(save_dir / 'pdp_analysis.png', dpi=300, bbox_inches='tight')
    plt.close()
    logging.info("✓ Single variable PDP plot saved")
    
    # 2. Dual variable PDP analysis (feature interaction)
    if n_top_features >= 2:
        logging.info("Generating dual variable PDP plot (feature interaction)...")
        
        # Select top 4 most important features for dual variable interaction analysis
        n_interaction_features = min(4, n_top_features)
        interaction_features = top_features[:n_interaction_features]
        
        # Create interaction feature pairs
        interaction_pairs = []
        for i in range(len(interaction_features)):
            for j in range(i+1, len(interaction_features)):
                interaction_pairs.append((interaction_features[i], interaction_features[j]))
        
        if len(interaction_pairs) > 0:
            from mpl_toolkits.mplot3d import Axes3D
            
            n_pairs = len(interaction_pairs)
            n_cols = 3
            n_rows = (n_pairs + 2) // 3
            fig = plt.figure(figsize=(18, 6*n_rows))
            
            excel_2d_data_dict = {}
            
            for idx, (feat1, feat2) in enumerate(interaction_pairs):
                try:
                    feat1_idx = feature_names.index(feat1)
                    feat2_idx = feature_names.index(feat2)
                    
                    # Manually calculate dual variable PDP (bypass sklearn's check mechanism)
                    # Get range of two feature values
                    feat1_values = X_background_df.iloc[:, feat1_idx].values
                    feat2_values = X_background_df.iloc[:, feat2_idx].values
                    feat1_min, feat1_max = feat1_values.min(), feat1_values.max()
                    feat2_min, feat2_max = feat2_values.min(), feat2_values.max()
                    
                    # Create grid points
                    grid_resolution = 30
                    grid_values_1 = np.linspace(feat1_min, feat1_max, grid_resolution)
                    grid_values_2 = np.linspace(feat2_min, feat2_max, grid_resolution)
                    
                    # Calculate average prediction value for each grid point
                    average = np.zeros((len(grid_values_2), len(grid_values_1)))
                    for i, val1 in enumerate(grid_values_1):
                        for j, val2 in enumerate(grid_values_2):
                            # Copy background data
                            X_temp = X_background_df.copy()
                            # Set two features to grid values
                            X_temp.iloc[:, feat1_idx] = val1
                            X_temp.iloc[:, feat2_idx] = val2
                            # Predict
                            preds = wrapped_model.predict(X_temp.values)
                            # Calculate average prediction value
                            average[j, i] = np.mean(preds)
                    
                    pd_result = {
                        'grid_values': [grid_values_1, grid_values_2],
                        'average': [average]
                    }
                    
                    grid_values_1 = pd_result['grid_values'][0]
                    grid_values_2 = pd_result['grid_values'][1]
                    average = pd_result['average'][0]
                    
                    # Ensure average is 2D
                    if average.ndim == 3:
                        average = average[0]
                    
                    # Create meshgrid
                    X1, X2 = np.meshgrid(grid_values_1, grid_values_2)
                    
                    # Ensure dimension matches
                    if average.shape != X1.shape:
                        if average.shape == (len(grid_values_1), len(grid_values_2)):
                            average = average.T
                    
                    # Save 2D PDP data
                    pair_name = f'{feat1}_vs_{feat2}'
                    excel_2d_data_dict[pair_name] = pd.DataFrame({
                        f'{feat1}': X1.flatten(),
                        f'{feat2}': X2.flatten(),
                        'Partial_Dependence': average.flatten()
                    })
                    
                    # Create 3D subplot
                    ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                    
                    # Plot 3D surface
                    surf = ax.plot_surface(X1, X2, average, cmap='viridis', alpha=0.9, 
                                          linewidth=0, antialiased=True, edgecolor='none')
                    
                    ax.set_xlabel(feat1, fontsize=10, labelpad=8)
                    ax.set_ylabel(feat2, fontsize=10, labelpad=8)
                    ax.set_zlabel('Partial dependence value', fontsize=10, labelpad=8)
                    ax.set_title(f'Interaction: {feat1} vs {feat2}', fontsize=12, fontweight='bold', pad=15)
                    
                    fig.colorbar(surf, ax=ax, shrink=0.6, aspect=20, pad=0.1)
                    
                    logging.info(f"  ✓ {pair_name}: Dual variable PDP calculation completed")
                    
                except Exception as e:
                    logging.warning(f"  ✗ {feat1} vs {feat2}: Dual variable PDP calculation failed: {e}")
                    ax = fig.add_subplot(n_rows, n_cols, idx+1, projection='3d')
                    ax.text(0.5, 0.5, 0.5, f'Calculation failed', 
                           ha='center', va='center', transform=ax.transAxes)
                    ax.set_axis_off()
            
            plt.tight_layout()
            plt.savefig(save_dir / 'pdp_2d_interaction.png', dpi=300, bbox_inches='tight')
            plt.close()
            logging.info("✓ Dual variable PDP plot saved")
            
            # Save 2D PDP data to Excel (will be saved in main function)
            excel_data_dict['_2d_interactions'] = excel_2d_data_dict
    
    logging.info("="*80 + "\n")
    
    return excel_data_dict


# ---------------------------------------------------------------------------
# Noise analysis and uncertainty analysis
# ---------------------------------------------------------------------------

def analyze_noise_robustness(
    model: nn.Module,
    X_test: np.ndarray,
    y_test: np.ndarray,
    target_scaler: MinMaxScaler,
    device: str,
    save_dir: Path,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Analyze model robustness to noise and outliers (similar to XGBoost noise analysis.py).
    
    Args:
        model: 训练好的PINN模型
        X_test: 测试集特征
        y_test: 测试集真实值
        target_scaler: 目标值标准化器
        device: 运行设备
        save_dir: 保存目录
    
    Returns:
        (robustness_df, outlier_df): 噪声鲁棒性和异常值鲁棒性结果DataFrame
    """
    logging.info("\n" + "="*80)
    logging.info("Performing noise robustness analysis")
    logging.info("="*80)
    
    # Set matplotlib font (ensure Chinese and negative sign display correctly)
    plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans', 'Arial Unicode MS']
    plt.rcParams['axes.unicode_minus'] = False
    
    model.eval()
    
    # 1. Add Gaussian noise test
    logging.info("Performing robustness analysis to Gaussian noise...")
    noise_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Noise level (percentage of feature standard deviation)
    noise_rmses = []
    noise_stds = []
    
    for noise_level in noise_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times to get standard deviation
            # Add Gaussian noise
            X_test_noisy = X_test.copy()
            for i in range(X_test.shape[1]):
                feature_std = np.std(X_test[:, i])
                noise = np.random.normal(0, noise_level * feature_std, size=len(X_test))
                X_test_noisy[:, i] = X_test[:, i] + noise
            
            # Predict
            with torch.no_grad():
                X_tensor = torch.from_numpy(X_test_noisy.astype(np.float32)).to(device)
                preds_scaled = model(X_tensor).cpu().numpy()
                preds = target_scaler.inverse_transform(preds_scaled).flatten()
            
            rmse = np.sqrt(np.mean((y_test - preds) ** 2))
            rmses.append(rmse)
        
        noise_rmses.append(np.mean(rmses))
        noise_stds.append(np.std(rmses))
        logging.info(f"  Noise level {noise_level*100:.0f}%: RMSE = {np.mean(rmses):.6f} ± {np.std(rmses):.6f}")
    
    # 2. Add outlier test
    logging.info("Performing robustness analysis to outliers...")
    outlier_levels = [0, 0.02, 0.04, 0.06, 0.08, 0.10]  # Outlier比例
    outlier_severity = 5.0  # Outlier severity (multiple of standard deviation)
    outlier_rmses = []
    outlier_stds = []
    
    for outlier_pct in outlier_levels:
        rmses = []
        for _ in range(10):  # Repeat 10 times
            # Add outliers
            X_test_outlier = X_test.copy()
            n_outliers = int(len(X_test) * outlier_pct)
            if n_outliers > 0:
                outlier_indices = np.random.choice(len(X_test), n_outliers, replace=False)
                for idx in outlier_indices:
                    # Randomly select a feature and add outlier
                    feature_idx = np.random.randint(X_test.shape[1])
                    feature_mean = np.mean(X_test[:, feature_idx])
                    feature_std = np.std(X_test[:, feature_idx])
                    X_test_outlier[idx, feature_idx] = feature_mean + outlier_severity * feature_std
            
            # Predict
            with torch.no_grad():
                X_tensor = torch.from_numpy(X_test_outlier.astype(np.float32)).to(device)
                preds_scaled = model(X_tensor).cpu().numpy()
                preds = target_scaler.inverse_transform(preds_scaled).flatten()
            
            rmse = np.sqrt(np.mean((y_test - preds) ** 2))
            rmses.append(rmse)
        
        outlier_rmses.append(np.mean(rmses))
        outlier_stds.append(np.std(rmses))
        logging.info(f"  Outlier percentage {outlier_pct*100:.0f}%: RMSE = {np.mean(rmses):.6f} ± {np.std(rmses):.6f}")
    
    # 3. Plot robustness analysis
    try:
        fig, axes = plt.subplots(1, 2, figsize=(15, 6))
        
        # Gaussian noise robustness
        axes[0].errorbar([l * 100 for l in noise_levels], noise_rmses, yerr=noise_stds, 
                         marker='o', capsize=5, capthick=2, linewidth=2, color='blue')
        axes[0].set_xlabel('Noise level (% feature standard deviation)', fontsize=12)
        axes[0].set_ylabel('RMSE', fontsize=12)
        axes[0].set_title('Robustness to additive Gaussian noise', fontsize=14, fontweight='bold')
        axes[0].grid(True, alpha=0.3)
        
        # Outlier robustness
        axes[1].errorbar([l * 100 for l in outlier_levels], outlier_rmses, yerr=outlier_stds, 
                         marker='o', capsize=5, capthick=2, linewidth=2, color='orange')
        axes[1].set_xlabel('Outlier injection percentage (% of sample number)', fontsize=12)
        axes[1].set_ylabel('RMSE', fontsize=12)
        axes[1].set_title(f'Robustness to outliers (severity={outlier_severity}×standard deviation)', fontsize=14, fontweight='bold')
        axes[1].grid(True, alpha=0.3)
        
        plt.tight_layout()
        plt.savefig(save_dir / 'robustness_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        logging.info("✓ Robustness analysis plot saved")
    except Exception as e:
        logging.warning(f"Robustness analysis plot generation failed: {e}")
    
    # Save results
    robustness_df = pd.DataFrame({
        'Noise level (%)': [l * 100 for l in noise_levels],
        'RMSE': noise_rmses,
        'Standard deviation': noise_stds
    })
    
    outlier_df = pd.DataFrame({
        'Outlier percentage (%)': [l * 100 for l in outlier_levels],
        'RMSE': outlier_rmses,
        'Standard deviation': outlier_stds
    })
    
    # Calculate performance degradation rate
    noise_degradation = (noise_rmses[-1] - noise_rmses[0]) / noise_rmses[0] * 100 if noise_rmses[0] > 0 else 0
    outlier_degradation = (outlier_rmses[-1] - outlier_rmses[0]) / outlier_rmses[0] * 100 if outlier_rmses[0] > 0 else 0
    
    logging.info(f"\nNoise robustness analysis results:")
    logging.info(f"  - 0% noise when RMSE: {noise_rmses[0]:.6f}")
    logging.info(f"  - 10% noise when RMSE: {noise_rmses[-1]:.6f} (performance degradation {noise_degradation:.1f}%)")
    logging.info(f"  - 0% outliers when RMSE: {outlier_rmses[0]:.6f}")
    logging.info(f"  - 10% outliers when RMSE: {outlier_rmses[-1]:.6f} (performance change {outlier_degradation:.1f}%)")
    
    logging.info("="*80 + "\n")
    
    return robustness_df, outlier_df


def analyze_prediction_uncertainty(
    model: nn.Module,
    X_test: np.ndarray,
    y_test: np.ndarray,
    target_scaler: MinMaxScaler,
    device: str,
    save_dir: Path,
    n_bootstrap: int = 100,
    bootstrap_noise_level: float = 0.01,
) -> pd.DataFrame:
    """Analyze prediction intervals and uncertainty (similar to XGBoost noise analysis.py format).
    
    Args:
        model: Trained PINN model
        X_test: Test set features
        y_test: Test set true values
        target_scaler: Target value scaler
        device: Running device
        save_dir: Save directory
        n_bootstrap: Bootstrap resampling次数
        bootstrap_noise_level: Bootstrap noise level (percentage of feature standard deviation)
    
    Returns:
        intervals_df: DataFrame containing prediction interval results
    """
    logging.info("\n" + "="*80)
    logging.info("Performing prediction uncertainty analysis (Bootstrap and Quantile Regression methods)")
    logging.info("="*80)
    logging.info(f"Bootstrap resampling times: {n_bootstrap}")
    logging.info(f"Noise level: {bootstrap_noise_level*100}% of feature std")
    
    model.eval()
    
    # 1. Bootstrap prediction interval
    logging.info("Generating Bootstrap prediction interval...")
    bootstrap_predictions = []
    
    for i in range(n_bootstrap):
        if (i + 1) % 20 == 0:
            logging.info(f"  Progress: {i+1}/{n_bootstrap}")
        
        # Add noise to input features (similar to Monte Carlo Dropout idea)
        X_test_noisy = X_test.copy()
        for j in range(X_test.shape[1]):
            feature_std = np.std(X_test[:, j])
            noise = np.random.normal(0, bootstrap_noise_level * feature_std, size=len(X_test))
            X_test_noisy[:, j] = X_test[:, j] + noise
        
        # Predict
        with torch.no_grad():
            X_tensor = torch.from_numpy(X_test_noisy.astype(np.float32)).to(device)
            preds_scaled = model(X_tensor).cpu().numpy()
            preds = target_scaler.inverse_transform(preds_scaled).flatten()
            bootstrap_predictions.append(preds)
    
    bootstrap_predictions = np.array(bootstrap_predictions)  # shape: (n_bootstrap, n_samples)
    bootstrap_median = np.median(bootstrap_predictions, axis=0)
    bootstrap_lower = np.percentile(bootstrap_predictions, 10, axis=0)  # 80% confidence interval
    bootstrap_upper = np.percentile(bootstrap_predictions, 90, axis=0)
    
    # 2. Quantile Regression prediction interval (using residual distribution)
    logging.info("Generating Quantile Regression prediction interval...")
    with torch.no_grad():
        X_tensor = torch.from_numpy(X_test.astype(np.float32)).to(device)
        preds_scaled = model(X_tensor).cpu().numpy()
        y_pred_test = target_scaler.inverse_transform(preds_scaled).flatten()
    
    residuals = y_test - y_pred_test
    residual_std = np.std(residuals)
    
    # Use the standard deviation of residuals to build the prediction interval (80% confidence interval)
    quantile_lower = y_pred_test - 1.28 * residual_std  # 10th percentile for 80% interval
    quantile_upper = y_pred_test + 1.28 * residual_std  # 90th percentile for 80% interval
    
    # Sort for visualization (sort by true values)
    sort_indices = np.argsort(y_test)
    y_test_sorted = y_test[sort_indices]
    bootstrap_median_sorted = bootstrap_median[sort_indices]
    bootstrap_lower_sorted = bootstrap_lower[sort_indices]
    bootstrap_upper_sorted = bootstrap_upper[sort_indices]
    quantile_lower_sorted = quantile_lower[sort_indices]
    quantile_upper_sorted = quantile_upper[sort_indices]
    y_pred_test_sorted = y_pred_test[sort_indices]
    
    # Calculate coverage
    bootstrap_coverage = np.mean((y_test >= bootstrap_lower) & (y_test <= bootstrap_upper))
    quantile_coverage = np.mean((y_test >= quantile_lower) & (y_test <= quantile_upper))
    
    # Calculate average interval width
    bootstrap_width = np.mean(bootstrap_upper - bootstrap_lower)
    quantile_width = np.mean(quantile_upper - quantile_lower)
    
    logging.info(f"\nPrediction interval analysis results:")
    logging.info(f"  Bootstrap 80% confidence interval coverage: {bootstrap_coverage*100:.1f}%")
    logging.info(f"  Quantile Regression 80% confidence interval coverage: {quantile_coverage*100:.1f}%")
    logging.info(f"  Bootstrap average interval width: {bootstrap_width:.6f}")
    logging.info(f"  Quantile Regression average interval width: {quantile_width:.6f}")
    
    # 3. Plot prediction interval comparison (similar to XGBoost format)
    try:
        # Set matplotlib font (ensure Chinese and negative sign display correctly)
        plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans', 'Arial Unicode MS']
        plt.rcParams['axes.unicode_minus'] = False
        
        fig = plt.figure(figsize=(20, 10))
        
        # Create grid layout: left two columns display prediction intervals, right column display performance comparison
        from matplotlib.gridspec import GridSpec
        gs = GridSpec(2, 2, figure=fig, hspace=0.3, wspace=0.3, 
                     width_ratios=[2, 1], height_ratios=[1, 1])
        
        # Bootstrap prediction interval (top left)
        ax1 = fig.add_subplot(gs[0, 0])
        ax1.scatter(range(len(y_test_sorted)), y_test_sorted, s=30, c='black', alpha=0.7, 
                   label='True values', zorder=3)
        ax1.plot(range(len(y_test_sorted)), bootstrap_median_sorted, 'r--', linewidth=2, 
                label='Predicted median', zorder=2)
        ax1.fill_between(range(len(y_test_sorted)), bootstrap_lower_sorted, bootstrap_upper_sorted, 
                        alpha=0.3, color='lightblue', label='80% prediction interval', zorder=1)
        ax1.set_xlabel('Test samples (sorted by true values)', fontsize=11)
        ax1.set_ylabel('Peak strain', fontsize=11)
        ax1.set_title('Bootstrap prediction interval', fontsize=13, fontweight='bold')
        ax1.legend(loc='lower right', fontsize=9)
        ax1.grid(True, alpha=0.3)
        
        # Quantile Regression prediction interval (bottom left)
        ax2 = fig.add_subplot(gs[1, 0])
        ax2.scatter(range(len(y_test_sorted)), y_test_sorted, s=30, c='black', alpha=0.7, 
                   label='True values', zorder=3)
        ax2.plot(range(len(y_test_sorted)), y_pred_test_sorted, 'r--', linewidth=2, 
                label='Predicted values', zorder=2)
        ax2.fill_between(range(len(y_test_sorted)), quantile_lower_sorted, quantile_upper_sorted, 
                         alpha=0.3, color='lightgreen', label='80% prediction interval', zorder=1)
        ax2.set_xlabel('Test samples (sorted by true values)', fontsize=11)
        ax2.set_ylabel('Peak strain', fontsize=11)
        ax2.set_title('Quantile Regression prediction interval', fontsize=13, fontweight='bold')
        ax2.legend(loc='lower right', fontsize=9)
        ax2.grid(True, alpha=0.3)
        
        # Prediction interval coverage analysis (right)
        ax3 = fig.add_subplot(gs[:, 1])
        
        methods = ['Bootstrap', 'Quantile Regression']
        coverages = [bootstrap_coverage * 100, quantile_coverage * 100]
        widths = [bootstrap_width, quantile_width]
        
        y_pos = np.arange(len(methods))
        height = 0.35
        
        # Create dual axis
        bars1 = ax3.barh(y_pos - height/2, coverages, height, 
                        label='Coverage (%)', color='skyblue', alpha=0.8)
        ax3_twin = ax3.twinx()
        bars2 = ax3_twin.barh(y_pos + height/2, widths, height, 
                             label='Average interval width', color='coral', alpha=0.8)
        
        ax3.set_ylabel('Method', fontsize=11)
        ax3.set_xlabel('Coverage (%)', fontsize=11, color='blue')
        ax3_twin.set_xlabel('Average interval width', fontsize=11, color='red')
        ax3.set_yticks(y_pos)
        ax3.set_yticklabels(methods)
        ax3.tick_params(axis='x', labelcolor='blue')
        ax3_twin.tick_params(axis='x', labelcolor='red')
        ax3.grid(True, alpha=0.3, axis='x')
        ax3.set_title('Performance comparison', fontsize=13, fontweight='bold', pad=20)
        
        # Add numerical labels
        for i, v in enumerate(coverages):
            ax3.text(v + 2, i - height/2, f'{v:.1f}%', ha='left', va='center', fontsize=9)
        for i, v in enumerate(widths):
            ax3_twin.text(v + max(widths)*0.05, i + height/2, f'{v:.4f}', ha='left', va='center', fontsize=9)
        
        # Add legend
        ax3.legend(loc='upper left', fontsize=9)
        ax3_twin.legend(loc='lower left', fontsize=9)
        
        # Invert y axis to make Bootstrap at the top
        ax3.invert_yaxis()
        
        plt.tight_layout()
        plt.savefig(save_dir / 'uncertainty_analysis.png', dpi=300, bbox_inches='tight')
        plt.close()
        logging.info("✓ Uncertainty analysis plot saved")
    except Exception as e:
        logging.warning(f"Failed to generate uncertainty analysis plot: {e}")
        import traceback
        traceback.print_exc()
    
    # Save prediction interval results
    intervals_df = pd.DataFrame({
        'Sample ID': range(len(y_test)),
        'True values': y_test,
        'Predicted values': y_pred_test,
        'Bootstrap lower bound': bootstrap_lower,
        'Bootstrap upper bound': bootstrap_upper,
        'Bootstrap median': bootstrap_median,
        'Quantile lower bound': quantile_lower,
        'Quantile upper bound': quantile_upper,
    })
    
    logging.info("="*80 + "\n")
    
    return intervals_df


# ---------------------------------------------------------------------------
# Main function
# ---------------------------------------------------------------------------

def find_project_root():
    """Automatically find project root directory"""
    current_dir = os.path.abspath(os.getcwd())
    for _ in range(5):
        if os.path.exists(os.path.join(current_dir, 'dataset')) or \
           os.path.exists(os.path.join(current_dir, 'SAVE')):
            return current_dir
        current_dir = os.path.dirname(current_dir)
    return os.path.abspath(os.getcwd())


def main() -> None:
    """Main function - automatically load model and perform SHAP analysis, noise analysis, and uncertainty analysis (similar to XGBoost noise analysis.py)"""
    configure_logging()
    
    logging.info("\n" + "="*80)
    logging.info("PINN peak strain prediction model - SHAP analysis and noise analysis")
    logging.info("="*80 + "\n")
    
    # Automatically find project root directory
    PROJECT_ROOT = find_project_root()
    SAVE_ROOT = os.path.join(PROJECT_ROOT, 'peak_strain', 'SAVE')
    DATASET_PATH = os.path.join(PROJECT_ROOT, 'dataset', 'dataset_final.xlsx')
    
    logging.info(f"Project root directory: {PROJECT_ROOT}")
    logging.info(f"Model save directory: {SAVE_ROOT}")
    logging.info(f"Dataset path: {DATASET_PATH}")
    
    # 1. Load trained model
    model_path = os.path.join(SAVE_ROOT, 'pinn_peak_strain.pt')
    if not os.path.exists(model_path):
        raise FileNotFoundError(f"Model file not found: {model_path}\nPlease run training script: peak_strain_model_minmax.py")
    
    logging.info("\n[1/5] Load trained model...")
    model_data = load_trained_model(SAVE_ROOT, device="cpu")
    
    # 2. Load and preprocess data
    logging.info("\n[2/5] Load dataset...")
    data = load_and_preprocess_data(
        DATASET_PATH,
        model_data["feature_cols"],
        model_data["feature_scaler"],
        target_col=model_data["base_config"]["target_column"],
        stress_col=model_data["base_config"]["stress_column"],
    )
    
    # Data分割（用于SHAP分析和噪声分析）
    df = data["df"]
    X_all = data["features_scaled"]
    y_all = data["target"]
    
    X_train_val = X_all[train_val_mask] if train_val_mask.sum() > 0 else X_all
    y_train_val = y_all[train_val_mask] if train_val_mask.sum() > 0 else y_all
    X_test = X_all[test_mask] if test_mask.sum() > 0 else X_all
    y_test = y_all[test_mask] if test_mask.sum() > 0 else y_all
    
    logging.info(f"  Training+validation set: {len(X_train_val)} samples")
    logging.info(f"  Test set: {len(X_test)} samples")
    
    # 3. PINN model prediction and Xiao formula prediction
    logging.info("\n[3/5] Model prediction and performance evaluation...")
    
    # PINN prediction
    pinn_preds_all = predict_with_pinn(
        model_data["model"],
        X_all,
        model_data["target_scaler"],
        "cpu",
    )
    pinn_preds_test = pinn_preds_all[test_mask] if test_mask.sum() > 0 else pinn_preds_all
    pinn_metrics = compute_metrics(y_test, pinn_preds_test)
    
    # Xiao formula prediction
    xiao_preds_all = compute_xiao_formula(data["stress"], data["r_values"])
    xiao_preds_test = xiao_preds_all[test_mask] if test_mask.sum() > 0 else xiao_preds_all
    xiao_metrics = compute_metrics(y_test, xiao_preds_test)
    
    logging.info("✓ PINN test set R² = %.4f", pinn_metrics["r2"])
    logging.info("✓ Xiao formula test set R² = %.4f", xiao_metrics["r2"])
    print_comparison_table(pinn_metrics, xiao_metrics)
    
    # 4. SHAP interpretability analysis and PDP analysis
    logging.info("\n[4/6] SHAP interpretability analysis and PDP analysis...")
    shap_importance_values = None
    pdp_data_dict = None
    
    if SHAP_AVAILABLE:
        try:
            # Use all data for SHAP analysis (X_all as background and explanation data)
            plot_shap_analysis(
                model_data["model"],
                X_all,  # Background dataset: all data
                X_all,  # Explanation dataset: all data
                model_data["feature_cols"],
                model_data["target_scaler"],
                "cpu",
                Path(SAVE_ROOT),
                max_background_samples=min(100, len(X_all)),  # Background data sampling
            )
            logging.info("✓ SHAP analysis completed (based on all data)")
            
            # Calculate SHAP importance values (for PDP analysis)
            # Note: Here we need to recalculate SHAP values to get importance, or read from saved results
            # To simplify, we use a simplified method: based on the results of SHAP analysis
            # In actual application, importance values can be returned from the SHAP analysis function
            try:
                # Use training+validation set to calculate SHAP importance (faster)
                predict_fn = predict_wrapper(model_data["model"], model_data["target_scaler"], "cpu")
                explainer = shap.KernelExplainer(predict_fn, X_train_val[:min(50, len(X_train_val))])
                shap_values_sample = explainer.shap_values(X_train_val[:min(20, len(X_train_val))], nsamples=50)
                shap_importance_values = np.abs(shap_values_sample).mean(axis=0)
                logging.info("✓ SHAP importance values calculated (for PDP analysis)")
            except Exception as e:
                logging.warning(f"⚠ Failed to calculate SHAP importance values, PDP analysis will use default feature order: {e}")
        except Exception as e:
            logging.warning(f"⚠ SHAP analysis failed: {e}")
            import traceback
            traceback.print_exc()
    else:
        logging.warning("⚠ SHAP not installed, skipping SHAP analysis")
    
    # PDP analysis (partial dependence plot)
    try:
        pdp_data_dict = plot_pdp_analysis(
            model_data["model"],
            X_train_val,  # Use training+validation set as background data
            model_data["feature_cols"],
            model_data["target_scaler"],
            "cpu",
            Path(SAVE_ROOT),
            n_top_features=6,
            shap_importance_values=shap_importance_values,
        )
        logging.info("✓ PDP analysis completed")
    except Exception as e:
        logging.warning(f"⚠ PDP analysis failed: {e}")
        import traceback
        traceback.print_exc()
    
    # 5. Noise analysis and uncertainty analysis
    logging.info("\n[5/6] Noise analysis and uncertainty analysis...")
    robustness_df = None
    outlier_df = None
    uncertainty_df = None
    
    # 5.1 Noise robustness analysis
    try:
        robustness_df, outlier_df = analyze_noise_robustness(
            model_data["model"],
            X_test,
            y_test,
            model_data["target_scaler"],
            "cpu",
            Path(SAVE_ROOT),
        )
        logging.info("✓ Noise robustness analysis completed")
    except Exception as e:
        logging.warning(f"⚠ Noise robustness analysis failed: {e}")
        import traceback
        traceback.print_exc()
    
    # 5.2 Uncertainty analysis (Bootstrap method)
    try:
        uncertainty_df = analyze_prediction_uncertainty(
            model_data["model"],
            X_test,
            y_test,
            model_data["target_scaler"],
            "cpu",
            Path(SAVE_ROOT),
            n_bootstrap=100,
            bootstrap_noise_level=0.01,
        )
        logging.info("✓ Uncertainty analysis completed")
    except Exception as e:
        logging.warning(f"⚠ Uncertainty analysis failed: {e}")
        import traceback
        traceback.print_exc()
    
    # 6. Save all results to an Excel file
    logging.info("\n" + "="*80)
    logging.info("Save all analysis results to an Excel file")
    logging.info("="*80)
    
    excel_path = os.path.join(SAVE_ROOT, 'pinn_analysis_results.xlsx')
    
    try:
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
            # 1. PINN vs Xiao detailed comparison
            comparison_df = prepare_comparison_data(
                data["df"],
                pinn_preds_all,
                xiao_preds_all,
                model_data["base_config"]["target_column"],
            )
            comparison_df.to_excel(writer, sheet_name='PINN_vs_Xiao comparison', index=False)
            logging.info("✓ Saved: PINN_vs_Xiao comparison")
            
            # 2. Prediction results (full dataset + prediction column)
            predictions_df = prepare_predictions_data(
                data["df"],
                pinn_preds_all,
                xiao_preds_all,
                model_data["base_config"]["target_column"],
            )
            predictions_df.to_excel(writer, sheet_name='Prediction results', index=False)
            logging.info("✓ Saved: Prediction results")
            
            # 3. Performance metrics
            metrics_df = prepare_metrics_data(pinn_metrics, xiao_metrics)
            if len(metrics_df) > 0:
                metrics_df.to_excel(writer, sheet_name='Performance metrics', index=False)
                logging.info("✓ Saved: Performance metrics")
            
            # 4. Noise robustness analysis results
            if robustness_df is not None:
                robustness_df.to_excel(writer, sheet_name='Noise robustness analysis', index=False)
                logging.info("✓ Saved: Noise robustness analysis")
            
            if outlier_df is not None:
                outlier_df.to_excel(writer, sheet_name='Outlier robustness analysis', index=False)
                logging.info("✓ Saved: Outlier robustness analysis")
            
            # 5. Uncertainty analysis results
            if uncertainty_df is not None:
                uncertainty_df.to_excel(writer, sheet_name='Uncertainty analysis', index=False)
                logging.info("✓ Saved: Uncertainty analysis")
            
            # 6. PDP analysis results (single variable)
            if pdp_data_dict is not None and '_2d_interactions' not in pdp_data_dict:
                for feature_name, data_df in pdp_data_dict.items():
                    if feature_name != '_2d_interactions':
                        sheet_name = f'PDP_{feature_name}'
                        if len(sheet_name) > 31:
                            sheet_name = f'PDP_{feature_name[:27]}'
                        # Replace Excel unsupported characters
                        sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                        try:
                            data_df.to_excel(writer, sheet_name=sheet_name, index=False)
                            logging.info(f"✓ Saved: {sheet_name}")
                        except Exception as e:
                            logging.warning(f"⚠ Failed to save PDP data {feature_name}: {e}")
            
            # 7. PDP analysis results (two variable interactions)
            if pdp_data_dict is not None and '_2d_interactions' in pdp_data_dict:
                excel_2d_data_dict = pdp_data_dict['_2d_interactions']
                for pair_name, data_df in excel_2d_data_dict.items():
                    sheet_name = f'PDP2D_{pair_name}'
                    if len(sheet_name) > 31:
                        sheet_name = f'PDP2D_{pair_name[:25]}'
                    # Replace Excel unsupported characters
                    sheet_name = sheet_name.replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_').replace(':', '_')
                    try:
                        data_df.to_excel(writer, sheet_name=sheet_name, index=False)
                        logging.info(f"✓ Saved: {sheet_name}")
                    except Exception as e:
                        logging.warning(f"⚠ Failed to save two variable PDP data {pair_name}: {e}")
        
        logging.info(f"\n✓ All results saved to: {excel_path}")
        
        # Count number of worksheets
        try:
            import openpyxl
            wb = openpyxl.load_workbook(excel_path)
            sheet_count = len(wb.sheetnames)
            logging.info(f"  Contains {sheet_count} worksheets: {', '.join(wb.sheetnames)}")
        except:
            logging.info(f"  Contains multiple worksheets")
    
    except Exception as e:
        logging.error(f"⚠ Failed to save Excel: {e}")
        import traceback
        traceback.print_exc()
    
    # 7. Predict all data and write to dataset_with_PINN_peak_strain.xlsx
    logging.info("\n" + "="*80)
    logging.info("Predict all data and write to dataset_with_PINN_peak_strain.xlsx")
    logging.info("="*80)
    
    try:
        # Create prediction result dictionary
        sample_ids = df.get("No_Customized", range(len(df))).values
        pred_dict = {}
        for idx, sample_id in enumerate(sample_ids):
            pred_dict[sample_id] = float(pinn_preds_all[idx])
        
        # Read original Excel file
        df_original = pd.read_excel(DATASET_PATH, sheet_name=0)
        
        # Use map method to match
        if 'No_Customized' in df_original.columns:
            df_original['PINN_peak_strain'] = df_original['No_Customized'].map(pred_dict)
        else:
            if len(pinn_preds_all) == len(df_original):
                df_original['PINN_peak_strain'] = pinn_preds_all
            else:
                raise ValueError(f"Prediction sample number ({len(pinn_preds_all)}) does not match the number of rows in the original data ({len(df_original)})")
        
        # Save results
        output_file = os.path.join(PROJECT_ROOT, "dataset", "dataset_with_PINN_peak_strain.xlsx")
        os.makedirs(os.path.dirname(output_file), exist_ok=True)
        df_original.to_excel(output_file, index=False)
        logging.info(f"✓ Prediction results written to: {output_file}")
        logging.info(f"  - New column: PINN_peak_strain (PINN model prediction value)")
        logging.info(f"  - Total prediction sample number: {len(pinn_preds_all)}")
        logging.info(f"  - Number of non-empty prediction values: {df_original['PINN_peak_strain'].notna().sum()}")
        
        # Backup to save directory
        backup_file = os.path.join(SAVE_ROOT, 'dataset_with_PINN_peak_strain_backup.xlsx')
        df_original.to_excel(backup_file, index=False)
        logging.info(f"✓ Backup file saved to: {backup_file}")
        
    except Exception as e:
        logging.error(f"⚠ Warning: Failed to write to Excel: {e}")
        import traceback
        traceback.print_exc()
    
    logging.info("\n" + "="*80)
    logging.info("✓ All analysis completed! Results saved to: %s", SAVE_ROOT)
    logging.info("="*80 + "\n")


if __name__ == "__main__":
    # Simplified version: run directly, no command line parameters (similar to XGBoost noise analysis.py)
    # If you need to customize parameters, you can modify the paths in the main function
    main()

